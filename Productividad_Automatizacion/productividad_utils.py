"""
Helpers del modulo Productividad.

El foco actual es dejar la base del proyecto lista y centralizar:
- reglas de rango
- paths historicos
- validacion del Excel descargado
- logging

La navegacion WMS sigue pendiente de confirmacion runtime.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
import os
from pathlib import Path
import sys
import hashlib
import json
import re
import unicodedata
from typing import Any, Dict, Iterable, List, Optional, Sequence

from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl import Workbook
import requests

from productividad_config import (
    CLIENTS,
    DOWNLOAD_DIR,
    EXPECTED_HEADERS,
    OFFICIAL_DESTINATION_MODE,
    LOCAL_HISTORICAL_REFERENCE_ROOT,
    LOG_DIR,
    MONTH_FOLDERS,
    NORMALIZED_DIR,
    QUARANTINE_DIR,
    RANGE_END_TIME,
    RANGE_START_TIME,
    SHAREPOINT_BACKUP_ROOT,
    SHAREPOINT_DOCUMENT_LIBRARY,
    SHAREPOINT_PRODUCTIVIDAD_ROOT,
    SHAREPOINT_SITE_NAME,
    SHAREPOINT_VERIFY_DIR,
    VALID_SHEET_NAMES,
)


@dataclass
class ReportingWindow:
    mode: str
    from_dt: datetime
    to_dt: datetime
    target_year: int
    target_month: int


@dataclass
class WorkbookInspection:
    path: Path
    sheet_name: str
    internal_cd: str
    internal_company: str
    internal_scope: str
    headers: List[str]
    has_data_rows: bool
    first_non_empty_after_header: Optional[str] = None
    source_format: str = "xlsx"
    max_non_empty_col: int = 0


@dataclass
class ValidationResult:
    ok: bool
    is_empty_valid: bool
    target_path: str
    target_mode: str = OFFICIAL_DESTINATION_MODE
    critical_errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    inspection: Optional[WorkbookInspection] = None


@dataclass
class HistoricalComparison:
    historical_headers: List[str]
    runtime_headers: List[str]
    headers_match_exact: bool
    headers_match_normalized: bool
    normalized_viable: bool
    normalization_rule: str
    notes: List[str] = field(default_factory=list)


@dataclass
class StructuralValidation:
    stage: str
    ok: bool
    expected_header_count: int
    observed_header_count: int
    expected_header_row: int
    observed_header_row: int
    expected_detail_start_row: int
    observed_detail_start_row: int
    portada_ok: bool
    exact_header_match: bool
    repeated_header_rows: List[int] = field(default_factory=list)
    short_rows: List[int] = field(default_factory=list)
    long_rows: List[int] = field(default_factory=list)
    suspicious_rows: List[int] = field(default_factory=list)
    issues: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class ChunkRange:
    chunk_id: str
    start_dt: datetime
    end_dt: datetime
    chunk_days: int
    level: int = 0


@dataclass
class ChunkCoverageAudit:
    ok: bool
    issues: List[str] = field(default_factory=list)


@dataclass
class ChunkConsolidationResult:
    ok: bool
    consolidated_path: Optional[Path]
    total_chunks: int
    chunks_with_rows: int
    empty_chunks: int
    total_rows_seen: int
    unique_rows: int
    duplicate_rows_removed: int
    duplicate_samples: List[str] = field(default_factory=list)
    issues: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class SharePointPublishPlan:
    client_key: str
    local_candidate: Path
    sharepoint_target_path: str
    sharepoint_folder_path: str
    sharepoint_filename: str
    remote_exists: Optional[bool] = None
    remote_backup_target_path: Optional[str] = None
    ready: bool = False
    issues: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class SharePointRemoteFileState:
    exists: bool
    item_id: str = ""
    name: str = ""
    size: int = 0
    last_modified: str = ""
    web_url: str = ""


@dataclass
class SharePointPostUploadVerification:
    ok: bool
    remote_state: SharePointRemoteFileState
    backup_state: Optional[SharePointRemoteFileState] = None
    local_size: int = 0
    remote_size: int = 0
    local_sha256: str = ""
    remote_sha256: str = ""
    local_semantic_sha256: str = ""
    remote_semantic_sha256: str = ""
    remote_verify_copy: Optional[Path] = None
    issues: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


def ensure_runtime_dirs() -> None:
    Path(LOG_DIR).mkdir(parents=True, exist_ok=True)
    Path(DOWNLOAD_DIR).mkdir(parents=True, exist_ok=True)
    Path(NORMALIZED_DIR).mkdir(parents=True, exist_ok=True)
    Path(QUARANTINE_DIR).mkdir(parents=True, exist_ok=True)
    Path(SHAREPOINT_VERIFY_DIR).mkdir(parents=True, exist_ok=True)


def build_log_path(prefix: str = "productividad") -> Path:
    ensure_runtime_dirs()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(LOG_DIR) / f"{prefix}_{stamp}.log"


def log(message: str, log_path: Optional[Path] = None) -> None:
    print(message, flush=True)
    if log_path:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as fh:
            fh.write(message + "\n")


def normalize_text(value: object) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip().upper()


def parse_clock(value: str) -> time:
    return datetime.strptime(value, "%H:%M:%S").time()


def month_folder_name(year: int, month: int) -> str:
    if month not in MONTH_FOLDERS:
        raise ValueError(f"Mes invalido: {month}")
    return MONTH_FOLDERS[month]


def build_reporting_window(
    *,
    mode: str,
    target_year: int,
    target_month: int,
    execution_dt: Optional[datetime] = None,
) -> ReportingWindow:
    execution_dt = execution_dt or datetime.now()
    start_dt = datetime.combine(date(target_year, target_month, 1), parse_clock(RANGE_START_TIME))

    if mode == "current":
        if execution_dt.year != target_year or execution_dt.month != target_month:
            raise ValueError("El modo 'current' requiere target_year/target_month del mes en ejecucion.")
        end_dt = datetime.combine(execution_dt.date(), parse_clock(RANGE_END_TIME))
    elif mode == "closed":
        if target_month == 12:
            next_month = date(target_year + 1, 1, 1)
        else:
            next_month = date(target_year, target_month + 1, 1)
        end_dt = datetime.combine(next_month, parse_clock(RANGE_END_TIME))
    else:
        raise ValueError("mode debe ser 'current' o 'closed'.")

    return ReportingWindow(
        mode=mode,
        from_dt=start_dt,
        to_dt=end_dt,
        target_year=target_year,
        target_month=target_month,
    )


def format_dt(dt_value: datetime) -> str:
    return dt_value.strftime("%d/%m/%Y %H:%M:%S")


def format_wms_date(dt_value: datetime) -> str:
    return dt_value.strftime("%d/%m/%y")


def build_operational_chunks(
    *,
    start_dt: datetime,
    end_dt: datetime,
    chunk_days: int,
    prefix: str = "C",
    level: int = 0,
) -> List[ChunkRange]:
    if chunk_days <= 0:
        raise ValueError("chunk_days debe ser mayor que 0.")
    if end_dt <= start_dt:
        raise ValueError("El rango operativo del chunk debe tener duracion positiva.")

    chunks: List[ChunkRange] = []
    cursor = start_dt
    index = 1
    end_clock = parse_clock(RANGE_END_TIME)

    while cursor < end_dt:
        candidate_end = datetime.combine(cursor.date() + timedelta(days=chunk_days), end_clock)
        if candidate_end <= cursor:
            candidate_end = cursor + timedelta(days=chunk_days)
        chunk_end = min(candidate_end, end_dt)
        if chunk_end <= cursor:
            raise RuntimeError("No se pudo construir un subrango operativo con duracion positiva.")
        chunks.append(
            ChunkRange(
                chunk_id=f"{prefix}{index:02d}",
                start_dt=cursor,
                end_dt=chunk_end,
                chunk_days=chunk_days,
                level=level,
            )
        )
        cursor = chunk_end
        index += 1

    return chunks


def audit_chunk_coverage(
    chunks: Sequence[ChunkRange],
    *,
    expected_start: datetime,
    expected_end: datetime,
) -> ChunkCoverageAudit:
    issues: List[str] = []
    if not chunks:
        issues.append("No se generaron chunks para cubrir el rango solicitado.")
        return ChunkCoverageAudit(ok=False, issues=issues)

    if chunks[0].start_dt != expected_start:
        issues.append(
            "La cobertura no inicia en el punto esperado: "
            f"observado {format_dt(chunks[0].start_dt)}, esperado {format_dt(expected_start)}."
        )

    if chunks[-1].end_dt != expected_end:
        issues.append(
            "La cobertura no termina en el punto esperado: "
            f"observado {format_dt(chunks[-1].end_dt)}, esperado {format_dt(expected_end)}."
        )

    for prev_chunk, next_chunk in zip(chunks, chunks[1:]):
        if prev_chunk.end_dt != next_chunk.start_dt:
            issues.append(
                "Se detecto hueco o solape entre chunks: "
                f"{prev_chunk.chunk_id} termina {format_dt(prev_chunk.end_dt)} y "
                f"{next_chunk.chunk_id} inicia {format_dt(next_chunk.start_dt)}."
            )
        if prev_chunk.end_dt < prev_chunk.start_dt or next_chunk.end_dt < next_chunk.start_dt:
            issues.append("Se detecto un chunk con duracion negativa.")

    return ChunkCoverageAudit(ok=not issues, issues=issues)


def build_local_historical_reference_path(client: dict, target_year: int, target_month: int) -> Path:
    month_folder = month_folder_name(target_year, target_month)
    filename = f"{client['alias_archivo']}.xlsx"
    return (
        Path(LOCAL_HISTORICAL_REFERENCE_ROOT)
        / client["carpeta_destino_historica"]
        / str(target_year)
        / month_folder
        / filename
    )


def build_sharepoint_target_path(client: dict, target_year: int, target_month: int) -> str:
    month_folder = month_folder_name(target_year, target_month)
    filename = f"{client['alias_archivo']}.xlsx"
    return "/".join(
        [
            SHAREPOINT_SITE_NAME,
            SHAREPOINT_DOCUMENT_LIBRARY,
            SHAREPOINT_PRODUCTIVIDAD_ROOT,
            client["carpeta_destino_historica"],
            str(target_year),
            month_folder,
            filename,
        ]
    )


def build_sharepoint_folder_path(client: dict, target_year: int, target_month: int) -> str:
    month_folder = month_folder_name(target_year, target_month)
    return "/".join(
        [
            SHAREPOINT_PRODUCTIVIDAD_ROOT,
            client["carpeta_destino_historica"],
            str(target_year),
            month_folder,
        ]
    )


def build_sharepoint_backup_folder_path(client: dict, target_year: int, target_month: int) -> str:
    month_folder = month_folder_name(target_year, target_month)
    return "/".join(
        [
            SHAREPOINT_PRODUCTIVIDAD_ROOT,
            SHAREPOINT_BACKUP_ROOT,
            client["carpeta_destino_historica"],
            str(target_year),
            month_folder,
            client["alias_archivo"],
        ]
    )


def expected_internal_cd(client: dict) -> str:
    return normalize_text(client.get("internal_cd_expected") or client["cd"]).replace("CD ", "")


def find_client(cd: str, alias_archivo: str) -> dict:
    cd_norm = normalize_text(cd)
    alias_norm = normalize_text(alias_archivo)
    for client in CLIENTS:
        if normalize_text(client["cd"]) == cd_norm and normalize_text(client["alias_archivo"]) == alias_norm:
            return client
    raise KeyError(f"No existe configuracion para cd='{cd}' alias='{alias_archivo}'.")


def inspect_workbook(path: Path) -> WorkbookInspection:
    wb = load_workbook(path, read_only=True, data_only=False)
    sheet_name = wb.sheetnames[0]
    if sheet_name in VALID_SHEET_NAMES:
        selected = sheet_name
    else:
        for candidate in wb.sheetnames:
            if candidate in VALID_SHEET_NAMES:
                selected = candidate
                break
        else:
            selected = sheet_name

    ws = wb[selected]
    headers_all = [str(ws.cell(row=9, column=i).value or "").strip() for i in range(1, 61)]
    last_non_empty = max((i for i, value in enumerate(headers_all, start=1) if value), default=0)
    headers = headers_all[:last_non_empty]

    has_data_rows = False
    first_non_empty_after_header = None
    for row_index in range(10, 25):
        cell_a = ws.cell(row=row_index, column=1).value
        cell_b = ws.cell(row=row_index, column=2).value
        if cell_a not in (None, "") or cell_b not in (None, ""):
            first_non_empty_after_header = normalize_text(cell_a)
            if str(cell_a or "").strip().isdigit():
                has_data_rows = True
            break

    return WorkbookInspection(
        path=path,
        sheet_name=selected,
        internal_cd=normalize_text(ws["A1"].value),
        internal_company=normalize_text(ws["A4"].value),
        internal_scope=normalize_text(ws["A5"].value),
        headers=headers,
        has_data_rows=has_data_rows,
        first_non_empty_after_header=first_non_empty_after_header,
        source_format="xlsx",
        max_non_empty_col=last_non_empty,
    )


def parse_legacy_excel_html(path: Path) -> Dict[str, Any]:
    html = path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    rows: List[List[str]] = []
    for table in tables:
        for tr in table.find_all("tr"):
            cells = [" ".join(td.get_text(" ", strip=True).split()) for td in tr.find_all(["td", "th"])]
            if cells:
                rows.append(cells)

    # Detectar cuando el WMS devuelve "EMPRESA - TODAS" en lugar de la empresa filtrada.
    # Esto ocurre por un race condition de sesión entre clientes consecutivos del mismo depósito.
    # Lanzamos un error específico para que el caller pueda reintentarlo.
    for row in rows[:8]:
        for cell in row:
            if "EMPRESA - TODAS" in cell.upper() or "EMPRESA- TODAS" in cell.upper():
                raise RuntimeError(
                    "WMS_EMPRESA_TODAS: el WMS devolvio datos de todas las empresas "
                    "en lugar de la empresa filtrada. Reintentando."
                )

    report_index = None
    for idx, row in enumerate(rows):
        if len(row) >= 18 and row[0] == "Comprobante" and row[1] == "Artículo":
            report_index = idx
            break

    if report_index is None:
        raise RuntimeError("No se encontro fila de encabezado historico en el .XLS HTML.")

    title_rows = rows[:report_index]
    headers = rows[report_index]
    detail_rows = rows[report_index + 1 :]
    detail_records = []
    trailing_messages = []

    for row in detail_rows:
        if len(row) == 1 and row[0].startswith("El reporte está ordenado"):
            trailing_messages.append(row[0])
            continue
        detail_records.append(row)

    internal_cd = normalize_text(title_rows[0][0] if len(title_rows) >= 1 and title_rows[0] else "")
    internal_company = normalize_text(title_rows[2][0] if len(title_rows) >= 3 and title_rows[2] else "")
    internal_scope = normalize_text(title_rows[3][0] if len(title_rows) >= 4 and title_rows[3] else "")

    return {
        "tables_count": len(tables),
        "title_rows": title_rows,
        "headers": headers,
        "detail_records": detail_records,
        "trailing_messages": trailing_messages,
        "internal_cd": internal_cd,
        "internal_company": internal_company,
        "internal_scope": internal_scope,
    }


def inspect_legacy_workbook_html(path: Path) -> WorkbookInspection:
    parsed = parse_legacy_excel_html(path)
    first_non_empty_after_header = None
    has_data_rows = False
    for row in parsed["detail_records"]:
        if any(cell for cell in row):
            first_non_empty_after_header = normalize_text(row[0])
            if str(row[0]).strip().isdigit():
                has_data_rows = True
            break

    headers = list(parsed["headers"])
    return WorkbookInspection(
        path=path,
        sheet_name="Reporte de Movimientos",
        internal_cd=parsed["internal_cd"],
        internal_company=parsed["internal_company"],
        internal_scope=parsed["internal_scope"],
        headers=headers,
        has_data_rows=has_data_rows,
        first_non_empty_after_header=first_non_empty_after_header,
        source_format="xls_html",
        max_non_empty_col=len(headers),
    )


def inspect_any_workbook(path: Path) -> WorkbookInspection:
    path = Path(path)
    if path.suffix.lower() == ".xls":
        return inspect_legacy_workbook_html(path)
    return inspect_workbook(path)


def load_historical_header(path: Path) -> List[str]:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [str(ws.cell(row=9, column=i).value or "").strip() for i in range(1, 61)]
    last_non_empty = max((i for i, value in enumerate(headers, start=1) if value), default=0)
    return headers[:last_non_empty]


def compare_runtime_to_historical(runtime_path: Path, historical_path: Path) -> HistoricalComparison:
    runtime_path = Path(runtime_path)
    historical_path = Path(historical_path)
    inspection = inspect_any_workbook(runtime_path)
    historical_headers = load_historical_header(historical_path)
    runtime_headers = inspection.headers

    hist_norm = [normalize_text(value) for value in historical_headers]
    run_norm = [normalize_text(value) for value in runtime_headers]

    exact = historical_headers == runtime_headers
    normalized = hist_norm == run_norm
    notes: List[str] = []
    normalized_viable = False
    normalization_rule = "none"

    if exact:
        normalized_viable = True
        normalization_rule = "copy_headers_exact"
        notes.append("La cabecera runtime coincide exactamente con la historica.")
    elif normalized:
        normalized_viable = True
        normalization_rule = "normalize_header_whitespace"
        notes.append("La cabecera runtime coincide con la historica tras normalizar espacios.")
    elif len(historical_headers) == len(runtime_headers) + 1:
        # Caso observado: la plantilla historica contiene una columna vacia intermedia.
        possible = []
        for insert_at in range(len(historical_headers)):
            candidate = runtime_headers[:insert_at] + [""] + runtime_headers[insert_at:]
            candidate_norm = [normalize_text(value) for value in candidate]
            if candidate_norm == hist_norm:
                possible.append(insert_at)
        if len(possible) == 1:
            normalized_viable = True
            normalization_rule = f"insert_blank_column_at_{possible[0] + 1}"
            notes.append(
                f"El runtime omite una columna vacia de la plantilla historica; "
                f"se puede reinsertar de forma deterministica en posicion {possible[0] + 1}."
            )
        else:
            notes.append("No se pudo identificar una unica insercion deterministica de columna vacia.")
    else:
        notes.append("La cabecera runtime no coincide limpiamente con la historica.")

    return HistoricalComparison(
        historical_headers=historical_headers,
        runtime_headers=runtime_headers,
        headers_match_exact=exact,
        headers_match_normalized=normalized,
        normalized_viable=normalized_viable,
        normalization_rule=normalization_rule,
        notes=notes,
    )


def _apply_normalization_rule(row: List[str], comparison: HistoricalComparison) -> List[str]:
    if comparison.normalization_rule.startswith("insert_blank_column_at_"):
        insert_at = int(comparison.normalization_rule.rsplit("_", 1)[-1]) - 1
        return row[:insert_at] + [""] + row[insert_at:]
    return list(row)


def normalize_legacy_html_to_xlsx(
    runtime_path: Path,
    historical_path: Path,
    target_alias: str,
    log_path: Optional[Path] = None,
) -> Dict[str, Any]:
    parsed = parse_legacy_excel_html(runtime_path)
    comparison = compare_runtime_to_historical(runtime_path, historical_path)
    if not comparison.normalized_viable:
        raise RuntimeError("La normalizacion al layout historico no es segura ni deterministica.")

    rows_out: List[List[str]] = []
    title_rows = parsed["title_rows"]
    detail_rows = parsed["detail_records"]
    header_row = _apply_normalization_rule(parsed["headers"], comparison)

    # Preserva la estructura posicional historica del reporte:
    # row1=row titulo CD, row2=titulo, row3=blank, row4..row7 metadata,
    # row8=blank, row9=headers, row10+=detalle.
    if len(title_rows) >= 1:
        rows_out.append(title_rows[0])
    if len(title_rows) >= 2:
        rows_out.append(title_rows[1])
    rows_out.append([])
    if len(title_rows) >= 3:
        rows_out.append(title_rows[2])
    if len(title_rows) >= 4:
        rows_out.append(title_rows[3])
    if len(title_rows) >= 5:
        rows_out.append(title_rows[4])
    if len(title_rows) >= 6:
        rows_out.append(title_rows[5])
    rows_out.append([])
    rows_out.append(header_row)
    for row in detail_rows:
        rows_out.append(_apply_normalization_rule(row, comparison))
    for message in parsed["trailing_messages"]:
        rows_out.append([message])

    out_path = Path(NORMALIZED_DIR) / f"{target_alias}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Movimientos"
    for r_idx, row in enumerate(rows_out, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value
    wb.save(out_path)

    if log_path:
        log(f"[NORMALIZACION] Archivo staging normalizado guardado: {out_path}", log_path)
    return {
        "normalized_path": out_path,
        "comparison": comparison,
        "row_count": len(rows_out),
    }


def save_legacy_html_as_raw_xlsx(
    runtime_path: Path,
    target_alias: str,
    log_path: Optional[Path] = None,
) -> Path:
    parsed = parse_legacy_excel_html(runtime_path)

    rows_out: List[List[str]] = []
    title_rows = parsed["title_rows"]
    detail_rows = parsed["detail_records"]

    # El bruto debe quedar como .xlsx, pero conservando la estructura runtime
    # observada, sin aplicar aun la normalizacion al layout historico.
    if len(title_rows) >= 1:
        rows_out.append(title_rows[0])
    if len(title_rows) >= 2:
        rows_out.append(title_rows[1])
    rows_out.append([])
    if len(title_rows) >= 3:
        rows_out.append(title_rows[2])
    if len(title_rows) >= 4:
        rows_out.append(title_rows[3])
    if len(title_rows) >= 5:
        rows_out.append(title_rows[4])
    if len(title_rows) >= 6:
        rows_out.append(title_rows[5])
    rows_out.append([])
    rows_out.append(list(parsed["headers"]))
    for row in detail_rows:
        rows_out.append(list(row))
    for message in parsed["trailing_messages"]:
        rows_out.append([message])

    out_path = Path(DOWNLOAD_DIR) / f"{target_alias}_bruto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Movimientos"
    for r_idx, row in enumerate(rows_out, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value
    wb.save(out_path)

    if log_path:
        log(f"[BRUTO] Archivo runtime convertido y guardado como .xlsx: {out_path}", log_path)
    return out_path


def _load_full_rows(path: Path, row_from: int, row_to: int, max_col: int = 80) -> List[List[str]]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    rows: List[List[str]] = []
    for row_idx in range(row_from, row_to + 1):
        row = [str(ws.cell(row=row_idx, column=col_idx).value or "").strip() for col_idx in range(1, max_col + 1)]
        rows.append(row)
    return rows


def _last_non_empty_index(row: Sequence[str]) -> int:
    return max((idx for idx, value in enumerate(row, start=1) if normalize_text(value)), default=0)


def _header_token(value: object) -> str:
    text = normalize_text(value)
    text = text.replace("ÃƒÂ", "").replace("Ã", "").replace("Â", "")
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def _is_header_row_marker(row: Sequence[str]) -> bool:
    if len(row) < 2:
        return False
    return _header_token(row[0]) == "COMPROBANTE" and _header_token(row[1]) == "ARTICULO"


def _row_values_to_strings(row: Sequence[Any], target_len: int) -> List[str]:
    values = [str(value or "").strip() for value in row]
    if len(values) < target_len:
        values.extend([""] * (target_len - len(values)))
    return values[:target_len]


def validate_structural_xlsx(
    workbook_path: Path,
    historical_path: Path,
    client: dict,
    stage: str,
) -> StructuralValidation:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    historical_headers = load_historical_header(historical_path)
    historical_norm = [normalize_text(value) for value in historical_headers]

    observed_header_row = 0
    observed_header_values: List[str] = []
    for row_idx in range(1, 16):
        row = [str(ws.cell(row=row_idx, column=col_idx).value or "").strip() for col_idx in range(1, 81)]
        if normalize_text(row[0]) == "COMPROBANTE" and normalize_text(row[1]) in {"ARTÍCULO", "ARTÃ­CULO", "ARTICULO"}:
            observed_header_row = row_idx
            observed_header_values = row
            break

    observed_header_count = _last_non_empty_index(observed_header_values) if observed_header_values else 0
    expected_header_count = len(historical_headers)
    comparison = compare_runtime_to_historical(workbook_path, historical_path)

    if stage == "raw":
        if comparison.normalization_rule.startswith("insert_blank_column_at_"):
            expected_effective_header_count = expected_header_count - 1
        else:
            expected_effective_header_count = expected_header_count
    else:
        expected_effective_header_count = expected_header_count

    portada_ok = all(
        [
            normalize_text(ws["A1"].value) == normalize_text(client["cd"]).replace("CD ", ""),
            "INFORME DE MOVIMIENTOS" in normalize_text(ws["A2"].value),
            normalize_text(ws["A4"].value) == normalize_text(client["empresa_wms"]),
            normalize_text(ws["A5"].value) in {
                normalize_text(client["deposito_wms_origen"]),
                normalize_text(client["cd"]).replace("CD ", ""),
            },
        ]
    )

    repeated_header_rows: List[int] = []
    short_rows: List[int] = []
    long_rows: List[int] = []
    suspicious_rows: List[int] = []
    issues: List[str] = []
    warnings: List[str] = []

    if observed_header_row != 9:
        issues.append(f"Encabezado fuera de fila esperada: observado fila {observed_header_row}, esperado fila 9.")

    if observed_header_count != expected_effective_header_count:
        issues.append(
            "Cantidad de columnas no coincide con lo esperado para la etapa "
            f"{stage}: observado {observed_header_count}, esperado {expected_effective_header_count}."
        )

    exact_header_match = False
    observed_header_trimmed = observed_header_values[:observed_header_count] if observed_header_values else []
    observed_header_norm = [normalize_text(value) for value in observed_header_trimmed]

    if stage == "normalized":
        exact_header_match = observed_header_trimmed == historical_headers
        if not exact_header_match:
            issues.append("El normalizado no coincide exactamente con la cabecera historica.")
    else:
        exact_header_match = comparison.normalized_viable
        if comparison.normalization_rule.startswith("insert_blank_column_at_"):
            insert_at = int(comparison.normalization_rule.rsplit("_", 1)[-1]) - 1
            candidate = observed_header_norm[:insert_at] + [""] + observed_header_norm[insert_at:]
            if candidate != historical_norm:
                issues.append("La columna vacia reinserta no reconstruye exactamente la cabecera historica.")
            elif insert_at + 1 != 21:
                issues.append(
                    f"La columna vacia no quedo en la posicion esperada: observado {insert_at + 1}, esperado 21."
                )
        elif observed_header_norm != historical_norm:
            issues.append("La cabecera del bruto no coincide ni permite una normalizacion segura al historico.")

    max_row = ws.max_row
    detail_start_row = 10
    for row_idx in range(detail_start_row, max_row + 1):
        row = [str(ws.cell(row=row_idx, column=col_idx).value or "").strip() for col_idx in range(1, 81)]
        last_non_empty = _last_non_empty_index(row)
        if last_non_empty == 0:
            continue

        row_norm = [normalize_text(value) for value in row]
        if row_norm[0] == "COMPROBANTE" and row_norm[1] in {"ARTÍCULO", "ARTÃ­CULO", "ARTICULO"}:
            repeated_header_rows.append(row_idx)
            continue

        if "EL REPORTE ESTÁ ORDENADO" in normalize_text(row[0]) or "EL REPORTE ESTÃ¡ ORDENADO" in normalize_text(row[0]):
            continue

        if last_non_empty > expected_effective_header_count:
            long_rows.append(row_idx)

        first_cell = normalize_text(row[0])
        if first_cell and not first_cell.isdigit():
            suspicious_rows.append(row_idx)
            continue

        if first_cell.isdigit():
            critical_cells = {
                2: normalize_text(row[1]),
                9: normalize_text(row[8]),
                10: normalize_text(row[9]),
                11: normalize_text(row[10]),
                12: normalize_text(row[11]),
            }
            if any(value == "" for value in critical_cells.values()):
                short_rows.append(row_idx)
                suspicious_rows.append(row_idx)

    if repeated_header_rows:
        issues.append(f"Se detectaron encabezados repetidos dentro del detalle: {repeated_header_rows}.")
    if short_rows:
        issues.append(f"Se detectaron filas cortas o corridas en el detalle: {short_rows[:10]}.")
    if long_rows:
        issues.append(f"Se detectaron filas con columnas extra/desfasadas: {long_rows[:10]}.")
    if suspicious_rows:
        issues.append(f"Se detectaron filas sospechosas o desplazadas: {suspicious_rows[:10]}.")
    if not portada_ok:
        issues.append("La portada no coincide con el layout esperado del cliente.")
    if ws["A10"].value in (None, "") and ws.max_row >= 10:
        warnings.append("La fila 10 no contiene primer valor en A; revisar si el detalle arranca con columnas desplazadas.")

    return StructuralValidation(
        stage=stage,
        ok=not issues,
        expected_header_count=expected_effective_header_count,
        observed_header_count=observed_header_count,
        expected_header_row=9,
        observed_header_row=observed_header_row,
        expected_detail_start_row=10,
        observed_detail_start_row=10,
        portada_ok=portada_ok,
        exact_header_match=exact_header_match,
        repeated_header_rows=repeated_header_rows,
        short_rows=short_rows,
        long_rows=long_rows,
        suspicious_rows=suspicious_rows,
        issues=issues,
        warnings=warnings,
    )


def validate_structural_xlsx(
    workbook_path: Path,
    historical_path: Path,
    client: dict,
    stage: str,
) -> StructuralValidation:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    historical_headers = load_historical_header(historical_path)
    historical_norm = [normalize_text(value) for value in historical_headers]
    scan_col_count = max(80, len(historical_headers))

    observed_header_row = 0
    observed_header_values: List[str] = []
    top_rows: Dict[int, List[str]] = {}
    for row_idx, row_values in enumerate(
        ws.iter_rows(min_row=1, max_row=15, max_col=scan_col_count, values_only=True),
        start=1,
    ):
        row = _row_values_to_strings(row_values, scan_col_count)
        top_rows[row_idx] = row
        if _is_header_row_marker(row):
            observed_header_row = row_idx
            observed_header_values = row
            break

    observed_header_count = _last_non_empty_index(observed_header_values) if observed_header_values else 0
    expected_header_count = len(historical_headers)
    comparison = compare_runtime_to_historical(workbook_path, historical_path)

    if stage == "raw":
        if comparison.normalization_rule.startswith("insert_blank_column_at_"):
            expected_effective_header_count = expected_header_count - 1
        else:
            expected_effective_header_count = expected_header_count
    else:
        expected_effective_header_count = expected_header_count

    portada_ok = all(
        [
            normalize_text(top_rows.get(1, [""])[0]) == expected_internal_cd(client),
            "INFORME DE MOVIMIENTOS" in normalize_text(top_rows.get(2, [""])[0]),
            normalize_text(top_rows.get(4, [""])[0]) == normalize_text(client["empresa_wms"]),
            normalize_text(top_rows.get(5, [""])[0]) in {
                normalize_text(client["deposito_wms_origen"]),
                expected_internal_cd(client),
            },
        ]
    )

    repeated_header_rows: List[int] = []
    short_rows: List[int] = []
    long_rows: List[int] = []
    suspicious_rows: List[int] = []
    issues: List[str] = []
    warnings: List[str] = []

    if observed_header_row != 9:
        issues.append(f"Encabezado fuera de fila esperada: observado fila {observed_header_row}, esperado fila 9.")

    if observed_header_count != expected_effective_header_count:
        issues.append(
            "Cantidad de columnas no coincide con lo esperado para la etapa "
            f"{stage}: observado {observed_header_count}, esperado {expected_effective_header_count}."
        )

    exact_header_match = False
    observed_header_trimmed = observed_header_values[:observed_header_count] if observed_header_values else []
    observed_header_norm = [normalize_text(value) for value in observed_header_trimmed]

    if stage == "normalized":
        exact_header_match = observed_header_trimmed == historical_headers
        if not exact_header_match:
            issues.append("El normalizado no coincide exactamente con la cabecera historica.")
    else:
        exact_header_match = comparison.normalized_viable
        if comparison.normalization_rule.startswith("insert_blank_column_at_"):
            insert_at = int(comparison.normalization_rule.rsplit("_", 1)[-1]) - 1
            candidate = observed_header_norm[:insert_at] + [""] + observed_header_norm[insert_at:]
            if candidate != historical_norm:
                issues.append("La columna vacia reinserta no reconstruye exactamente la cabecera historica.")
            elif insert_at + 1 != 21:
                issues.append(
                    f"La columna vacia no quedo en la posicion esperada: observado {insert_at + 1}, esperado 21."
                )
        elif observed_header_norm != historical_norm:
            issues.append("La cabecera del bruto no coincide ni permite una normalizacion segura al historico.")

    max_row = ws.max_row
    detail_start_row = 10
    first_detail_row: Optional[List[str]] = None
    for row_idx, row_values in enumerate(
        ws.iter_rows(
            min_row=detail_start_row,
            max_row=max_row,
            max_col=scan_col_count,
            values_only=True,
        ),
        start=detail_start_row,
    ):
        row = _row_values_to_strings(row_values, scan_col_count)
        if first_detail_row is None:
            first_detail_row = row
        last_non_empty = _last_non_empty_index(row)
        if last_non_empty == 0:
            continue

        if _is_header_row_marker(row):
            repeated_header_rows.append(row_idx)
            continue

        row0_norm = normalize_text(row[0])
        if "EL REPORTE" in row0_norm and "ORDENADO" in row0_norm:
            continue

        if last_non_empty > expected_effective_header_count:
            long_rows.append(row_idx)

        first_cell = _header_token(row[0])
        if first_cell and not first_cell.isdigit():
            suspicious_rows.append(row_idx)
            continue

        if first_cell.isdigit():
            critical_cells = {
                2: normalize_text(row[1]),
                9: normalize_text(row[8]),
                10: normalize_text(row[9]),
                11: normalize_text(row[10]),
                12: normalize_text(row[11]),
            }
            if any(value == "" for value in critical_cells.values()):
                short_rows.append(row_idx)
                suspicious_rows.append(row_idx)

    if repeated_header_rows:
        issues.append(f"Se detectaron encabezados repetidos dentro del detalle: {repeated_header_rows}.")
    if short_rows:
        issues.append(f"Se detectaron filas cortas o corridas en el detalle: {short_rows[:10]}.")
    if long_rows:
        issues.append(f"Se detectaron filas con columnas extra/desfasadas: {long_rows[:10]}.")
    if suspicious_rows:
        issues.append(f"Se detectaron filas sospechosas o desplazadas: {suspicious_rows[:10]}.")
    if not portada_ok:
        issues.append("La portada no coincide con el layout esperado del cliente.")
    if first_detail_row is not None and first_detail_row[0] == "" and max_row >= 10:
        warnings.append("La fila 10 no contiene primer valor en A; revisar si el detalle arranca con columnas desplazadas.")

    return StructuralValidation(
        stage=stage,
        ok=not issues,
        expected_header_count=expected_effective_header_count,
        observed_header_count=observed_header_count,
        expected_header_row=9,
        observed_header_row=observed_header_row,
        expected_detail_start_row=10,
        observed_detail_start_row=10,
        portada_ok=portada_ok,
        exact_header_match=exact_header_match,
        repeated_header_rows=repeated_header_rows,
        short_rows=short_rows,
        long_rows=long_rows,
        suspicious_rows=suspicious_rows,
        issues=issues,
        warnings=warnings,
    )


def _read_normalized_sections(path: Path, expected_header_count: int) -> Dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    title_rows: List[List[str]] = []
    for row_values in ws.iter_rows(min_row=1, max_row=8, max_col=expected_header_count, values_only=True):
        title_rows.append(_row_values_to_strings(row_values, expected_header_count))

    header_row_values = next(
        ws.iter_rows(min_row=9, max_row=9, max_col=expected_header_count, values_only=True),
        tuple("" for _ in range(expected_header_count)),
    )
    header_row = _row_values_to_strings(header_row_values, expected_header_count)

    detail_rows: List[List[str]] = []
    trailing_messages: List[str] = []
    for row_values in ws.iter_rows(min_row=10, max_row=ws.max_row, max_col=expected_header_count, values_only=True):
        row = _row_values_to_strings(row_values, expected_header_count)
        last_non_empty = _last_non_empty_index(row)
        if last_non_empty == 0:
            continue
        if _is_header_row_marker(row):
            continue
        row0_norm = normalize_text(row[0])
        if "EL REPORTE" in row0_norm and "ORDENADO" in row0_norm:
            trailing_messages.append(row[0])
            continue
        detail_rows.append(row[:expected_header_count])

    return {
        "title_rows": title_rows,
        "header_row": header_row,
        "detail_rows": detail_rows,
        "trailing_messages": trailing_messages,
    }


def build_detail_row_fingerprint(row: Sequence[str], expected_header_count: int) -> str:
    normalized_values = [normalize_text(value) for value in list(row)[:expected_header_count]]
    payload = "\u241f".join(normalized_values)
    return sha256_bytes(payload.encode("utf-8"))


def consolidate_normalized_chunks(
    *,
    chunk_paths: Sequence[Path],
    historical_path: Path,
    target_alias: str,
    log_path: Optional[Path] = None,
) -> ChunkConsolidationResult:
    historical_headers = load_historical_header(historical_path)
    expected_header_count = len(historical_headers)
    issues: List[str] = []
    warnings: List[str] = []
    duplicate_samples: List[str] = []
    ordered_unique_rows: List[List[str]] = []
    seen_fingerprints: set[str] = set()
    total_rows_seen = 0
    chunks_with_rows = 0
    empty_chunks = 0
    duplicate_rows_removed = 0
    title_rows: Optional[List[List[str]]] = None
    trailing_messages: List[str] = []

    if not chunk_paths:
        return ChunkConsolidationResult(
            ok=False,
            consolidated_path=None,
            total_chunks=0,
            chunks_with_rows=0,
            empty_chunks=0,
            total_rows_seen=0,
            unique_rows=0,
            duplicate_rows_removed=0,
            issues=["No existen chunks normalizados para consolidar."],
            warnings=[],
        )

    for chunk_path in chunk_paths:
        section = _read_normalized_sections(Path(chunk_path), expected_header_count)
        if section["header_row"] != historical_headers:
            issues.append(f"El chunk {Path(chunk_path).name} no coincide exactamente con la cabecera historica.")
            continue

        if title_rows is None:
            title_rows = section["title_rows"]
            trailing_messages = list(section["trailing_messages"])

        detail_rows = section["detail_rows"]
        if detail_rows:
            chunks_with_rows += 1
        else:
            empty_chunks += 1

        for row in detail_rows:
            total_rows_seen += 1
            fingerprint = build_detail_row_fingerprint(row, expected_header_count)
            if fingerprint in seen_fingerprints:
                duplicate_rows_removed += 1
                if len(duplicate_samples) < 10:
                    duplicate_samples.append(
                        f"{Path(chunk_path).name} | comprobante={row[0]} | articulo={row[1]} | fecha={row[8]} {row[9]}"
                    )
                continue
            seen_fingerprints.add(fingerprint)
            ordered_unique_rows.append(row)

    if title_rows is None:
        issues.append("No se pudo capturar la portada desde ningun chunk normalizado.")

    if duplicate_rows_removed and log_path:
        log(
            (
                "[DERCO][DEDUP] Filas duplicadas descartadas en consolidacion: "
                f"{duplicate_rows_removed}"
            ),
            log_path,
        )
        for sample in duplicate_samples:
            log(f"[DERCO][DEDUP] {sample}", log_path)

    if issues:
        return ChunkConsolidationResult(
            ok=False,
            consolidated_path=None,
            total_chunks=len(chunk_paths),
            chunks_with_rows=chunks_with_rows,
            empty_chunks=empty_chunks,
            total_rows_seen=total_rows_seen,
            unique_rows=len(ordered_unique_rows),
            duplicate_rows_removed=duplicate_rows_removed,
            duplicate_samples=duplicate_samples,
            issues=issues,
            warnings=warnings,
        )

    rows_out: List[List[str]] = []
    for row in title_rows or []:
        rows_out.append(row[:expected_header_count])
    rows_out.append(list(historical_headers))
    rows_out.extend([row[:expected_header_count] for row in ordered_unique_rows])
    if not ordered_unique_rows and trailing_messages:
        rows_out.append([trailing_messages[0]])

    out_path = Path(NORMALIZED_DIR) / f"{target_alias}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Movimientos"
    for row_idx, row in enumerate(rows_out, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
    wb.save(out_path)

    return ChunkConsolidationResult(
        ok=True,
        consolidated_path=out_path,
        total_chunks=len(chunk_paths),
        chunks_with_rows=chunks_with_rows,
        empty_chunks=empty_chunks,
        total_rows_seen=total_rows_seen,
        unique_rows=len(ordered_unique_rows),
        duplicate_rows_removed=duplicate_rows_removed,
        duplicate_samples=duplicate_samples,
        issues=issues,
        warnings=warnings,
    )


def log_structural_validation(result: StructuralValidation, log_path: Optional[Path] = None) -> None:
    log(
        (
            f"[ESTRUCTURA][{result.stage}] ok={result.ok} "
            f"header_row={result.observed_header_row}/{result.expected_header_row} "
            f"header_cols={result.observed_header_count}/{result.expected_header_count} "
            f"portada_ok={result.portada_ok} exact_header_match={result.exact_header_match}"
        ),
        log_path,
    )
    for warning in result.warnings:
        log(f"[WARN][ESTRUCTURA][{result.stage}] {warning}", log_path)
    for issue in result.issues:
        log(f"[CRITICO][ESTRUCTURA][{result.stage}] {issue}", log_path)


def _extract_candidate_stamp(path: Path, alias_archivo: str) -> Optional[datetime]:
    pattern = re.compile(rf"^{re.escape(alias_archivo)}_(\d{{8}}_\d{{6}})\.xlsx$", re.IGNORECASE)
    match = pattern.match(path.name)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y%m%d_%H%M%S")
    except ValueError:
        return None


def find_latest_normalized_candidate(alias_archivo: str) -> Optional[Path]:
    directory = Path(NORMALIZED_DIR)
    if not directory.exists():
        return None
    candidates = [path for path in directory.glob(f"{alias_archivo}_*.xlsx") if path.is_file()]
    if not candidates:
        return None

    stamped_candidates = []
    fallback_candidates = []
    for path in candidates:
        stamp = _extract_candidate_stamp(path, alias_archivo)
        if stamp is not None:
            stamped_candidates.append((stamp, path.stat().st_mtime, path))
        else:
            fallback_candidates.append((path.stat().st_mtime, path))

    if stamped_candidates:
        stamped_candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
        return stamped_candidates[0][2]

    fallback_candidates.sort(key=lambda item: item[0], reverse=True)
    return fallback_candidates[0][1] if fallback_candidates else None


def load_azure_graph_module():
    helper_dir = Path(r"C:\ClaudeWork\WMS_Automatizacion")
    if str(helper_dir) not in sys.path:
        sys.path.insert(0, str(helper_dir))
    import azure_graph  # type: ignore

    return azure_graph


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(Path(path).read_bytes())


def workbook_semantic_sha256(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    entries: List[str] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value in (None, ""):
                continue
            if hasattr(value, "isoformat"):
                try:
                    normalized = value.isoformat(sep=" ")
                except TypeError:
                    normalized = value.isoformat()
            else:
                normalized = str(value).strip()
            entries.append(f"{cell.row}|{cell.column}|{normalized}")
    return sha256_bytes("\n".join(entries).encode("utf-8"))


def upload_file_to_sharepoint(
    token: str,
    drive_id: str,
    folder_path: str,
    local_path: Path,
    remote_name: str,
) -> bool:
    azure_graph = load_azure_graph_module()
    url = f"{azure_graph.GRAPH_BASE}/drives/{drive_id}/root:/{folder_path}/{remote_name}:/content"
    with open(local_path, "rb") as fh:
        data = fh.read()
    response = requests.put(
        url,
        data=data,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream",
        },
        timeout=180,
    )
    return response.status_code in (200, 201)


def upload_bytes_to_sharepoint(
    token: str,
    drive_id: str,
    folder_path: str,
    remote_name: str,
    payload: bytes,
) -> bool:
    azure_graph = load_azure_graph_module()
    url = f"{azure_graph.GRAPH_BASE}/drives/{drive_id}/root:/{folder_path}/{remote_name}:/content"
    response = requests.put(
        url,
        data=payload,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream",
        },
        timeout=180,
    )
    return response.status_code in (200, 201)


def get_sharepoint_file_state(
    token: str,
    drive_id: str,
    folder_path: str,
    remote_name: str,
) -> SharePointRemoteFileState:
    azure_graph = load_azure_graph_module()
    url = f"{azure_graph.GRAPH_BASE}/drives/{drive_id}/root:/{folder_path}:/children"
    response = requests.get(url, headers=azure_graph._gh(token), timeout=60)
    if response.status_code == 404:
        return SharePointRemoteFileState(exists=False)
    response.raise_for_status()
    for item in response.json().get("value", []):
        if item.get("name") == remote_name and "file" in item:
            return SharePointRemoteFileState(
                exists=True,
                item_id=item.get("id", ""),
                name=item.get("name", ""),
                size=int(item.get("size", 0)),
                last_modified=item.get("lastModifiedDateTime", ""),
                web_url=item.get("webUrl", ""),
            )
    return SharePointRemoteFileState(exists=False)


def download_sharepoint_file_bytes(
    token: str,
    drive_id: str,
    remote_state: SharePointRemoteFileState,
) -> bytes:
    if not remote_state.exists or not remote_state.item_id:
        raise RuntimeError("No existe archivo remoto para descargar.")
    azure_graph = load_azure_graph_module()
    url = f"{azure_graph.GRAPH_BASE}/drives/{drive_id}/items/{remote_state.item_id}/content"
    response = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=180)
    response.raise_for_status()
    return response.content


def prepare_sharepoint_publish_plan(
    *,
    client_key: str,
    client: dict,
    candidate_path: Path,
    historical_path: Path,
    target_year: int,
    target_month: int,
) -> SharePointPublishPlan:
    candidate_path = Path(candidate_path)
    issues: List[str] = []
    warnings: List[str] = []

    if not candidate_path.exists():
        issues.append(f"No existe el candidato local: {candidate_path}")
        return SharePointPublishPlan(
            client_key=client_key,
            local_candidate=candidate_path,
            sharepoint_target_path=build_sharepoint_target_path(client, target_year, target_month),
            sharepoint_folder_path=build_sharepoint_folder_path(client, target_year, target_month),
            sharepoint_filename=f"{client['alias_archivo']}.xlsx",
            ready=False,
            issues=issues,
            warnings=warnings,
        )

    validation = validate_downloaded_workbook(candidate_path, client, target_year, target_month)
    structural = validate_structural_xlsx(candidate_path, historical_path, client, "normalized")
    comparison = compare_runtime_to_historical(candidate_path, historical_path)

    issues.extend(validation.critical_errors)
    issues.extend(structural.issues)
    warnings.extend(validation.warnings)
    warnings.extend(structural.warnings)

    if not comparison.headers_match_exact:
        issues.append("El candidato no coincide exactamente con el layout historico esperado.")

    return SharePointPublishPlan(
        client_key=client_key,
        local_candidate=candidate_path,
        sharepoint_target_path=build_sharepoint_target_path(client, target_year, target_month),
        sharepoint_folder_path=build_sharepoint_folder_path(client, target_year, target_month),
        sharepoint_filename=f"{client['alias_archivo']}.xlsx",
        remote_backup_target_path=(
            build_sharepoint_backup_folder_path(client, target_year, target_month)
            + "/"
            + f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{client['alias_archivo']}.xlsx"
        ),
        ready=not issues,
        issues=issues,
        warnings=warnings,
    )


def resolve_sharepoint_remote_state(
    plan: SharePointPublishPlan,
    *,
    token: Optional[str] = None,
    drive_id: Optional[str] = None,
) -> SharePointPublishPlan:
    azure_graph = load_azure_graph_module()
    token = token or azure_graph.get_token()
    drive_id = drive_id or azure_graph.get_drive_id(token)
    remote_state = get_sharepoint_file_state(token, drive_id, plan.sharepoint_folder_path, plan.sharepoint_filename)
    plan.remote_exists = remote_state.exists
    return plan


def create_sharepoint_remote_backup(
    *,
    token: str,
    drive_id: str,
    client: dict,
    plan: SharePointPublishPlan,
    target_year: int,
    target_month: int,
) -> Optional[SharePointRemoteFileState]:
    remote_state = get_sharepoint_file_state(token, drive_id, plan.sharepoint_folder_path, plan.sharepoint_filename)
    if not remote_state.exists:
        return None

    payload = download_sharepoint_file_bytes(token, drive_id, remote_state)
    backup_folder = build_sharepoint_backup_folder_path(client, target_year, target_month)
    backup_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{plan.sharepoint_filename}"
    ok = upload_bytes_to_sharepoint(token, drive_id, backup_folder, backup_name, payload)
    if not ok:
        raise RuntimeError("La creacion del backup remoto retorno False.")

    backup_state = get_sharepoint_file_state(token, drive_id, backup_folder, backup_name)
    if not backup_state.exists:
        raise RuntimeError("No se pudo verificar el backup remoto recien creado.")

    plan.remote_backup_target_path = "/".join(
        [SHAREPOINT_SITE_NAME, SHAREPOINT_DOCUMENT_LIBRARY, backup_folder, backup_name]
    )
    return backup_state


def verify_sharepoint_upload(
    *,
    token: str,
    drive_id: str,
    client: dict,
    plan: SharePointPublishPlan,
    historical_path: Path,
    target_year: int,
    target_month: int,
    remote_before: Optional[SharePointRemoteFileState] = None,
    backup_state: Optional[SharePointRemoteFileState] = None,
) -> SharePointPostUploadVerification:
    ensure_runtime_dirs()
    issues: List[str] = []
    warnings: List[str] = []

    remote_after = get_sharepoint_file_state(token, drive_id, plan.sharepoint_folder_path, plan.sharepoint_filename)
    if not remote_after.exists:
        issues.append("El archivo remoto no aparece en SharePoint despues de la subida.")
        return SharePointPostUploadVerification(
            ok=False,
            remote_state=remote_after,
            backup_state=backup_state,
            issues=issues,
            warnings=warnings,
        )

    local_bytes = plan.local_candidate.read_bytes()
    remote_bytes = download_sharepoint_file_bytes(token, drive_id, remote_after)
    local_size = len(local_bytes)
    remote_size = len(remote_bytes)
    local_hash = sha256_bytes(local_bytes)
    remote_hash = sha256_bytes(remote_bytes)

    if remote_size != local_size:
        warnings.append(f"Tamano remoto difiere del candidato local: remoto={remote_size}, local={local_size}.")
    if remote_hash != local_hash:
        warnings.append("El hash binario SHA256 remoto no coincide con el candidato local.")
    if not remote_after.last_modified:
        issues.append("SharePoint no devolvio fecha de modificacion del archivo remoto.")
    if remote_before and remote_before.exists and remote_after.last_modified == remote_before.last_modified:
        warnings.append("La fecha de modificacion remota no cambio respecto del estado previo.")

    verify_copy = Path(SHAREPOINT_VERIFY_DIR) / f"{plan.sharepoint_filename.replace('.xlsx', '')}_remote_verify_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    verify_copy.write_bytes(remote_bytes)
    local_semantic_hash = workbook_semantic_sha256(plan.local_candidate)
    remote_semantic_hash = workbook_semantic_sha256(verify_copy)
    if remote_semantic_hash != local_semantic_hash:
        issues.append("La huella semantica del workbook remoto no coincide con el candidato local.")

    validation = validate_downloaded_workbook(verify_copy, client, target_year, target_month)
    structural = validate_structural_xlsx(verify_copy, historical_path, client, "normalized")
    comparison = compare_runtime_to_historical(verify_copy, historical_path)

    issues.extend(validation.critical_errors)
    issues.extend(structural.issues)
    warnings.extend(validation.warnings)
    warnings.extend(structural.warnings)
    if not comparison.headers_match_exact:
        issues.append("La relectura del remoto no coincide exactamente con el layout historico.")

    return SharePointPostUploadVerification(
        ok=not issues,
        remote_state=remote_after,
        backup_state=backup_state,
        local_size=local_size,
        remote_size=remote_size,
        local_sha256=local_hash,
        remote_sha256=remote_hash,
        local_semantic_sha256=local_semantic_hash,
        remote_semantic_sha256=remote_semantic_hash,
        remote_verify_copy=verify_copy,
        issues=issues,
        warnings=warnings,
    )


def _infer_alias_candidates(internal_cd: str, internal_company: str) -> List[str]:
    matches = []
    for client in CLIENTS:
        if normalize_text(client["empresa_wms"]) != internal_company:
            continue
        if expected_internal_cd(client) != internal_cd:
            continue
        matches.append(client["alias_archivo"])
    return matches


def validate_downloaded_workbook(
    workbook_path: Path,
    client: dict,
    target_year: int,
    target_month: int,
) -> ValidationResult:
    inspection = inspect_any_workbook(workbook_path)
    target_path = build_sharepoint_target_path(client, target_year, target_month)
    critical_errors: List[str] = []
    warnings: List[str] = []

    expected_cd_internal = expected_internal_cd(client)
    expected_company = normalize_text(client["empresa_wms"])
    expected_headers = [normalize_text(value) for value in EXPECTED_HEADERS]
    actual_headers = [normalize_text(value) for value in inspection.headers[: len(EXPECTED_HEADERS)]]

    if inspection.sheet_name not in VALID_SHEET_NAMES:
        critical_errors.append(
            f"Hoja invalida: '{inspection.sheet_name}'. Se esperaba una de {list(VALID_SHEET_NAMES)}."
        )

    if inspection.internal_cd != expected_cd_internal:
        critical_errors.append(
            f"CD inconsistente: esperado '{expected_cd_internal}', obtenido '{inspection.internal_cd}'."
        )

    if inspection.internal_company != expected_company:
        critical_errors.append(
            f"Empresa inconsistente: esperado '{expected_company}', obtenido '{inspection.internal_company}'."
        )

    inferred_aliases = _infer_alias_candidates(inspection.internal_cd, inspection.internal_company)
    if inferred_aliases and client["alias_archivo"] not in inferred_aliases:
        critical_errors.append(
            f"Alias inconsistente: esperado '{client['alias_archivo']}', inferido {inferred_aliases} desde el Excel."
        )
    elif not inferred_aliases:
        warnings.append("No se pudo inferir alias unico desde el contenido interno del Excel.")

    if actual_headers != expected_headers:
        critical_errors.append("Encabezados historicos no coinciden con la estructura esperada de Productividad.")

    expected_scope = normalize_text(client["deposito_wms_origen"])
    if inspection.internal_scope and inspection.internal_scope not in {expected_scope, expected_cd_internal}:
        warnings.append(
            f"Scope interno '{inspection.internal_scope}' no coincide limpiamente con el deposito esperado '{expected_scope}'."
        )

    if not inspection.has_data_rows:
        warnings.append("Archivo valido sin movimientos: tratar como vacio, no como fallo.")

    return ValidationResult(
        ok=not critical_errors,
        is_empty_valid=not critical_errors and not inspection.has_data_rows,
        target_path=target_path,
        critical_errors=critical_errors,
        warnings=warnings,
        inspection=inspection,
    )


def build_catalog_table_rows(clients: Iterable[dict]) -> List[str]:
    rows = []
    for client in clients:
        rows.append(
            " | ".join(
                [
                    client["cd"],
                    client["alias_archivo"],
                    client["empresa_wms"],
                    client["deposito_wms_origen"],
                    client["carpeta_destino_historica"],
                    str(client["active"]),
                ]
            )
        )
    return rows


def format_window(window: ReportingWindow) -> str:
    return (
        f"{window.mode}: {window.from_dt.strftime('%Y-%m-%d %H:%M:%S')} -> "
        f"{window.to_dt.strftime('%Y-%m-%d %H:%M:%S')}"
    )


def _dedupe_emails(values: Sequence[str]) -> List[str]:
    seen: set[str] = set()
    output: List[str] = []
    for value in values:
        email = (value or "").strip()
        if not email:
            continue
        key = email.lower()
        if key in seen:
            continue
        seen.add(key)
        output.append(email)
    return output


def get_productividad_notification_recipients() -> tuple[str, List[str], List[str]]:
    """Retorna (sender, to_recipients, cc_recipients).
    Lee EMAIL_DESTINO y EMAIL_CC del .env local de Productividad (fallback: .env raíz).
    """
    local_env = Path(__file__).resolve().parent / ".env"
    if local_env.exists():
        load_dotenv(local_env, override=False)
    load_dotenv(Path(r"C:\ClaudeWork\.env"), override=False)

    sender = (
        (os.getenv("PRODUCTIVIDAD_EMAIL_FROM") or "").strip()
        or (os.getenv("SHAREPOINT_USER") or "").strip()
    )

    def _parse(var: str) -> List[str]:
        raw = os.getenv(var, "")
        return [e.strip() for e in raw.replace(";", ",").split(",") if e.strip()]

    # EMAIL_DESTINO (local .env) tiene prioridad sobre PRODUCTIVIDAD_EMAIL_DESTINOS (raíz)
    to_list = _parse("EMAIL_DESTINO") or _parse("PRODUCTIVIDAD_EMAIL_DESTINOS")
    cc_list = _parse("EMAIL_CC")

    to_recipients = _dedupe_emails(([sender] if sender else []) + to_list)
    cc_recipients = _dedupe_emails([e for e in cc_list if e not in to_recipients])
    return sender, to_recipients, cc_recipients


def send_html_notification(
    *,
    subject: str,
    html_body: str,
    log_path: Optional[Path] = None,
    recipients_override: Optional[List[str]] = None,
) -> bool:
    sender, configured_recipients, cc_recipients = get_productividad_notification_recipients()
    recipients = recipients_override if recipients_override else configured_recipients
    if not sender:
        if log_path:
            log("[NOTIF] PRODUCTIVIDAD_EMAIL_FROM/SHAREPOINT_USER no configurado. Correo no enviado.", log_path)
        return False
    if not recipients:
        if log_path:
            log("[NOTIF] No hay destinatarios configurados para Productividad. Correo no enviado.", log_path)
        return False

    try:
        azure_graph = load_azure_graph_module()
        # Un solo envío con todos los TO y CC
        ok = azure_graph.enviar_email(
            from_email=sender,
            to_email=recipients[0],
            extra_to_emails=recipients[1:],
            cc_emails=cc_recipients if not recipients_override else None,
            asunto=subject,
            html_body=html_body,
        )
        if ok:
            if log_path:
                log(f"[NOTIF] Correo enviado via Graph API | TO: {', '.join(recipients)}" +
                    (f" | CC: {', '.join(cc_recipients)}" if cc_recipients else ""), log_path)
            return True
        if log_path:
            log("[NOTIF] Graph API retorno False. Intentando Outlook Desktop...", log_path)
    except Exception as exc:
        if log_path:
            log(f"[NOTIF] Graph API no disponible: {exc}. Intentando Outlook Desktop...", log_path)

    try:
        import win32com.client  # type: ignore

        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = "; ".join(recipients)
        if cc_recipients and not recipients_override:
            mail.CC = "; ".join(cc_recipients)
        mail.Subject = subject
        mail.HTMLBody = html_body
        mail.Send()
        if log_path:
            log("[NOTIF] Correo Productividad enviado via Outlook Desktop.", log_path)
        return True
    except Exception as exc:
        if log_path:
            log(f"[NOTIF] No se pudo enviar correo Productividad (Graph ni Outlook): {exc}", log_path)
        return False


def _productividad_status_chip(text: str, value: str) -> str:
    return (
        '<td style="padding:8px 14px;border:1px solid #cbd5e1;border-radius:18px;'
        f'background:#fff;font-size:12px;font-weight:bold;color:#243b53">{text}: {value}</td>'
    )


def count_normalized_rows(alias_archivo: str) -> int:
    """Returns data row count from latest normalized xlsx for alias_archivo.
    Returns -1 if no candidate found or on read error."""
    candidate = find_latest_normalized_candidate(alias_archivo)
    if candidate is None:
        return -1
    try:
        wb = load_workbook(candidate, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        _FOOTER_PREFIX = "El reporte"
        count = 0
        for row in ws.iter_rows(min_row=10, values_only=True):
            non_none = [c for c in row if c is not None]
            if not non_none:
                continue
            first = str(non_none[0]).strip()
            if first.startswith(_FOOTER_PREFIX):
                continue
            count += 1
        wb.close()
        return count
    except Exception:
        return -1


def build_productividad_summary_table(rows: Sequence[Dict[str, Any]]) -> str:
    body = ""
    for row in rows:
        estado = row["estado"]
        movimientos = row.get("movimientos")

        if estado == "FALLO":
            icono, bg = "&#10060; Fallo", "#fdecea"
            mov_html = "&#8212;"
        elif estado == "AL_DIA":
            icono, bg = "&#9989; Al d&iacute;a", "#eafaf1"
            mov_html = "&#8212;"
        elif estado == "SIN_DATOS":
            icono, bg = "&#8212; Sin movimientos", "#f5f5f5"
            mov_html = "0"
        elif estado == "PARCIAL":
            icono, bg = "&#9888;&#65039; Con observaciones", "#fef9e7"
            mov_html = f"{movimientos:,}".replace(",", ".") if movimientos is not None else "&#8212;"
        else:
            icono, bg = "&#9989; OK", "#eafaf1"
            mov_html = f"{movimientos:,}".replace(",", ".") if movimientos is not None else "&#8212;"

        body += f"""
        <tr style="background:{bg}">
          <td style="padding:8px 12px;border-bottom:1px solid #ddd;font-family:Calibri;font-size:13px;width:34%">{row['cliente']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #ddd;font-family:Calibri;font-size:13px;width:22%">{row.get('cd', '')}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #ddd;font-family:Calibri;font-size:13px;text-align:right;width:16%">{mov_html}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #ddd;font-family:Calibri;font-size:13px;font-weight:bold;width:28%">{icono}</td>
        </tr>"""

    return f"""<table style="border-collapse:collapse;width:100%;max-width:720px;table-layout:fixed">
      <colgroup>
        <col style="width:34%">
        <col style="width:22%">
        <col style="width:16%">
        <col style="width:28%">
      </colgroup>
      <thead>
        <tr style="background:#2c3e50;color:#fff">
          <th style="padding:10px 12px;text-align:left;font-family:Calibri;font-size:13px">Cliente</th>
          <th style="padding:10px 12px;text-align:left;font-family:Calibri;font-size:13px">CD</th>
          <th style="padding:10px 12px;text-align:right;font-family:Calibri;font-size:13px">Movimientos</th>
          <th style="padding:10px 12px;text-align:left;font-family:Calibri;font-size:13px">Resultado</th>
        </tr>
      </thead>
      <tbody>{body}</tbody>
    </table>"""


def build_productividad_closure_email(
    *,
    summary_rows: Sequence[Dict[str, Any]],
    active_clients_closed: int,
    log_file: Optional[Path] = None,
    generated_at: Optional[datetime] = None,
) -> tuple[str, str, Dict[str, Any]]:
    generated_at = generated_at or datetime.now()
    any_failures = any(row["estado"] in {"FALLO", "PARCIAL", "ERROR"} for row in summary_rows)
    overall_status = "CON_FALLOS" if any_failures else "OK"
    header_color = "#c0392b" if any_failures else "#27ae60"
    header_text = "&#10060; Proceso con incidencias relevantes" if any_failures else "&#9989; Proceso finalizado correctamente"
    subject = (
        f"[Productividad] Proceso con incidencias {generated_at.strftime('%d/%m/%Y')}"
        if any_failures
        else f"[Productividad] Proceso finalizado correctamente {generated_at.strftime('%d/%m/%Y')}"
    )
    table_html = build_productividad_summary_table(summary_rows)
    incidencias = [row["detalle"] for row in summary_rows if row["estado"] in {"FALLO", "PARCIAL"} and row.get("detalle")]
    incidencias_html = ""
    if incidencias:
        items = "".join(
            f"<li style='margin:0 0 6px 0'>{detalle}</li>" for detalle in incidencias
        )
        incidencias_html = f"""
        <div style="margin-top:14px;padding:12px 14px;background:#fff8e1;border:1px solid #f3d19c;border-radius:6px">
          <div style="font-size:13px;font-weight:bold;color:#8a5d00;margin-bottom:6px">Incidencias relevantes:</div>
          <ul style="margin:0 0 0 18px;padding:0;color:#5f370e;font-size:13px">{items}</ul>
        </div>"""

    html = f"""
    <html><body style="margin:0;padding:0;background:#f4f4f4;font-family:Calibri,Arial,sans-serif">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f4">
      <tr><td align="center" style="padding:16px">
        <table width="760" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:6px;border:1px solid #ddd">
          <tr>
            <td style="background:{header_color};padding:16px 20px;border-radius:6px 6px 0 0">
              <span style="color:#fff;font-size:18px;font-weight:bold">Modulo Productividad</span><br>
              <span style="color:#fff;font-size:14px">{header_text} &nbsp;|&nbsp; {generated_at.strftime('%d/%m/%Y')}</span>
            </td>
          </tr>
          <tr>
            <td style="padding:20px">
              <p style="margin:0 0 8px 0;color:#243b53;font-size:14px">
                <strong>Estado general:</strong> {"Proceso finalizado correctamente." if not any_failures else "Proceso con incidencias relevantes."}
              </p>
              <p style="margin:0 0 14px 0;color:#243b53;font-size:14px">
                Se complet&oacute; la actualizaci&oacute;n de la informaci&oacute;n de productividad de los clientes activos,
                dejando los archivos disponibles en SharePoint para su uso operativo y de gesti&oacute;n.
              </p>
              {table_html}
              {incidencias_html}
              <p style="color:#6b7280;font-size:11px;margin-top:16px">Notificación automática generada por Sistema Automatizado WMS Egakat.</p>
            </td>
          </tr>
        </table>
      </td></tr>
    </table>
    </body></html>"""

    payload = {
        "fecha": generated_at.strftime("%d/%m/%Y"),
        "hora": generated_at.strftime("%H:%M:%S"),
        "estado_global": overall_status,
        "clientes_activos_cerrados": active_clients_closed,
        "destino_oficial": "SharePoint",
        "staging": "Local separado",
        "tabla_html": table_html,
        "filas": list(summary_rows),
        "log": str(log_file) if log_file else "",
    }
    return subject, html, payload


def save_productividad_email_artifacts(
    *,
    subject: str,
    html_body: str,
    payload: Dict[str, Any],
) -> Dict[str, Path]:
    ensure_runtime_dirs()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_path = Path(LOG_DIR) / f"productividad_email_{stamp}.html"
    json_path = Path(LOG_DIR) / f"productividad_email_{stamp}.json"
    html_path.write_text(html_body, encoding="utf-8")
    json_path.write_text(
        json.dumps({"subject": subject, **payload}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return {"html_path": html_path, "json_path": json_path}


def quarantine_file(path: Path, reason: str, log_path: Optional[Path] = None) -> Path:
    ensure_runtime_dirs()
    target = Path(QUARANTINE_DIR) / f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
    target.write_bytes(path.read_bytes())
    if log_path:
        log(f"[CUARENTENA] Archivo enviado a cuarentena por '{reason}': {target}", log_path)
    return target
