"""
Helpers del modulo Productividad.

El foco actual es dejar la base del proyecto lista y centralizar:
- reglas de rango
- paths historicos
- validacion del Excel descargado
- logging

La navegacion WMS sigue pendiente de confirmacion runtime.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable, List, Optional, Sequence

from openpyxl import load_workbook

from productividad_config import (
    CLIENTS,
    DOWNLOAD_DIR,
    EXPECTED_HEADERS,
    LOG_DIR,
    MONTH_FOLDERS,
    PRODUCTIVIDAD_ROOT,
    RANGE_END_TIME,
    RANGE_START_TIME,
    VALID_SHEET_NAMES,
)


@dataclass
class ReportingWindow:
    mode: str
    from_dt: datetime
    to_dt: datetime
    target_year: int
    target_month: int


@dataclass
class WorkbookInspection:
    path: Path
    sheet_name: str
    internal_cd: str
    internal_company: str
    internal_scope: str
    headers: List[str]
    has_data_rows: bool
    first_non_empty_after_header: Optional[str] = None


@dataclass
class ValidationResult:
    ok: bool
    is_empty_valid: bool
    target_path: Path
    critical_errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    inspection: Optional[WorkbookInspection] = None


def ensure_runtime_dirs() -> None:
    Path(LOG_DIR).mkdir(parents=True, exist_ok=True)
    Path(DOWNLOAD_DIR).mkdir(parents=True, exist_ok=True)


def build_log_path(prefix: str = "productividad") -> Path:
    ensure_runtime_dirs()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(LOG_DIR) / f"{prefix}_{stamp}.log"


def log(message: str, log_path: Optional[Path] = None) -> None:
    print(message, flush=True)
    if log_path:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as fh:
            fh.write(message + "\n")


def normalize_text(value: object) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip().upper()


def parse_clock(value: str) -> time:
    return datetime.strptime(value, "%H:%M:%S").time()


def month_folder_name(year: int, month: int) -> str:
    if month not in MONTH_FOLDERS:
        raise ValueError(f"Mes invalido: {month}")
    return MONTH_FOLDERS[month]


def build_reporting_window(
    *,
    mode: str,
    target_year: int,
    target_month: int,
    execution_dt: Optional[datetime] = None,
) -> ReportingWindow:
    execution_dt = execution_dt or datetime.now()
    start_dt = datetime.combine(date(target_year, target_month, 1), parse_clock(RANGE_START_TIME))

    if mode == "current":
        if execution_dt.year != target_year or execution_dt.month != target_month:
            raise ValueError("El modo 'current' requiere target_year/target_month del mes en ejecucion.")
        end_dt = datetime.combine(execution_dt.date(), parse_clock(RANGE_END_TIME))
    elif mode == "closed":
        if target_month == 12:
            next_month = date(target_year + 1, 1, 1)
        else:
            next_month = date(target_year, target_month + 1, 1)
        end_dt = datetime.combine(next_month, parse_clock(RANGE_END_TIME))
    else:
        raise ValueError("mode debe ser 'current' o 'closed'.")

    return ReportingWindow(
        mode=mode,
        from_dt=start_dt,
        to_dt=end_dt,
        target_year=target_year,
        target_month=target_month,
    )


def build_target_path(client: dict, target_year: int, target_month: int) -> Path:
    month_folder = month_folder_name(target_year, target_month)
    filename = f"{client['alias_archivo']}.xlsx"
    return Path(PRODUCTIVIDAD_ROOT) / client["carpeta_destino_historica"] / str(target_year) / month_folder / filename


def find_client(cd: str, alias_archivo: str) -> dict:
    cd_norm = normalize_text(cd)
    alias_norm = normalize_text(alias_archivo)
    for client in CLIENTS:
        if normalize_text(client["cd"]) == cd_norm and normalize_text(client["alias_archivo"]) == alias_norm:
            return client
    raise KeyError(f"No existe configuracion para cd='{cd}' alias='{alias_archivo}'.")


def inspect_workbook(path: Path) -> WorkbookInspection:
    wb = load_workbook(path, read_only=True, data_only=False)
    sheet_name = wb.sheetnames[0]
    if sheet_name in VALID_SHEET_NAMES:
        selected = sheet_name
    else:
        for candidate in wb.sheetnames:
            if candidate in VALID_SHEET_NAMES:
                selected = candidate
                break
        else:
            selected = sheet_name

    ws = wb[selected]
    headers = [str(ws.cell(row=9, column=i).value or "").strip() for i in range(1, len(EXPECTED_HEADERS) + 1)]

    has_data_rows = False
    first_non_empty_after_header = None
    for row_index in range(10, 25):
        cell_a = ws.cell(row=row_index, column=1).value
        cell_b = ws.cell(row=row_index, column=2).value
        if cell_a not in (None, "") or cell_b not in (None, ""):
            first_non_empty_after_header = normalize_text(cell_a)
            if str(cell_a or "").strip().isdigit():
                has_data_rows = True
            break

    return WorkbookInspection(
        path=path,
        sheet_name=selected,
        internal_cd=normalize_text(ws["A1"].value),
        internal_company=normalize_text(ws["A4"].value),
        internal_scope=normalize_text(ws["A5"].value),
        headers=headers,
        has_data_rows=has_data_rows,
        first_non_empty_after_header=first_non_empty_after_header,
    )


def _infer_alias_candidates(internal_cd: str, internal_company: str) -> List[str]:
    matches = []
    for client in CLIENTS:
        if normalize_text(client["empresa_wms"]) != internal_company:
            continue
        if normalize_text(client["cd"]).replace("CD ", "") != internal_cd:
            continue
        matches.append(client["alias_archivo"])
    return matches


def validate_downloaded_workbook(
    workbook_path: Path,
    client: dict,
    target_year: int,
    target_month: int,
) -> ValidationResult:
    inspection = inspect_workbook(workbook_path)
    target_path = build_target_path(client, target_year, target_month)
    critical_errors: List[str] = []
    warnings: List[str] = []

    expected_cd_internal = normalize_text(client["cd"]).replace("CD ", "")
    expected_company = normalize_text(client["empresa_wms"])
    expected_headers = [normalize_text(value) for value in EXPECTED_HEADERS]
    actual_headers = [normalize_text(value) for value in inspection.headers]

    if inspection.sheet_name not in VALID_SHEET_NAMES:
        critical_errors.append(
            f"Hoja invalida: '{inspection.sheet_name}'. Se esperaba una de {list(VALID_SHEET_NAMES)}."
        )

    if inspection.internal_cd != expected_cd_internal:
        critical_errors.append(
            f"CD inconsistente: esperado '{expected_cd_internal}', obtenido '{inspection.internal_cd}'."
        )

    if inspection.internal_company != expected_company:
        critical_errors.append(
            f"Empresa inconsistente: esperado '{expected_company}', obtenido '{inspection.internal_company}'."
        )

    inferred_aliases = _infer_alias_candidates(inspection.internal_cd, inspection.internal_company)
    if inferred_aliases and client["alias_archivo"] not in inferred_aliases:
        critical_errors.append(
            f"Alias inconsistente: esperado '{client['alias_archivo']}', inferido {inferred_aliases} desde el Excel."
        )
    elif not inferred_aliases:
        warnings.append("No se pudo inferir alias unico desde el contenido interno del Excel.")

    if actual_headers != expected_headers:
        critical_errors.append("Encabezados historicos no coinciden con la estructura esperada de Productividad.")

    expected_scope = normalize_text(client["deposito_wms_origen"])
    if inspection.internal_scope and inspection.internal_scope not in {expected_scope, expected_cd_internal}:
        warnings.append(
            f"Scope interno '{inspection.internal_scope}' no coincide limpiamente con el deposito esperado '{expected_scope}'."
        )

    if not inspection.has_data_rows:
        warnings.append("Archivo valido sin movimientos: tratar como vacio, no como fallo.")

    return ValidationResult(
        ok=not critical_errors,
        is_empty_valid=not critical_errors and not inspection.has_data_rows,
        target_path=target_path,
        critical_errors=critical_errors,
        warnings=warnings,
        inspection=inspection,
    )


def build_catalog_table_rows(clients: Iterable[dict]) -> List[str]:
    rows = []
    for client in clients:
        rows.append(
            " | ".join(
                [
                    client["cd"],
                    client["alias_archivo"],
                    client["empresa_wms"],
                    client["deposito_wms_origen"],
                    client["carpeta_destino_historica"],
                    str(client["active"]),
                ]
            )
        )
    return rows


def format_window(window: ReportingWindow) -> str:
    return (
        f"{window.mode}: {window.from_dt.strftime('%Y-%m-%d %H:%M:%S')} -> "
        f"{window.to_dt.strftime('%Y-%m-%d %H:%M:%S')}"
    )
