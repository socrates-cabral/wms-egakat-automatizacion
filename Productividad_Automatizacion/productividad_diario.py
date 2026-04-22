"""
productividad_diario.py — Descarga diaria incremental con append+dedup a SharePoint.

Flujo por cliente:
  1. Leer checkpoint → calcular ventana from_dt/to_dt
  2. Descargar Excel WMS para esa ventana (Playwright, reutiliza logica de productividad_descarga.py)
  3. Extraer filas de datos del chunk descargado
  4. Descargar archivo existente en SharePoint
  5. Concatenar + dedup por clave compuesta
  6. Subir resultado de vuelta a SharePoint
  7. Actualizar checkpoint

Manejo de mes cruzado: si la ventana cruza fin de mes, las filas se segregan
por Fecha.month y se suben a los archivos del mes correspondiente.
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

import io
import json
import os
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright

# .env desde la raiz del repo (no leer el del directorio actual)
load_dotenv(dotenv_path=Path(__file__).parent.parent / ".env")

from productividad_config import (
    CONTROLLED_LIGHTWEIGHT_CLIENTS,
    CONTROLLED_HEAVY_CLIENTS,
    DOWNLOAD_DIR,
    LOG_DIR,
    MONTH_FOLDERS,
    RANGE_END_TIME,
    RANGE_START_TIME,
    SHAREPOINT_PRODUCTIVIDAD_ROOT,
    WMS_LOGIN_URL,
)
from productividad_utils import (
    build_log_path,
    build_sharepoint_folder_path,
    download_sharepoint_file_bytes,
    get_sharepoint_file_state,
    load_azure_graph_module,
    log,
    normalize_text,
    parse_legacy_excel_html,
    upload_bytes_to_sharepoint,
    send_html_notification,
    build_productividad_closure_email,
    save_productividad_email_artifacts,
)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

# Modo prueba: True = solo envia correo a TESTING_EMAIL, no a destinatarios reales.
# Cambiar a False cuando el script esté validado en producción.
TESTING_MODE = True
TESTING_EMAIL = "socrates.cabral@egakat.cl"

_SCRIPT_DIR = Path(__file__).resolve().parent
_LOGS_DIR = _SCRIPT_DIR / "logs"
_LOCKFILE = _LOGS_DIR / "productividad_diario_run.lock"
_CHECKPOINT_FILE = _LOGS_DIR / "productividad_diario_checkpoint.json"
_DOWNLOAD_DIR = Path(DOWNLOAD_DIR)

# Clave compuesta que identifica un movimiento unico para dedup.
# Combinacion de comprobante + articulo + fecha + hora + numero es suficientemente selectiva.
_DEDUP_KEY = ["Comprobante", "Comprobante externo", "Artículo", "Fecha", "Hora", "Número"]

# Clientes que procesa este script: todos los livianos + derco (heavy).
# DERCO se maneja igual que los livianos porque la ventana diaria es siempre pequena.
_DAILY_CLIENTS: Dict[str, dict] = {
    **CONTROLLED_LIGHTWEIGHT_CLIENTS,
    **CONTROLLED_HEAVY_CLIENTS,
}

# Checkpoint inicial para el primer arranque.
# Formato: {client_key: "YYYY-MM-DD"} — fecha del ultimo dia procesado exitosamente.
# La ventana de la primera ejecucion sera: esta_fecha_08:00 → hoy_06:00.
_CHECKPOINT_SEED = {
    "abinbev":            "2026-04-20",
    "bha":                "2026-04-20",
    "daikin":             "2026-04-17",
    "pochteca":           "2026-04-20",
    "mascota_quilicura":  "2026-04-20",
    "derco":              "2026-04-21",
    "barentz":            "2026-04-20",
    "buraschi":           "2026-04-20",
    "cepas_chile":        "2026-04-20",
    "collico":            "2026-04-20",
    "delibest":           "2026-04-20",
    "intime":             "2026-04-20",
    "runo":               "2026-04-20",
    "tresmontes":         "2026-04-20",
    "unilever":           "2026-04-20",
    "nativo_drinks":      "2026-04-21",
    "omnitech":           "2026-04-21",
}

# ---------------------------------------------------------------------------
# Lock — evita ejecuciones simultaneas
# ---------------------------------------------------------------------------

def _pid_alive(pid: int) -> bool:
    try:
        import subprocess
        result = subprocess.run(
            ["tasklist", "/FI", f"PID eq {pid}", "/NH"],
            capture_output=True, text=True, timeout=5,
        )
        return str(pid) in result.stdout
    except Exception:
        return False


def _acquire_lock(log_path: Path) -> bool:
    _LOGS_DIR.mkdir(exist_ok=True)
    if _LOCKFILE.exists():
        try:
            pid = int(_LOCKFILE.read_text().strip())
            if _pid_alive(pid):
                log(f"[ERROR] Ya hay una instancia corriendo (PID {pid}). Abortando.", log_path)
                return False
        except Exception:
            pass
        _LOCKFILE.unlink(missing_ok=True)
    _LOCKFILE.write_text(str(os.getpid()))
    return True


def _release_lock() -> None:
    try:
        _LOCKFILE.unlink(missing_ok=True)
    except Exception:
        pass

# ---------------------------------------------------------------------------
# Checkpoint — fecha de ultimo run exitoso por cliente
# ---------------------------------------------------------------------------

def _load_checkpoint() -> Dict[str, str]:
    """Retorna {client_key: "YYYY-MM-DD"}.
    Si el archivo no existe, crea uno con los valores semilla.
    """
    if not _CHECKPOINT_FILE.exists():
        _LOGS_DIR.mkdir(parents=True, exist_ok=True)
        _CHECKPOINT_FILE.write_text(
            json.dumps(_CHECKPOINT_SEED, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return dict(_CHECKPOINT_SEED)
    try:
        return json.loads(_CHECKPOINT_FILE.read_text(encoding="utf-8"))
    except Exception:
        return dict(_CHECKPOINT_SEED)


def _save_checkpoint_client(client_key: str, run_date: date) -> None:
    """Persiste la fecha exitosa de to_dt.date() para el cliente.
    Solo se llama si todo el pipeline del cliente termino sin error.
    """
    data = _load_checkpoint()
    data[client_key] = run_date.isoformat()
    _CHECKPOINT_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

# ---------------------------------------------------------------------------
# Feriados Chile
# ---------------------------------------------------------------------------

def _load_feriados() -> Dict[date, str]:
    """Lee Tabla Feriados.xlsx desde OneDrive, retorna {fecha: nombre_festividad}."""
    feriados_path = (
        r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA"
        r"\Datos para Dashboard - Productividad\Tabla Feriados.xlsx"
    )
    try:
        df = pd.read_excel(feriados_path, sheet_name="T_Feriados")
        result: Dict[date, str] = {}
        for _, row in df.iterrows():
            dia = row["Día"]
            festividad = str(row.get("Festividad", "")).strip()
            if pd.isna(dia):
                continue
            if isinstance(dia, datetime):
                dia = dia.date()
            elif not isinstance(dia, date):
                try:
                    dia = pd.to_datetime(dia).date()
                except Exception:
                    continue
            result[dia] = festividad
        return result
    except Exception:
        return {}


def _is_feriado(target_date: date, feriados: Dict[date, str]) -> Optional[str]:
    """Retorna nombre de festividad si target_date es feriado, None si no."""
    return feriados.get(target_date)

# ---------------------------------------------------------------------------
# Ventana de descarga
# ---------------------------------------------------------------------------

def _first_business_day_of_month(year: int, month: int) -> date:
    """Primer dia habil (lun-vie) del mes — fallback si el cliente no tiene checkpoint."""
    d = date(year, month, 1)
    while d.weekday() >= 5:  # 5=sabado, 6=domingo
        d += timedelta(days=1)
    return d


def _compute_window(client_key: str, checkpoint: Dict[str, str]) -> Tuple[datetime, datetime]:
    """Calcula from_dt y to_dt para la ventana de descarga.

    from_dt = last_run_date del checkpoint a las 08:00:00
    to_dt   = hoy a las 06:00:00

    El gap 06:00-08:00 es intencional: zona de transicion de turno donde el WMS
    puede estar en estado inconsistente.

    Para el lunes: si last_run fue viernes, from_dt = viernes 08:00 y
    to_dt = lunes 06:00, lo que cubre automaticamente sabado y domingo.
    El mismo mecanismo cubre feriados anteriores — el gap se cierra solo.
    """
    today = datetime.now().date()
    start_clock = datetime.strptime(RANGE_START_TIME, "%H:%M:%S").time()
    end_clock = datetime.strptime(RANGE_END_TIME, "%H:%M:%S").time()

    if client_key in checkpoint:
        last_run = date.fromisoformat(checkpoint[client_key])
    else:
        last_run = _first_business_day_of_month(today.year, today.month)

    from_dt = datetime.combine(last_run, start_clock)
    to_dt = datetime.combine(today, end_clock)
    return from_dt, to_dt

# ---------------------------------------------------------------------------
# Playwright — descarga desde WMS
# ---------------------------------------------------------------------------

def _load_wms_credentials() -> Tuple[str, str]:
    wms_user = os.getenv("WMS_USUARIO") or "SCABRAL"
    wms_password = os.getenv("WMS_PASSWORD") or os.getenv("WMS_CLAVE") or ""
    if not wms_password:
        raise RuntimeError("No se encontro credencial WMS en .env.")
    return wms_user, wms_password


def _select_option_by_label(page, label: str, field_name: str, log_path: Path) -> None:
    selects = page.locator("select")
    observed_options = []
    for index in range(selects.count()):
        locator = selects.nth(index)
        options = [" ".join(t.split()).strip() for t in locator.locator("option").all_inner_texts()]
        observed_options.append(f"select[{index}] -> {options}")
        if label in options:
            locator.select_option(label=label)
            log(f"[OK] Selector {field_name}: select[{index}] -> {label}", log_path)
            return
    for observed in observed_options:
        log(f"[DEBUG] Opciones {field_name}: {observed}", log_path)
    raise RuntimeError(f"No se encontro opcion '{label}' para {field_name}.")


def _wait_for_option(page, label: str, field_name: str, log_path: Path,
                     poll_ms: int = 500, timeout_ms: int = 45_000) -> None:
    """Polling hasta que la opcion aparezca en algun <select> (dependencia AJAX del WMS)."""
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        selects = page.locator("select")
        for index in range(selects.count()):
            options = [" ".join(t.split()).strip()
                       for t in selects.nth(index).locator("option").all_inner_texts()]
            if label in options:
                log(f"[OK] Opcion '{label}' disponible en select[{index}] tras espera AJAX.", log_path)
                return
        page.wait_for_timeout(poll_ms)
    log(f"[WARN] Timeout {timeout_ms}ms esperando '{label}' en {field_name}. Intentando igual.", log_path)


def _fill_first(page, selectors: List[str], value: str, field_name: str, log_path: Path) -> None:
    for selector in selectors:
        locator = page.locator(selector)
        if locator.count() == 0:
            continue
        locator.first.fill("")
        locator.first.fill(value)
        locator.first.press("Tab")
        log(f"[OK] Campo {field_name}: {selector} -> {value}", log_path)
        return
    raise RuntimeError(f"No se encontro campo para {field_name}.")


def _download_wms_export(
    *,
    client: dict,
    from_dt: datetime,
    to_dt: datetime,
    log_path: Path,
    alias_prefix: str,
) -> Path:
    """Navega el WMS via Playwright y descarga el Excel de movimientos.
    Retorna la ruta del archivo .xls HTML descargado.
    """
    wms_user, wms_password = _load_wms_credentials()
    from_date = from_dt.strftime("%d/%m/%y")
    to_date = to_dt.strftime("%d/%m/%y")
    from_time = from_dt.strftime("%H:%M:%S")
    to_time = to_dt.strftime("%H:%M:%S")

    html_name = f"diario_{alias_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
    html_path = _DOWNLOAD_DIR / html_name
    _DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    is_heavy = bool(client.get("heavy_client"))
    download_timeout_ms = int(client.get("download_timeout_ms", 180_000))

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True, slow_mo=0)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        page.set_default_timeout(60_000)

        try:
            page.goto(WMS_LOGIN_URL, wait_until="load", timeout=60_000)
            page.fill("input[name='vUSR']", wms_user)
            page.fill("input[name='vPASSWORD']", wms_password)
            page.click("input[name='BUTTON3']")
            page.wait_for_load_state("load", timeout=60_000)
            page.wait_for_timeout(1500)
            log("[OK] Login WMS completado.", log_path)

            _select_option_by_label(page, client["deposito_wms_origen"], "deposito inicial", log_path)
            page.click("input[value='Aceptar']")
            page.wait_for_load_state("load", timeout=60_000)
            page.wait_for_timeout(1500)

            page.click("text=Procesos WMS")
            page.wait_for_load_state("load", timeout=60_000)
            page.wait_for_timeout(1000)

            page.get_by_role("link", name="Movimientos por Operación").click()
            page.wait_for_load_state("load", timeout=60_000)
            page.wait_for_timeout(1500)
            log("[OK] Pantalla Movimientos por Operacion cargada.", log_path)

            _select_option_by_label(page, client["deposito_wms_origen"], "deposito", log_path)
            page.wait_for_timeout(1500)
            _select_option_by_label(page, client["empresa_wms"], "empresa", log_path)
            # WMS recarga tipo-operacion y cuenta via AJAX tras cambio de empresa
            page.wait_for_timeout(2000)
            _select_option_by_label(page, "ORDEN DE PREP. C/STOCK", "tipo de operacion", log_path)
            _wait_for_option(page, "Stock Físico", "cuenta", log_path)
            _select_option_by_label(page, "Stock Físico", "cuenta", log_path)
            page.wait_for_timeout(500)

            _fill_first(page, ["input[name='vFECDESDE']"], from_date, "desde fecha", log_path)
            _fill_first(page, ["input[name='vFECHASTA']"], to_date, "hasta fecha", log_path)
            _fill_first(page, ["input[name='vHORADES']"], from_time, "desde hora", log_path)
            _fill_first(page, ["input[name='vHORAHAS']"], to_time, "hasta hora", log_path)

            page.click("input[name='CONFIRMAR']")
            # Espera dinamica + minimo 15s: el WMS puede tardar en habilitar el boton Excel.
            try:
                page.wait_for_load_state("networkidle", timeout=120_000)
            except Exception:
                pass
            page.wait_for_timeout(15_000)
            log(f"[OK] Consulta ejecutada: {from_date} {from_time} -> {to_date} {to_time}", log_path)

            if is_heavy:
                # DERCO: WMS sirve el Excel via URL directa (no como attachment).
                # Interceptamos la request post-click y descargamos directamente.
                xls_urls: List[str] = []

                def _on_request(req) -> None:
                    url: str = req.url
                    if any(x in url.lower() for x in (".xls", "excel", "salidaexcel", "download", "export")):
                        xls_urls.append(url)

                context.on("request", _on_request)
                page.locator("img#W0155SALIDAEXCEL").click()

                for _ in range(600):  # poll hasta 60s
                    if xls_urls:
                        break
                    page.wait_for_timeout(100)

                context.remove_listener("request", _on_request)

                if not xls_urls:
                    raise RuntimeError("No se capturo URL del Excel tras 60s — WMS no genero el archivo.")

                url_excel = xls_urls[-1]
                log(f"[OK] URL Excel interceptada: {url_excel}", log_path)
                response = page.request.get(url_excel, timeout=download_timeout_ms)
                html_path.write_bytes(response.body())
            else:
                with page.expect_download(timeout=download_timeout_ms) as dl_info:
                    page.locator("img#W0155SALIDAEXCEL").click()
                dl_info.value.save_as(str(html_path))

            log(f"[OK] Export WMS descargado: {html_path}", log_path)
            return html_path

        finally:
            context.close()
            browser.close()

# ---------------------------------------------------------------------------
# Extraccion de DataFrame desde el .xls HTML del WMS
# ---------------------------------------------------------------------------

def _parse_wms_html_to_df(html_path: Path, log_path: Path) -> pd.DataFrame:
    """Parsea el .xls HTML y retorna DataFrame con filas de detalle.
    Reutiliza parse_legacy_excel_html de productividad_utils (misma funcion que usa descarga.py).
    """
    parsed = parse_legacy_excel_html(html_path)
    headers = parsed["headers"]
    detail = parsed["detail_records"]

    if not detail:
        log("[OK] Chunk WMS sin filas de detalle (ventana sin movimientos).", log_path)
        return pd.DataFrame(columns=headers)

    # Deduplicar nombres de columna — el WMS repite "Artículo" en pos 1 y 21.
    # Pandas devuelve DataFrame en vez de Series al acceder columnas duplicadas,
    # lo que rompe .str en pasos posteriores.
    seen: dict = {}
    unique_headers = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}.{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)

    # Ajustar columnas si el detalle tiene distinto ancho que el header
    n_cols = len(unique_headers)
    aligned = [row[:n_cols] + [""] * max(0, n_cols - len(row)) for row in detail]
    df = pd.DataFrame(aligned, columns=unique_headers)

    # Fecha como datetime para poder filtrar por mes en el cruce
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")

    # Normalizar strings de clave de dedup
    for col in ["Comprobante", "Número", "Hora", "Artículo"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    rows_total = len(df)
    try:
        col = df["Comprobante"]
        if isinstance(col, pd.DataFrame):
            col = col.iloc[:, 0]
        rows_valid = col.astype(str).str.match(r"^\d+$").sum()
    except Exception:
        rows_valid = rows_total
    log(f"[OK] Chunk WMS: {rows_total} filas | {rows_valid} con Comprobante numerico.", log_path)
    return df

# ---------------------------------------------------------------------------
# SharePoint: descargar archivo existente como DataFrame
# ---------------------------------------------------------------------------

def _download_sharepoint_df(
    *,
    token: str,
    drive_id: str,
    folder_path: str,
    filename: str,
    log_path: Path,
) -> Tuple[Optional[pd.DataFrame], Optional[bytes]]:
    """Descarga el archivo xlsx de SharePoint y lo parsea.
    Retorna (DataFrame_filas_datos, bytes_crudos).
    Si el archivo no existe retorna (None, None).
    """
    remote_state = get_sharepoint_file_state(token, drive_id, folder_path, filename)
    if not remote_state.exists:
        log(f"[INFO] No existe en SharePoint: {folder_path}/{filename}", log_path)
        return None, None

    raw = download_sharepoint_file_bytes(token, drive_id, remote_state)
    log(f"[OK] Descargado de SharePoint: {filename} ({len(raw):,} bytes)", log_path)

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # Header fila 9 via iter_rows — evita random-access lento en read_only
        header_row = next(ws.iter_rows(min_row=9, max_row=9, values_only=True), ())
        header_raw = [str(v or "").strip() for v in header_row]
        last_h = max((i for i, v in enumerate(header_raw, 1) if v), default=0) or 50
        raw_headers = header_raw[:last_h]

        # Deduplicar nombres igual que en _parse_wms_html_to_df
        seen: dict = {}
        headers = []
        for h in raw_headers:
            if h in seen:
                seen[h] += 1
                headers.append(f"{h}.{seen[h]}")
            else:
                seen[h] = 0
                headers.append(h)

        # iter_rows en vez de range(ws.max_row) — evita iterar 1M filas vacías
        # Corte automático tras 5 filas vacías consecutivas
        rows_data = []
        empty_streak = 0
        for row_vals in ws.iter_rows(min_row=10, max_col=last_h, values_only=True):
            row = list(row_vals)
            non_none = [v for v in row if v is not None and str(v).strip() != ""]
            if not non_none:
                empty_streak += 1
                if empty_streak >= 5:
                    break
                continue
            empty_streak = 0
            first_str = str(row[0] or "").strip()
            if first_str.startswith("El reporte"):
                continue
            rows_data.append(row)

        wb.close()

        df = pd.DataFrame(rows_data, columns=headers) if rows_data else pd.DataFrame(columns=headers)

        if "Fecha" in df.columns:
            df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")

        log(f"[OK] DataFrame SharePoint leido: {len(df)} filas existentes.", log_path)
        return df, raw

    except Exception as exc:
        log(f"[WARN] No se pudo parsear Excel de SharePoint: {exc}. Se tratara como vacio.", log_path)
        return pd.DataFrame(), raw

# ---------------------------------------------------------------------------
# Dedup y merge
# ---------------------------------------------------------------------------

def _dedup_merge(
    df_existing: Optional[pd.DataFrame],
    df_new: pd.DataFrame,
    log_path: Path,
) -> pd.DataFrame:
    """Concatena df_existing + df_new y elimina duplicados por _DEDUP_KEY.

    keep='first' significa que las filas existentes tienen prioridad sobre las nuevas,
    lo cual es correcto: si un movimiento ya esta en SharePoint, no se altera.
    """
    if df_existing is None or df_existing.empty:
        return df_new.copy()
    if df_new.empty:
        return df_existing.copy()

    # Alinear columnas — el nuevo chunk puede omitir columnas vacias
    all_cols = list(df_existing.columns)
    for col in df_new.columns:
        if col not in all_cols:
            all_cols.append(col)

    df_combined = pd.concat(
        [df_existing.reindex(columns=all_cols), df_new.reindex(columns=all_cols)],
        ignore_index=True,
    )

    dedup_cols = [c for c in _DEDUP_KEY if c in df_combined.columns]
    if dedup_cols:
        # Normalizar a string para evitar comparaciones de tipo mixto (datetime vs string)
        key_df = df_combined[dedup_cols].astype(str)
        before = len(df_combined)
        mask_dup = key_df.duplicated(keep="first")
        df_combined = df_combined[~mask_dup].reset_index(drop=True)
        removed = before - len(df_combined)
        if removed:
            log(f"[DEDUP] {removed} filas duplicadas descartadas (clave: {dedup_cols}).", log_path)
    else:
        log("[WARN] Ninguna columna de dedup encontrada — concatenando sin dedup.", log_path)

    return df_combined

# ---------------------------------------------------------------------------
# Serializar DataFrame a xlsx con portada del WMS
# ---------------------------------------------------------------------------

def _build_portada_rows(client: dict, year: int, month: int) -> List[List]:
    """Genera las 8 filas de portada (filas 1-8) con el mismo formato que produce el WMS."""
    mes_nombre = MONTH_FOLDERS.get(month, str(month))
    return [
        [normalize_text(client["deposito_wms_origen"])],   # row 1
        ["INFORME DE MOVIMIENTOS"],                         # row 2
        [],                                                 # row 3 vacia
        [client["empresa_wms"]],                            # row 4
        [client["deposito_wms_origen"]],                    # row 5
        [f"{mes_nombre} {year}"],                           # row 6 periodo
        [],                                                 # row 7 vacia
        [],                                                 # row 8 vacia
    ]


def _df_to_xlsx_bytes(df: pd.DataFrame, client: dict, year: int, month: int) -> bytes:
    """Serializa el DataFrame a bytes xlsx con estructura WMS:
    filas 1-8 portada, fila 9 headers, fila 10+ datos.
    Fechas se escriben como string DD/MM/YYYY para mantener compatibilidad con Power Query.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Movimientos"

    for r_idx, row_vals in enumerate(_build_portada_rows(client, year, month), start=1):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx).value = val

    for c_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=9, column=c_idx).value = col_name

    for r_offset, (_, row) in enumerate(df.iterrows(), start=0):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, (datetime, pd.Timestamp)) and not pd.isna(val):
                val = val.strftime("%d/%m/%Y")
            ws.cell(row=10 + r_offset, column=c_idx).value = val

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ---------------------------------------------------------------------------
# Subida a SharePoint con backup previo
# ---------------------------------------------------------------------------

def _upload_df_to_sharepoint(
    *,
    token: str,
    drive_id: str,
    folder_path: str,
    filename: str,
    df: pd.DataFrame,
    client: dict,
    year: int,
    month: int,
    existing_bytes: Optional[bytes],
    log_path: Path,
) -> bool:
    payload = _df_to_xlsx_bytes(df, client, year, month)
    ok = upload_bytes_to_sharepoint(token, drive_id, folder_path, filename, payload)
    if ok:
        log(f"[OK] Subido: {folder_path}/{filename} ({len(payload):,} bytes | {len(df):,} filas)", log_path)
    else:
        log(f"[FALLO] upload_bytes_to_sharepoint fallo: {folder_path}/{filename}", log_path)
    return ok

# ---------------------------------------------------------------------------
# Manejo de mes cruzado
# ---------------------------------------------------------------------------

def _split_by_month(df: pd.DataFrame) -> Dict[Tuple[int, int], pd.DataFrame]:
    """Segrega el DataFrame por (year, month) segun la columna Fecha.
    Retorna {(year, month): df_subset}.
    Filas con Fecha nula se omiten (no se pueden asignar a un mes).
    """
    if "Fecha" not in df.columns or df.empty:
        return {}

    result: Dict[Tuple[int, int], pd.DataFrame] = {}
    df_valid = df.dropna(subset=["Fecha"]).copy()
    for (y, m), subset in df_valid.groupby(
        [df_valid["Fecha"].dt.year, df_valid["Fecha"].dt.month]
    ):
        result[(int(y), int(m))] = subset.reset_index(drop=True)
    return result

# ---------------------------------------------------------------------------
# Proceso completo de un cliente
# ---------------------------------------------------------------------------

def _process_client(
    *,
    client_key: str,
    client: dict,
    checkpoint: Dict[str, str],
    feriados: Dict[date, str],
    token: str,
    drive_id: str,
    log_path: Path,
) -> dict:
    """Ejecuta el pipeline completo para un cliente.
    Retorna {"ok": bool, "filas_nuevas": int, "filas_totales": int, "detalle": str}.
    """
    today = datetime.now().date()
    from_dt, to_dt = _compute_window(client_key, checkpoint)

    if from_dt >= to_dt:
        log(f"[SKIP] {client_key}: ventana vacia (from={from_dt} >= to={to_dt}).", log_path)
        return {"ok": True, "filas_nuevas": 0, "filas_totales": 0, "detalle": "ventana vacia"}

    log(
        f"[DIARIO] {client['empresa_wms']} | "
        f"{from_dt.strftime('%Y-%m-%d %H:%M')} → {to_dt.strftime('%Y-%m-%d %H:%M')}",
        log_path,
    )

    # 1) Descargar del WMS con hasta 3 intentos
    html_path: Optional[Path] = None
    MAX_ATTEMPTS = 3
    RETRY_WAIT_S = 30

    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            html_path = _download_wms_export(
                client=client,
                from_dt=from_dt,
                to_dt=to_dt,
                log_path=log_path,
                alias_prefix=client["alias_archivo"],
            )
            break
        except Exception as exc:
            retriable = (
                "WMS_EMPRESA_TODAS" in str(exc)
                or "TimeoutError" in type(exc).__name__
                or "Timeout" in str(exc)
                or "No se encontro" in str(exc)
            )
            if retriable and attempt < MAX_ATTEMPTS:
                log(
                    f"[WARN] Intento {attempt}/{MAX_ATTEMPTS} fallo (retriable): {exc}. "
                    f"Reintentando en {RETRY_WAIT_S}s...",
                    log_path,
                )
                time.sleep(RETRY_WAIT_S)
            else:
                log(f"[FALLO] {client_key} intento {attempt}: {exc}", log_path)
                return {"ok": False, "filas_nuevas": 0, "filas_totales": 0, "detalle": str(exc)}

    if html_path is None:
        return {"ok": False, "filas_nuevas": 0, "filas_totales": 0, "detalle": "descarga fallida"}

    # 2) Parsear chunk
    try:
        df_new = _parse_wms_html_to_df(html_path, log_path)
    except Exception as exc:
        log(f"[FALLO] Parseando HTML WMS: {exc}", log_path)
        return {"ok": False, "filas_nuevas": 0, "filas_totales": 0, "detalle": str(exc)}
    finally:
        if html_path and html_path.exists():
            html_path.unlink(missing_ok=True)

    filas_nuevas_total = len(df_new)

    # 3) Determinar si la ventana cruza un mes.
    # Agrupamos las filas nuevas por mes para escribirlas al archivo correcto.
    # Si no cruza mes, todo va al archivo del mes de from_dt.
    m_from = (from_dt.year, from_dt.month)
    m_to = (to_dt.year, to_dt.month)
    months_to_process = [m_from] if m_from == m_to else [m_from, m_to]

    new_by_month = _split_by_month(df_new) if not df_new.empty else {}
    # Si no hay columna Fecha o no se pudo segmentar, poner todo en el mes de from_dt
    if not new_by_month and not df_new.empty:
        new_by_month = {m_from: df_new}

    filas_totales_acum = 0
    all_ok = True

    for (year, month) in months_to_process:
        df_new_month = new_by_month.get((year, month), pd.DataFrame(columns=df_new.columns))
        folder_path = build_sharepoint_folder_path(client, year, month)
        filename = f"{client['alias_archivo']}.xlsx"

        # 4) Descargar existente de SharePoint
        df_existing, existing_bytes = _download_sharepoint_df(
            token=token,
            drive_id=drive_id,
            folder_path=folder_path,
            filename=filename,
            log_path=log_path,
        )

        # 5) Merge + dedup
        df_merged = _dedup_merge(df_existing, df_new_month, log_path)
        filas_totales_acum = max(filas_totales_acum, len(df_merged))

        if df_merged.empty:
            log(f"[SKIP] {client_key} {year}/{month:02d}: DataFrame vacio tras merge.", log_path)
            continue

        # 6) Subir resultado
        ok = _upload_df_to_sharepoint(
            token=token,
            drive_id=drive_id,
            folder_path=folder_path,
            filename=filename,
            df=df_merged,
            client=client,
            year=year,
            month=month,
            existing_bytes=existing_bytes,
            log_path=log_path,
        )
        if not ok:
            all_ok = False
            log(f"[FALLO] {client_key} {year}/{month:02d}: fallo la subida a SharePoint.", log_path)

    # 7) Actualizar checkpoint solo si todo salio bien
    if all_ok:
        _save_checkpoint_client(client_key, to_dt.date())
        log(
            f"[OK] {client['empresa_wms']}: {filas_nuevas_total} filas nuevas | "
            f"{filas_totales_acum} filas totales | checkpoint={to_dt.date()}",
            log_path,
        )
        return {
            "ok": True,
            "filas_nuevas": filas_nuevas_total,
            "filas_totales": filas_totales_acum,
            "detalle": "",
        }
    else:
        return {
            "ok": False,
            "filas_nuevas": filas_nuevas_total,
            "filas_totales": filas_totales_acum,
            "detalle": "fallo en subida SharePoint",
        }

# ---------------------------------------------------------------------------
# Email de cierre
# ---------------------------------------------------------------------------

def _send_final_email(results: Dict[str, dict], log_path: Path) -> None:
    summary_rows = []
    for client_key, r in results.items():
        client = _DAILY_CLIENTS.get(client_key, {})
        filas = r.get("filas_nuevas", 0)
        if not r["ok"]:
            estado = "FALLO"
        elif filas == 0:
            estado = "SIN_DATOS"
        else:
            estado = "OK"
        # Distinguir ventana vacía (ya procesado hoy) de genuinamente sin movimientos
        detalle_raw = r.get("detalle", "")
        if detalle_raw == "ventana vacia":
            estado = "AL_DIA"

        detalle_limpio = "Error en descarga o procesamiento" if estado == "FALLO" else ""

        summary_rows.append({
            "cliente": client.get("empresa_wms", client_key),
            "cd": client.get("cd", ""),
            "movimientos": filas if filas > 0 else None,
            "estado": estado,
            "detalle": detalle_limpio,
        })

    active_ok = sum(1 for r in results.values() if r["ok"])
    has_failures = any(not r["ok"] for r in results.values())

    subject, html_body, payload = build_productividad_closure_email(
        summary_rows=summary_rows,
        active_clients_closed=active_ok,
        log_file=log_path,
    )
    if has_failures:
        subject = f"[FALLO PARCIAL] {subject}"

    artifacts = save_productividad_email_artifacts(subject=subject, html_body=html_body, payload=payload)
    log(f"[EMAIL] Preview: {artifacts['html_path']}", log_path)

    if TESTING_MODE:
        log(f"[EMAIL] MODO PRUEBA — enviando solo a {TESTING_EMAIL}", log_path)
        recipients = [TESTING_EMAIL]
    else:
        recipients = None  # usa destinatarios del .env

    ok = send_html_notification(
        subject=subject,
        html_body=html_body,
        log_path=log_path,
        recipients_override=recipients,
    )
    if ok:
        log("[EMAIL] Correo de cierre enviado.", log_path)
    else:
        log("[EMAIL] No se pudo enviar el correo de cierre.", log_path)

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    log_path = build_log_path("productividad_diario")

    if not _acquire_lock(log_path):
        return 1

    try:
        feriados = _load_feriados()
        today = datetime.now().date()

        nombre_feriado = _is_feriado(today, feriados)
        if nombre_feriado:
            log(f"[SKIP] Feriado: {nombre_feriado} ({today}). No se ejecuta la descarga.", log_path)
            return 0

        log(f"[INICIO] productividad_diario.py | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", log_path)
        log(f"[INFO] {len(feriados)} feriados cargados desde OneDrive.", log_path)

        checkpoint = _load_checkpoint()
        log(f"[CHECKPOINT] {json.dumps(checkpoint, ensure_ascii=False)}", log_path)

        # Cargar módulo Graph una sola vez; token se refresca por cliente
        try:
            azure_graph = load_azure_graph_module()
            token = azure_graph.get_token()
            drive_id = azure_graph.get_drive_id(token)
            log("[OK] Token Azure Graph obtenido.", log_path)
        except Exception as exc:
            log(f"[CRITICO] No se pudo obtener token Azure Graph: {exc}", log_path)
            return 2

        results: Dict[str, dict] = {}

        for client_key, client in _DAILY_CLIENTS.items():
            log(f"\n[>>> CLIENTE] {client['empresa_wms']} ({client_key})", log_path)
            # Refrescar token antes de cada cliente — los tokens Graph expiran en ~1h
            # y el run completo puede superar ese tiempo.
            try:
                token = azure_graph.get_token()
            except Exception as exc:
                log(f"[WARN] No se pudo refrescar token para {client_key}: {exc}", log_path)
            try:
                result = _process_client(
                    client_key=client_key,
                    client=client,
                    checkpoint=checkpoint,
                    feriados=feriados,
                    token=token,
                    drive_id=drive_id,
                    log_path=log_path,
                )
                results[client_key] = result
            except Exception as exc:
                log(f"[FALLO] {client_key}: excepcion no capturada: {exc}", log_path)
                results[client_key] = {
                    "ok": False, "filas_nuevas": 0, "filas_totales": 0, "detalle": str(exc)
                }
            # Recargar checkpoint para reflejar actualizaciones de clientes previos
            checkpoint = _load_checkpoint()

        # Resumen en log
        log("\n[RESUMEN] ─────────────────────────────────────────────", log_path)
        log(f"{'Cliente':<22} {'Filas nuevas':>12} {'Filas totales':>13} Estado", log_path)
        for client_key, r in results.items():
            client = _DAILY_CLIENTS.get(client_key, {})
            estado = "OK" if r["ok"] else "FALLO"
            log(
                f"{client.get('empresa_wms', client_key):<22} "
                f"{r['filas_nuevas']:>12,} "
                f"{r['filas_totales']:>13,} "
                f"{estado}",
                log_path,
            )

        ok_count = sum(1 for r in results.values() if r["ok"])
        fail_count = len(results) - ok_count
        log(f"\n[TOTALES] ok={ok_count} | fallos={fail_count} | clientes={len(results)}", log_path)

        _send_final_email(results, log_path)

        return 0 if fail_count == 0 else 2

    finally:
        _release_lock()


if __name__ == "__main__":
    raise SystemExit(main())
