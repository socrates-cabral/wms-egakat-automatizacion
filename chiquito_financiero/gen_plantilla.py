import sys
sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent / "Plantilla_ChiquitoFinanzas.xlsx"

# ─── Colores ────────────────────────────────────────────────────────────────
COLOR_HEADER_DARK  = "2D333B"   # fondo header principal
COLOR_HEADER_MED   = "444C56"   # fondo header secundario
COLOR_ACCENT       = "F78166"   # acento rojo/naranja
COLOR_ACCENT2      = "79C0FF"   # acento azul claro
COLOR_FILA_PAR     = "F6F8FA"   # fila par suave
COLOR_POSITIVO     = "2EA043"   # verde positivo
COLOR_NEGATIVO     = "DA3633"   # rojo negativo

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_white(size=11):
    return Font(bold=True, color="FFFFFF", size=size, name="Calibri")

def normal(size=10, color="000000"):
    return Font(size=size, name="Calibri", color=color)

def thin_border():
    side = Side(style="thin", color="D0D7DE")
    return Border(left=side, right=side, top=side, bottom=side)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def header_row(ws, row, headers, col_start=1, fill_color=COLOR_HEADER_DARK):
    """Escribe una fila de headers con formato oscuro."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = bold_white(11)
        cell.fill = fill(fill_color)
        cell.alignment = center()
        cell.border = thin_border()

def data_row(ws, row, values, col_start=1, alt=False):
    """Escribe una fila de datos con formato alternado."""
    bg = COLOR_FILA_PAR if alt else "FFFFFF"
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        cell.font = normal(10)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()

# ─── Hoja 1: Instrucciones ──────────────────────────────────────────────────
def crear_instrucciones(wb):
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "Plantilla Financiera — Chiquito Mueblería"
    c.font = Font(bold=True, size=18, color="FFFFFF", name="Calibri")
    c.fill = fill(COLOR_HEADER_DARK)
    c.alignment = center()
    ws.row_dimensions[2].height = 40

    # Subtítulo
    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = "Control de Ingresos · Gastos · Deudas · Costos Fijos"
    c.font = Font(bold=False, size=12, color="FFFFFF", name="Calibri")
    c.fill = fill(COLOR_HEADER_MED)
    c.alignment = center()

    instrucciones = [
        ("", ""),
        ("HOJA", "QUÉ CONTIENE / CÓMO USARLA"),
        ("Caja_Mensual",
         "Resumen mes a mes: Ingresos totales, Gastos totales y Resultado neto. "
         "Actualiza las columnas Ingresos_Total y Gastos_Total cada fin de mes."),
        ("Detalle_Gastos",
         "Registro de cada gasto: fecha, categoría, descripción y monto. "
         "Agrega una fila por cada compra o pago realizado."),
        ("Deudas",
         "Lista de compromisos financieros (banco, tarjeta, leasing, préstamos). "
         "Actualiza Saldo_Pendiente cada vez que realizas un abono."),
        ("Costos_Fijos",
         "Gastos que se repiten cada mes con monto similar: arriendo, luz, agua, etc. "
         "Revisa y ajusta los montos cuando cambien."),
        ("", ""),
        ("CONSEJO",
         "Registra los gastos apenas ocurren — no los acumules al fin de mes. "
         "Un registro diario toma menos de 2 minutos y te da control real de tu negocio."),
        ("FRECUENCIA SUGERIDA",
         "Diaria: Detalle_Gastos  |  Mensual: Caja_Mensual  |  Trimestral: Deudas"),
        ("MONEDA",
         "Todos los montos en Pesos Chilenos (CLP) sin puntos de miles en la celda. "
         "El formato numérico agrega el separador automáticamente."),
    ]

    row = 5
    for i, (hoja, desc) in enumerate(instrucciones):
        ws.row_dimensions[row].height = 28
        if hoja == "HOJA":
            ws.merge_cells(f"B{row}:B{row}")
            ws.merge_cells(f"C{row}:G{row}")
            ws[f"B{row}"].value = hoja
            ws[f"B{row}"].font = bold_white(10)
            ws[f"B{row}"].fill = fill(COLOR_HEADER_MED)
            ws[f"B{row}"].alignment = center()
            ws[f"B{row}"].border = thin_border()
            ws[f"C{row}"].value = desc
            ws[f"C{row}"].font = bold_white(10)
            ws[f"C{row}"].fill = fill(COLOR_HEADER_MED)
            ws[f"C{row}"].alignment = left()
            ws[f"C{row}"].border = thin_border()
        elif hoja == "":
            pass
        else:
            bg = COLOR_FILA_PAR if i % 2 == 0 else "FFFFFF"
            ws[f"B{row}"].value = hoja
            ws[f"B{row}"].font = Font(bold=True, size=10, name="Calibri", color=COLOR_ACCENT if hoja in ("CONSEJO", "FRECUENCIA SUGERIDA", "MONEDA") else "000000")
            ws[f"B{row}"].fill = fill(bg)
            ws[f"B{row}"].alignment = center()
            ws[f"B{row}"].border = thin_border()
            ws.merge_cells(f"C{row}:G{row}")
            ws[f"C{row}"].value = desc
            ws[f"C{row}"].font = Font(size=10, name="Calibri")
            ws[f"C{row}"].fill = fill(bg)
            ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"C{row}"].border = thin_border()
            ws.row_dimensions[row].height = 40
        row += 1

    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 22)
    set_col_width(ws, "C", 70)
    for col in ["D", "E", "F", "G"]:
        set_col_width(ws, col, 5)

    ws.sheet_properties.tabColor = "79C0FF"


# ─── Hoja 2: Caja_Mensual ───────────────────────────────────────────────────
def crear_caja_mensual(wb):
    ws = wb.create_sheet("Caja_Mensual")
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "Caja Mensual — Resumen de Ingresos y Gastos"
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill = fill(COLOR_HEADER_DARK)
    c.alignment = center()
    ws.row_dimensions[2].height = 35

    headers = ["Mes", "Ingresos_Total", "Gastos_Total", "Resultado", "Observacion"]
    header_row(ws, 4, headers, col_start=2)
    ws.row_dimensions[4].height = 28

    # Datos ficticios realistas — taller de mueblería ~$1.8M ingresos
    meses = [
        ("Enero 2026",      1_750_000, 1_580_000),
        ("Febrero 2026",    1_820_000, 1_620_000),
        ("Marzo 2026",      1_900_000, 1_650_000),
        ("Abril 2026",      1_780_000, 1_600_000),
        ("Mayo 2026",       1_850_000, 1_670_000),
        ("Junio 2026",      1_760_000, 1_590_000),
        ("Julio 2026",      1_830_000, 1_640_000),
        ("Agosto 2026",     1_950_000, 1_700_000),
        ("Septiembre 2026", 1_800_000, 1_610_000),
        ("Octubre 2026",    1_870_000, 1_660_000),
        ("Noviembre 2026",  2_100_000, 1_800_000),
        ("Diciembre 2026",  2_300_000, 1_950_000),
    ]

    observaciones = [
        "Mes tranquilo — temporada baja",
        "Sin novedad",
        "Pedido especial cocina",
        "",
        "Pago cuota maquinaria",
        "",
        "",
        "Clientes nuevos",
        "",
        "Feria Diseño Hogar",
        "Temporada alta — Navidad",
        "Cierre de año — ventas altas",
    ]

    NUM_FMT = '#,##0'  # miles con coma

    for i, ((mes, ing, gas), obs) in enumerate(zip(meses, observaciones)):
        row = 5 + i
        ws.row_dimensions[row].height = 22
        alt = i % 2 == 0
        bg = COLOR_FILA_PAR if alt else "FFFFFF"

        # Mes
        c_mes = ws.cell(row=row, column=2, value=mes)
        c_mes.font = normal(10)
        c_mes.fill = fill(bg)
        c_mes.alignment = left()
        c_mes.border = thin_border()

        # Ingresos
        c_ing = ws.cell(row=row, column=3, value=ing)
        c_ing.font = Font(size=10, name="Calibri", color=COLOR_POSITIVO)
        c_ing.fill = fill(bg)
        c_ing.alignment = Alignment(horizontal="right", vertical="center")
        c_ing.border = thin_border()
        c_ing.number_format = NUM_FMT

        # Gastos
        c_gas = ws.cell(row=row, column=4, value=gas)
        c_gas.font = Font(size=10, name="Calibri", color=COLOR_NEGATIVO)
        c_gas.fill = fill(bg)
        c_gas.alignment = Alignment(horizontal="right", vertical="center")
        c_gas.border = thin_border()
        c_gas.number_format = NUM_FMT

        # Resultado (fórmula)
        resultado = ing - gas  # valor directo para que sea calculable offline
        col_ing = get_column_letter(3)  # C
        col_gas = get_column_letter(4)  # D
        c_res = ws.cell(row=row, column=5, value=f"={col_ing}{row}-{col_gas}{row}")
        res_color = COLOR_POSITIVO if (ing - gas) >= 0 else COLOR_NEGATIVO
        c_res.font = Font(size=10, name="Calibri", bold=True, color=res_color)
        c_res.fill = fill(bg)
        c_res.alignment = Alignment(horizontal="right", vertical="center")
        c_res.border = thin_border()
        c_res.number_format = NUM_FMT

        # Observacion
        c_obs = ws.cell(row=row, column=6, value=obs)
        c_obs.font = Font(size=10, name="Calibri", color="444C56", italic=True)
        c_obs.fill = fill(bg)
        c_obs.alignment = left()
        c_obs.border = thin_border()

    # Fila TOTAL
    total_row = 5 + 12
    ws.row_dimensions[total_row].height = 26
    c_tot_label = ws.cell(row=total_row, column=2, value="TOTAL AÑO")
    c_tot_label.font = bold_white(11)
    c_tot_label.fill = fill(COLOR_HEADER_DARK)
    c_tot_label.alignment = center()
    c_tot_label.border = thin_border()

    for col, col_letter in [(3, "C"), (4, "D"), (5, "E")]:
        formula = f"=SUM({col_letter}5:{col_letter}{total_row - 1})"
        c = ws.cell(row=total_row, column=col, value=formula)
        c.font = bold_white(11)
        c.fill = fill(COLOR_HEADER_DARK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = thin_border()
        c.number_format = NUM_FMT

    ws.cell(row=total_row, column=6).fill = fill(COLOR_HEADER_DARK)
    ws.cell(row=total_row, column=6).border = thin_border()

    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 20)
    set_col_width(ws, "C", 18)
    set_col_width(ws, "D", 18)
    set_col_width(ws, "E", 18)
    set_col_width(ws, "F", 30)

    ws.sheet_properties.tabColor = "2EA043"


# ─── Hoja 3: Detalle_Gastos ─────────────────────────────────────────────────
def crear_detalle_gastos(wb):
    ws = wb.create_sheet("Detalle_Gastos")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "Detalle de Gastos — Registro de Transacciones"
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill = fill(COLOR_HEADER_DARK)
    c.alignment = center()
    ws.row_dimensions[2].height = 35

    headers = ["Fecha", "Mes", "Categoria", "Descripcion", "Monto", "Nota"]
    header_row(ws, 4, headers, col_start=2)
    ws.row_dimensions[4].height = 28

    from datetime import date
    gastos = [
        (date(2026, 1, 5),  "Enero 2026",     "Materiales",       "Madera MDF 18mm x 10 planchas",         185_000, "Proveedor Maderas Sur"),
        (date(2026, 1, 8),  "Enero 2026",     "Servicios Basicos","Electricidad taller Enero",              42_500,  ""),
        (date(2026, 1, 10), "Enero 2026",     "Materiales",       "Tornillos, bisagras y herrajes varios",  28_900,  ""),
        (date(2026, 1, 15), "Enero 2026",     "Arriendo",         "Arriendo taller Enero",                 380_000, "Pago mensual fijo"),
        (date(2026, 1, 20), "Enero 2026",     "Transporte",       "Gasolina camioneta (2 cargas)",          55_000,  ""),
        (date(2026, 1, 25), "Enero 2026",     "Sueldos",          "Sueldo ayudante Manuel",                350_000, ""),
        (date(2026, 2, 3),  "Febrero 2026",   "Materiales",       "Tablones pino 3x5 metros",              210_000, ""),
        (date(2026, 2, 8),  "Febrero 2026",   "Servicios Basicos","Agua + basura Febrero",                  18_200,  ""),
        (date(2026, 2, 12), "Febrero 2026",   "Herramientas",     "Disco sierra circular repuesto x3",      35_000,  ""),
        (date(2026, 2, 15), "Febrero 2026",   "Arriendo",         "Arriendo taller Febrero",               380_000, ""),
        (date(2026, 2, 20), "Febrero 2026",   "Marketing",        "Publicación en Marketplace Facebook",     5_000,  "Boost 7 días"),
        (date(2026, 2, 25), "Febrero 2026",   "Sueldos",          "Sueldo ayudante Manuel",                350_000, ""),
        (date(2026, 3, 5),  "Marzo 2026",     "Materiales",       "Laca, sellador y barniz",                62_000,  "Ferretería Los Andes"),
        (date(2026, 3, 10), "Marzo 2026",     "Servicios Basicos","Electricidad taller Marzo",              48_000,  "Mes con horas extra"),
        (date(2026, 3, 15), "Marzo 2026",     "Arriendo",         "Arriendo taller Marzo",                 380_000, ""),
        (date(2026, 3, 18), "Marzo 2026",     "Cuota Deuda",      "Cuota crédito banco Marzo",             120_000, "Banco Estado"),
        (date(2026, 3, 22), "Marzo 2026",     "Transporte",       "Envío muebles cliente — flete",          35_000,  ""),
        (date(2026, 3, 25), "Marzo 2026",     "Sueldos",          "Sueldo ayudante Manuel",                350_000, ""),
        (date(2026, 3, 28), "Marzo 2026",     "Materiales",       "Tela tapizado sofá por encargo",         95_000,  "Cliente Contreras"),
        (date(2026, 3, 30), "Marzo 2026",     "Varios",           "Almuerzo reunión con cliente",           18_500,  ""),
    ]

    NUM_FMT = '#,##0'
    DATE_FMT = 'DD-MM-YYYY'

    for i, (fecha, mes, cat, desc, monto, nota) in enumerate(gastos):
        row = 5 + i
        ws.row_dimensions[row].height = 22
        alt = i % 2 == 0
        bg = COLOR_FILA_PAR if alt else "FFFFFF"

        c_fecha = ws.cell(row=row, column=2, value=fecha)
        c_fecha.font = normal(10)
        c_fecha.fill = fill(bg)
        c_fecha.alignment = center()
        c_fecha.border = thin_border()
        c_fecha.number_format = DATE_FMT

        for col_idx, val in [(3, mes), (4, cat), (5, desc)]:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = normal(10)
            c.fill = fill(bg)
            c.alignment = left()
            c.border = thin_border()

        c_monto = ws.cell(row=row, column=6, value=monto)
        c_monto.font = Font(size=10, name="Calibri", color=COLOR_NEGATIVO)
        c_monto.fill = fill(bg)
        c_monto.alignment = Alignment(horizontal="right", vertical="center")
        c_monto.border = thin_border()
        c_monto.number_format = NUM_FMT

        c_nota = ws.cell(row=row, column=7, value=nota)
        c_nota.font = Font(size=9, name="Calibri", color="6E7681", italic=True)
        c_nota.fill = fill(bg)
        c_nota.alignment = left()
        c_nota.border = thin_border()

    # Total
    total_row = 5 + len(gastos)
    ws.row_dimensions[total_row].height = 26
    ws.merge_cells(f"B{total_row}:E{total_row}")
    c_lbl = ws[f"B{total_row}"]
    c_lbl.value = "TOTAL GASTOS (muestra)"
    c_lbl.font = bold_white(11)
    c_lbl.fill = fill(COLOR_HEADER_DARK)
    c_lbl.alignment = center()
    c_lbl.border = thin_border()

    c_tot = ws.cell(row=total_row, column=6, value=f"=SUM(F5:F{total_row - 1})")
    c_tot.font = bold_white(11)
    c_tot.fill = fill(COLOR_HEADER_DARK)
    c_tot.alignment = Alignment(horizontal="right", vertical="center")
    c_tot.border = thin_border()
    c_tot.number_format = NUM_FMT

    ws.cell(row=total_row, column=7).fill = fill(COLOR_HEADER_DARK)
    ws.cell(row=total_row, column=7).border = thin_border()

    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 14)
    set_col_width(ws, "C", 16)
    set_col_width(ws, "D", 20)
    set_col_width(ws, "E", 38)
    set_col_width(ws, "F", 16)
    set_col_width(ws, "G", 28)

    ws.sheet_properties.tabColor = "DA3633"


# ─── Hoja 4: Deudas ─────────────────────────────────────────────────────────
def crear_deudas(wb):
    ws = wb.create_sheet("Deudas")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = "Control de Deudas y Compromisos Financieros"
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill = fill(COLOR_HEADER_DARK)
    c.alignment = center()
    ws.row_dimensions[2].height = 35

    headers = ["Acreedor", "Tipo", "Saldo_Pendiente", "Cuota_Mensual",
               "Tasa_Mensual_%", "Cuotas_Rest.", "Vencimiento"]
    header_row(ws, 4, headers, col_start=2)
    ws.row_dimensions[4].height = 28

    from datetime import date
    deudas = [
        ("Banco Estado",      "Crédito bancario", 3_600_000, 120_000, 0.9,  30, date(2028, 9, 1)),
        ("Tarjeta CMR Falabella", "Tarjeta crédito", 480_000,  96_000, 2.3,   5, date(2026, 6, 15)),
        ("Leasing maquinaria", "Leasing",         8_200_000, 215_000, 0.7,  38, date(2029, 5, 1)),
        ("Préstamo familiar",  "Familiar",          600_000,  50_000, 0.0,  12, date(2027, 3, 1)),
    ]

    NUM_FMT  = '#,##0'
    DATE_FMT = 'DD-MM-YYYY'

    for i, (acreedor, tipo, saldo, cuota, tasa, cuotas, venc) in enumerate(deudas):
        row = 5 + i
        ws.row_dimensions[row].height = 24
        alt = i % 2 == 0
        bg = COLOR_FILA_PAR if alt else "FFFFFF"

        vals = [acreedor, tipo, saldo, cuota, tasa, cuotas, venc]
        for j, val in enumerate(vals):
            c = ws.cell(row=row, column=2 + j, value=val)
            c.font = normal(10)
            c.fill = fill(bg)
            c.border = thin_border()

            if j in (2, 3):  # montos
                c.number_format = NUM_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font = Font(size=10, name="Calibri",
                              color=(COLOR_NEGATIVO if saldo > 1_000_000 else "D4A72C"))
            elif j == 4:  # tasa
                c.number_format = '0.00"%"'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif j == 5:  # cuotas restantes
                c.alignment = center()
            elif j == 6:  # fecha
                c.number_format = DATE_FMT
                c.alignment = center()
            else:
                c.alignment = left()

    # Totales
    total_row = 5 + len(deudas)
    ws.row_dimensions[total_row].height = 26
    ws.merge_cells(f"B{total_row}:C{total_row}")
    c_lbl = ws[f"B{total_row}"]
    c_lbl.value = "TOTAL DEUDA"
    c_lbl.font = bold_white(11)
    c_lbl.fill = fill(COLOR_HEADER_DARK)
    c_lbl.alignment = center()
    c_lbl.border = thin_border()

    c_saldo = ws.cell(row=total_row, column=4, value=f"=SUM(D5:D{total_row - 1})")
    c_saldo.font = bold_white(11)
    c_saldo.fill = fill(COLOR_HEADER_DARK)
    c_saldo.alignment = Alignment(horizontal="right", vertical="center")
    c_saldo.border = thin_border()
    c_saldo.number_format = NUM_FMT

    c_cuota = ws.cell(row=total_row, column=5, value=f"=SUM(E5:E{total_row - 1})")
    c_cuota.font = bold_white(11)
    c_cuota.fill = fill(COLOR_HEADER_DARK)
    c_cuota.alignment = Alignment(horizontal="right", vertical="center")
    c_cuota.border = thin_border()
    c_cuota.number_format = NUM_FMT

    for col in [6, 7, 8]:
        c = ws.cell(row=total_row, column=col)
        c.fill = fill(COLOR_HEADER_DARK)
        c.border = thin_border()

    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 22)
    set_col_width(ws, "C", 20)
    set_col_width(ws, "D", 18)
    set_col_width(ws, "E", 18)
    set_col_width(ws, "F", 14)
    set_col_width(ws, "G", 14)
    set_col_width(ws, "H", 16)

    ws.sheet_properties.tabColor = "F78166"


# ─── Hoja 5: Costos_Fijos ───────────────────────────────────────────────────
def crear_costos_fijos(wb):
    ws = wb.create_sheet("Costos_Fijos")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "Costos Fijos Mensuales del Taller"
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill = fill(COLOR_HEADER_DARK)
    c.alignment = center()
    ws.row_dimensions[2].height = 35

    headers = ["Concepto", "Monto_Mensual", "Tipo", "Nota"]
    header_row(ws, 4, headers, col_start=2)
    ws.row_dimensions[4].height = 28

    costos = [
        ("Arriendo taller",          380_000, "Fijo",     "Contrato anual — vence Dic 2026"),
        ("Electricidad",              45_000,  "Variable", "Promedio últimos 6 meses"),
        ("Agua y alcantarillado",     15_000,  "Fijo",     ""),
        ("Internet + teléfono",       28_000,  "Fijo",     "Plan combo VTR"),
        ("Gasolina camioneta",        55_000,  "Variable", "Estimado — varía con precio combustible"),
        ("Sueldo ayudante",          350_000,  "Fijo",     "Manuel — contrato indefinido"),
        ("Seguro taller y equipos",   18_000,  "Fijo",     "Pago anual prorrateado"),
        ("Gastos varios / imprevistos",30_000,  "Variable", "Fondo para emergencias menores"),
    ]

    NUM_FMT = '#,##0'

    for i, (concepto, monto, tipo, nota) in enumerate(costos):
        row = 5 + i
        ws.row_dimensions[row].height = 24
        alt = i % 2 == 0
        bg = COLOR_FILA_PAR if alt else "FFFFFF"

        c_conc = ws.cell(row=row, column=2, value=concepto)
        c_conc.font = normal(10)
        c_conc.fill = fill(bg)
        c_conc.alignment = left()
        c_conc.border = thin_border()

        c_monto = ws.cell(row=row, column=3, value=monto)
        c_monto.font = Font(size=10, name="Calibri", color=COLOR_NEGATIVO)
        c_monto.fill = fill(bg)
        c_monto.alignment = Alignment(horizontal="right", vertical="center")
        c_monto.border = thin_border()
        c_monto.number_format = NUM_FMT

        tipo_color = "2EA043" if tipo == "Fijo" else "D4A72C"
        c_tipo = ws.cell(row=row, column=4, value=tipo)
        c_tipo.font = Font(size=10, name="Calibri", bold=True, color=tipo_color)
        c_tipo.fill = fill(bg)
        c_tipo.alignment = center()
        c_tipo.border = thin_border()

        c_nota = ws.cell(row=row, column=5, value=nota)
        c_nota.font = Font(size=9, name="Calibri", color="6E7681", italic=True)
        c_nota.fill = fill(bg)
        c_nota.alignment = left()
        c_nota.border = thin_border()

    # Total
    total_row = 5 + len(costos)
    ws.row_dimensions[total_row].height = 26
    c_lbl = ws.cell(row=total_row, column=2, value="TOTAL COSTOS FIJOS/MES")
    c_lbl.font = bold_white(11)
    c_lbl.fill = fill(COLOR_HEADER_DARK)
    c_lbl.alignment = center()
    c_lbl.border = thin_border()

    c_tot = ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row - 1})")
    c_tot.font = bold_white(11)
    c_tot.fill = fill(COLOR_HEADER_DARK)
    c_tot.alignment = Alignment(horizontal="right", vertical="center")
    c_tot.border = thin_border()
    c_tot.number_format = NUM_FMT

    for col in [4, 5]:
        c = ws.cell(row=total_row, column=col)
        c.fill = fill(COLOR_HEADER_DARK)
        c.border = thin_border()

    # Nota al pie
    ws.row_dimensions[total_row + 2].height = 20
    c_pie = ws.cell(row=total_row + 2, column=2,
                    value="* Revisa y actualiza estos montos cada 3 meses para reflejar cambios reales.")
    c_pie.font = Font(size=9, name="Calibri", italic=True, color="6E7681")
    ws.merge_cells(f"B{total_row + 2}:E{total_row + 2}")

    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 28)
    set_col_width(ws, "C", 18)
    set_col_width(ws, "D", 12)
    set_col_width(ws, "E", 38)

    ws.sheet_properties.tabColor = "D4A72C"


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    crear_instrucciones(wb)
    crear_caja_mensual(wb)
    crear_detalle_gastos(wb)
    crear_deudas(wb)
    crear_costos_fijos(wb)

    wb.save(OUTPUT_PATH)
    print(f"Plantilla generada: {OUTPUT_PATH}")
    print(f"Hojas creadas: {[ws.title for ws in wb.worksheets]}")
    size_kb = OUTPUT_PATH.stat().st_size / 1024
    print(f"Tamaño: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
