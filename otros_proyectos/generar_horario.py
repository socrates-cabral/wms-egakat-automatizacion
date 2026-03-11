from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colores ───────────────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"
C_BLUE     = "2E75B6"
C_GOLD     = "B8860B"
C_GREEN    = "375623"
C_BGGREEN  = "E2EFDA"
C_BGGOLD   = "FFF2CC"
C_BGBLUE   = "D6E4F0"
C_BGRED    = "FCE4D6"
C_PURPLE   = "7030A0"
C_BGPURPLE = "EAD1FF"
C_GRAY     = "F5F5F5"
C_GRAY2    = "E8E8E8"
C_RED      = "C00000"
C_ORANGE   = "C55A11"
C_TEAL     = "1F7070"
C_BGTEAL   = "D9F0F0"

def st(ws, row, col, value=None, bold=False, bg=None, fc="000000",
       align="center", wrap=False, size=11, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fc, name="Calibri", size=size, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    thin = Side(style="thin", color="CCCCCC")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def title_row(ws, row, text, cols, bg=C_HEADER, size=13, fc="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color=fc, name="Calibri", size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

def section_header(ws, row, text, cols, bg=C_BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def note_row(ws, row, text, cols, bg=C_BGGOLD, fc="555555"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(italic=True, color=fc, name="Calibri", size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[row].height = 16

def block_schedule(ws, start_row, rows_data, col_count=3):
    """
    rows_data: list of (hora, actividad, nota, bg, bold, fc_act)
    """
    r = start_row
    for hora, acti, nota, bg, bold, fc_act in rows_data:
        st(ws, r, 1, hora,  bold=bold, bg=bg, fc="444444", align="center", size=10)
        st(ws, r, 2, acti,  bold=bold, bg=bg, fc=fc_act,  align="left",   size=11)
        if col_count >= 3:
            st(ws, r, 3, nota, bold=False, bg=C_GRAY, fc="666666", align="left", size=10, italic=True)
        ws.row_dimensions[r].height = 20
        r += 1
    return r

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1 — Mañana (bloque fijo todos los días)
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Rutina Mañana"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 40
ws1.column_dimensions["C"].width = 32

title_row(ws1, 1, "RUTINA DE MAÑANA  —  Lunes a Viernes  (bloque fijo)", 3)
note_row(ws1, 2,
    "Ventana de 40 minutos: ejercicio + ducha + salida. Sin margen para estudio matutino.",
    3, bg="EBF3FB", fc="1F4E79")

# Encabezados
row = 3
for col, h in enumerate(["Hora", "Actividad", "Detalle"], 1):
    st(ws1, row, col, h, bold=True, bg=C_BLUE, fc="FFFFFF")
ws1.row_dimensions[row].height = 22

morning = [
    ("05:30",        "⏰  Despertar",                          "Alarma única — sin snooze",                      C_BGGOLD,  True,  "B8860B"),
    ("05:30 – 05:45","💪  Ejercicio  (15 min)",                "Abdominales + flexiones — cuidado hombro",       C_BGGREEN, True,  C_GREEN),
    ("05:45 – 06:05","🚿  Ducha + vestirse",                   "Ropa y bolso preparados la noche anterior",      C_BGBLUE,  False, "000000"),
    ("06:05 – 06:10","🚪  Salida del departamento",            "Desayuno ya preparado — llevarlo o en ruta",     C_BGGOLD,  True,  C_ORANGE),
]

row = block_schedule(ws1, 4, morning)

# Nota lesión
row += 1
section_header(ws1, row, "NOTA — Lesión de hombro", 3, bg=C_ORANGE)
row += 1
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
c = ws1.cell(row=row, column=1,
    value="Evitar ejercicios que carguen el hombro. Priorizar abdominales y flexiones con apoyo de rodillas si es necesario. Retomar progresión cuando el médico lo indique.")
c.font      = Font(italic=True, color="C55A11", name="Calibri", size=11)
c.fill      = PatternFill("solid", fgColor=C_BGRED)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
thin = Side(style="thin", color="CCCCCC")
c.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
ws1.row_dimensions[row].height = 40

# Prep noche anterior
row += 2
section_header(ws1, row, "PREPARACIÓN NOCHE ANTERIOR  (5 min antes de dormir)", 3, bg=C_TEAL)
row += 1
preps = [
    ("✔", "Dejar ropa del día siguiente lista"),
    ("✔", "Preparar desayuno y cena del día siguiente"),
    ("✔", "Cargar celular y mochila"),
    ("✔", "Revisar agenda del día siguiente"),
]
for icono, texto in preps:
    st(ws1, row, 1, icono,  bold=True, bg=C_BGTEAL, fc=C_TEAL, align="center")
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws1.cell(row=row, column=2, value=texto)
    c.font      = Font(name="Calibri", size=11, color="000000")
    c.fill      = PatternFill("solid", fgColor=C_BGTEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws1.row_dimensions[row].height = 20
    row += 1

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2 — Tardes (3 plantillas)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Rutina Tarde-Noche")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 36
ws2.column_dimensions["C"].width = 34

title_row(ws2, 1, "RUTINA TARDE-NOCHE  —  3 Plantillas según el día", 3)
note_row(ws2, 2,
    "El bloque de estudio siempre va después de la actividad principal. Bed: 22:30 innegociable.",
    3, bg="EBF3FB", fc="1F4E79")

row = 3

# ── Plantilla A: Día con hijo ────────────────────────────────────────────────
section_header(ws2, row, "PLANTILLA A  —  Día de visita al hijo  (~días alternos)", 3, bg=C_ORANGE)
row += 1
dias_hijo = [
    ("18:00 – 18:30","Traslado a casa / al hijo",              "Descompresión en ruta",                         C_BGRED,   False, "000000"),
    ("18:30 – 20:30","👨‍👦  Tiempo con el hijo",                 "Prioridad absoluta — sin interrupciones",        C_BGRED,   True,  C_RED),
    ("20:30 – 21:30","📚  ESTUDIO  (60 min)",                   "Sesión enfocada — SQL Server o IA",             C_BGGREEN, True,  C_GREEN),
    ("21:30 – 22:00","Cena + cierre del día",                  "Ya preparada — solo calentar",                  C_GRAY,    False, "000000"),
    ("22:00 – 22:30","Wind down — sin pantallas",              "Lectura, respiración, preparar mañana",         C_BGBLUE,  False, "000000"),
    ("22:30",        "🌙  Luces apagadas",                     "7 horas de sueño",                              C_HEADER,  True,  "FFFFFF"),
]
row = block_schedule(ws2, row, dias_hijo)

row += 1

# ── Plantilla B: Día de gimnasio ─────────────────────────────────────────────
section_header(ws2, row, "PLANTILLA B  —  Día de gimnasio  (3-4 veces/semana, lesión permita)", 3, bg=C_TEAL)
row += 1
dias_gym = [
    ("18:00 – 18:30","Traslado al gimnasio",                   "",                                              C_BGTEAL,  False, "000000"),
    ("18:30 – 19:30","🏋️  Gimnasio",                           "Rutina adaptada — hombro en recuperación",      C_BGTEAL,  True,  C_TEAL),
    ("19:30 – 20:00","Ducha + cena",                           "Cena ya preparada",                             C_GRAY,    False, "000000"),
    ("20:00 – 21:30","📚  ESTUDIO  (90 min)",                   "Sesión principal — SQL Server o IA",            C_BGGREEN, True,  C_GREEN),
    ("21:30 – 22:00","Wind down — sin pantallas",              "Preparar ropa y comida del día siguiente",      C_BGBLUE,  False, "000000"),
    ("22:30",        "🌙  Luces apagadas",                     "7 horas de sueño",                              C_HEADER,  True,  "FFFFFF"),
]
row = block_schedule(ws2, row, dias_gym)

row += 1

# ── Plantilla C: Día libre ───────────────────────────────────────────────────
section_header(ws2, row, "PLANTILLA C  —  Día libre  (sin hijo ni gimnasio)", 3, bg=C_BLUE)
row += 1
dias_libre = [
    ("18:00 – 18:30","Traslado a casa",                        "Descompresión — sin pantallas de estudio",      C_BGBLUE,  False, "000000"),
    ("18:30 – 19:00","Cena tranquila",                         "Ya preparada",                                  C_GRAY,    False, "000000"),
    ("19:00 – 20:30","📚  ESTUDIO  (90 min)",                   "Sesión principal — SQL Server o IA",            C_BGGREEN, True,  C_GREEN),
    ("20:30 – 22:00","Tiempo personal",                        "Serie, lectura, llamadas — sin culpa",          C_BGBLUE,  False, "000000"),
    ("22:00 – 22:30","Wind down",                              "Preparar ropa y comida del día siguiente",      C_GRAY,    False, "000000"),
    ("22:30",        "🌙  Luces apagadas",                     "7 horas de sueño",                              C_HEADER,  True,  "FFFFFF"),
]
row = block_schedule(ws2, row, dias_libre)

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3 — Plan Semanal
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Plan Semanal")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 13
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 30

title_row(ws3, 1, "PLAN SEMANAL  —  SQL Server + IA + Vida", 5)
note_row(ws3, 2,
    "Hijo: días alternos  |  Gimnasio: días sin hijo (3-4x/semana incl. finde)  |  Lesión hombro: rutina adaptada",
    5, bg="EBF3FB", fc="1F4E79")

row = 3
for col, h in enumerate(["Día", "Tema estudio", "Plantilla tarde", "Bloque / Duración", "Enfoque"], 1):
    st(ws3, row, col, h, bold=True, bg=C_BLUE, fc="FFFFFF")
ws3.row_dimensions[row].height = 22

COLOR_SQL = "D6E4F0"
COLOR_IA  = "E2EFDA"
COLOR_REP = "FFF2CC"
COLOR_PRO = C_BGPURPLE
COLOR_DES = C_GRAY2

# Semana ejemplo (hijo días impares: lun, mie, vie / gym: mar, jue + sábado)
weekly = [
    ("Lunes",     "SQL Server", "A — Hijo",      "18:30-20:30 hijo\n20:30-21:30 estudio (1h)", "Consultas, JOINs, filtros",                   COLOR_SQL),
    ("Martes",    "IA",         "B — Gimnasio",  "18:30-19:30 gym\n20:00-21:30 estudio (90m)", "Claude API, conceptos, práctica",             COLOR_IA),
    ("Miércoles", "SQL Server", "A — Hijo",      "18:30-20:30 hijo\n20:30-21:30 estudio (1h)", "Índices, optimización, vistas",               COLOR_SQL),
    ("Jueves",    "IA",         "B — Gimnasio",  "18:30-19:30 gym\n20:00-21:30 estudio (90m)", "Python + datos, automatización",              COLOR_IA),
    ("Viernes",   "Repaso",     "A o C",         "Según el día  (1h)",                         "Reforzar lo más débil de la semana",          COLOR_REP),
    ("Sábado",    "Proyecto",   "B — Gimnasio",  "08:00-11:00 proyecto (3h)",                  "SQL + IA + Python — proyecto integrador",     COLOR_PRO),
    ("Domingo",   "Descanso",   "Hijo o libre",  "—",                                          "Sin agenda. Recuperación total.",             COLOR_DES),
]

row = 4
for cols in weekly:
    dia, tema, plantilla, bloque, enfoque, bg = cols
    bold_row = dia == "Sábado"
    fc_tema = C_PURPLE if dia == "Sábado" else ("888888" if dia == "Domingo" else "1F4E79")
    st(ws3, row, 1, dia,       bold=bold_row, bg=bg, fc="000000",  align="center")
    st(ws3, row, 2, tema,      bold=bold_row, bg=bg, fc=fc_tema,   align="center")
    st(ws3, row, 3, plantilla, bold=False,    bg=bg, fc="555555",  align="center", size=10)
    st(ws3, row, 4, bloque,    bold=False,    bg=bg, fc="444444",  align="center", size=10, wrap=True)
    st(ws3, row, 5, enfoque,   bold=False,    bg=bg, fc="000000",  align="left",   wrap=True)
    ws3.row_dimensions[row].height = 32
    row += 1

# Proyectos integradores sábado
row += 1
section_header(ws3, row, "PROYECTOS INTEGRADORES — Sábados 08:00-11:00", 5, bg=C_PURPLE)
row += 1
proyectos = [
    "1.  Cargar inventario.csv a SQL Server y calcular IRA / ILA con queries",
    "2.  Automatizar el reporte Excel con datos frescos desde SQL Server",
    "3.  Construir agente con Claude API que lea el inventario y genere alertas",
    "4.  Dashboard de KPIs logísticos: SQL + Python + Excel integrados",
]
for p in proyectos:
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws3.cell(row=row, column=1, value=p)
    c.font      = Font(name="Calibri", size=11, color="000000")
    c.fill      = PatternFill("solid", fgColor=C_BGPURPLE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    thin = Side(style="thin", color="CCCCCC")
    c.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws3.row_dimensions[row].height = 20
    row += 1

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4 — Alternancia SQL / IA
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Alternancia SQL-IA")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 16
ws4.column_dimensions["B"].width = 14
ws4.column_dimensions["C"].width = 36
ws4.column_dimensions["D"].width = 36

title_row(ws4, 1, "PLAN DE CONTENIDOS  —  SQL Server + IA  (8 semanas)", 4)
note_row(ws4, 2,
    "Cada sesión de 60-90 min. Un tema por día, alternando. Sábados: proyecto que integra ambos.",
    4, bg="EBF3FB", fc="1F4E79")

row = 3
for col, h in enumerate(["Semana", "Día", "SQL Server", "IA / Python"], 1):
    st(ws4, row, col, h, bold=True, bg=C_BLUE, fc="FFFFFF")
ws4.row_dimensions[row].height = 22

contenidos = [
    ("Sem 1", "Lun/Mié/Vie", "SELECT, WHERE, ORDER BY, tipos de datos",          "Introducción a LLMs, Claude API, primeras llamadas"),
    ("Sem 2", "Lun/Mié/Vie", "JOINs (INNER, LEFT, RIGHT, FULL)",                 "Prompts básicos, temperatura, roles del sistema"),
    ("Sem 3", "Lun/Mié/Vie", "GROUP BY, HAVING, funciones de agregación",        "Llamadas con Python, leer CSV y enviar a Claude"),
    ("Sem 4", "Lun/Mié/Vie", "Subconsultas y CTEs",                              "Automatización: generar reportes con IA"),
    ("Sem 5", "Lun/Mié/Vie", "Índices, EXPLAIN, optimización de queries",        "Agentes simples: tools, function calling"),
    ("Sem 6", "Lun/Mié/Vie", "Procedimientos almacenados y funciones",           "Integrar SQL + Python + Claude en un pipeline"),
    ("Sem 7", "Lun/Mié/Vie", "Vistas, triggers, transacciones",                  "Proyecto: análisis de inventario con IA"),
    ("Sem 8", "Lun/Mié/Vie", "Backup, seguridad, SQL Server Agent",             "Proyecto: dashboard automático con alertas IA"),
]

row = 4
for i, (sem, dias, sql, ia) in enumerate(contenidos):
    bg = C_BGBLUE if i % 2 == 0 else C_GRAY
    st(ws4, row, 1, sem,  bold=True,  bg=bg, fc="1F4E79", align="center")
    st(ws4, row, 2, dias, bold=False, bg=bg, fc="555555", align="center", size=10)
    st(ws4, row, 3, sql,  bold=False, bg=C_BGBLUE, fc="000000", align="left", wrap=True)
    st(ws4, row, 4, ia,   bold=False, bg=C_BGGREEN, fc="000000", align="left", wrap=True)
    ws4.row_dimensions[row].height = 28
    row += 1

# ════════════════════════════════════════════════════════════════════════════
# HOJA 5 — Reglas
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Reglas del Sistema")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 6
ws5.column_dimensions["B"].width = 64

title_row(ws5, 1, "REGLAS QUE PROTEGEN EL SISTEMA", 2)
ws5.row_dimensions[1].height = 30

reglas = [
    (C_BGGREEN, C_GREEN,   "1", "El estudio va en la tarde-noche, DESPUÉS del hijo o el gimnasio. No al revés."),
    (C_BGGOLD,  C_GOLD,    "2", "22:30 es innegociable — 7 horas de sueño son parte del plan de estudio."),
    (C_BGRED,   C_RED,     "3", "El hombro manda. Nada de ejercicios que lo carguen hasta recuperación completa."),
    (C_BGBLUE,  C_BLUE,    "4", "El tiempo con el hijo NO es negociable. El estudio se adapta a él, no al revés."),
    (C_BGPURPLE,C_PURPLE,  "5", "El sábado por la mañana es el bloque de proyecto — donde todo se conecta."),
    (C_GRAY2,   "555555",  "6", "Si un día fallas, no recuperes al día siguiente. Sigue el ritmo normal."),
]

row = 3
thin = Side(style="thin", color="CCCCCC")
for bg, fc, num, texto in reglas:
    c1 = ws5.cell(row=row, column=1, value=num)
    c1.font      = Font(bold=True, color=fc, name="Calibri", size=14)
    c1.fill      = PatternFill("solid", fgColor=bg)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    c2 = ws5.cell(row=row, column=2, value=texto)
    c2.font      = Font(name="Calibri", size=11, color="000000")
    c2.fill      = PatternFill("solid", fgColor=bg)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    c2.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws5.row_dimensions[row].height = 36
    row += 1

# ── Guardar ───────────────────────────────────────────────────────────────────
path = "C:/ClaudeWork/horario_estudio.xlsx"
wb.save(path)
print(f"Guardado: {path}")
