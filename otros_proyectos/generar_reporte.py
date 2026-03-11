import csv
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Leer CSV ──────────────────────────────────────────────────────────────────
productos = []
with open("C:/ClaudeWork/inventario.csv", newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        row["Stock_Sistema"] = int(row["Stock_Sistema"])
        row["Stock_Fisico"]  = int(row["Stock_Fisico"])
        productos.append(row)

# ── Calcular métricas ─────────────────────────────────────────────────────────
for p in productos:
    ss, sf = p["Stock_Sistema"], p["Stock_Fisico"]
    p["Diferencia"]  = sf - ss
    p["IRA_pct"]     = round(min(ss, sf) / ss * 100, 2)
    p["Linea_Exacta"] = "SI" if sf == ss else "NO"
    if p["IRA_pct"] < 98:
        p["Estado"] = "CRITICO"
    elif p["Diferencia"] > 0:
        p["Estado"] = "SOBRANTE"
    elif p["Diferencia"] == 0:
        p["Estado"] = "EXACTO"
    else:
        p["Estado"] = "OK"

total_ss   = sum(p["Stock_Sistema"] for p in productos)
total_sf   = sum(p["Stock_Fisico"]  for p in productos)
total_min  = sum(min(p["Stock_Sistema"], p["Stock_Fisico"]) for p in productos)
ira_total  = round(total_min / total_ss * 100, 2)
lineas_ok  = sum(1 for p in productos if p["Linea_Exacta"] == "SI")
ila_total  = round(lineas_ok / len(productos) * 100, 2)

# ── Estilos ───────────────────────────────────────────────────────────────────
COLOR_HEADER   = "1F4E79"   # azul oscuro
COLOR_SUBHEAD  = "2E75B6"   # azul medio
COLOR_CRITICO  = "FF4444"   # rojo
COLOR_SOBRANTE = "FF9900"   # naranja
COLOR_EXACTO   = "00B050"   # verde
COLOR_OK       = "E2EFDA"   # verde claro (fondo fila)
COLOR_SUMMARY  = "D6E4F0"   # azul pálido

def cell_style(ws, row, col, value=None, bold=False, bg=None, font_color="000000",
               align="center", wrap=False, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=font_color,
                       name="Calibri", size=11)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    thin = Side(style="thin", color="AAAAAA")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if num_format:
        c.number_format = num_format
    return c

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1 — Detalle
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Detalle Inventario"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

# Título
ws1.merge_cells("A1:H1")
t = ws1["A1"]
t.value     = "REPORTE DE EXACTITUD DE INVENTARIO  —  Bodega Logística"
t.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
t.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

# Encabezados columna
headers = ["SKU", "Descripcion", "Stock\nSistema", "Stock\nFisico",
           "Diferencia", "IRA %", "Linea\nExacta", "Estado"]
col_widths = [16, 34, 12, 12, 12, 10, 10, 12]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell_style(ws1, 2, col, h, bold=True, bg=COLOR_SUBHEAD,
               font_color="FFFFFF", wrap=True)
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.row_dimensions[2].height = 32

# Filas de datos
for i, p in enumerate(productos, start=3):
    estado = p["Estado"]
    row_bg = None
    if estado == "CRITICO":
        row_bg = "FFE0E0"
    elif estado == "EXACTO":
        row_bg = "E2F0D9"

    cell_style(ws1, i, 1, p["SKU"],          align="left",  bg=row_bg)
    cell_style(ws1, i, 2, p["Descripcion"],  align="left",  bg=row_bg)
    cell_style(ws1, i, 3, p["Stock_Sistema"], bg=row_bg, num_format="#,##0")
    cell_style(ws1, i, 4, p["Stock_Fisico"],  bg=row_bg, num_format="#,##0")

    diff = p["Diferencia"]
    dc = cell_style(ws1, i, 5, diff, bg=row_bg, num_format="+#,##0;-#,##0;0")
    if diff < 0:
        dc.font = Font(color="CC0000", bold=True, name="Calibri", size=11)
    elif diff > 0:
        dc.font = Font(color="FF6600", bold=True, name="Calibri", size=11)

    ira = p["IRA_pct"]
    ic = cell_style(ws1, i, 6, ira / 100, bg=row_bg, num_format="0.00%")
    if ira < 98:
        ic.font = Font(color="CC0000", bold=True, name="Calibri", size=11)

    cell_style(ws1, i, 7, p["Linea_Exacta"], bg=row_bg)

    # Chip de estado
    COLOR_CHIP = {"CRITICO": COLOR_CRITICO, "SOBRANTE": COLOR_SOBRANTE,
                  "EXACTO": COLOR_EXACTO,   "OK": "4472C4"}
    ec = cell_style(ws1, i, 8, estado, bold=True,
                    bg=COLOR_CHIP.get(estado, "AAAAAA"),
                    font_color="FFFFFF")

    ws1.row_dimensions[i].height = 18

# Fila TOTAL
row_t = len(productos) + 3
for col in range(1, 9):
    cell_style(ws1, row_t, col, bg="1F4E79", font_color="FFFFFF", bold=True)

ws1.cell(row=row_t, column=1).value = "TOTAL"
ws1.cell(row=row_t, column=1).alignment = Alignment(horizontal="center", vertical="center")

cell_style(ws1, row_t, 3, total_ss,              bold=True, bg=COLOR_HEADER, font_color="FFFFFF", num_format="#,##0")
cell_style(ws1, row_t, 4, total_sf,              bold=True, bg=COLOR_HEADER, font_color="FFFFFF", num_format="#,##0")
cell_style(ws1, row_t, 5, total_sf - total_ss,   bold=True, bg=COLOR_HEADER, font_color="FFFFFF", num_format="+#,##0;-#,##0;0")
cell_style(ws1, row_t, 6, ira_total / 100,       bold=True, bg=COLOR_HEADER, font_color="FFFFFF", num_format="0.00%")
ws1.row_dimensions[row_t].height = 22

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2 — Resumen Ejecutivo
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resumen Ejecutivo")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 40

# Título
ws2.merge_cells("A1:C1")
t2 = ws2["A1"]
t2.value     = "RESUMEN EJECUTIVO — EXACTITUD DE INVENTARIO"
t2.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
t2.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

# Fecha
ws2.merge_cells("A2:C2")
f2 = ws2["A2"]
f2.value     = "Fecha de análisis: 2026-03-06"
f2.font      = Font(italic=True, size=10, color="555555", name="Calibri")
f2.alignment = Alignment(horizontal="center")
ws2.row_dimensions[2].height = 16

# KPIs globales
kpis = [
    ("Indicador",        "Valor",        "Referencia / Meta"),
    ("IRA Total",        f"{ira_total}%", "Meta: ≥ 99.5%"),
    ("ILA Total",        f"{ila_total}%", "Meta: ≥ 80%"),
    ("Total uds sistema",f"{total_ss:,}", ""),
    ("Total uds físicas",f"{total_sf:,}", ""),
    ("Diferencia neta",  f"{total_sf - total_ss:+,} uds", ""),
    ("Líneas exactas",   f"{lineas_ok} de {len(productos)}", ""),
    ("Líneas críticas",  f"{sum(1 for p in productos if p['Estado']=='CRITICO')} de {len(productos)}", "IRA < 98%"),
]

row = 4
for i, (label, valor, ref) in enumerate(kpis):
    is_header = (i == 0)
    bg = COLOR_SUBHEAD if is_header else (COLOR_SUMMARY if i % 2 == 0 else "FFFFFF")
    fc = "FFFFFF" if is_header else "000000"
    cell_style(ws2, row, 1, label, bold=is_header, bg=bg, font_color=fc, align="left")
    cell_style(ws2, row, 2, valor, bold=is_header, bg=bg, font_color=fc)
    cell_style(ws2, row, 3, ref,   bold=False,     bg=bg, font_color="555555", align="left")
    ws2.row_dimensions[row].height = 20
    row += 1

# Sección críticos
row += 1
ws2.merge_cells(f"A{row}:C{row}")
h = ws2[f"A{row}"]
h.value     = "DIFERENCIAS CRITICAS (IRA < 98%)"
h.font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
h.fill      = PatternFill("solid", fgColor=COLOR_CRITICO)
h.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[row].height = 24
row += 1

criticos = [p for p in productos if p["Estado"] == "CRITICO"]
if criticos:
    for p in criticos:
        cell_style(ws2, row, 1, p["SKU"],          bold=True, bg="FFE0E0", align="left")
        cell_style(ws2, row, 2, f"IRA: {p['IRA_pct']}%", bold=True, bg="FFE0E0", font_color="CC0000")
        cell_style(ws2, row, 3, p["Descripcion"],  bg="FFE0E0", align="left")
        ws2.row_dimensions[row].height = 18
        row += 1
else:
    ws2.merge_cells(f"A{row}:C{row}")
    ws2[f"A{row}"].value = "Sin diferencias críticas."
    row += 1

# Sección sobrantes
row += 1
ws2.merge_cells(f"A{row}:C{row}")
h2 = ws2[f"A{row}"]
h2.value     = "SOBRANTES DETECTADOS"
h2.font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
h2.fill      = PatternFill("solid", fgColor=COLOR_SOBRANTE)
h2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[row].height = 24
row += 1

sobrantes = [p for p in productos if p["Estado"] == "SOBRANTE"]
if sobrantes:
    for p in sobrantes:
        cell_style(ws2, row, 1, p["SKU"],             bold=True, bg="FFF2CC", align="left")
        cell_style(ws2, row, 2, f"+{p['Diferencia']} uds", bold=True, bg="FFF2CC", font_color="CC6600")
        cell_style(ws2, row, 3, p["Descripcion"],     bg="FFF2CC", align="left")
        ws2.row_dimensions[row].height = 18
        row += 1

# ── Guardar ───────────────────────────────────────────────────────────────────
output_path = "C:/ClaudeWork/reporte_inventario.xlsx"
wb.save(output_path)
print(f"Archivo guardado: {output_path}")
