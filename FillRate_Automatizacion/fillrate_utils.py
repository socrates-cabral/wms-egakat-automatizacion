"""
Helpers operativos para FillRate.

Este modulo concentra:
- logging
- fechas del periodo actual
- compatibilidad de variables de entorno
- autenticacion y operaciones de Microsoft Graph API
- lectura y procesamiento de Excel
- replace del mes actual en workbook acumulado
- construccion de correo resumen
"""

from __future__ import annotations

import os
import re
import threading
import time
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
from dotenv import load_dotenv
from msal import ConfidentialClientApplication
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fillrate_config import (
    BASE_SHEET_NAME,
    DOWNLOAD_BASENAME_PREFIXES,
    DOWNLOAD_SUFFIXES,
    ESTADOS_ALERTA,
    TARGET_SHEET_NAME,
    WARNING_MAX_DAYS,
    get_sharepoint_relative_path,
)


ROOT_DIR = Path(__file__).resolve().parent
LOG_DIR = ROOT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPE = ["https://graph.microsoft.com/.default"]
DEFAULT_SP_HOST = "egakatcom.sharepoint.com"
DEFAULT_SP_SITE_PATH = "/sites/DatosparaDashboard"
DEFAULT_SP_LIBRARY = "Documentos"

DATA_COLUMNS = 26
DATA_START_ROW = 2
DATE_COLUMN_INDEX = 9  # columna I
STATE_COLUMN_INDEX = 7  # columna G
ORDER_COLUMN_INDEX = 5  # columna E
OBSERVATION_COLUMN_INDEX = 22  # V
FORMULA_START_COL = 27  # AA
MAX_SUPPORTED_FORMULA_COL = 47  # AU
COL_ENTREGADO_TIEMPO = 40     # "Entregado a tiempo?"
COL_ENTREGADO_COMPLETO = 41   # "Entregado completo y sin daños?"
OTIF_UMBRAL_VERDE = 95.0
OTIF_UMBRAL_AMARILLO = 85.0
FORMULA_TEMPLATE_ROW = 2
MANUAL_OVERRIDE_COLUMNS = (22, 40, 41)  # V, AN, AO

_token_cache: Dict[str, Any] = {"token": None, "expires_at": 0.0}
_token_lock = threading.Lock()
_graph_location_cache: Dict[str, str] = {"site_id": "", "drive_id": ""}

load_dotenv(ROOT_DIR / ".env", encoding="utf-8-sig")
load_dotenv(ROOT_DIR.parent / ".env", encoding="utf-8-sig")


def _now() -> datetime:
    return datetime.now()


def build_log_path(execution_dt: Optional[datetime] = None) -> Path:
    execution_dt = execution_dt or _now()
    return LOG_DIR / f"fillrate_{execution_dt.strftime('%Y-%m-%d')}.log"


def sanitize_log_message(message: str) -> str:
    sanitized = str(message)
    replacements = {
        "→": "->",
        "✓": "[OK]",
        "✔": "[OK]",
        "✗": "[FALLO]",
        "❌": "[FALLO]",
        "⚠": "[WARN]",
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "Á": "A",
        "É": "E",
        "Í": "I",
        "Ó": "O",
        "Ú": "U",
        "ñ": "n",
        "Ñ": "N",
    }
    for original, replacement in replacements.items():
        sanitized = sanitized.replace(original, replacement)
    return sanitized


def log(message: str, log_path: Optional[Path] = None) -> None:
    log_path = log_path or build_log_path()
    line = f"[{_now().strftime('%H:%M:%S')}] {sanitize_log_message(message)}"
    print(line, flush=True)
    for attempt in range(3):
        try:
            with open(log_path, "a", encoding="utf-8") as handle:
                handle.write(line + "\n")
            return
        except PermissionError:
            if attempt < 2:
                time.sleep(1)


def get_env_value(*names: str, required: bool = False) -> str:
    for name in names:
        value = os.getenv(name, "").strip()
        if value:
            return value
    if required:
        joined = ", ".join(names)
        raise RuntimeError(f"Variable(s) de entorno requerida(s) no disponible(s): {joined}")
    return ""


def get_wms_credentials(user_override: Optional[str] = None) -> Tuple[str, str]:
    user = (user_override or "").strip() or get_env_value("WMS_USUARIO", "WMS_USER", "WMS_USERNAME", required=True)
    password = get_env_value("WMS_CLAVE", "WMS_PASSWORD", "WMS_PASSWORD2", required=True)
    return user, password


def get_graph_settings() -> Dict[str, str]:
    tenant_id = get_env_value("TENANT_ID", "Directory_(tenant)_ID", required=True)
    client_id = get_env_value("CLIENT_ID", "Application_(client)_ID", required=True)
    client_secret = get_env_value("CLIENT_SECRET", "Client_Secret_Value", required=True)
    site_id = get_env_value("SHAREPOINT_SITE_ID")
    drive_id = get_env_value("SHAREPOINT_DRIVE_ID")
    sender = get_env_value("SHAREPOINT_USER", required=True)
    return {
        "tenant_id": tenant_id,
        "client_id": client_id,
        "client_secret": client_secret,
        "site_id": site_id,
        "drive_id": drive_id,
        "sender": sender,
        "sp_host": get_env_value("SHAREPOINT_HOST") or DEFAULT_SP_HOST,
        "sp_site_path": get_env_value("SHAREPOINT_SITE_PATH") or DEFAULT_SP_SITE_PATH,
        "sp_library": get_env_value("SHAREPOINT_LIBRARY") or DEFAULT_SP_LIBRARY,
    }


def get_email_recipients() -> Tuple[List[str], List[str]]:
    to_raw = get_env_value("EMAIL_DESTINO", "SHAREPOINT_USER")
    cc_raw = get_env_value("EMAIL_CC")
    to_list = [item.strip() for item in to_raw.split(";") if item.strip()]
    cc_list = [item.strip() for item in cc_raw.split(";") if item.strip()]
    return _dedupe_preserve_order(to_list), _dedupe_preserve_order(cc_list)


def _dedupe_preserve_order(values: Iterable[str]) -> List[str]:
    output: List[str] = []
    seen = set()
    for value in values:
        lower = value.lower()
        if lower in seen:
            continue
        seen.add(lower)
        output.append(value)
    return output


def get_graph_token(force_refresh: bool = False) -> str:
    settings = get_graph_settings()
    with _token_lock:
        now_ts = datetime.now().timestamp()
        if not force_refresh and _token_cache["token"] and _token_cache["expires_at"] > now_ts + 300:
            return _token_cache["token"]

        app = ConfidentialClientApplication(
            client_id=settings["client_id"],
            client_credential=settings["client_secret"],
            authority=f"https://login.microsoftonline.com/{settings['tenant_id']}",
        )
        result = app.acquire_token_for_client(scopes=GRAPH_SCOPE)
        token = result.get("access_token")
        if not token:
            error = result.get("error_description") or result.get("error") or "No se pudo obtener token Graph"
            raise RuntimeError(error)

        _token_cache["token"] = token
        _token_cache["expires_at"] = now_ts + int(result.get("expires_in", 3600))
        return token


def _graph_headers(token: str, extra: Optional[Dict[str, str]] = None) -> Dict[str, str]:
    headers = {"Authorization": f"Bearer {token}"}
    if extra:
        headers.update(extra)
    return headers


def resolve_graph_location_ids(log_path: Optional[Path] = None) -> Tuple[str, str]:
    if _graph_location_cache["site_id"] and _graph_location_cache["drive_id"]:
        return _graph_location_cache["site_id"], _graph_location_cache["drive_id"]

    settings = get_graph_settings()
    site_id = settings.get("site_id", "").strip()
    drive_id = settings.get("drive_id", "").strip()
    if site_id and drive_id:
        _graph_location_cache["site_id"] = site_id
        _graph_location_cache["drive_id"] = drive_id
        return site_id, drive_id

    if not site_id:
        site_url = f"{GRAPH_BASE}/sites/{settings['sp_host']}:{settings['sp_site_path']}"
        site_resp = graph_request("GET", site_url, expected_statuses=(200,), retries=1, timeout=30)
        site_id = site_resp.json()["id"]
        log(
            f"[WARN] SHAREPOINT_SITE_ID ausente; se resolvio dinamicamente desde {settings['sp_host']}{settings['sp_site_path']}.",
            log_path,
        )

    if not drive_id:
        drives_url = f"{GRAPH_BASE}/sites/{site_id}/drives"
        drives_resp = graph_request("GET", drives_url, expected_statuses=(200,), retries=1, timeout=30)
        for drive in drives_resp.json().get("value", []):
            if drive.get("name") == settings["sp_library"]:
                drive_id = drive["id"]
                break
        if not drive_id:
            raise RuntimeError(
                f"No se encontro la biblioteca SharePoint '{settings['sp_library']}' para el sitio configurado."
            )
        log(
            f"[WARN] SHAREPOINT_DRIVE_ID ausente; se resolvio dinamicamente para biblioteca '{settings['sp_library']}'.",
            log_path,
        )

    _graph_location_cache["site_id"] = site_id
    _graph_location_cache["drive_id"] = drive_id
    return site_id, drive_id


def graph_request(
    method: str,
    url: str,
    *,
    expected_statuses: Sequence[int],
    retries: int = 1,
    timeout: int = 60,
    log_path: Optional[Path] = None,
    **kwargs: Any,
) -> requests.Response:
    last_error: Optional[Exception] = None
    headers = dict(kwargs.pop("headers", {}) or {})

    for attempt in range(retries + 1):
        token = get_graph_token(force_refresh=(attempt > 0))
        request_headers = {
            "Authorization": f"Bearer {token}",
            **headers,
        }
        try:
            response = requests.request(method, url, headers=request_headers, timeout=timeout, **kwargs)
            if response.status_code in expected_statuses:
                return response
            if attempt < retries:
                if response.status_code == 423:
                    log(
                        f"[WARN] Graph API intento {attempt + 1}/{retries + 1} devolvio 423 (archivo bloqueado); "
                        f"esperando 60s antes de reintentar.",
                        log_path,
                    )
                    time.sleep(60)
                else:
                    log(
                        f"[WARN] Graph API intento {attempt + 1}/{retries + 1} devolvio {response.status_code}; reintentando.",
                        log_path,
                    )
                continue
            response.raise_for_status()
        except Exception as exc:
            last_error = exc
            if attempt < retries:
                log(
                    f"[WARN] Graph API intento {attempt + 1}/{retries + 1} fallo; reintentando una vez.",
                    log_path,
                )
                continue
            raise

    if last_error:
        raise last_error
    raise RuntimeError("Graph API no retorno respuesta valida.")


def download_sharepoint_file(relative_path: str, log_path: Optional[Path] = None) -> Path:
    settings = get_graph_settings()
    site_id, drive_id = resolve_graph_location_ids(log_path=log_path)
    encoded_path = relative_path.replace("#", "%23")
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{encoded_path}:/content"
    )
    response = graph_request("GET", url, expected_statuses=(200,), retries=3, timeout=120, log_path=log_path)
    suffix = Path(relative_path).suffix or ".xlsx"
    with NamedTemporaryFile(delete=False, suffix=suffix) as handle:
        handle.write(response.content)
        return Path(handle.name)


def upload_sharepoint_file(local_path: Path, relative_path: str, log_path: Optional[Path] = None) -> None:
    site_id, drive_id = resolve_graph_location_ids(log_path=log_path)
    encoded_path = relative_path.replace("#", "%23")
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{encoded_path}:/content"
    )
    with open(local_path, "rb") as handle:
        graph_request(
            "PUT",
            url,
            expected_statuses=(200, 201),
            retries=3,
            timeout=180,
            data=handle.read(),
            headers={"Content-Type": "application/octet-stream"},
            log_path=log_path,
        )


def trigger_excel_recalculation(relative_path: str, log_path: Optional[Path] = None) -> bool:
    """
    Abre una sesion Excel Online para que Graph API recalcule las formulas del workbook.
    Necesario para que ET/EC (cols 40-41) tengan valores cacheados tras nuestro upload.
    Retorna True si el recalculo fue exitoso, False si no esta disponible (no critico).
    """
    try:
        site_id, drive_id = resolve_graph_location_ids(log_path=log_path)
        encoded_path = relative_path.replace("#", "%23")
        token = get_graph_token()

        # 1. Obtener item ID del archivo
        item_url = f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/root:/{encoded_path}"
        item_resp = graph_request("GET", item_url, expected_statuses=(200,), retries=2, timeout=30, log_path=log_path)
        item_id = item_resp.json()["id"]

        # 2. Crear sesion Excel (persistChanges=True para que guarde el recalculo)
        session_url = f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/items/{item_id}/workbook/createSession"
        session_resp = graph_request(
            "POST",
            session_url,
            expected_statuses=(201,),
            retries=2,
            timeout=60,
            json={"persistChanges": True},
            log_path=log_path,
        )
        session_id = session_resp.json().get("id", "")
        if not session_id:
            log("[WARN] Excel session: no se obtuvo session ID.", log_path)
            return False

        session_headers = {"workbook-session-id": session_id}

        # 3. Forzar recalculo completo
        calc_url = f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/items/{item_id}/workbook/application/calculate"
        graph_request(
            "POST",
            calc_url,
            expected_statuses=(200, 204),
            retries=2,
            timeout=120,
            json={"calculationType": "Full"},
            headers=session_headers,
            log_path=log_path,
        )

        # 4. Cerrar sesion (commit)
        close_url = f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/items/{item_id}/workbook/closeSession"
        try:
            graph_request(
                "POST",
                close_url,
                expected_statuses=(204,),
                retries=1,
                timeout=30,
                headers=session_headers,
                log_path=log_path,
            )
        except Exception:
            pass  # No critico si falla el cierre

        log("[SP] Recalculo Excel Online OK.", log_path)
        return True

    except Exception as exc:
        log(f"[WARN] Recalculo Excel Online no disponible (no critico): {exc}", log_path)
        return False


def recalculation_settle_wait(file_size_kb: int, log_path: Optional[Path] = None) -> None:
    """
    Espera inicial proporcional al tamaño del archivo para que Excel Online comience
    a persistir el recalculo. Los reintentos usan su propio wait adicional.
    """
    if file_size_kb >= 5_000:
        wait_s = 20
    elif file_size_kb >= 500:
        wait_s = 10
    else:
        wait_s = 5
    log(f"[SP] Esperando {wait_s}s para que Excel Online persista el recalculo ({file_size_kb} KB)...", log_path)
    time.sleep(wait_s)


def send_summary_email(asunto: str, html_body: str, log_path: Optional[Path] = None) -> bool:
    settings = get_graph_settings()
    to_list, cc_list = get_email_recipients()
    if not to_list:
        log("[WARN] No hay destinatarios configurados para el correo final.", log_path)
        return False

    payload = {
        "message": {
            "subject": asunto,
            "body": {"contentType": "HTML", "content": html_body},
            "toRecipients": [{"emailAddress": {"address": email}} for email in to_list],
            "ccRecipients": [{"emailAddress": {"address": email}} for email in cc_list],
        },
        "saveToSentItems": True,
    }
    url = f"{GRAPH_BASE}/users/{settings['sender']}/sendMail"
    graph_request("POST", url, expected_statuses=(202,), retries=3, timeout=60, json=payload, log_path=log_path)
    return True


def get_reporting_window(reference_date: Optional[date] = None) -> Tuple[date, date]:
    """
    Calcula ventana de fechas para descarga WMS.

    Reglas:
    - Backfill (--mes): Mes completo especificado hasta último día
    - Días 1-10 del mes: Ventana desde INICIO DEL MES ANTERIOR hasta HOY
      -> Captura pedidos del mes anterior que se completan en primeros 10 días
      -> Permite actualizar OTIF con pedidos que se finalizan tarde
    - Día 11+: Ventana desde INICIO DEL MES ACTUAL hasta HOY
      -> Operación normal (mes actual acumulado)
    """
    import calendar as _cal
    today = reference_date or date.today()
    real_today = date.today()

    # BACKFILL (--mes): Mes completo hasta último día
    if reference_date and (today.year, today.month) < (real_today.year, real_today.month):
        start = date(today.year, today.month, 1)
        last_day = _cal.monthrange(today.year, today.month)[1]
        return start, date(today.year, today.month, last_day)

    # DÍAS 1-10: Ventana desde inicio de mes ANTERIOR hasta HOY
    # Captura pedidos del mes anterior que se completan en la primera semana
    if 1 <= real_today.day <= 10 and reference_date is None:
        last_month_end = date(real_today.year, real_today.month, 1) - timedelta(days=1)
        last_month_start = date(last_month_end.year, last_month_end.month, 1)
        return last_month_start, real_today

    # DÍA 11+: Ventana normal (solo mes actual hasta HOY)
    start = date(today.year, today.month, 1)
    return start, real_today


def format_wms_date(value: date) -> str:
    return value.strftime("%d/%m/%y")


def slugify_filename(value: str) -> str:
    cleaned = sanitize_log_message(value)
    cleaned = cleaned.replace(" ", "_").replace("/", "_").replace("\\", "_")
    return re.sub(r"[^A-Za-z0-9_\-\.]", "", cleaned) or "cliente"


def find_downloaded_file(download_dir: Path, created_after: datetime) -> Optional[Path]:
    candidates: List[Path] = []
    for item in download_dir.iterdir():
        if not item.is_file():
            continue
        if item.suffix.lower() not in DOWNLOAD_SUFFIXES:
            continue
        if not any(item.name.startswith(prefix) for prefix in DOWNLOAD_BASENAME_PREFIXES):
            continue
        modified = datetime.fromtimestamp(item.stat().st_mtime)
        if modified >= created_after - timedelta(seconds=2):
            candidates.append(item)
    if not candidates:
        return None
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return candidates[0]


def move_download_to_log(source_path: Path, client_name: str) -> Path:
    extension = source_path.suffix or ".xlsx"
    destination = LOG_DIR / f"temp_fillrate_{slugify_filename(client_name)}{extension}"
    if destination.exists():
        destination.unlink()
    source_path.replace(destination)
    return destination


def coerce_datetime(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        patterns = (
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%y %H:%M:%S",
            "%d/%m/%y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d/%m/%y",
        )
        for pattern in patterns:
            try:
                return datetime.strptime(raw, pattern)
            except ValueError:
                continue
    return None


def normalize_state(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_fillrate_rows(download_path: Path) -> List[List[Any]]:
    """
    Lee las filas del Excel descargado del WMS.
    Retorna filas extendidas hasta MAX_SUPPORTED_FORMULA_COL para capturar
    ET/EC (cols 40-41) si el reporte WMS los incluye como valores cacheados.
    deduplicate_exact_rows y la escritura a SharePoint usan solo las primeras
    DATA_COLUMNS (26) columnas.
    """
    workbook = load_workbook(download_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: List[List[Any]] = []
    for excel_row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        data = list(excel_row[:DATA_COLUMNS])
        if not any(value not in (None, "") for value in data):
            continue
        if len(data) < DATA_COLUMNS:
            data.extend([None] * (DATA_COLUMNS - len(data)))
        # Extender hasta MAX_SUPPORTED_FORMULA_COL para ET/EC si el WMS los incluye
        extended = list(excel_row[:MAX_SUPPORTED_FORMULA_COL])
        if len(extended) < MAX_SUPPORTED_FORMULA_COL:
            extended.extend([None] * (MAX_SUPPORTED_FORMULA_COL - len(extended)))
        rows.append(extended)
    workbook.close()
    return rows


def compute_pending_from_wms_rows(rows: List[List[Any]], month: int, year: int) -> Dict[str, Any]:
    """Calcula pedidos pendientes desde las filas crudas del WMS (sin necesidad del SharePoint)."""
    en_preparacion: set = set()
    preparados: set = set()
    fechas: List[datetime] = []
    for row in rows:
        fecha = row[DATE_COLUMN_INDEX - 1]
        if not isinstance(fecha, (datetime, date)):
            continue
        f = fecha if isinstance(fecha, datetime) else datetime.combine(fecha, datetime.min.time())
        if f.month != month or f.year != year:
            continue
        estado = str(row[STATE_COLUMN_INDEX - 1]).strip() if row[STATE_COLUMN_INDEX - 1] else ""
        nro = row[ORDER_COLUMN_INDEX - 1]
        if estado in ("En Preparacion", "En Preparación"):
            en_preparacion.add(nro)
            fechas.append(f)
        elif estado in ("Preparado", "Preparados"):
            preparados.add(nro)
            fechas.append(f)
    return {
        "en_preparacion": len(en_preparacion),
        "preparados": len(preparados),
        "total": len(en_preparacion) + len(preparados),
        "mas_antiguo": min(fechas).date() if fechas else None,
    }


def compute_otif_from_wms_rows(rows: List[List[Any]], month: int, year: int) -> Dict[str, Any]:
    """
    Calcula OTIF desde las filas crudas del WMS usando ET (col 40) y EC (col 41).
    Requiere que el reporte WMS exporte esas columnas con valores cacheados.
    Si no hay valores (WMS no las exporta), retorna pedidos=0.
    """
    pedidos: Dict[Any, Dict[str, bool]] = {}
    for row in rows:
        fecha = row[DATE_COLUMN_INDEX - 1]
        if not isinstance(fecha, (datetime, date)):
            continue
        f = fecha if isinstance(fecha, datetime) else datetime.combine(fecha, datetime.min.time())
        if f.month != month or f.year != year:
            continue
        et_raw = row[COL_ENTREGADO_TIEMPO - 1] if len(row) >= COL_ENTREGADO_TIEMPO else None
        if et_raw is None:
            continue
        nro = row[ORDER_COLUMN_INDEX - 1]
        et = str(et_raw).strip().upper() == "SI"
        ec_raw = row[COL_ENTREGADO_COMPLETO - 1] if len(row) >= COL_ENTREGADO_COMPLETO else None
        ec = str(ec_raw).strip().upper() == "SI" if ec_raw is not None else False
        if nro not in pedidos:
            pedidos[nro] = {"on_time": et, "in_full": ec}
        else:
            pedidos[nro]["on_time"] = pedidos[nro]["on_time"] and et
            pedidos[nro]["in_full"] = pedidos[nro]["in_full"] and ec
    total = len(pedidos)
    if total == 0:
        return {"pedidos": 0, "pct_on_time": None, "pct_in_full": None, "pct_otif": None}
    ot = sum(1 for p in pedidos.values() if p["on_time"])
    inf = sum(1 for p in pedidos.values() if p["in_full"])
    otif = sum(1 for p in pedidos.values() if p["on_time"] and p["in_full"])
    return {
        "pedidos": total,
        "pct_on_time": round(ot / total * 100, 1),
        "pct_in_full": round(inf / total * 100, 1),
        "pct_otif": round(otif / total * 100, 1),
    }


def deduplicate_exact_rows(rows: Sequence[Sequence[Any]]) -> Tuple[List[List[Any]], int]:
    """
    Elimina duplicados exactos del bloque mensual descargado.
    La comparacion usa las 26 columnas de datos en el orden original.
    """
    unique_rows: List[List[Any]] = []
    seen = set()
    duplicates_removed = 0

    for row in rows:
        row_key = tuple(row[:DATA_COLUMNS])
        if row_key in seen:
            duplicates_removed += 1
            continue
        seen.add(row_key)
        unique_rows.append(list(row[:DATA_COLUMNS]))

    return unique_rows, duplicates_removed


def build_month_row_key(row_values: Sequence[Any]) -> Tuple[Any, ...]:
    """
    Llave de negocio para preservar overrides manuales al reescribir el mes actual.
    Evita columnas que cambian por formula o por estados derivados.
    """
    key_indexes = (1, 2, 4, 5, 6, 8, 9)  # A, B, D, E, F, H, I
    normalized: List[Any] = []
    for column_number in key_indexes:
        value = row_values[column_number - 1] if len(row_values) >= column_number else None
        if isinstance(value, datetime):
            normalized.append(value.isoformat())
        elif isinstance(value, date):
            normalized.append(datetime.combine(value, datetime.min.time()).isoformat())
        else:
            normalized.append(value)
    return tuple(normalized)


def collect_manual_overrides_for_month(
    target_sheet: Worksheet,
    row_indexes: Sequence[int],
) -> Dict[Tuple[Any, ...], Dict[int, Any]]:
    overrides: Dict[Tuple[Any, ...], Dict[int, Any]] = {}

    for row_idx in row_indexes:
        row_values = [target_sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, DATA_COLUMNS + 1)]
        row_key = build_month_row_key(row_values)
        row_overrides: Dict[int, Any] = {}
        for col_idx in MANUAL_OVERRIDE_COLUMNS:
            cell_value = target_sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_value, str) and cell_value.startswith("="):
                continue
            if cell_value in (None, ""):
                continue
            row_overrides[col_idx] = cell_value
        if row_overrides:
            overrides[row_key] = row_overrides

    return overrides


def normalize_header_name(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
    }
    for original, replacement in replacements.items():
        text = text.replace(original, replacement)
    return text


def find_header_column(target_sheet: Worksheet, header_name: str) -> Optional[int]:
    wanted = normalize_header_name(header_name)
    for col_idx in range(1, target_sheet.max_column + 1):
        current = normalize_header_name(target_sheet.cell(row=1, column=col_idx).value)
        if current == wanted:
            return col_idx
    return None


def fill_corte_column(target_sheet: Worksheet, corte_col_idx: int, meses_corte: Dict[int, str]) -> None:
    for row_idx in range(DATA_START_ROW, target_sheet.max_row + 1):
        fecha_gen = coerce_datetime(target_sheet.cell(row=row_idx, column=DATE_COLUMN_INDEX).value)
        target_sheet.cell(row=row_idx, column=corte_col_idx).value = (
            calcular_corte(fecha_gen.date(), meses_corte) if fecha_gen is not None else None
        )


def ensure_corte_column(
    target_sheet: Worksheet,
    meses_corte: Dict[int, str],
    log_path: Optional[Path] = None,
) -> int:
    existing_col = find_header_column(target_sheet, "Corte")
    if existing_col:
        fill_corte_column(target_sheet, existing_col, meses_corte)
        return existing_col

    new_col_idx = target_sheet.max_column + 1
    header_cell = target_sheet.cell(row=1, column=new_col_idx)
    target_cell_template = target_sheet.cell(row=FORMULA_TEMPLATE_ROW, column=new_col_idx)

    source_header = target_sheet.cell(row=1, column=new_col_idx - 1)
    source_template = target_sheet.cell(row=FORMULA_TEMPLATE_ROW, column=new_col_idx - 1)

    set_cell_value_like_template(header_cell, source_header, "Corte")
    set_cell_value_like_template(target_cell_template, source_template, None)

    source_letter = get_column_letter(new_col_idx - 1)
    target_letter = get_column_letter(new_col_idx)
    target_sheet.column_dimensions[target_letter].width = target_sheet.column_dimensions[source_letter].width

    for row_idx in range(DATA_START_ROW, target_sheet.max_row + 1):
        source_cell = target_sheet.cell(row=row_idx, column=new_col_idx - 1)
        target_cell = target_sheet.cell(row=row_idx, column=new_col_idx)
        fecha_gen = coerce_datetime(target_sheet.cell(row=row_idx, column=DATE_COLUMN_INDEX).value)
        set_cell_value_like_template(
            target_cell,
            source_cell,
            calcular_corte(fecha_gen.date(), meses_corte) if fecha_gen is not None else None,
        )

    log(
        f"[WARN] Se agrego columna 'Corte' al archivo '{target_sheet.title}' en posicion {new_col_idx}.",
        log_path,
    )
    return new_col_idx


def build_warnings(client_name: str, rows: Sequence[Sequence[Any]], today: Optional[date] = None) -> List[Dict[str, Any]]:
    today = today or date.today()
    warnings: List[Dict[str, Any]] = []
    seen_pedidos: set = set()
    for row in rows:
        estado = normalize_state(row[STATE_COLUMN_INDEX - 1])
        if estado not in ESTADOS_ALERTA:
            continue
        fecha_gen = coerce_datetime(row[DATE_COLUMN_INDEX - 1])
        if fecha_gen is None:
            continue
        dias = (today - fecha_gen.date()).days
        if dias > WARNING_MAX_DAYS:
            nro_pedido = row[ORDER_COLUMN_INDEX - 1]
            if nro_pedido in seen_pedidos:
                continue
            seen_pedidos.add(nro_pedido)
            warnings.append(
                {
                    "cliente": client_name,
                    "nro_pedido": nro_pedido,
                    "estado": estado,
                    "dias_transcurridos": dias,
                }
            )
    return warnings


def compute_pending_summary(ws, month: int, year: int) -> Dict[str, Any]:
    """
    Cuenta pedidos únicos en estado Preparado / En Preparación para el mes dado.
    Filtra por col 9 (Fecha y hora de Ingreso) dentro del mes/año.
    """
    en_preparacion: set = set()
    preparados: set = set()
    fechas: List[datetime] = []

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row[0]:
            continue
        fecha = row[DATE_COLUMN_INDEX - 1]
        if not isinstance(fecha, (datetime, date)):
            continue
        f = fecha if isinstance(fecha, datetime) else datetime.combine(fecha, datetime.min.time())
        if f.month != month or f.year != year:
            continue
        estado = str(row[STATE_COLUMN_INDEX - 1]).strip() if row[STATE_COLUMN_INDEX - 1] else ""
        nro = row[ORDER_COLUMN_INDEX - 1]
        if estado in ("En Preparacion", "En Preparación"):
            en_preparacion.add(nro)
            fechas.append(f)
        elif estado in ("Preparado", "Preparados"):
            preparados.add(nro)
            fechas.append(f)

    return {
        "en_preparacion": len(en_preparacion),
        "preparados": len(preparados),
        "total": len(en_preparacion) + len(preparados),
        "mas_antiguo": min(fechas).date() if fechas else None,
    }


def compute_otif_summary(ws, month: int, year: int) -> Dict[str, Any]:
    """
    Calcula % On Time, % In Full y % OTIF agrupado por Nro Pedido (col 5)
    para el mes dado. Solo incluye pedidos con ET no nulo (entregados).
    """
    pedidos: Dict[Any, Dict[str, bool]] = {}

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row[0]:
            continue
        fecha = row[DATE_COLUMN_INDEX - 1]
        if not isinstance(fecha, (datetime, date)):
            continue
        f = fecha if isinstance(fecha, datetime) else datetime.combine(fecha, datetime.min.time())
        if f.month != month or f.year != year:
            continue
        et_raw = row[COL_ENTREGADO_TIEMPO - 1]
        if et_raw is None:
            continue  # pendiente, excluir del cálculo
        nro = row[ORDER_COLUMN_INDEX - 1]
        et = str(et_raw).strip().upper() == "SI"
        ec_raw = row[COL_ENTREGADO_COMPLETO - 1]
        ec = str(ec_raw).strip().upper() == "SI" if ec_raw is not None else False
        if nro not in pedidos:
            pedidos[nro] = {"on_time": et, "in_full": ec}
        else:
            pedidos[nro]["on_time"] = pedidos[nro]["on_time"] and et
            pedidos[nro]["in_full"] = pedidos[nro]["in_full"] and ec

    total = len(pedidos)
    if total == 0:
        return {"pedidos": 0, "pct_on_time": None, "pct_in_full": None, "pct_otif": None}

    ot = sum(1 for p in pedidos.values() if p["on_time"])
    inf = sum(1 for p in pedidos.values() if p["in_full"])
    otif = sum(1 for p in pedidos.values() if p["on_time"] and p["in_full"])
    return {
        "pedidos": total,
        "pct_on_time": round(ot / total * 100, 1),
        "pct_in_full": round(inf / total * 100, 1),
        "pct_otif": round(otif / total * 100, 1),
    }


def ajustar_formula(formula_template: str, fila_template: int, fila_nueva: int) -> str:
    if not isinstance(formula_template, str) or not formula_template.startswith("="):
        return formula_template
    pattern = rf"(?<!\$)([A-Z]{{1,3}})(?<!\$)({fila_template})(?!\d)"
    return re.sub(pattern, lambda match: f"{match.group(1)}{fila_nueva}", formula_template)


def get_target_sheet(workbook, log_path: Optional[Path] = None) -> Tuple[Worksheet, bool]:
    lower_map = {name.strip().lower(): name for name in workbook.sheetnames}
    if TARGET_SHEET_NAME.lower() in lower_map:
        return workbook[lower_map[TARGET_SHEET_NAME.lower()]], False

    fallback_name = workbook.sheetnames[0]
    if fallback_name.strip().lower() == BASE_SHEET_NAME.lower() and len(workbook.sheetnames) > 1:
        fallback_name = workbook.sheetnames[1]
    log(
        f"[WARN] Hoja '{TARGET_SHEET_NAME}' no existe; se usara fallback '{fallback_name}'.",
        log_path,
    )
    return workbook[fallback_name], True


def should_remove_row_by_month(row_date: Optional[datetime], month: int, year: int) -> bool:
    return bool(row_date and row_date.month == month and row_date.year == year)


def copy_formula_templates(target_sheet: Worksheet, formula_end_col: int) -> Dict[int, Any]:
    templates: Dict[int, Any] = {}
    for col_idx in range(FORMULA_START_COL, formula_end_col + 1):
        templates[col_idx] = target_sheet.cell(row=FORMULA_TEMPLATE_ROW, column=col_idx).value
    return templates


def set_cell_value_like_template(target_cell, template_cell, value: Any) -> None:
    target_cell.value = value
    if template_cell.has_style:
        target_cell._style = copy(template_cell._style)
    if template_cell.number_format:
        target_cell.number_format = template_cell.number_format
    if template_cell.font:
        target_cell.font = copy(template_cell.font)
    if template_cell.fill:
        target_cell.fill = copy(template_cell.fill)
    if template_cell.border:
        target_cell.border = copy(template_cell.border)
    if template_cell.alignment:
        target_cell.alignment = copy(template_cell.alignment)
    if template_cell.protection:
        target_cell.protection = copy(template_cell.protection)


def calcular_corte(fecha_valor: date, meses_es: Dict[int, str]) -> str:
    if fecha_valor.day >= 16:
        mes_1 = fecha_valor.month
        mes_2 = (fecha_valor.month % 12) + 1
        return f"{meses_es[mes_1]}-{meses_es[mes_2]} {fecha_valor.year}"
    mes_2 = fecha_valor.month
    mes_1 = 12 if mes_2 == 1 else mes_2 - 1
    return f"{meses_es[mes_1]}-{meses_es[mes_2]} {fecha_valor.year}"


def update_sharepoint_workbook(
    client: Dict[str, Any],
    new_rows: Sequence[Sequence[Any]],
    *,
    log_path: Optional[Path] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    meses_corte: Optional[Dict[int, str]] = None,
) -> Dict[str, Any]:
    if not new_rows:
        return {
            "relative_path": get_sharepoint_relative_path(client),
            "target_sheet": None,
            "used_fallback_sheet": False,
            "replaced_rows": 0,
            "new_rows": 0,
            "duplicates_removed": 0,
        }

    unique_rows, duplicates_removed = deduplicate_exact_rows(new_rows)
    sample_date = coerce_datetime(unique_rows[0][DATE_COLUMN_INDEX - 1])
    if sample_date is None and (month is None or year is None):
        raise RuntimeError("No se pudo inferir mes y anio desde los datos nuevos.")

    month = month or sample_date.month
    year = year or sample_date.year
    meses_corte = meses_corte or {}

    relative_path = get_sharepoint_relative_path(client)
    log(f"[SP] Descargando archivo SharePoint: {relative_path}", log_path)
    local_copy = download_sharepoint_file(relative_path, log_path=log_path)
    log(f"[SP] Archivo descargado OK ({local_copy.stat().st_size // 1024} KB).", log_path)

    try:
        log("[SP] Cargando workbook para modificacion...", log_path)
        workbook = load_workbook(local_copy)
        log("[SP] Workbook cargado OK.", log_path)
        target_sheet, used_fallback_sheet = get_target_sheet(workbook, log_path)
        corte_col_idx = ensure_corte_column(target_sheet, meses_corte, log_path=log_path)
        formula_end_col = min(target_sheet.max_column, MAX_SUPPORTED_FORMULA_COL)
        formula_templates = copy_formula_templates(target_sheet, formula_end_col)

        rows_to_delete: List[int] = []
        for row_idx in range(DATA_START_ROW, target_sheet.max_row + 1):
            row_date = coerce_datetime(target_sheet.cell(row=row_idx, column=DATE_COLUMN_INDEX).value)
            if should_remove_row_by_month(row_date, month, year):
                rows_to_delete.append(row_idx)

        manual_overrides = collect_manual_overrides_for_month(target_sheet, rows_to_delete)

        for row_idx in reversed(rows_to_delete):
            target_sheet.delete_rows(row_idx, 1)

        template_cells = {
            col_idx: target_sheet.cell(row=FORMULA_TEMPLATE_ROW, column=col_idx)
            for col_idx in range(1, formula_end_col + 1)
        }

        start_row = target_sheet.max_row + 1
        for offset, row_values in enumerate(unique_rows):
            row_number = start_row + offset
            for col_idx in range(1, DATA_COLUMNS + 1):
                template_cell = template_cells[col_idx]
                target_cell = target_sheet.cell(row=row_number, column=col_idx)
                set_cell_value_like_template(target_cell, template_cell, row_values[col_idx - 1])

            for col_idx in range(FORMULA_START_COL, formula_end_col + 1):
                template_cell = template_cells[col_idx]
                target_cell = target_sheet.cell(row=row_number, column=col_idx)
                template_value = formula_templates.get(col_idx)
                value_to_write = ajustar_formula(template_value, FORMULA_TEMPLATE_ROW, row_number)
                if col_idx == corte_col_idx:
                    fecha_gen = coerce_datetime(row_values[DATE_COLUMN_INDEX - 1])
                    value_to_write = (
                        calcular_corte(fecha_gen.date(), meses_corte)
                        if fecha_gen is not None
                        else None
                    )
                set_cell_value_like_template(target_cell, template_cell, value_to_write)

            row_key = build_month_row_key(row_values)
            if row_key in manual_overrides:
                for col_idx, override_value in manual_overrides[row_key].items():
                    template_cell = template_cells[col_idx]
                    target_cell = target_sheet.cell(row=row_number, column=col_idx)
                    set_cell_value_like_template(target_cell, template_cell, override_value)

        log("[SP] Guardando workbook modificado...", log_path)
        workbook.save(local_copy)
        log(f"[SP] Workbook guardado ({local_copy.stat().st_size // 1024} KB). Subiendo a SharePoint...", log_path)
        upload_sharepoint_file(local_copy, relative_path, log_path=log_path)

        # Forzar recalculo Excel Online para que ET/EC (cols 40-41) tengan valores cacheados
        recalc_ok = trigger_excel_recalculation(relative_path, log_path=log_path)

        # Leer OTIF/pendientes del archivo recalculado con verificacion de persistencia.
        # En lugar de esperar un tiempo fijo, verificamos que ET no sea todo None
        # (señal de que el recalculo ya persisto). Hasta 3 intentos con backoff.
        pending_data: Dict[str, Any] = {"en_preparacion": 0, "preparados": 0, "total": 0, "mas_antiguo": None}
        otif_data: Dict[str, Any] = {"pedidos": 0, "pct_on_time": None, "pct_in_full": None, "pct_otif": None}
        if recalc_ok:
            file_size_kb = local_copy.stat().st_size // 1024
            max_attempts = 3
            retry_wait_s = 15 if file_size_kb >= 5_000 else 10
            recalc_confirmed = False

            for attempt in range(1, max_attempts + 1):
                recalculation_settle_wait(file_size_kb, log_path=log_path)
                recalc_copy: Optional[Path] = None
                try:
                    recalc_copy = download_sharepoint_file(relative_path, log_path=log_path)
                    wb_recalc = load_workbook(recalc_copy, data_only=True, read_only=True)
                    ws_recalc, _ = get_target_sheet(wb_recalc, log_path)

                    # Verificar que al menos una celda ET no sea None (recalculo persistido)
                    et_found = False
                    for row in ws_recalc.iter_rows(min_row=DATA_START_ROW, values_only=True):
                        if len(row) >= COL_ENTREGADO_TIEMPO and row[COL_ENTREGADO_TIEMPO - 1] is not None:
                            et_found = True
                            break

                    if not et_found:
                        wb_recalc.close()
                        if attempt < max_attempts:
                            log(
                                f"[SP] Recalculo aun no persiste (intento {attempt}/{max_attempts}). "
                                f"Reintentando en {retry_wait_s}s...",
                                log_path,
                            )
                            time.sleep(retry_wait_s)
                            continue
                        else:
                            log(
                                f"[WARN] Recalculo no persistio tras {max_attempts} intentos. "
                                "OTIF quedara sin datos para este cliente.",
                                log_path,
                            )
                            break

                    pending_data = compute_pending_summary(ws_recalc, month, year)
                    otif_data = compute_otif_summary(ws_recalc, month, year)
                    wb_recalc.close()
                    recalc_confirmed = True
                    log(
                        f"[SP] OTIF post-recalculo (intento {attempt}) — "
                        f"pendientes={pending_data['total']}, otif_pedidos={otif_data['pedidos']}.",
                        log_path,
                    )
                    break

                except Exception as exc:
                    log(f"[WARN] Lectura OTIF intento {attempt}/{max_attempts} fallo: {exc}", log_path)
                    if attempt < max_attempts:
                        time.sleep(retry_wait_s)
                finally:
                    if recalc_copy:
                        try:
                            recalc_copy.unlink(missing_ok=True)
                        except Exception:
                            pass

            if not recalc_confirmed:
                log("[WARN] OTIF no pudo verificarse post-recalculo (no critico).", log_path)

        return {
            "relative_path": relative_path,
            "target_sheet": target_sheet.title,
            "used_fallback_sheet": used_fallback_sheet,
            "replaced_rows": len(rows_to_delete),
            "new_rows": len(unique_rows),
            "duplicates_removed": duplicates_removed,
            "pendientes": pending_data,
            "otif": otif_data,
        }
    finally:
        try:
            local_copy.unlink(missing_ok=True)
        except Exception:
            pass


@dataclass
class ClientExecutionResult:
    cliente: str
    cd: str
    estado: str
    filas_nuevas: Optional[int] = None
    filas_reemplazadas: Optional[int] = None
    advertencias: int = 0
    detalle: str = ""
    used_fallback_sheet: bool = False
    target_sheet: str = ""
    retried_success: bool = False
    pendientes: Optional[Dict[str, Any]] = None
    otif: Optional[Dict[str, Any]] = None


def build_summary_html(results: Sequence[ClientExecutionResult], warnings: Sequence[Dict[str, Any]]) -> str:
    total_clientes = len(results)
    ok_count = sum(1 for item in results if item.estado in ("OK", "Ya descargado"))
    sin_datos_count = sum(1 for item in results if item.estado == "Sin datos")
    omitidos_count = sum(1 for item in results if item.estado == "Omitido")
    error_count = sum(1 for item in results if item.estado == "Error")
    warning_count = len(warnings)

    if error_count > 0:
        overall_state = "CON FALLOS"
        header_color = "#c0392b"
        header_bg = "#fff3f0"
        header_icon = "&#10060;"
        header_text = "La corrida termino con errores en uno o mas clientes."
    elif warning_count > 0:
        overall_state = "CON ADVERTENCIAS"
        header_color = "#d97706"
        header_bg = "#fff7ed"
        header_icon = "&#9888;&#65039;"
        header_text = "La corrida termino correctamente, pero hay pedidos pendientes a revisar."
    else:
        overall_state = "TODO OK"
        header_color = "#1f7a4c"
        header_bg = "#eefaf3"
        header_icon = "&#9989;"
        header_text = "La corrida termino correctamente y sin observaciones relevantes."

    def render_status_badge(item: ClientExecutionResult) -> Tuple[str, str]:
        if item.estado == "OK":
            return ("&#9989; OK &#8635;" if item.retried_success else "&#9989; OK"), "#ecfdf3"
        if item.estado == "Ya descargado":
            return "&#9989; OK", "#ecfdf3"
        if item.estado == "Sin datos":
            return "&#128196; Sin datos", "#f8fafc"
        if item.estado == "Omitido":
            return "&#9209; Omitido", "#f8fafc"
        return "&#10060; Error", "#fff1f2"

    rows_html = []
    for item in results:
        filas_nuevas = item.filas_nuevas if item.filas_nuevas is not None else "—"
        filas_reemplazadas = item.filas_reemplazadas if item.filas_reemplazadas is not None else "—"
        warning_text = str(item.advertencias) if item.advertencias is not None else "—"
        detail_bits: List[str] = []
        if item.used_fallback_sheet and item.target_sheet:
            detail_bits.append(f"Fallback hoja: {item.target_sheet}")
        if item.detalle:
            detail_bits.append(item.detalle)
        detalle = " | ".join(detail_bits)
        detalle_html = f"<br><span style='font-size:11px;color:#475569'>{detalle}</span>" if detalle else ""
        status_html, row_bg = render_status_badge(item)
        rows_html.append(
            f"""
            <tr style="background:{row_bg}">
              <td style="padding:8px 10px;border:1px solid #d9e2ec">{item.cliente}{detalle_html}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec">{item.cd}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec;font-weight:bold">{status_html}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:right">{filas_nuevas}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:right">{filas_reemplazadas}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:right">{warning_text}</td>
            </tr>
            """
        )

    # ── Tabla 2: Pendientes del mes ──────────────────────────────────────────
    mes_label = datetime.now().strftime("%B %Y").capitalize()
    pendientes_rows = [r for r in results if r.pendientes and r.pendientes.get("total", 0) > 0]
    if pendientes_rows:
        pend_rows_html = []
        for item in pendientes_rows:
            p = item.pendientes
            # mas_antiguo puede ser datetime.date o str (desde checkpoint JSON)
            ma = p.get("mas_antiguo")
            if ma:
                if isinstance(ma, str):
                    # Ya es string formato YYYY-MM-DD, convertir a DD/MM/YYYY
                    try:
                        # datetime ya está importado a nivel de módulo
                        parsed = datetime.strptime(ma, "%Y-%m-%d")
                        mas_antiguo = parsed.strftime("%d/%m/%Y")
                    except ValueError:
                        mas_antiguo = ma  # Usar tal cual si no se puede parsear
                else:
                    # Es datetime.date
                    mas_antiguo = ma.strftime("%d/%m/%Y")
            else:
                mas_antiguo = "—"
            total = p.get("total", 0)
            row_bg = "#fff7ed" if total > 0 else "#ffffff"
            pend_rows_html.append(f"""
            <tr style="background:{row_bg}">
              <td style="padding:8px 10px;border:1px solid #d9e2ec">{item.cliente}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec">{item.cd}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:center">{p.get("en_preparacion", 0)}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:center">{p.get("preparados", 0)}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:center;font-weight:bold">{total}</td>
              <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:center">{mas_antiguo}</td>
            </tr>""")
        pendientes_html = f"""
        <div style="margin-top:22px">
          <div style="font-family:Calibri,Arial,sans-serif;font-size:16px;font-weight:bold;color:#1e3a5f;margin-bottom:10px">
            Pedidos pendientes — {mes_label}
          </div>
          <table style="border-collapse:collapse;width:100%;background:#ffffff;border:1px solid #d9e2ec">
            <thead>
              <tr style="background:#2f4358;color:#ffffff">
                <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:left">Cliente</th>
                <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:left">CD</th>
                <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:center">En Preparacion</th>
                <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:center">Preparados</th>
                <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:center">Total</th>
                <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:center">Mas antiguo</th>
              </tr>
            </thead>
            <tbody>{''.join(pend_rows_html)}</tbody>
          </table>
        </div>"""
    else:
        pendientes_html = ""

    # ── Tabla 3: OTIF del mes ─────────────────────────────────────────────────
    def _otif_color(pct: Optional[float]) -> Tuple[str, str]:
        if pct is None:
            return "#ffffff", "#334155"
        if pct >= OTIF_UMBRAL_VERDE:
            return "#ecfdf3", "#1f7a4c"
        if pct >= OTIF_UMBRAL_AMARILLO:
            return "#fff7ed", "#d97706"
        return "#fff3f0", "#c0392b"

    def _fmt_pct(pct: Optional[float]) -> str:
        return f"{pct:.1f}%" if pct is not None else "—"

    otif_rows_html = []
    for item in results:
        if item.estado == "Omitido":
            continue
        o = item.otif
        if o and o.get("pedidos", 0) > 0:
            ot_bg, ot_fg = _otif_color(o.get("pct_on_time"))
            if_bg, if_fg = _otif_color(o.get("pct_in_full"))
            otif_bg, otif_fg = _otif_color(o.get("pct_otif"))
            pedidos_cell = str(o["pedidos"])
        else:
            ot_bg = if_bg = otif_bg = "#f8fafc"
            ot_fg = if_fg = otif_fg = "#94a3b8"
            pedidos_cell = "—"
            o = {}

        otif_rows_html.append(f"""
        <tr>
          <td style="padding:8px 10px;border:1px solid #d9e2ec">{item.cliente}</td>
          <td style="padding:8px 10px;border:1px solid #d9e2ec">{item.cd}</td>
          <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:center">{pedidos_cell}</td>
          <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:center;background:{ot_bg};color:{ot_fg};font-weight:bold">{_fmt_pct(o.get("pct_on_time"))}</td>
          <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:center;background:{if_bg};color:{if_fg};font-weight:bold">{_fmt_pct(o.get("pct_in_full"))}</td>
          <td style="padding:8px 10px;border:1px solid #d9e2ec;text-align:center;background:{otif_bg};color:{otif_fg};font-weight:bold">{_fmt_pct(o.get("pct_otif"))}</td>
        </tr>""")

    otif_html = f"""
    <div style="margin-top:22px">
      <div style="font-family:Calibri,Arial,sans-serif;font-size:16px;font-weight:bold;color:#1e3a5f;margin-bottom:10px">
        OTIF — {mes_label}
        <span style="font-size:11px;font-weight:normal;color:#5b6b7f;margin-left:8px">
          (pedidos entregados del mes agrupados por Nro Pedido)
        </span>
      </div>
      <table style="border-collapse:collapse;width:100%;background:#ffffff;border:1px solid #d9e2ec">
        <thead>
          <tr style="background:#2f4358;color:#ffffff">
            <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:left">Cliente</th>
            <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:left">CD</th>
            <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:center">Pedidos</th>
            <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:center">% On Time</th>
            <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:center">% In Full</th>
            <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:center">% OTIF</th>
          </tr>
        </thead>
        <tbody>{''.join(otif_rows_html)}</tbody>
      </table>
      <div style="font-size:11px;color:#5b6b7f;margin-top:6px">
        &#9646; &ge;95% &nbsp;&#9646; 85&#8211;94% &nbsp;&#9646; &lt;85%
        &nbsp;&nbsp;|&nbsp;&nbsp;
        <span style="color:#1f7a4c;font-weight:bold">&#9646;</span> Verde &ge;95%
        &nbsp;<span style="color:#d97706;font-weight:bold">&#9646;</span> Amarillo 85&#8211;94%
        &nbsp;<span style="color:#c0392b;font-weight:bold">&#9646;</span> Rojo &lt;85%
      </div>
    </div>"""

    warnings_html = ""
    if warnings:
        items = []
        for item in warnings:
            items.append(
                f"<li>{item['cliente']} | Nro Pedido: {item['nro_pedido']} | Estado: {item['estado']} | "
                f"{item['dias_transcurridos']} dias desde generacion</li>"
            )
        warnings_html = (
            "<div style='margin-top:22px;border:1px solid #f5d0a9;border-radius:14px;background:#fff7ed;padding:18px 20px'>"
            "<div style='font-family:Calibri,Arial,sans-serif;font-size:22px;font-weight:bold;color:#9a3412;margin-bottom:10px'>"
            "Pedidos con estados pendientes &gt; 7 dias</div>"
            f"<ul style='font-family:Calibri,Arial,sans-serif;color:#334155;margin:0 0 0 18px;padding:0;line-height:1.7'>{''.join(items)}</ul>"
            "</div>"
        )

    resumen_cards = f"""
    <table cellpadding="0" cellspacing="0" style="margin:18px 0 18px 0">
      <tr>
        <td style="padding:10px 16px;border:1px solid #dbe5f0;border-radius:20px;background:#ffffff;font-size:12px;font-weight:bold;color:#243b53">Clientes: {total_clientes}</td>
        <td width="8"></td>
        <td style="padding:10px 16px;border:1px solid #dbe5f0;border-radius:20px;background:#ffffff;font-size:12px;font-weight:bold;color:#243b53">OK: {ok_count}</td>
        <td width="8"></td>
        <td style="padding:10px 16px;border:1px solid #dbe5f0;border-radius:20px;background:#ffffff;font-size:12px;font-weight:bold;color:#243b53">Sin datos: {sin_datos_count}</td>
        <td width="8"></td>
        <td style="padding:10px 16px;border:1px solid #dbe5f0;border-radius:20px;background:#ffffff;font-size:12px;font-weight:bold;color:#243b53">Omitidos: {omitidos_count}</td>
        <td width="8"></td>
        <td style="padding:10px 16px;border:1px solid #dbe5f0;border-radius:20px;background:#ffffff;font-size:12px;font-weight:bold;color:#243b53">Errores: {error_count}</td>
        <td width="8"></td>
        <td style="padding:10px 16px;border:1px solid #dbe5f0;border-radius:20px;background:#ffffff;font-size:12px;font-weight:bold;color:#243b53">Advertencias: {warning_count}</td>
      </tr>
    </table>
    """

    return f"""
    <html>
      <body style="margin:0;padding:0;background:#eef3f8;font-family:Calibri,Arial,sans-serif;color:#1f2937">
        <table width="100%" cellpadding="0" cellspacing="0" style="background:#eef3f8;padding:20px 0">
          <tr>
            <td align="center">
              <table width="860" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;border:1px solid #d9e2ec">
                <tr>
                  <td style="background:{header_color};padding:18px 22px;color:#ffffff">
                    <div style="font-size:16px;font-weight:bold;letter-spacing:0.2px">WMS Egakat — NNSS Diario</div>
                    <div style="font-size:28px;font-weight:bold;margin-top:6px">{header_icon} {overall_state}</div>
                    <div style="font-size:14px;margin-top:8px">{header_text}</div>
                  </td>
                </tr>
                <tr>
                  <td style="padding:18px 22px;background:{header_bg}">
                    {resumen_cards}
                    <table style="border-collapse:collapse;width:100%;background:#ffffff;border:1px solid #d9e2ec">
                      <thead>
                        <tr style="background:#2f4358;color:#ffffff">
                          <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:left">Cliente</th>
                          <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:left">CD</th>
                          <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:left">Estado</th>
                          <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:right">Filas nuevas</th>
                          <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:right">Reemplazadas</th>
                          <th style="padding:11px 10px;border:1px solid #d9e2ec;text-align:right">Advertencias</th>
                        </tr>
                      </thead>
                      <tbody>
                        {''.join(rows_html)}
                      </tbody>
                    </table>
                    {pendientes_html}
                    {otif_html}
                    {warnings_html}
                    <div style="font-size:12px;color:#5b6b7f;margin-top:18px">
                      Notificacion automatica generada por Sistema Automatizado WMS Egakat.
                    </div>
                  </td>
                </tr>
              </table>
            </td>
          </tr>
        </table>
      </body>
    </html>
    """
