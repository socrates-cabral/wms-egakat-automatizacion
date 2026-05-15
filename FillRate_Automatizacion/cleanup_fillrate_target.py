"""
cleanup_fillrate_target.py — Limpia un archivo target de FillRate:
  1. Borra filas "contaminantes" (Empresa != Empresa esperada del cliente)
  2. Borra duplicados de identidad, dejando UNA copia (la primera, fila menor)

Read-only por defecto. Con --apply hace los cambios + backup automático.

Después de correr esto, las filas FALTANTES (faltantes legítimos de mayo, etc.) se
recuperan mediante `py fillrate_descarga.py --client X --mes MM/YYYY` que usa la
lógica probada con fórmulas/templates correctos.

Uso:
    py cleanup_fillrate_target.py \\
        --target "C:\\...\\data Omnitech.xlsx" \\
        --wms-export "C:\\...\\Reporte_..._OMNITECH.xlsx" \\
        [--apply]
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import argparse
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# Mismas constantes que fillrate_utils.py — NO cambiar sin entender el porqué
DATA_START_ROW = 2
EMPRESA_COL_IDX = 1       # col B
NRO_APLICA_COL_IDX = 3    # col D
FECHA_INGRESO_COL_IDX = 8 # col I


def _norm_empresa(v):
    if v is None:
        return None
    return str(v).strip().upper()


def _to_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y", "%d-%m-%Y %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _row_identity(row_values):
    empresa = _norm_empresa(row_values[EMPRESA_COL_IDX]) if len(row_values) > EMPRESA_COL_IDX else None
    nro_aplica = row_values[NRO_APLICA_COL_IDX] if len(row_values) > NRO_APLICA_COL_IDX else None
    fecha = _to_dt(row_values[FECHA_INGRESO_COL_IDX]) if len(row_values) > FECHA_INGRESO_COL_IDX else None
    mes_year = (fecha.year, fecha.month) if fecha else None
    return (empresa, nro_aplica, mes_year)


def detect_empresa_esperada(wms_export_path: Path):
    """Identifica la Empresa correcta mirando el WMS export (modo mayoritario)."""
    wb = load_workbook(wms_export_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    empresas = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if len(row) > EMPRESA_COL_IDX and row[EMPRESA_COL_IDX] is not None:
            empresas.append(_norm_empresa(row[EMPRESA_COL_IDX]))
    wb.close()
    if not empresas:
        return None
    return Counter(empresas).most_common(1)[0][0]


def main():
    p = argparse.ArgumentParser(description="Limpieza de target FillRate: contaminacion + dupes.")
    p.add_argument("--target", required=True, help="Path al archivo OneDrive (data Cliente.xlsx)")
    p.add_argument("--wms-export", required=True, help="Path al WMS export (para detectar empresa esperada)")
    p.add_argument("--apply", action="store_true", help="Aplicar cambios (default: dry-run)")
    args = p.parse_args()

    target_path = Path(args.target)
    wms_path = Path(args.wms_export)
    if not target_path.exists():
        sys.exit(f"ERROR: no existe --target: {target_path}")
    if not wms_path.exists():
        sys.exit(f"ERROR: no existe --wms-export: {wms_path}")

    empresa_esperada = detect_empresa_esperada(wms_path)
    if not empresa_esperada:
        sys.exit("ERROR: WMS export vacio o sin Empresa detectable")
    print(f"Empresa esperada (segun WMS export): '{empresa_esperada}'")

    # Cargar target read_only primero para identificar qué borrar
    wb_r = load_workbook(target_path, read_only=True, data_only=True)
    ws_r = None
    for name in ("Seguimiento de pedidos", "Datos", "Sheet1", "Hoja1"):
        if name in wb_r.sheetnames:
            ws_r = wb_r[name]
            break
    if ws_r is None:
        ws_r = wb_r[wb_r.sheetnames[0]]
    print(f"Target sheet: '{ws_r.title}' (max_row={ws_r.max_row})")

    rows_data = []
    for row_idx, row in enumerate(ws_r.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        if len(row) <= NRO_APLICA_COL_IDX or row[NRO_APLICA_COL_IDX] is None:
            continue
        empresa = _norm_empresa(row[EMPRESA_COL_IDX]) if len(row) > EMPRESA_COL_IDX else None
        ident = _row_identity(row)
        rows_data.append({"row_idx": row_idx, "empresa": empresa, "ident": ident})
    wb_r.close()

    # 1. Contaminación: empresa != esperada
    contaminantes = [r for r in rows_data if r["empresa"] != empresa_esperada]

    # 2. Dupes: para cada identidad con N>1, marcar todas menos la primera (row_idx menor)
    by_ident = defaultdict(list)
    for r in rows_data:
        if r["empresa"] == empresa_esperada:  # excluir contaminantes (ya marcadas)
            by_ident[r["ident"]].append(r["row_idx"])

    dupes_a_borrar = []
    for ident, idxs in by_ident.items():
        if len(idxs) > 1:
            idxs_sorted = sorted(idxs)
            for idx in idxs_sorted[1:]:  # mantener el primero, borrar el resto
                dupes_a_borrar.append({"row_idx": idx, "ident": ident})

    # Set total de filas a borrar
    rows_a_borrar = set(r["row_idx"] for r in contaminantes) | set(d["row_idx"] for d in dupes_a_borrar)

    print()
    print("=" * 80)
    print("PLAN DE LIMPIEZA")
    print("=" * 80)
    print(f"Contaminantes (Empresa != '{empresa_esperada}'): {len(contaminantes)} filas")
    for r in contaminantes[:10]:
        print(f"  fila {r['row_idx']}: Empresa='{r['empresa']}', ident={r['ident']}")
    if len(contaminantes) > 10:
        print(f"  ... y {len(contaminantes) - 10} más")

    print()
    print(f"Duplicados de identidad: {len(dupes_a_borrar)} filas a borrar (manteniendo primera)")
    for d in dupes_a_borrar[:10]:
        print(f"  fila {d['row_idx']}: ident={d['ident']}")
    if len(dupes_a_borrar) > 10:
        print(f"  ... y {len(dupes_a_borrar) - 10} más")

    print()
    print(f"TOTAL filas a borrar: {len(rows_a_borrar)}")

    if not rows_a_borrar:
        print("\n✓ Nada que limpiar. Target sano.")
        return

    if not args.apply:
        print()
        print("=" * 80)
        print("MODO DRY-RUN — no se aplicaron cambios.")
        print("Para ejecutar la limpieza, agregar --apply al comando.")
        print("=" * 80)
        return

    # APPLY: hacer backup y borrar
    backup_path = target_path.with_suffix(f".backup-{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    print()
    print(f"Backup: {backup_path}")
    shutil.copy2(target_path, backup_path)

    print(f"Cargando workbook editable...")
    wb = load_workbook(target_path)
    ws = None
    for name in ("Seguimiento de pedidos", "Datos", "Sheet1", "Hoja1"):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    # Borrar en orden descendente para no invalidar row indices
    import itertools, operator
    sorted_rows = sorted(rows_a_borrar)
    groups = []
    for _, grp in itertools.groupby(enumerate(sorted_rows), lambda x: x[0] - x[1]):
        g = list(map(operator.itemgetter(1), grp))
        groups.append((g[0], len(g)))
    print(f"Borrando {len(rows_a_borrar)} filas en {len(groups)} rangos consecutivos (bulk delete)...")
    for start_row, count in reversed(groups):
        ws.delete_rows(start_row, count)

    print(f"Guardando target...")
    wb.save(target_path)
    print(f"✓ Listo. {len(rows_a_borrar)} filas borradas. Backup en: {backup_path.name}")
    print()
    print("=" * 80)
    print("SIGUIENTE PASO")
    print("=" * 80)
    print("Para agregar filas faltantes (si las había), correr:")
    nombre_cliente = target_path.stem.replace("data ", "")
    print(f'  py FillRate_Automatizacion\\fillrate_descarga.py --client "{nombre_cliente}" --mes 05/2026')
    print()
    print("Y luego volver a correr reconcile_fillrate.py para verificar.")


if __name__ == "__main__":
    main()
