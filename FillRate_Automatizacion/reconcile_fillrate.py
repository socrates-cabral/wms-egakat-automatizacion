"""
reconcile_fillrate.py — Reporta diferencias entre un export WMS manual del año y
el archivo target en OneDrive (data Cliente.xlsx).

Read-only por diseño. No modifica nada. Usa la misma identidad compuesta que el
fix de fillrate_utils.py: (Empresa normalizada, Nro Aplica, Mes-Year Fecha Ingreso).

Uso:
    py reconcile_fillrate.py \
        --wms-export "C:\\ruta\\export_anual_omnitech.xlsx" \
        --target "C:\\Users\\Socrates Cabral\\OneDrive - EGA KAT LOGISTICA SPA\\Datos para Dashboard - NNSS Operacional\\Pudahuel\\data Omnitech.xlsx"

Output:
  - Cuántas filas hay en cada archivo (con identidad única).
  - Cuántas faltan en target (presentes en WMS, no en target) — agruparlas por mes.
  - Cuántas hay en target que no están en WMS (posibles overrides manuales / legacy).
  - Comandos FillRate sugeridos para recuperar mes por mes.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import argparse
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# Mismas constantes que fillrate_utils.py — NO cambiar sin entender el porqué
DATA_START_ROW = 2
EMPRESA_COL_IDX = 1       # col B (índice 1 base 0)
NRO_APLICA_COL_IDX = 3    # col D
FECHA_INGRESO_COL_IDX = 8 # col I


def _norm_empresa(v):
    if v is None:
        return None
    return str(v).strip().upper()


def _to_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y", "%d-%m-%Y %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _row_identity(row):
    empresa = _norm_empresa(row[EMPRESA_COL_IDX]) if len(row) > EMPRESA_COL_IDX else None
    nro_aplica = row[NRO_APLICA_COL_IDX] if len(row) > NRO_APLICA_COL_IDX else None
    fecha = _to_dt(row[FECHA_INGRESO_COL_IDX]) if len(row) > FECHA_INGRESO_COL_IDX else None
    mes_year = (fecha.year, fecha.month) if fecha else None
    return (empresa, nro_aplica, mes_year)


def load_rows(path: Path, label: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    # Buscar sheet de datos (preferir "Seguimiento de pedidos", caer al primero)
    ws = None
    for name in ("Seguimiento de pedidos", "Datos", "Sheet1", "Hoja1"):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    print(f"[{label}] archivo: {path.name}")
    print(f"[{label}] sheet: '{ws.title}' (max_row={ws.max_row})")

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        # Saltar filas vacías o template (sin Nro Aplica)
        if len(row) <= NRO_APLICA_COL_IDX or row[NRO_APLICA_COL_IDX] is None:
            continue
        rows.append((row_idx, row))
    wb.close()
    print(f"[{label}] filas con datos reales: {len(rows)}")
    return rows


def main():
    p = argparse.ArgumentParser(description="Reconciliacion read-only FillRate: WMS export vs OneDrive target.")
    p.add_argument("--wms-export", required=True, help="Path al Excel descargado manualmente del WMS (idealmente del año entero).")
    p.add_argument("--target", required=True, help="Path al archivo target OneDrive (data Cliente.xlsx).")
    args = p.parse_args()

    wms_path = Path(args.wms_export)
    target_path = Path(args.target)
    if not wms_path.exists():
        sys.exit(f"ERROR: no existe --wms-export: {wms_path}")
    if not target_path.exists():
        sys.exit(f"ERROR: no existe --target: {target_path}")

    print("=" * 80)
    print(f"RECONCILIACION FILLRATE")
    print("=" * 80)
    wms_rows = load_rows(wms_path, "WMS    ")
    print()
    target_rows = load_rows(target_path, "TARGET ")
    print()

    # Mapas identidad → row_idx (si hay duplicados de identidad en un archivo, se queda el primero)
    wms_identidades = {}
    wms_dupes = 0
    for row_idx, row in wms_rows:
        ident = _row_identity(row)
        if ident in wms_identidades:
            wms_dupes += 1
        else:
            wms_identidades[ident] = (row_idx, row)

    target_identidades = {}
    target_dupes = 0
    for row_idx, row in target_rows:
        ident = _row_identity(row)
        if ident in target_identidades:
            target_dupes += 1
        else:
            target_identidades[ident] = (row_idx, row)

    if wms_dupes:
        print(f"[WMS    ] {wms_dupes} duplicados de identidad detectados (mismo Empresa+Aplica+Mes).")
    if target_dupes:
        print(f"[TARGET ] {target_dupes} duplicados de identidad detectados.")

    # Faltantes: identidades en WMS no en target
    faltantes = [(ident, row) for ident, (row_idx, row) in wms_identidades.items() if ident not in target_identidades]
    # Extras: identidades en target no en WMS (posibles manual overrides, legacy)
    extras = [(ident, row) for ident, (row_idx, row) in target_identidades.items() if ident not in wms_identidades]
    # Coincidencias: están en ambos (target ya tiene esa OP — el target podría estar desactualizado pero existe)
    matched = sum(1 for ident in wms_identidades if ident in target_identidades)

    print()
    print("=" * 80)
    print("RESUMEN")
    print("=" * 80)
    print(f"  WMS export:             {len(wms_identidades):>6} filas con identidad única")
    print(f"  Target OneDrive:        {len(target_identidades):>6} filas con identidad única")
    print(f"  Coincidencias (ambos):  {matched:>6} → target tiene la OP (puede o no estar actualizada)")
    print(f"  Faltantes en target:    {len(faltantes):>6} → presentes en WMS, NO en target")
    print(f"  Extras en target:       {len(extras):>6} → en target, NO en WMS (override manual / legacy)")

    # Breakdown faltantes por mes
    faltantes_por_mes = defaultdict(int)
    for ident, _ in faltantes:
        _, _, mes_year = ident
        if mes_year:
            faltantes_por_mes[mes_year] += 1
        else:
            faltantes_por_mes[("?", "?")] += 1

    if faltantes_por_mes:
        print()
        print("Faltantes por mes (Fecha de Ingreso):")
        for key in sorted(faltantes_por_mes.keys()):
            count = faltantes_por_mes[key]
            if key == ("?", "?"):
                print(f"  sin fecha:   {count} filas")
            else:
                y, m = key
                print(f"  {y}-{m:02d}:     {count} filas")

    if extras:
        print()
        print(f"Muestra de extras (primeros 5):")
        for ident, row in extras[:5]:
            print(f"  identidad={ident}")

    print()
    print("=" * 80)
    print("ACCION RECOMENDADA")
    print("=" * 80)
    if not faltantes:
        print("✓ Target está completo. No falta nada.")
        if extras:
            print(f"  Sin embargo, hay {len(extras)} filas en target que no están en el WMS export.")
            print("  Pueden ser overrides manuales legítimos, o legacy de meses anteriores al WMS export.")
            print("  Validar si querés conservarlos o no.")
    else:
        nombre_cliente = target_path.stem.replace("data ", "")
        print(f"Para recuperar las {len(faltantes)} filas faltantes en {target_path.name}:")
        print()
        print("OPCIÓN A — Re-correr FillRate mes por mes (con el fix de identidad compuesta YA activo):")
        for (y, m) in sorted(faltantes_por_mes.keys()):
            if (y, m) == ("?", "?"):
                continue
            print(f"  py fillrate_descarga.py --client \"{nombre_cliente}\" --mes {m:02d}/{y}")
        print()
        print("OPCIÓN B — Manual: copiá las filas del WMS export al target en Excel y guardá.")
        print()
        print("Validar luego corriendo este script otra vez: faltantes debería ser 0.")


if __name__ == "__main__":
    main()
