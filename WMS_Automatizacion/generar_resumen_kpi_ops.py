"""
generar_resumen_kpi_ops.py
Genera un resumen JSON liviano con KPIs operativos de NNSS y productividad.

Alcance de Fase 2B.1:
- Solo lectura de Excel y archivos locales.
- No ejecuta descargas ni automatizaciones productivas.
- No modifica la API ni los scripts existentes.

MIGRACIÓN SERVIDOR 24/7:
- Configurar ONEDRIVE_ROOT en .env para cambiar usuario
- Fallback a path actual si no está configurado (deprecado)
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

# Modulo compartido con canal_derco_auto.py para que ambos clasifiquen Rack/Est igual.
from pathlib import Path as _PathBoot  # noqa: E402
sys.path.insert(0, str(_PathBoot(__file__).parent))
from canal_derco_utils import (  # noqa: E402
    canal_principal_derco,
    cargar_base_ces,
    clasificar_ubicacion_dim,
    resolver_canal_con_ces,
)

import argparse
import calendar
import json
import math
import os
import re
import unicodedata
import warnings
from collections import defaultdict, Counter
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

try:
    import pandas as pd
except ImportError:  # Específico, no Exception genérico
    pd = None

from openpyxl import load_workbook
from dotenv import load_dotenv
from productividad_usuarios import (
    calcular_por_usuario,
    calcular_por_usuario_canal,
    calcular_por_usuario_cliente,
)
from recepciones_kpi import construir_payload_recepciones


BASE_DIR = Path(__file__).resolve().parent
CLAUDEWORK_DIR = BASE_DIR.parent
LOGDIR = CLAUDEWORK_DIR / "logs"

# Cargar .env para ONEDRIVE_ROOT
load_dotenv(CLAUDEWORK_DIR / ".env")

# ── Paths OneDrive - Migración servidor 24/7 ──────────────────────────────
# Prioridad: ONEDRIVE_ROOT desde .env (para servidor)
# Fallback: Path actual hardcodeado (DEPRECADO - solo laptop)
_ONEDRIVE_ROOT_ENV = os.getenv("ONEDRIVE_ROOT")
if _ONEDRIVE_ROOT_ENV:
    _ONEDRIVE_ROOT = Path(_ONEDRIVE_ROOT_ENV)
    print(f"[INFO] Usando ONEDRIVE_ROOT desde .env: {_ONEDRIVE_ROOT}")
else:
    _ONEDRIVE_ROOT = Path(r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA")
    print(f"[WARN] ONEDRIVE_ROOT no configurado en .env - usando path hardcodeado (DEPRECADO)")
    print(f"       Agregar a .env: ONEDRIVE_ROOT={_ONEDRIVE_ROOT}")

NNSS_DIR = _ONEDRIVE_ROOT / "Datos para Dashboard - NNSS Operacional"
# Base CES vive en la carpeta de Productividad (no NNSS).
BASE_CES_PATH = _ONEDRIVE_ROOT / "Datos para Dashboard - Productividad" / "Archivos Soporte" / "Base CES.xlsx"
PRODUCTIVIDAD_ROOT_OFICIAL = _ONEDRIVE_ROOT / "Datos para Dashboard - Productividad"
DIMENSIONES_ROOT = _ONEDRIVE_ROOT / "datos para Dashboard EK" / "Productividad"
DIMENSIONES_FILENAME = "Tablas dimensiones.xlsx"
STOCK_WMS_ROOT = _ONEDRIVE_ROOT / "Datos para Dashboard - Stock WMS Semanal"
STAGING_ROOT = _ONEDRIVE_ROOT / "Datos para Dashboard - Stagin IN- OUT"
POSICIONES_ROOT = _ONEDRIVE_ROOT / "Datos para Dashboard - Consulta de Posiciones"
INVENTARIO_DIM_ROOT = _ONEDRIVE_ROOT / "datos para Dashboard EK" / "Inventario"
INVENTARIO_DIM_FILENAME = "Tabla Ubicaciones CDs.xlsx"
RECEPCIONES_ROOT_OFICIAL = _ONEDRIVE_ROOT / "Datos para Dashboard - Clientes EK"
CONTEOS_OFICIAL_ROOT = _ONEDRIVE_ROOT / "Datos para Dashboard - Registros de conteos"
CONTEOS_INVENTARIO_FILENAME = "Registros de conteo ciclico.xlsx"
CONTEOS_INVENTARIO_VARIANTES = {
    "REGISTROS DE CONTEO CICLICO.XLSX",
    "REGISTROS DE CONTEO CÍCLICO.XLSX",
    "REGISTROS DE CONTEOS.XLSX",
    "REGISTROS DE CONTEOS CICLICOS.XLSX",
    "REGISTROS DE CONTEOS CÍCLICOS.XLSX",
}
LOCACIONES_OCUPACION_EXCLUIDAS = {"ANALISIS INV.", "OTRAS", "REFRESCAMIENTO BATERIA"}
ESTADOS_STAGING_VALIDOS = {"STAGING IN", "STAGING OUT"}

ESTADOS_PENDIENTES = {
    "POR ACEPTAR",
    "EN PREPARACION",
    "PREPARADO",
    "PREPARADOS",
}
ESTADOS_NO_PENDIENTES = {
    "REMITIDO",
    "REMITIDOS",
    "DESPACHADO",
    "DESPACHADOS",
    "CON SALIDA",
}
ESTADOS_PREPARACION = {"EN PREPARACION"}
ESTADOS_PREPARADOS = {"PREPARADO", "PREPARADOS", "CON SALIDA", "REMITIDO", "REMITIDOS"}
ESTADOS_DESPACHADOS = {"DESPACHADO", "DESPACHADOS"}
OBJETIVOS_MENSUALES_CONTEO = {
    (2026, 4): 25288,
}
MESES_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}
RANK_ESTADO = {
    "POR ACEPTAR": 1,
    "EN PREPARACION": 2,
    "PREPARADO": 3,
    "PREPARADOS": 3,
    "REMITIDO": 4,
    "REMITIDOS": 4,
    "DESPACHADO": 5,
    "DESPACHADOS": 5,
    "CON SALIDA": 6,
}
COLUMNAS_NNSS_REQUERIDAS = [
    "Deposito",
    "Empresa",
    "Nro Aplica",
    "Nro Pedido",
    "Estado Pedido",
    "Cliente",
    "Fecha y hora de Ingreso",
    "Fecha y hora de Inicio Preparacion",
    "Cantidad Original",
    "Cantidad Preparada",
    "Cantidad Despachada",
    "Fecha de Ingreso",
    "FR (%)",
    "Fecha Entrega",
    "Lead time (Dias)",
    "Tiempo entrega real",
    "Entregado a tiempo?",
    "Entregado completo y sin danos?",
    "Motivos de Diferencias",
]
COLUMNAS_PRODUCTIVIDAD_BASE = [
    "Centro",
    "Cliente",
    "Comprobante",
    "Contenedor",
    "Ubicacion",
    "Fecha Vto.",
    "Lote",
    "Fecha",
    "Hora",
    "Tipo de operacion",
    "Naturaleza",
    "Numero",
    "Registro",
    "Salida",
    "Entrada",
    "Saldo",
    "Comprobante externo",
    "Destino",
    "Observaciones",
    "SKU",
    "Descripcion",
    "Unidades por Contenedor",
    "Trabajo",
    "Pedido Fecha Preparado",
    "Pedido Fecha Remitido",
    "Pedido Fecha Despachado",
    "Pedido Fecha Salida",
    "Viaje Nro. SGL",
    "Viaje Status",
    "Direccion",
    "Nro. de Doc. Externo",
]


@dataclass
class FuenteDetectada:
    ruta: Path
    hoja: str | None = None
    cliente: str | None = None
    centro: str | None = None


def normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    return " ".join(texto.split())


def quitar_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(c)
    )


def normalizar_etiqueta(valor: Any) -> str:
    return quitar_acentos(normalizar_texto(valor)).upper()


def normalizar_mayusculas(valor: Any) -> str:
    return normalizar_texto(valor).upper()


def normalizar_clave(valor: Any) -> str:
    texto = normalizar_texto(valor)
    if not texto:
        return ""
    if texto.endswith(".0"):
        texto = texto[:-2]
    return texto


def iso_mtime(path: Path) -> str | None:
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def es_archivo_respaldo(path: Path) -> bool:
    nombre = path.name.lower()
    stem = path.stem.lower()
    if nombre.startswith("~$"):
        return True
    tokens = ("backup", "respaldo", "copia", "copy", "old", "tmp")
    return any(token in stem for token in tokens)


def normalizar_numero_pedido(valor: Any) -> str:
    texto = normalizar_clave(valor)
    if not texto:
        return ""
    if "E+" in texto.upper():
        try:
            return str(int(float(texto)))
        except Exception:
            return texto
    return texto


def parse_numero(valor: Any) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        if isinstance(valor, float) and math.isnan(valor):
            return 0.0
        return float(valor)
    texto = normalizar_texto(valor)
    if not texto:
        return 0.0
    texto = texto.replace("%", "").replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return 0.0


def parse_fecha(valor: Any) -> datetime | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime.combine(valor, time.min)
    if pd is not None:
        try:
            convertido = pd.to_datetime(valor, errors="coerce", dayfirst=True)
            if pd.notna(convertido):
                if hasattr(convertido, "to_pydatetime"):
                    return convertido.to_pydatetime()
                return convertido
        except Exception:
            pass
    texto = normalizar_texto(valor)
    if not texto:
        return None
    formatos = [
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ]
    for formato in formatos:
        try:
            return datetime.strptime(texto, formato)
        except ValueError:
            continue
    return None


def serializar_fecha(valor: Any) -> str | None:
    if valor is None:
        return None
    if pd is not None and pd.isna(valor):
        return None
    if pd is not None and isinstance(valor, pd.Timestamp):
        valor = valor.to_pydatetime()
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(valor, date):
        return datetime.combine(valor, time.min).strftime("%Y-%m-%d %H:%M:%S")
    return None


def round_safe(valor: float | int | None, ndigits: int = 2) -> float:
    if valor is None:
        return 0.0
    try:
        return round(float(valor), ndigits)
    except Exception:
        return 0.0


def porcentaje_safe(numerador: float, denominador: float) -> float:
    if not denominador:
        return 0.0
    return round_safe((numerador / denominador) * 100.0, 2)


def porcentaje_safe_nullable(numerador: float, denominador: float) -> float | None:
    if not denominador:
        return None
    return porcentaje_safe(numerador, denominador)


def serializar_numero_cantidad(valor: float) -> int | float:
    if abs(valor - round(valor)) < 1e-9:
        return int(round(valor))
    return round_safe(valor, 2)


def division_lineal_nullable(numerador: float, denominador: float) -> float | None:
    if not denominador:
        return None
    return round_safe(float(numerador) / float(denominador), 2)


def es_fecha_valida(valor: Any) -> bool:
    if valor is None:
        return False
    if pd is not None and pd.isna(valor):
        return False
    return isinstance(valor, (datetime, date))


def primer_visible_xlsx(path: Path) -> str | None:
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return None
    try:
        for ws in wb.worksheets:
            if ws.sheet_state == "visible":
                return ws.title
    finally:
        wb.close()
    return None


def descubrir_fuentes_nnss(verbose: bool = False) -> list[FuenteDetectada]:
    fuentes: list[FuenteDetectada] = []
    if not NNSS_DIR.exists():
        return fuentes
    for path in sorted(NNSS_DIR.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        hoja = primer_visible_xlsx(path)
        if not hoja:
            continue
        fuentes.append(FuenteDetectada(ruta=path, hoja=hoja))
        if verbose:
            print(f"[NNSS] Fuente detectada: {path} | hoja={hoja}")
    return fuentes


def normalizar_empresa_nnss(empresa: Any, deposito: Any) -> str:
    empresa_txt = normalizar_mayusculas(empresa)
    deposito_txt = normalizar_mayusculas(deposito)
    if empresa_txt == "MASCOTAS LATINAS" and deposito_txt == "PUDAHUEL":
        return "MASCOTAS LATINAS_P"
    if deposito_txt == "PUDAHUEL UNITARIO" and empresa_txt:
        return f"{empresa_txt}_U"
    return empresa_txt


def leer_consulta_fr(fuentes: list[FuenteDetectada], year: int, month: int, verbose: bool = False) -> tuple[pd.DataFrame | None, dict[str, Any]]:
    if pd is None:
        return None, {
            "disponible": False,
            "ruta": str(NNSS_DIR),
            "archivos_leidos": 0,
            "mensaje": "pandas no está disponible en este entorno.",
        }

    frames: list[pd.DataFrame] = []
    archivos_leidos = 0
    errores: list[str] = []

    for fuente in fuentes:
        try:
            df = pd.read_excel(
                fuente.ruta,
                sheet_name=fuente.hoja,
                engine="openpyxl",
            )
            df.columns = [quitar_acentos(normalizar_texto(col)) for col in df.columns]
            columnas_presentes = [c for c in COLUMNAS_NNSS_REQUERIDAS if c in df.columns]
            if not columnas_presentes or "Nro Pedido" not in df.columns or "Estado Pedido" not in df.columns:
                continue
            df = df.copy()
            for columna in COLUMNAS_NNSS_REQUERIDAS:
                if columna not in df.columns:
                    df[columna] = None
            df["Deposito"] = df["Deposito"].map(normalizar_texto)
            df["Empresa"] = [
                normalizar_empresa_nnss(emp, dep)
                for emp, dep in zip(df["Empresa"], df["Deposito"])
            ]
            df["Cliente"] = df["Cliente"].map(normalizar_texto)
            df["Nro Aplica"] = df["Nro Aplica"].map(normalizar_clave)
            df["Nro Pedido"] = df["Nro Pedido"].map(normalizar_numero_pedido)
            df["Estado Pedido"] = df["Estado Pedido"].map(normalizar_texto)
            df["Fecha y hora de Ingreso"] = df["Fecha y hora de Ingreso"].map(parse_fecha)
            df["Fecha y hora de Inicio Preparacion"] = df["Fecha y hora de Inicio Preparacion"].map(parse_fecha)
            df["Fecha de Ingreso"] = df["Fecha de Ingreso"].map(parse_fecha)
            df["Fecha Entrega"] = df["Fecha Entrega"].map(parse_fecha)
            for columna in ("Cantidad Original", "Cantidad Preparada", "Cantidad Despachada"):
                df[columna] = df[columna].map(parse_numero)
            df["Key"] = df["Empresa"] + "|" + df["Nro Aplica"]
            df["pedido_key"] = (
                df["Deposito"].map(normalizar_mayusculas)
                + "|"
                + df["Empresa"]
                + "|"
                + df["Nro Pedido"]
            )
            df["__archivo_fuente"] = str(fuente.ruta)
            df["__hoja_fuente"] = fuente.hoja
            df = df[df["Nro Pedido"] != ""].copy()
            if df.empty:
                continue
            fe_col = df["Fecha Entrega"].map(parse_fecha) if "Fecha Entrega" in df.columns else None
            if fe_col is not None and fe_col.notna().any():
                df_periodo = df[
                    fe_col.map(lambda x: bool(x) and x.year == year and x.month == month)
                ].copy()
            elif "Ano" in df.columns and "Mes" in df.columns:
                anos = pd.to_numeric(df["Ano"], errors="coerce").fillna(0).astype(int)
                meses = pd.to_numeric(df["Mes"], errors="coerce").fillna(0).astype(int)
                df_periodo = df[(anos == year) & (meses == month)].copy()
            else:
                fechas = df["Fecha de Ingreso"].fillna(df["Fecha y hora de Ingreso"])
                df_periodo = df[
                    fechas.map(lambda x: bool(x) and x.year == year and x.month == month)
                ].copy()
            if df_periodo.empty:
                if verbose:
                    print(f"[NNSS] Sin filas para el periodo en {fuente.ruta.name}")
                continue
            frames.append(df_periodo)
            archivos_leidos += 1
        except Exception as exc:
            errores.append(f"{fuente.ruta.name}: {exc}")

    if not frames:
        return None, {
            "disponible": False,
            "ruta": str(NNSS_DIR),
            "archivos_leidos": 0,
            "mensaje": "No se encontraron filas NNSS para el periodo solicitado.",
            "errores": errores[:5],
        }

    combinado = pd.concat(frames, ignore_index=True)
    return combinado, {
        "disponible": True,
        "ruta": str(NNSS_DIR),
        "archivos_leidos": archivos_leidos,
        "archivos_fuente": sorted({Path(v).name for v in combinado["__archivo_fuente"].dropna().tolist()}),
        "errores": errores[:5],
    }


def calcular_fr_fila(estado: Any, cantidad_original: Any, cantidad_preparada: Any, cantidad_despachada: Any) -> float:
    estado_txt = normalizar_etiqueta(estado)
    original = parse_numero(cantidad_original)
    if original <= 0:
        return 0.0
    if estado_txt in ESTADOS_PREPARACION:
        return 100.0
    if estado_txt in ESTADOS_PREPARADOS:
        return porcentaje_safe(parse_numero(cantidad_preparada), original)
    if estado_txt in ESTADOS_DESPACHADOS:
        return porcentaje_safe(parse_numero(cantidad_despachada), original)
    return 0.0


def normalizar_bool_si_no(valor: Any) -> str:
    texto = normalizar_etiqueta(valor)
    if texto == "SI":
        return "SI"
    if texto == "NO":
        return "NO"
    return ""


def consolidar_estado_pedido(estados: list[str]) -> str:
    estados_validos = [normalizar_etiqueta(e) for e in estados if normalizar_texto(e)]
    if not estados_validos:
        return "SIN ESTADO"
    return min(estados_validos, key=lambda e: RANK_ESTADO.get(e, 999))


def es_fillrate_evaluable_row(estado: Any, cantidad_original: Any, cantidad_preparada: Any, cantidad_despachada: Any) -> bool:
    estado_txt = normalizar_etiqueta(estado)
    original = parse_numero(cantidad_original)
    if original <= 0:
        return False
    if estado_txt in ESTADOS_PREPARACION:
        return True
    if estado_txt in ESTADOS_PREPARADOS:
        return True
    if estado_txt in ESTADOS_DESPACHADOS:
        return True
    return False


def resumir_pedidos_nnss(df_pedidos: pd.DataFrame, fecha_consulta: datetime) -> list[dict[str, Any]]:
    pedidos_resumen: list[dict[str, Any]] = []

    for pedido_key, grupo in df_pedidos.groupby("pedido_key"):
        empresa = normalizar_texto(grupo["Empresa"].iloc[0]) or "SIN_EMPRESA"
        nro_pedido = normalizar_numero_pedido(grupo["Nro Pedido"].iloc[0])
        estado_consolidado = consolidar_estado_pedido(grupo["Estado Pedido"].tolist())
        fecha_inicio = min([f for f in grupo["Fecha y hora de Inicio Preparacion"].tolist() if es_fecha_valida(f)], default=None)
        fecha_ingreso = min(
            [
                f
                for f in (
                    grupo["Fecha y hora de Ingreso"].tolist()
                    + grupo["Fecha de Ingreso"].tolist()
                )
                if es_fecha_valida(f)
            ],
            default=None,
        )
        fecha_entrega = min([f for f in grupo["Fecha Entrega"].tolist() if es_fecha_valida(f)], default=None)
        es_arrastre = bool(
            fecha_ingreso is not None
            and fecha_entrega is not None
            and hasattr(fecha_ingreso, "month")
            and hasattr(fecha_entrega, "month")
            and (fecha_ingreso.year != fecha_entrega.year or fecha_ingreso.month != fecha_entrega.month)
        )
        base_antiguedad = fecha_inicio or fecha_ingreso
        dias_abierto = None
        if base_antiguedad:
            base_fecha = base_antiguedad.date() if isinstance(base_antiguedad, datetime) else base_antiguedad
            dias_abierto = (fecha_consulta.date() - base_fecha).days

        lineas = int(len(grupo))
        unidades = round_safe(grupo["Cantidad Original"].sum(), 2)
        et_vals = [normalizar_bool_si_no(v) for v in grupo["Entregado a tiempo?"].tolist()]
        ec_vals = [normalizar_bool_si_no(v) for v in grupo["Entregado completo y sin danos?"].tolist()]
        evaluable_otif = bool(et_vals and ec_vals) and all(v in {"SI", "NO"} for v in et_vals) and all(v in {"SI", "NO"} for v in ec_vals)
        fillrate_evaluable = any(
            es_fillrate_evaluable_row(
                row.get("Estado Pedido"),
                row.get("Cantidad Original"),
                row.get("Cantidad Preparada"),
                row.get("Cantidad Despachada"),
            )
            for _, row in grupo.iterrows()
        )
        pedido_on_time = all(v == "SI" for v in et_vals) if evaluable_otif else False
        pedido_in_full = all(v == "SI" for v in ec_vals) if evaluable_otif else False
        pedido_otif = pedido_on_time and pedido_in_full if evaluable_otif else False
        es_pendiente = estado_consolidado in ESTADOS_PENDIENTES

        motivos_no_if: list[str] = []
        if evaluable_otif and not pedido_in_full:
            ec_norm = grupo["Entregado completo y sin danos?"].map(normalizar_bool_si_no)
            motivos_raw = grupo["Motivos de Diferencias"].map(normalizar_texto)
            motivos_no_if = [
                (m if (m and m.lower() != "nan") else "Sin motivo registrado")
                for ec, m in zip(ec_norm, motivos_raw)
                if ec == "NO"
            ]

        if es_pendiente:
            motivo_no_evaluable = "Pendiente / sin entrega evaluable"
        else:
            motivo_no_evaluable = "Sin datos suficientes para evaluar entrega."

        deposito_raw = normalizar_texto(grupo["Deposito"].iloc[0]) if "Deposito" in grupo.columns else ""
        deposito_lower = deposito_raw.lower()
        if "pudahuel" in deposito_lower:
            deposito_cd = "PUDAHUEL"
        elif "quilicura" in deposito_lower:
            deposito_cd = "QUILICURA"
        else:
            deposito_cd = deposito_raw.upper() if deposito_raw else "SIN_CD"

        pedidos_resumen.append(
            {
                "pedido_key": pedido_key,
                "cliente": empresa,
                "cd": deposito_cd,
                "nro_pedido": nro_pedido,
                "estado": estado_consolidado,
                "lineas": lineas,
                "unidades": unidades,
                "fecha_inicio_preparacion": serializar_fecha(fecha_inicio),
                "fecha_ingreso": serializar_fecha(fecha_ingreso),
                "dias_abierto": dias_abierto,
                "evaluable_otif": evaluable_otif,
                "on_time": pedido_on_time,
                "in_full": pedido_in_full,
                "otif": pedido_otif,
                "es_pendiente": es_pendiente,
                "fillrate_evaluable": fillrate_evaluable,
                "motivo_no_evaluable": motivo_no_evaluable,
                "es_arrastre": es_arrastre,
                "motivos_no_in_full": motivos_no_if,
            }
        )

    return pedidos_resumen


def calcular_otif_por_pedido(pedidos_resumen: list[dict[str, Any]]) -> dict[str, Any]:
    pedidos_por_cliente: dict[str, dict[str, int]] = defaultdict(
        lambda: {"pedidos_evaluados": 0, "pedidos_on_time": 0, "pedidos_in_full": 0, "pedidos_otif": 0, "pedidos_no_evaluables": 0, "arrastres": 0, "arrastres_on_time": 0}
    )
    pedidos_por_cd: dict[str, dict[str, int]] = defaultdict(
        lambda: {"pedidos_evaluados": 0, "pedidos_on_time": 0, "pedidos_in_full": 0, "pedidos_otif": 0, "pedidos_no_evaluables": 0}
    )
    no_evaluables_por_cliente: dict[str, list[dict[str, Any]]] = defaultdict(list)
    detalle_no_ot_por_cd: dict[str, list[dict[str, Any]]] = defaultdict(list)
    detalle_no_if_por_cd: dict[str, list[dict[str, Any]]] = defaultdict(list)
    motivos_no_if_por_cd: dict[str, list[str]] = defaultdict(list)
    detalle_no_ot_por_cliente: dict[str, list[dict[str, Any]]] = defaultdict(list)
    detalle_no_if_por_cliente: dict[str, list[dict[str, Any]]] = defaultdict(list)
    evaluados = 0
    on_time = 0
    in_full = 0
    otif = 0
    no_evaluable = 0
    arrastres_global = {"total": 0, "on_time": 0}
    motivos_no_if_por_cliente: dict[str, list[str]] = defaultdict(list)
    motivos_no_if_global: list[str] = []

    for pedido in pedidos_resumen:
        empresa = pedido["cliente"]
        cd = pedido.get("cd", "SIN_CD")
        if not pedido["evaluable_otif"]:
            pedidos_por_cliente[empresa]["pedidos_no_evaluables"] += 1
            pedidos_por_cd[cd]["pedidos_no_evaluables"] += 1
            no_evaluables_por_cliente[empresa].append(pedido)
            no_evaluable += 1
            continue

        pedido_on_time = bool(pedido["on_time"])
        pedido_in_full = bool(pedido["in_full"])
        pedido_otif = bool(pedido["otif"])

        evaluados += 1
        on_time += int(pedido_on_time)
        in_full += int(pedido_in_full)
        otif += int(pedido_otif)

        pedidos_por_cliente[empresa]["pedidos_evaluados"] += 1
        pedidos_por_cliente[empresa]["pedidos_on_time"] += int(pedido_on_time)
        pedidos_por_cliente[empresa]["pedidos_in_full"] += int(pedido_in_full)
        pedidos_por_cliente[empresa]["pedidos_otif"] += int(pedido_otif)

        pedidos_por_cd[cd]["pedidos_evaluados"] += 1
        pedidos_por_cd[cd]["pedidos_on_time"] += int(pedido_on_time)
        pedidos_por_cd[cd]["pedidos_in_full"] += int(pedido_in_full)
        pedidos_por_cd[cd]["pedidos_otif"] += int(pedido_otif)

        if pedido.get("es_arrastre"):
            pedidos_por_cliente[empresa]["arrastres"] += 1
            if pedido_on_time:
                pedidos_por_cliente[empresa]["arrastres_on_time"] += 1
            arrastres_global["total"] += 1
            if pedido_on_time:
                arrastres_global["on_time"] += 1

        if not pedido_in_full:
            for m in pedido.get("motivos_no_in_full") or []:
                motivos_no_if_por_cliente[empresa].append(m)
                motivos_no_if_por_cd[cd].append(m)
                motivos_no_if_global.append(m)

        if not pedido_on_time:
            _det_ot = {
                "nro_pedido": pedido["nro_pedido"],
                "cliente": pedido["cliente"],
                "estado": pedido.get("estado", ""),
                "es_arrastre": bool(pedido.get("es_arrastre")),
            }
            detalle_no_ot_por_cd[cd].append(_det_ot)
            detalle_no_ot_por_cliente[empresa].append(_det_ot)
        if not pedido_in_full:
            _det_if = {
                "nro_pedido": pedido["nro_pedido"],
                "cliente": pedido["cliente"],
                "estado": pedido.get("estado", ""),
                "motivos": [
                    {"motivo": m, "lineas": c}
                    for m, c in Counter(pedido.get("motivos_no_in_full") or []).most_common()
                ],
            }
            detalle_no_if_por_cd[cd].append(_det_if)
            detalle_no_if_por_cliente[empresa].append(_det_if)

    por_cliente = []
    for cliente, payload in sorted(pedidos_por_cliente.items()):
        evaluados_cliente = payload["pedidos_evaluados"]
        arrastres_c = payload.get("arrastres", 0)
        arrastres_ot_c = payload.get("arrastres_on_time", 0)
        pedidos_no_ot = evaluados_cliente - payload["pedidos_on_time"]
        pedidos_no_if = evaluados_cliente - payload["pedidos_in_full"]
        entry = {
            "cliente": cliente,
            "pedidos_evaluados": payload["pedidos_evaluados"],
            "pedidos_on_time": payload["pedidos_on_time"],
            "pedidos_no_on_time": pedidos_no_ot,
            "pedidos_in_full": payload["pedidos_in_full"],
            "pedidos_no_in_full": pedidos_no_if,
            "pedidos_otif": payload["pedidos_otif"],
            "pedidos_no_evaluables": payload["pedidos_no_evaluables"],
            "pct_on_time": porcentaje_safe(payload["pedidos_on_time"], evaluados_cliente),
            "pct_in_full": porcentaje_safe(payload["pedidos_in_full"], evaluados_cliente),
            "pct_otif": porcentaje_safe(payload["pedidos_otif"], evaluados_cliente),
        }
        if arrastres_c > 0:
            arrastres_tardios = arrastres_c - arrastres_ot_c
            entry["arrastres"] = {
                "total": arrastres_c,
                "on_time": arrastres_ot_c,
                "arrastres_tardios": arrastres_tardios,
                "pct_on_time_arrastres": porcentaje_safe(arrastres_ot_c, arrastres_c),
                "nota": f"De los {pedidos_no_ot} pedidos no on time, {arrastres_tardios} son arrastres de mes anterior (ingresados mes previo, Fecha Entrega mes en curso)",
            }
        motivos_c = motivos_no_if_por_cliente.get(cliente, [])
        if motivos_c:
            entry["motivos_no_in_full"] = [
                {"motivo": m, "lineas": c}
                for m, c in Counter(motivos_c).most_common(10)
            ]
        elif pedidos_no_if > 0:
            entry["motivos_no_in_full"] = []
        if pedidos_no_ot > 0:
            entry["detalle_no_on_time"] = detalle_no_ot_por_cliente.get(cliente, [])
        if pedidos_no_if > 0:
            entry["detalle_no_in_full"] = detalle_no_if_por_cliente.get(cliente, [])
        por_cliente.append(entry)

    def detalle_no_evaluable_payload(cliente: str, pedido: dict[str, Any]) -> dict[str, Any]:
        return {
            "cliente": cliente,
            "nro_pedido": pedido["nro_pedido"],
            "estado": pedido["estado"],
            "dias_abierto": pedido["dias_abierto"],
            "lineas": pedido["lineas"],
            "unidades": pedido["unidades"],
            "motivo": pedido["motivo_no_evaluable"],
            "fecha_inicio_preparacion": pedido["fecha_inicio_preparacion"],
        }

    def detalle_no_evaluable_sort_key(item: dict[str, Any]) -> tuple[Any, ...]:
        dias_abierto = item.get("dias_abierto")
        fecha_inicio = item.get("fecha_inicio_preparacion") or "9999-12-31 23:59:59"
        return (
            -(dias_abierto if dias_abierto is not None else -1),
            fecha_inicio,
            item.get("cliente") or "",
            item.get("nro_pedido") or "",
        )

    clientes_no_evaluables = []
    pedidos_no_evaluables_detalle = []
    pedidos_no_evaluables_detalle_por_cliente = []
    for cliente, pedidos in sorted(no_evaluables_por_cliente.items()):
        estados = defaultdict(int)
        detalle_cliente = []
        for pedido in pedidos:
            estados[pedido["estado"]] += 1
            detalle_item = detalle_no_evaluable_payload(cliente, pedido)
            pedidos_no_evaluables_detalle.append(detalle_item)
            detalle_cliente.append(
                {
                    "nro_pedido": detalle_item["nro_pedido"],
                    "estado": detalle_item["estado"],
                    "dias_abierto": detalle_item["dias_abierto"],
                    "lineas": detalle_item["lineas"],
                    "unidades": detalle_item["unidades"],
                    "motivo": detalle_item["motivo"],
                    "fecha_inicio_preparacion": detalle_item["fecha_inicio_preparacion"],
                }
            )

        pedidos_ordenados = sorted(
            pedidos,
            key=lambda item: (
                0 if (item["dias_abierto"] is not None and item["dias_abierto"] > 7) else 1,
                -(item["dias_abierto"] or -1),
                -item["unidades"],
                item["cliente"],
                item["nro_pedido"],
            ),
        )
        mas_antiguo = pedidos_ordenados[0] if pedidos_ordenados else None
        clientes_no_evaluables.append(
            {
                "cliente": cliente,
                "pedidos_no_evaluables": len(pedidos),
                "lineas_no_evaluables": sum(int(pedido["lineas"]) for pedido in pedidos),
                "unidades_no_evaluables": round_safe(sum(float(pedido["unidades"]) for pedido in pedidos), 2),
                "estados": dict(sorted(estados.items())),
                "motivo_principal": "Pedidos pendientes o sin datos suficientes para evaluar entrega.",
                "pedido_mas_antiguo": (
                    {
                        "nro_pedido": mas_antiguo["nro_pedido"],
                        "estado": mas_antiguo["estado"],
                        "dias_abierto": mas_antiguo["dias_abierto"],
                        "fecha_inicio_preparacion": mas_antiguo["fecha_inicio_preparacion"],
                    }
                    if mas_antiguo
                    else None
                ),
            }
        )
        detalle_cliente.sort(
            key=lambda item: detalle_no_evaluable_sort_key(
                {
                    "cliente": cliente,
                    **item,
                }
            )
        )
        pedidos_no_evaluables_detalle_por_cliente.append(
            {
                "cliente": cliente,
                "total_pedidos": len(pedidos),
                "lineas": sum(int(pedido["lineas"]) for pedido in pedidos),
                "unidades": round_safe(sum(float(pedido["unidades"]) for pedido in pedidos), 2),
                "detalle": detalle_cliente,
            }
        )

    clientes_no_evaluables = sorted(
        clientes_no_evaluables,
        key=lambda item: (-item["pedidos_no_evaluables"], -item["unidades_no_evaluables"], item["cliente"]),
    )
    pedidos_no_evaluables_detalle = sorted(
        pedidos_no_evaluables_detalle,
        key=detalle_no_evaluable_sort_key,
    )
    pedidos_no_evaluables_detalle_total = len(pedidos_no_evaluables_detalle)
    pedidos_no_evaluables_detalle_mostrados = len(pedidos_no_evaluables_detalle)
    pedidos_no_evaluables_detalle_truncado = pedidos_no_evaluables_detalle_mostrados < pedidos_no_evaluables_detalle_total
    pedidos_no_evaluables_detalle_por_cliente.sort(key=lambda item: (-item["total_pedidos"], -item["unidades"], item["cliente"]))

    arrastres_payload: dict[str, Any] | None = None
    if arrastres_global["total"] > 0:
        arrastres_tardios_global = arrastres_global["total"] - arrastres_global["on_time"]
        arrastres_payload = {
            "total": arrastres_global["total"],
            "on_time": arrastres_global["on_time"],
            "arrastres_tardios": arrastres_tardios_global,
            "pct_on_time_arrastres": porcentaje_safe(arrastres_global["on_time"], arrastres_global["total"]),
            "nota": f"Pedidos ingresados en mes anterior con Fecha Entrega en el mes en curso. {arrastres_tardios_global} de {arrastres_global['total']} arrastres no llegaron a tiempo.",
        }

    motivos_no_if_global_top = (
        [{"motivo": m, "lineas": c} for m, c in Counter(motivos_no_if_global).most_common(10)]
        if motivos_no_if_global else []
    )

    por_cd = []
    for cd_nombre, payload_cd in sorted(pedidos_por_cd.items()):
        evaluados_cd = payload_cd["pedidos_evaluados"]
        no_ot_cd = evaluados_cd - payload_cd["pedidos_on_time"]
        no_if_cd = evaluados_cd - payload_cd["pedidos_in_full"]
        entry_cd = {
            "cd": cd_nombre,
            "pedidos_evaluados": evaluados_cd,
            "pedidos_on_time": payload_cd["pedidos_on_time"],
            "pedidos_no_on_time": no_ot_cd,
            "pedidos_in_full": payload_cd["pedidos_in_full"],
            "pedidos_no_in_full": no_if_cd,
            "pedidos_otif": payload_cd["pedidos_otif"],
            "pedidos_no_evaluables": payload_cd["pedidos_no_evaluables"],
            "pct_on_time": porcentaje_safe(payload_cd["pedidos_on_time"], evaluados_cd),
            "pct_in_full": porcentaje_safe(payload_cd["pedidos_in_full"], evaluados_cd),
            "pct_otif": porcentaje_safe(payload_cd["pedidos_otif"], evaluados_cd),
        }
        if no_ot_cd > 0:
            entry_cd["detalle_no_on_time"] = detalle_no_ot_por_cd.get(cd_nombre, [])
        if no_if_cd > 0:
            entry_cd["detalle_no_in_full"] = detalle_no_if_por_cd.get(cd_nombre, [])
            motivos_cd = motivos_no_if_por_cd.get(cd_nombre, [])
            entry_cd["motivos_no_in_full"] = (
                [{"motivo": m, "lineas": c} for m, c in Counter(motivos_cd).most_common(10)]
                if motivos_cd else []
            )
        por_cd.append(entry_cd)

    return {
        "pedidos_evaluados": evaluados,
        "pedidos_no_evaluables": no_evaluable,
        "pct_on_time": porcentaje_safe(on_time, evaluados),
        "pct_in_full": porcentaje_safe(in_full, evaluados),
        "pct_otif": porcentaje_safe(otif, evaluados),
        "por_cliente": por_cliente,
        "por_cd": por_cd,
        "arrastres": arrastres_payload,
        "motivos_no_in_full_global": motivos_no_if_global_top,
        "clientes_no_evaluables": clientes_no_evaluables,
        "pedidos_no_evaluables_detalle_total": pedidos_no_evaluables_detalle_total,
        "pedidos_no_evaluables_detalle_mostrados": pedidos_no_evaluables_detalle_mostrados,
        "pedidos_no_evaluables_detalle_truncado": pedidos_no_evaluables_detalle_truncado,
        "pedidos_no_evaluables_detalle": pedidos_no_evaluables_detalle,
        "pedidos_no_evaluables_detalle_por_cliente": pedidos_no_evaluables_detalle_por_cliente,
        "criterio_calculo": "OTIF se calcula solo sobre pedidos evaluables. Los pedidos no evaluables quedan fuera del denominador.",
    }


def calcular_pendientes(pedidos_resumen: list[dict[str, Any]]) -> dict[str, Any]:
    pendientes_registros = []
    por_cliente = defaultdict(lambda: {"total_pedidos": 0, "unidades_pendientes": 0.0, "lineas_pendientes": 0, "pedido_mas_antiguo": None})
    mayores_7_dias = []

    for pedido in pedidos_resumen:
        if not pedido["es_pendiente"]:
            continue

        pendientes_registros.append(
            {
                "pedido_key": pedido["pedido_key"],
                "cliente": pedido["cliente"],
                "nro_pedido": pedido["nro_pedido"],
                "estado": pedido["estado"],
                "unidades": pedido["unidades"],
                "lineas": pedido["lineas"],
                "fecha_inicio_preparacion": pedido["fecha_inicio_preparacion"],
                "fecha_ingreso": pedido["fecha_ingreso"],
                "dias_abierto": pedido["dias_abierto"],
            }
        )

        por_cliente[pedido["cliente"]]["total_pedidos"] += 1
        por_cliente[pedido["cliente"]]["unidades_pendientes"] += pedido["unidades"]
        por_cliente[pedido["cliente"]]["lineas_pendientes"] += pedido["lineas"]
        candidato_antiguo = pedido["fecha_inicio_preparacion"] or pedido["fecha_ingreso"]
        actual_antiguo = por_cliente[pedido["cliente"]]["pedido_mas_antiguo"]
        if candidato_antiguo and (actual_antiguo is None or candidato_antiguo < actual_antiguo):
            por_cliente[pedido["cliente"]]["pedido_mas_antiguo"] = candidato_antiguo

        if pedido["dias_abierto"] is not None and pedido["dias_abierto"] > 7:
            mayores_7_dias.append(
                {
                    "cliente": pedido["cliente"],
                    "nro_pedido": pedido["nro_pedido"],
                    "estado": pedido["estado"],
                    "dias_abierto": pedido["dias_abierto"],
                    "fecha_inicio_preparacion": pedido["fecha_inicio_preparacion"],
                    "lineas": pedido["lineas"],
                    "unidades": pedido["unidades"],
                }
            )

    pendientes_registros.sort(key=lambda item: (-1 if item["dias_abierto"] is None else -item["dias_abierto"], item["cliente"], item["nro_pedido"]))
    mayores_7_dias.sort(key=lambda item: (-item["dias_abierto"], item["cliente"], item["nro_pedido"]))

    resumen_clientes = []
    for cliente, payload in sorted(por_cliente.items()):
        resumen_clientes.append(
            {
                "cliente": cliente,
                "total_pedidos": payload["total_pedidos"],
                "unidades_pendientes": round_safe(payload["unidades_pendientes"], 2),
                "lineas_pendientes": payload["lineas_pendientes"],
                "pedido_mas_antiguo": payload["pedido_mas_antiguo"],
            }
        )

    pedido_mas_antiguo = None
    if pendientes_registros:
        ordenables = [item for item in pendientes_registros if item["fecha_inicio_preparacion"] or item["fecha_ingreso"]]
        if ordenables:
            pedido_mas_antiguo = max(ordenables, key=lambda item: item["dias_abierto"] if item["dias_abierto"] is not None else -1)

    return {
        "total_pedidos": len(pendientes_registros),
        "total_unidades": round_safe(sum(item["unidades"] for item in pendientes_registros), 2),
        "total_lineas": sum(item["lineas"] for item in pendientes_registros),
        "pedido_mas_antiguo": pedido_mas_antiguo,
        "por_cliente": resumen_clientes,
        "mayores_7_dias": mayores_7_dias[:100],
    }


def calcular_cobertura_nnss(
    clientes_periodo: list[str],
    otif: dict[str, Any],
    clientes_fillrate: list[str],
) -> dict[str, Any]:
    clientes_otif = sorted(
        {
            normalizar_texto(item.get("cliente"))
            for item in (otif.get("por_cliente") or [])
            if int(item.get("pedidos_evaluados") or 0) > 0
        }
    )
    clientes_sin_otif = sorted(set(clientes_periodo) - set(clientes_otif))
    clientes_sin_fillrate = sorted(set(clientes_periodo) - set(clientes_fillrate))

    return {
        "clientes_periodo_total": len(clientes_periodo),
        "clientes_periodo": clientes_periodo,
        "clientes_con_otif_evaluable_total": len(clientes_otif),
        "clientes_con_otif_evaluable": clientes_otif,
        "clientes_sin_otif_evaluable_total": len(clientes_sin_otif),
        "clientes_sin_otif_evaluable": clientes_sin_otif,
        "clientes_con_fillrate_total": len(clientes_fillrate),
        "clientes_con_fillrate": clientes_fillrate,
        "clientes_sin_fillrate_evaluable_total": len(clientes_sin_fillrate),
        "clientes_sin_fillrate_evaluable": clientes_sin_fillrate,
        "nota": "Clientes sin OTIF o Fill Rate evaluable pueden tener pedidos pendientes, sin salida o datos incompletos para el indicador.",
    }


def construir_fillrate_cobertura(
    clientes_periodo: list[str],
    clientes_con_datos: list[str],
    pedidos_resumen: list[dict[str, Any]],
) -> dict[str, Any]:
    pedidos_por_cliente = defaultdict(int)
    for pedido in pedidos_resumen:
        pedidos_por_cliente[pedido["cliente"]] += 1

    clientes_sin_datos = []
    for cliente in sorted(set(clientes_periodo) - set(clientes_con_datos)):
        clientes_sin_datos.append(
            {
                "cliente": cliente,
                "motivo": "Sin registros con cantidades validas para calcular Fill Rate en el periodo.",
                "pedidos_periodo": int(pedidos_por_cliente.get(cliente, 0)),
            }
        )

    return {
        "criterio_calculo": "Fill Rate se calcula sobre registros con cantidades validas segun el estado del pedido.",
        "clientes_con_datos": clientes_con_datos,
        "clientes_sin_datos_evaluables": clientes_sin_datos,
        "todos_los_clientes_periodo_con_fillrate": len(clientes_sin_datos) == 0,
    }


def calcular_nnss(df_nnss: pd.DataFrame, year: int, month: int, fecha_consulta: datetime) -> tuple[dict[str, Any], list[str], list[str]]:
    if df_nnss is None or df_nnss.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontraron datos NNSS para el periodo solicitado.",
        }, [], []

    df = df_nnss.copy()
    df["fr_calculado"] = df.apply(
        lambda row: calcular_fr_fila(
            row.get("Estado Pedido"),
            row.get("Cantidad Original"),
            row.get("Cantidad Preparada"),
            row.get("Cantidad Despachada"),
        ),
        axis=1,
    )
    df["fr_evaluable"] = df.apply(
        lambda row: es_fillrate_evaluable_row(
            row.get("Estado Pedido"),
            row.get("Cantidad Original"),
            row.get("Cantidad Preparada"),
            row.get("Cantidad Despachada"),
        ),
        axis=1,
    )
    df["empresa_resumen"] = df["Empresa"].map(normalizar_texto)
    clientes_periodo = sorted(
        {
            normalizar_texto(cliente)
            for cliente in df["empresa_resumen"].dropna().tolist()
            if normalizar_texto(cliente)
        }
    )
    estados_resumen = []
    for estado, grupo in df.groupby(df["Estado Pedido"].map(normalizar_mayusculas)):
        estado = normalizar_etiqueta(estado)
        pedidos = grupo["pedido_key"].nunique()
        estados_resumen.append(
            {
                "estado": estado or "SIN ESTADO",
                "pedidos": int(pedidos),
                "lineas": int(len(grupo)),
                "unidades": round_safe(grupo["Cantidad Original"].sum(), 2),
            }
        )
    estados_resumen.sort(key=lambda item: item["estado"])

    fillrate_por_cliente = []
    for cliente, grupo in df.groupby("empresa_resumen"):
        fillrate_por_cliente.append(
            {
                "cliente": cliente or "SIN_EMPRESA",
                "promedio_fr": round_safe(grupo["fr_calculado"].mean(), 2),
                "lineas": int(len(grupo)),
                "pedidos": int(grupo["pedido_key"].nunique()),
                "tiene_datos_evaluables": bool(grupo["fr_evaluable"].any()),
            }
        )
    fillrate_por_cliente.sort(key=lambda item: (-item["promedio_fr"], item["cliente"]))
    clientes_fillrate = sorted(
        {
            item["cliente"]
            for item in fillrate_por_cliente
            if bool(item.get("tiene_datos_evaluables"))
        }
    )

    pedidos_resumen = resumir_pedidos_nnss(df, fecha_consulta=fecha_consulta)
    pendientes = calcular_pendientes(pedidos_resumen)
    otif = calcular_otif_por_pedido(pedidos_resumen)
    cobertura = calcular_cobertura_nnss(clientes_periodo, otif, clientes_fillrate)
    fillrate_cobertura = construir_fillrate_cobertura(clientes_periodo, clientes_fillrate, pedidos_resumen)

    alertas = []
    recomendaciones = []
    if pendientes["total_pedidos"] > 0:
        alertas.append(f"Hay {pendientes['total_pedidos']} pedidos pendientes en NNSS para el periodo consultado.")
    if pendientes["mayores_7_dias"]:
        alertas.append(
            f"Se detectaron {len(pendientes['mayores_7_dias'])} pedidos pendientes con mas de 7 dias desde inicio de preparacion."
        )
        recomendaciones.append("Priorizar revision de pedidos pendientes con mas de 7 dias.")
    if otif["pedidos_no_evaluables"] > 0:
        recomendaciones.append(
            f"Hay {otif['pedidos_no_evaluables']} pedidos sin datos suficientes para evaluar OTIF."
        )

    return {
        "disponible": True,
        "periodo": {"anio": year, "mes": month},
        "pedidos_por_estado": estados_resumen,
        "pendientes": pendientes,
        "cobertura": cobertura,
        "otif": otif,
        "fillrate": {
            "promedio_fr": round_safe(df["fr_calculado"].mean(), 2),
            "por_cliente": fillrate_por_cliente,
            **fillrate_cobertura,
        },
    }, alertas, recomendaciones


def construir_fila_otif_historico(
    year: int,
    month: int,
    cliente: str,
    clientes_otif: dict[str, dict[str, Any]],
    clientes_periodo: set[str],
    fuentes_por_cliente: dict[str, list[str]],
    motivo_base: str | None = None,
) -> dict[str, Any]:
    payload_cliente = clientes_otif.get(cliente)
    if payload_cliente:
        pedidos_evaluados = int(payload_cliente.get("pedidos_evaluados") or 0)
        pedidos_on_time = int(payload_cliente.get("pedidos_on_time") or 0)
        pedidos_in_full = int(payload_cliente.get("pedidos_in_full") or 0)
        pedidos_otif = int(payload_cliente.get("pedidos_otif") or 0)
        pedidos_no_evaluables = int(payload_cliente.get("pedidos_no_evaluables") or 0)
        disponible = pedidos_evaluados > 0
        if disponible:
            motivo = ""
        elif pedidos_no_evaluables > 0:
            motivo = "sin pedidos OTIF evaluables; solo hay pedidos no evaluables en el período"
        else:
            motivo = "sin datos válidos para el período"
        return {
            "anio": year,
            "mes": month,
            "mes_nombre": MESES_ES.get(month, f"Mes {month}"),
            "cliente": cliente,
            "pedidos_evaluados": pedidos_evaluados,
            "pedidos_on_time": pedidos_on_time,
            "pedidos_in_full": pedidos_in_full,
            "pedidos_otif": pedidos_otif,
            "pedidos_no_evaluables": pedidos_no_evaluables,
            "pct_on_time": porcentaje_safe_nullable(pedidos_on_time, pedidos_evaluados),
            "pct_in_full": porcentaje_safe_nullable(pedidos_in_full, pedidos_evaluados),
            "pct_otif": porcentaje_safe_nullable(pedidos_otif, pedidos_evaluados),
            "disponible": disponible,
            "motivo": motivo,
            "fuentes": sorted(fuentes_por_cliente.get(cliente, [])),
        }

    if cliente in clientes_periodo:
        motivo = "sin OTIF estructurado para el cliente en el período"
    else:
        motivo = motivo_base or "sin datos válidos para el período"
    return {
        "anio": year,
        "mes": month,
        "mes_nombre": MESES_ES.get(month, f"Mes {month}"),
        "cliente": cliente,
        "pedidos_evaluados": 0,
        "pedidos_on_time": 0,
        "pedidos_in_full": 0,
        "pedidos_otif": 0,
        "pedidos_no_evaluables": 0,
        "pct_on_time": None,
        "pct_in_full": None,
        "pct_otif": None,
        "disponible": False,
        "motivo": motivo,
        "fuentes": sorted(fuentes_por_cliente.get(cliente, [])),
    }


def construir_otif_ytd_desde_mensual(
    otif_mensual: list[dict[str, Any]],
    year: int,
    hasta_mes: int,
) -> list[dict[str, Any]]:
    filas_por_cliente: dict[str, dict[int, dict[str, Any]]] = defaultdict(dict)
    for fila in otif_mensual:
        if int(fila.get("anio") or 0) != year:
            continue
        mes = int(fila.get("mes") or 0)
        cliente = normalizar_texto(fila.get("cliente"))
        if not cliente or mes < 1 or mes > hasta_mes:
            continue
        filas_por_cliente[cliente][mes] = fila

    otif_ytd = []
    for cliente in sorted(filas_por_cliente):
        pedidos_evaluados_acum = 0
        pedidos_on_time_acum = 0
        pedidos_in_full_acum = 0
        pedidos_otif_acum = 0
        pedidos_no_evaluables_acum = 0
        meses_incluidos = []
        meses_sin_datos = []

        for mes in range(1, hasta_mes + 1):
            fila = filas_por_cliente[cliente].get(mes)
            if not fila:
                meses_sin_datos.append(mes)
                continue

            pedidos_evaluados_mes = int(fila.get("pedidos_evaluados") or 0)
            pedidos_on_time_mes = int(fila.get("pedidos_on_time") or 0)
            pedidos_in_full_mes = int(fila.get("pedidos_in_full") or 0)
            pedidos_otif_mes = int(fila.get("pedidos_otif") or 0)
            pedidos_no_evaluables_mes = int(fila.get("pedidos_no_evaluables") or 0)

            pedidos_evaluados_acum += pedidos_evaluados_mes
            pedidos_on_time_acum += pedidos_on_time_mes
            pedidos_in_full_acum += pedidos_in_full_mes
            pedidos_otif_acum += pedidos_otif_mes
            pedidos_no_evaluables_acum += pedidos_no_evaluables_mes

            if any(
                [
                    pedidos_evaluados_mes,
                    pedidos_on_time_mes,
                    pedidos_in_full_mes,
                    pedidos_otif_mes,
                    pedidos_no_evaluables_mes,
                ]
            ):
                meses_incluidos.append(mes)

            if not bool(fila.get("disponible")):
                meses_sin_datos.append(mes)

        disponible = pedidos_evaluados_acum > 0
        if disponible:
            motivo = None
        elif meses_incluidos:
            motivo = f"sin pedidos OTIF evaluables acumulados entre enero y {MESES_ES.get(hasta_mes, f'mes {hasta_mes}')} de {year}"
        else:
            motivo = f"sin datos validos acumulados entre enero y {MESES_ES.get(hasta_mes, f'mes {hasta_mes}')} de {year}"

        otif_ytd.append(
            {
                "anio": year,
                "desde_mes": 1,
                "hasta_mes": hasta_mes,
                "cliente": cliente,
                "pedidos_evaluados_acum": pedidos_evaluados_acum,
                "pedidos_on_time_acum": pedidos_on_time_acum,
                "pedidos_in_full_acum": pedidos_in_full_acum,
                "pedidos_otif_acum": pedidos_otif_acum,
                "pedidos_no_evaluables_acum": pedidos_no_evaluables_acum,
                "pct_on_time": porcentaje_safe_nullable(pedidos_on_time_acum, pedidos_evaluados_acum),
                "pct_in_full": porcentaje_safe_nullable(pedidos_in_full_acum, pedidos_evaluados_acum),
                "pct_otif": porcentaje_safe_nullable(pedidos_otif_acum, pedidos_evaluados_acum),
                "disponible": disponible,
                "motivo": motivo,
                "meses_incluidos": meses_incluidos,
                "meses_sin_datos": meses_sin_datos,
                "origen_valor": "cierre_recalculado",
                "criterio_historico": "recalculado_desde_fuente_viva",
            }
        )

    return otif_ytd


def calcular_componentes_fillrate_row(
    estado: Any,
    cantidad_original: Any,
    cantidad_preparada: Any,
    cantidad_despachada: Any,
) -> tuple[float, float, bool]:
    estado_txt = normalizar_etiqueta(estado)
    original = parse_numero(cantidad_original)
    if original <= 0:
        return 0.0, 0.0, False
    if estado_txt in ESTADOS_PREPARACION:
        return original, original, True
    if estado_txt in ESTADOS_PREPARADOS:
        return parse_numero(cantidad_preparada), original, True
    if estado_txt in ESTADOS_DESPACHADOS:
        return parse_numero(cantidad_despachada), original, True
    return 0.0, 0.0, False


def resumir_fillrate_historico_por_cliente(
    df_nnss: pd.DataFrame,
) -> tuple[dict[str, dict[str, Any]], set[str], dict[str, list[str]]]:
    clientes_fillrate: dict[str, dict[str, Any]] = {}
    clientes_periodo: set[str] = set()
    fuentes_por_cliente: dict[str, list[str]] = defaultdict(list)

    if df_nnss is None or df_nnss.empty:
        return clientes_fillrate, clientes_periodo, fuentes_por_cliente

    df = df_nnss.copy()
    componentes_fillrate = df.apply(
        lambda row: calcular_componentes_fillrate_row(
            row.get("Estado Pedido"),
            row.get("Cantidad Original"),
            row.get("Cantidad Preparada"),
            row.get("Cantidad Despachada"),
        ),
        axis=1,
        result_type="expand",
    )
    componentes_fillrate.columns = ["fr_numerador_base", "fr_denominador_base", "fr_evaluable_historico"]
    df[["fr_numerador_base", "fr_denominador_base", "fr_evaluable_historico"]] = componentes_fillrate

    for cliente, grupo in df.groupby("Empresa", dropna=False):
        cliente_txt = normalizar_texto(cliente)
        if not cliente_txt:
            continue

        clientes_periodo.add(cliente_txt)
        fuentes_por_cliente[cliente_txt] = sorted(
            {
                str(path)
                for path in grupo["__archivo_fuente"].dropna().tolist()
                if normalizar_texto(path)
            }
        )

        grupo_evaluable = grupo[grupo["fr_evaluable_historico"] == True]
        fillrate_numerador = float(grupo_evaluable["fr_numerador_base"].sum()) if not grupo_evaluable.empty else 0.0
        fillrate_denominador = float(grupo_evaluable["fr_denominador_base"].sum()) if not grupo_evaluable.empty else 0.0
        documentos_evaluables = int(grupo_evaluable["pedido_key"].nunique()) if not grupo_evaluable.empty else 0
        disponible = fillrate_denominador > 0
        if disponible:
            motivo = None
        else:
            motivo = "sin lineas Fill Rate evaluables en el periodo"

        clientes_fillrate[cliente_txt] = {
            "fillrate_numerador": serializar_numero_cantidad(fillrate_numerador),
            "fillrate_denominador": serializar_numero_cantidad(fillrate_denominador),
            "documentos_evaluables": documentos_evaluables,
            "pct_fillrate": porcentaje_safe_nullable(fillrate_numerador, fillrate_denominador),
            "disponible": disponible,
            "motivo": motivo,
            "fuentes": fuentes_por_cliente[cliente_txt],
            "origen_valor": "cierre_recalculado",
            "criterio_historico": "recalculado_desde_fuente_viva",
        }

    return clientes_fillrate, clientes_periodo, fuentes_por_cliente


def construir_fila_fillrate_historico(
    year: int,
    month: int,
    cliente: str,
    clientes_fillrate: dict[str, dict[str, Any]],
    clientes_periodo: set[str],
    fuentes_por_cliente: dict[str, list[str]],
    motivo_base: str | None = None,
) -> dict[str, Any]:
    payload_cliente = clientes_fillrate.get(cliente)
    if payload_cliente:
        return {
            "anio": year,
            "mes": month,
            "mes_nombre": MESES_ES.get(month, f"Mes {month}"),
            "cliente": cliente,
            "fillrate_numerador": payload_cliente.get("fillrate_numerador", 0),
            "fillrate_denominador": payload_cliente.get("fillrate_denominador", 0),
            "documentos_evaluables": int(payload_cliente.get("documentos_evaluables") or 0),
            "pct_fillrate": payload_cliente.get("pct_fillrate"),
            "disponible": bool(payload_cliente.get("disponible")),
            "motivo": payload_cliente.get("motivo"),
            "fuentes": sorted(payload_cliente.get("fuentes") or fuentes_por_cliente.get(cliente, [])),
            "origen_valor": payload_cliente.get("origen_valor", "cierre_recalculado"),
            "criterio_historico": payload_cliente.get("criterio_historico", "recalculado_desde_fuente_viva"),
        }

    if cliente in clientes_periodo:
        motivo = "sin Fill Rate estructurado para el cliente en el periodo"
    else:
        motivo = motivo_base or "sin datos validos para el periodo"
    return {
        "anio": year,
        "mes": month,
        "mes_nombre": MESES_ES.get(month, f"Mes {month}"),
        "cliente": cliente,
        "fillrate_numerador": 0,
        "fillrate_denominador": 0,
        "documentos_evaluables": 0,
        "pct_fillrate": None,
        "disponible": False,
        "motivo": motivo,
        "fuentes": sorted(fuentes_por_cliente.get(cliente, [])),
        "origen_valor": "cierre_recalculado",
        "criterio_historico": "recalculado_desde_fuente_viva",
    }


def construir_fillrate_ytd_desde_mensual(
    fillrate_mensual: list[dict[str, Any]],
    year: int,
    hasta_mes: int,
) -> list[dict[str, Any]]:
    filas_por_cliente: dict[str, dict[int, dict[str, Any]]] = defaultdict(dict)
    for fila in fillrate_mensual:
        if int(fila.get("anio") or 0) != year:
            continue
        mes = int(fila.get("mes") or 0)
        cliente = normalizar_texto(fila.get("cliente"))
        if not cliente or mes < 1 or mes > hasta_mes:
            continue
        filas_por_cliente[cliente][mes] = fila

    fillrate_ytd = []
    for cliente in sorted(filas_por_cliente):
        fillrate_numerador_acum = 0.0
        fillrate_denominador_acum = 0.0
        documentos_evaluables_acum = 0
        meses_incluidos = []
        meses_sin_datos = []

        for mes in range(1, hasta_mes + 1):
            fila = filas_por_cliente[cliente].get(mes)
            if not fila:
                meses_sin_datos.append(mes)
                continue

            fillrate_numerador_mes = parse_numero(fila.get("fillrate_numerador"))
            fillrate_denominador_mes = parse_numero(fila.get("fillrate_denominador"))
            documentos_evaluables_mes = int(fila.get("documentos_evaluables") or 0)

            fillrate_numerador_acum += fillrate_numerador_mes
            fillrate_denominador_acum += fillrate_denominador_mes
            documentos_evaluables_acum += documentos_evaluables_mes

            if any([fillrate_numerador_mes, fillrate_denominador_mes, documentos_evaluables_mes]):
                meses_incluidos.append(mes)

            if not bool(fila.get("disponible")):
                meses_sin_datos.append(mes)

        disponible = fillrate_denominador_acum > 0
        if disponible:
            motivo = None
        elif meses_incluidos:
            motivo = f"sin denominador Fill Rate acumulado entre enero y {MESES_ES.get(hasta_mes, f'mes {hasta_mes}')} de {year}"
        else:
            motivo = f"sin datos validos acumulados de Fill Rate entre enero y {MESES_ES.get(hasta_mes, f'mes {hasta_mes}')} de {year}"

        fillrate_ytd.append(
            {
                "anio": year,
                "desde_mes": 1,
                "hasta_mes": hasta_mes,
                "cliente": cliente,
                "fillrate_numerador_acum": serializar_numero_cantidad(fillrate_numerador_acum),
                "fillrate_denominador_acum": serializar_numero_cantidad(fillrate_denominador_acum),
                "documentos_evaluables_acum": documentos_evaluables_acum,
                "pct_fillrate": porcentaje_safe_nullable(fillrate_numerador_acum, fillrate_denominador_acum),
                "disponible": disponible,
                "motivo": motivo,
                "meses_incluidos": meses_incluidos,
                "meses_sin_datos": meses_sin_datos,
                "origen_valor": "cierre_recalculado",
                "criterio_historico": "recalculado_desde_fuente_viva",
            }
        )

    return fillrate_ytd


def inferir_cd_productividad_desde_fuente(path_fuente: Any) -> str:
    texto_path = normalizar_mayusculas(path_fuente)
    if "\\CD QUILICURA\\" in texto_path:
        return "QUILICURA"
    if "\\CD PUDAHUEL\\" in texto_path:
        return "PUDAHUEL"
    return ""


def cargar_dataframe_productividad_historico(
    fuentes: list[FuenteDetectada],
    ubicaciones_map: dict[str, dict[str, str]],
) -> tuple[pd.DataFrame | None, dict[str, Any]]:
    if pd is None:
        return None, {
            "disponible": False,
            "mensaje": "pandas no esta disponible en este entorno.",
            "archivos": [],
            "errores": [],
        }

    if not fuentes:
        return None, {
            "disponible": False,
            "mensaje": "No se encontro una fuente local de productividad para el periodo solicitado.",
            "archivos": [],
            "errores": [],
        }

    frames = []
    errores = []
    archivos_leidos = []
    for fuente in fuentes:
        try:
            df_fuente, info_fuente = limpiar_productividad_excel(fuente)
            if not df_fuente.empty:
                frames.append(df_fuente)
            archivos_leidos.append(info_fuente)
        except Exception as exc:
            errores.append(f"{fuente.ruta.name}: {exc}")

    if not frames:
        return None, {
            "disponible": False,
            "mensaje": "No se encontro una fuente local de productividad para el periodo solicitado.",
            "archivos": archivos_leidos,
            "errores": errores[:5],
        }

    df = pd.concat(frames, ignore_index=True)
    df["Ubicacion_Clave"] = df["Ubicacion"].map(normalizar_ubicacion_lookup)
    # Fase 2: clasificacion Rack/Est unificada con canal_derco_auto.py via canal_derco_utils.
    # Antes: lookup contra Tabla Ubicaciones CDs.xlsx (DimUbicaciones) -> SIN_DIM si no estaba.
    # Ahora: regla de prefijos sobre Ubicacion (mismo criterio que data Derco).
    df["Tipo_Ubicacion_Dim"] = df["Ubicacion"].map(clasificar_ubicacion_dim)
    df["Cliente_Meta"] = df.apply(lambda row: cliente_meta_productividad(row.get("Cliente"), row.get("Centro")), axis=1)
    df["Fecha_Turno"] = df["timestamp_operacion"].map(calcular_fecha_turno)
    df["Turno"] = df["timestamp_operacion"].map(calcular_turno)
    df["Hora_Operativa"] = df["timestamp_operacion"].map(calcular_hora_operativa)
    # Fase 2 CES: inyectar columna es_ces_destino antes de clasificar canal,
    # para que MY+destino_CES se reclasifique como canal CES (mismo criterio FillRate).
    df = _aplicar_es_ces_destino(df)
    canales = df.apply(ajustar_canal_detalle_derco, axis=1, result_type="expand")
    canales.columns = [
        "Canal_Principal",
        "Canal_Detalle",
        "Tipo_Ubicacion_Meta",
        "Tipo_Ubicacion_Corregida",
        "Canal_Detalle_Metodo",
    ]
    df = pd.concat([df, canales], axis=1)
    # Canal mayorista = CAP + MY + SG + CES (CES son concesionarios, parte del mayorista
    # según regla de negocio confirmada por usuario 2026-05-15). En canales_originales CES
    # sigue apareciendo separado.
    df["Canal_Agrupado"] = df["Canal_Principal"].map(lambda c: "CAP-MY-SG-CES" if c in {"CAP", "MY", "SG", "CES"} else c)
    df["Cliente_Norm"] = df["Cliente"].map(normalizar_texto_seguro)
    df["Centro_Norm"] = df["Centro"].map(normalizar_texto_seguro)
    df.loc[df["Centro_Norm"] == "", "Centro_Norm"] = df.loc[df["Centro_Norm"] == "", "__archivo_fuente"].map(
        inferir_cd_productividad_desde_fuente
    )
    df["Fecha_Turno_Texto"] = df["Fecha_Turno"].map(serializar_fecha_productividad)
    df["Es_Derco_Historico"] = df["Cliente"].map(normalizar_mayusculas).isin({"DERCO", "GRUPO PLANET"})

    return df, {
        "disponible": True,
        "archivos": archivos_leidos,
        "errores": errores[:5],
    }


def calcular_horas_trabajadas_subset_productividad(df_subset: pd.DataFrame | None) -> float:
    if df_subset is None or df_subset.empty:
        return 0.0
    horas_grupo = (
        df_subset.groupby(["Fecha_Turno", "Registro", "Turno"])["timestamp_operacion"]
        .agg(["min", "max"])
        .reset_index()
    )
    horas_total = 0.0
    for _, row in horas_grupo.iterrows():
        delta_horas = (row["max"] - row["min"]).total_seconds() / 3600.0
        if delta_horas > 1:
            delta_horas -= 1
        horas_total += max(delta_horas, 0.0)
    return horas_total


def resumir_metricas_productividad_subset(df_subset: pd.DataFrame | None) -> dict[str, Any]:
    if df_subset is None or df_subset.empty:
        return {
            "lineas": 0,
            "unidades": 0.0,
            "pedidos_unicos": 0,
            "dias_trabajados": 0,
            "horas_trabajadas_total": 0.0,
            "lineas_dia": None,
            "unidades_dia": None,
            "lineas_hora": None,
            "unidades_hora": None,
        }

    lineas = int(len(df_subset))
    unidades = round_safe(df_subset["Salida"].sum(), 2)
    pedidos_unicos = contar_pedidos_validos(df_subset["Nro. de Doc. Externo"])
    dias_trabajados = int(df_subset["Fecha_Turno"].dropna().nunique())
    horas_trabajadas_total = round_safe(calcular_horas_trabajadas_subset_productividad(df_subset), 2)
    return {
        "lineas": lineas,
        "unidades": unidades,
        "pedidos_unicos": pedidos_unicos,
        "dias_trabajados": dias_trabajados,
        "horas_trabajadas_total": horas_trabajadas_total,
        "lineas_dia": division_lineal_nullable(lineas, dias_trabajados),
        "unidades_dia": division_lineal_nullable(unidades, dias_trabajados),
        "lineas_hora": division_lineal_nullable(lineas, horas_trabajadas_total),
        "unidades_hora": division_lineal_nullable(unidades, horas_trabajadas_total),
    }


def construir_fila_productividad_historica(
    year: int,
    month: int,
    cliente: str,
    cd: str,
    df_subset: pd.DataFrame | None,
    fuentes: list[str] | None,
    motivo_base: str | None = None,
) -> dict[str, Any]:
    metricas = resumir_metricas_productividad_subset(df_subset)
    disponible = metricas["lineas"] > 0
    motivo = None if disponible else (motivo_base or "sin datos validos para el periodo")
    return {
        "anio": year,
        "mes": month,
        "mes_nombre": MESES_ES.get(month, f"Mes {month}"),
        "cliente": cliente,
        "cd": cd,
        "lineas": metricas["lineas"],
        "unidades": metricas["unidades"],
        "pedidos_unicos": metricas["pedidos_unicos"],
        "dias_trabajados": metricas["dias_trabajados"],
        "horas_trabajadas_total": metricas["horas_trabajadas_total"],
        "lineas_dia": metricas["lineas_dia"],
        "unidades_dia": metricas["unidades_dia"],
        "lineas_hora": metricas["lineas_hora"],
        "unidades_hora": metricas["unidades_hora"],
        "disponible": disponible,
        "motivo": motivo,
        "fuentes": sorted(fuentes or []),
        "origen_valor": "cierre_recalculado",
        "criterio_historico": "recalculado_desde_fuente_viva",
    }


def construir_fila_productividad_ytd(
    year: int,
    hasta_mes: int,
    cliente: str,
    cd: str,
    df_subset_ytd: pd.DataFrame | None,
    meses_incluidos: list[int],
    meses_sin_datos: list[int],
) -> dict[str, Any]:
    metricas = resumir_metricas_productividad_subset(df_subset_ytd)
    disponible = metricas["lineas"] > 0
    if disponible:
        motivo = None
    elif meses_incluidos:
        motivo = f"sin productividad acumulada valida entre enero y {MESES_ES.get(hasta_mes, f'mes {hasta_mes}')} de {year}"
    else:
        motivo = f"sin datos validos acumulados de productividad entre enero y {MESES_ES.get(hasta_mes, f'mes {hasta_mes}')} de {year}"
    return {
        "anio": year,
        "desde_mes": 1,
        "hasta_mes": hasta_mes,
        "cliente": cliente,
        "cd": cd,
        "lineas_acum": metricas["lineas"],
        "unidades_acum": metricas["unidades"],
        "pedidos_unicos_acum": metricas["pedidos_unicos"],
        "dias_trabajados_acum": metricas["dias_trabajados"],
        "horas_trabajadas_total_acum": metricas["horas_trabajadas_total"],
        "lineas_dia": metricas["lineas_dia"],
        "unidades_dia": metricas["unidades_dia"],
        "lineas_hora": metricas["lineas_hora"],
        "unidades_hora": metricas["unidades_hora"],
        "disponible": disponible,
        "motivo": motivo,
        "meses_incluidos": meses_incluidos,
        "meses_sin_datos": meses_sin_datos,
        "origen_valor": "cierre_recalculado",
        "criterio_historico": "recalculado_desde_fuente_viva",
    }


def construir_fila_derco_ap_historica(
    year: int,
    month: int,
    segmento_ap: str,
    cd: str,
    df_subset: pd.DataFrame | None,
    fuentes: list[str] | None,
    motivo_base: str | None = None,
) -> dict[str, Any]:
    fila = construir_fila_productividad_historica(
        year=year,
        month=month,
        cliente="DERCO",
        cd=cd,
        df_subset=df_subset,
        fuentes=fuentes,
        motivo_base=motivo_base,
    )
    fila["segmento_ap"] = segmento_ap
    return fila


def construir_fila_derco_ap_ytd(
    year: int,
    hasta_mes: int,
    segmento_ap: str,
    cd: str,
    df_subset_ytd: pd.DataFrame | None,
    meses_incluidos: list[int],
    meses_sin_datos: list[int],
) -> dict[str, Any]:
    fila = construir_fila_productividad_ytd(
        year=year,
        hasta_mes=hasta_mes,
        cliente="DERCO",
        cd=cd,
        df_subset_ytd=df_subset_ytd,
        meses_incluidos=meses_incluidos,
        meses_sin_datos=meses_sin_datos,
    )
    fila["segmento_ap"] = segmento_ap
    return fila


def filtrar_subset_productividad_cliente_cd(
    df_productividad: pd.DataFrame | None,
    cliente: str,
    cd: str,
) -> pd.DataFrame | None:
    if df_productividad is None or df_productividad.empty:
        return None
    subset = df_productividad[
        (df_productividad["Cliente_Norm"] == cliente)
        & (df_productividad["Centro_Norm"] == cd)
    ].copy()
    return subset if not subset.empty else None


def filtrar_subset_derco_ap_productividad(
    df_productividad: pd.DataFrame | None,
    segmento_ap: str,
    cd: str,
) -> pd.DataFrame | None:
    if df_productividad is None or df_productividad.empty:
        return None
    subset = df_productividad[
        df_productividad["Es_Derco_Historico"]
        & (df_productividad["Centro_Norm"] == cd)
    ].copy()
    if subset.empty:
        return None
    if segmento_ap == "AP Total":
        subset = subset[subset["Canal_Principal"] == "AP"].copy()
    elif segmento_ap == "AP Rack":
        subset = subset[subset["Canal_Detalle"] == "AP Rack"].copy()
    elif segmento_ap == "AP Estanteria":
        subset = subset[subset["Canal_Detalle"] == "AP Estanteria"].copy()
    else:
        subset = subset.iloc[0:0].copy()
    return subset if not subset.empty else None


def construir_historico_otif_mensual(
    nnss_fuentes: list[FuenteDetectada],
    year: int,
    hasta_mes: int,
    fecha_generacion: str,
    ubicaciones_map: dict[str, dict[str, str]],
) -> dict[str, Any]:
    registros_por_mes: dict[int, dict[str, Any]] = {}
    universo_clientes: set[str] = set()
    universo_productividad: set[tuple[str, str]] = set()
    universo_derco_ap: set[tuple[str, str]] = set()

    for month in range(1, hasta_mes + 1):
        fecha_consulta = fecha_corte_periodo(year, month)
        df_nnss, fuente_nnss = leer_consulta_fr(nnss_fuentes, year, month, verbose=False)
        nnss_payload, _, _ = calcular_nnss(
            df_nnss,
            year=year,
            month=month,
            fecha_consulta=fecha_consulta,
        )

        clientes_otif = {
            normalizar_texto(item.get("cliente")): item
            for item in (nnss_payload.get("otif", {}) or {}).get("por_cliente", [])
            if normalizar_texto(item.get("cliente"))
        }
        clientes_periodo = {
            normalizar_texto(cliente)
            for cliente in (nnss_payload.get("cobertura", {}) or {}).get("clientes_periodo", [])
            if normalizar_texto(cliente)
        }
        fuentes_por_cliente: dict[str, list[str]] = defaultdict(list)
        if df_nnss is not None and not df_nnss.empty:
            for cliente, grupo in df_nnss.groupby("Empresa", dropna=False):
                cliente_txt = normalizar_texto(cliente)
                if not cliente_txt:
                    continue
                fuentes = sorted(
                    {
                        str(path)
                        for path in grupo["__archivo_fuente"].dropna().tolist()
                        if normalizar_texto(path)
                    }
                )
                fuentes_por_cliente[cliente_txt] = fuentes
        clientes_fillrate, clientes_fillrate_periodo, fuentes_fillrate_por_cliente = resumir_fillrate_historico_por_cliente(df_nnss)
        fuentes_productividad_mes = descubrir_fuentes_productividad(year, month, verbose=False)
        df_productividad_mes, info_productividad_mes = cargar_dataframe_productividad_historico(
            fuentes_productividad_mes,
            ubicaciones_map=ubicaciones_map,
        )
        if df_productividad_mes is not None and not df_productividad_mes.empty:
            for cliente_prod, cd_prod in (
                df_productividad_mes[["Cliente_Norm", "Centro_Norm"]]
                .drop_duplicates()
                .itertuples(index=False, name=None)
            ):
                if cliente_prod and cd_prod:
                    universo_productividad.add((cliente_prod, cd_prod))
            derco_ap_cd = (
                df_productividad_mes.loc[
                    df_productividad_mes["Es_Derco_Historico"]
                    & df_productividad_mes["Canal_Principal"].eq("AP"),
                    "Centro_Norm",
                ]
                .dropna()
                .tolist()
            )
            for cd_derco in sorted({normalizar_texto(cd) for cd in derco_ap_cd if normalizar_texto(cd)}):
                universo_derco_ap.add(("AP Total", cd_derco))
                universo_derco_ap.add(("AP Rack", cd_derco))
                universo_derco_ap.add(("AP Estanteria", cd_derco))
        universo_clientes.update(clientes_periodo)
        universo_clientes.update(clientes_otif.keys())
        universo_clientes.update(clientes_fillrate_periodo)
        universo_clientes.update(clientes_fillrate.keys())
        registros_por_mes[month] = {
            "disponible_periodo": bool(nnss_payload.get("disponible")),
            "motivo_periodo": nnss_payload.get("mensaje", "sin datos NNSS para el período"),
            "clientes_otif": clientes_otif,
            "clientes_periodo": clientes_periodo,
            "fuentes_por_cliente": fuentes_por_cliente,
            "clientes_fillrate": clientes_fillrate,
            "clientes_fillrate_periodo": clientes_fillrate_periodo,
            "fuentes_fillrate_por_cliente": fuentes_fillrate_por_cliente,
            "fuente_info": fuente_nnss,
            "df_productividad": df_productividad_mes,
            "info_productividad": info_productividad_mes,
        }

    otif_mensual = []
    fillrate_mensual = []
    productividad_mensual_cliente = []
    derco_ap_mensual = []
    derco_canales_mensual = []
    for month in range(1, hasta_mes + 1):
        datos_mes = registros_por_mes[month]
        clientes_mes = sorted(
            universo_clientes
            | set(datos_mes["clientes_otif"].keys())
            | set(datos_mes["clientes_periodo"])
            | set(datos_mes["clientes_fillrate"].keys())
            | set(datos_mes["clientes_fillrate_periodo"])
        )
        for cliente in clientes_mes:
            otif_mensual.append(
                construir_fila_otif_historico(
                    year=year,
                    month=month,
                    cliente=cliente,
                    clientes_otif=datos_mes["clientes_otif"],
                    clientes_periodo=datos_mes["clientes_periodo"],
                    fuentes_por_cliente=datos_mes["fuentes_por_cliente"],
                    motivo_base=datos_mes["motivo_periodo"],
                )
            )
            fillrate_mensual.append(
                construir_fila_fillrate_historico(
                    year=year,
                    month=month,
                    cliente=cliente,
                    clientes_fillrate=datos_mes["clientes_fillrate"],
                    clientes_periodo=datos_mes["clientes_fillrate_periodo"],
                    fuentes_por_cliente=datos_mes["fuentes_fillrate_por_cliente"],
                    motivo_base=datos_mes["motivo_periodo"],
                )
            )

        motivo_productividad = (
            (datos_mes.get("info_productividad") or {}).get("mensaje")
            or "sin datos validos para el periodo"
        )
        for cliente_prod, cd_prod in sorted(universo_productividad):
            subset_prod = filtrar_subset_productividad_cliente_cd(
                datos_mes.get("df_productividad"),
                cliente=cliente_prod,
                cd=cd_prod,
            )
            fuentes_prod = []
            if subset_prod is not None and not subset_prod.empty:
                fuentes_prod = sorted(
                    {
                        str(path)
                        for path in subset_prod["__archivo_fuente"].dropna().tolist()
                        if normalizar_texto(path)
                    }
                )
            productividad_mensual_cliente.append(
                construir_fila_productividad_historica(
                    year=year,
                    month=month,
                    cliente=cliente_prod,
                    cd=cd_prod,
                    df_subset=subset_prod,
                    fuentes=fuentes_prod,
                    motivo_base=motivo_productividad,
                )
            )

        for segmento_ap, cd_derco in sorted(universo_derco_ap):
            subset_derco_ap = filtrar_subset_derco_ap_productividad(
                datos_mes.get("df_productividad"),
                segmento_ap=segmento_ap,
                cd=cd_derco,
            )
            fuentes_derco_ap = []
            if subset_derco_ap is not None and not subset_derco_ap.empty:
                fuentes_derco_ap = sorted(
                    {
                        str(path)
                        for path in subset_derco_ap["__archivo_fuente"].dropna().tolist()
                        if normalizar_texto(path)
                    }
                )
            derco_ap_mensual.append(
                construir_fila_derco_ap_historica(
                    year=year,
                    month=month,
                    segmento_ap=segmento_ap,
                    cd=cd_derco,
                    df_subset=subset_derco_ap,
                    fuentes=fuentes_derco_ap,
                    motivo_base=motivo_productividad,
                )
            )

        # Canal DERCO historico — MY, SG, CAP, GT, AP por mes
        df_prod_mes = datos_mes.get("df_productividad")
        if df_prod_mes is not None and not df_prod_mes.empty:
            df_derco_mes = df_prod_mes[df_prod_mes["Es_Derco_Historico"]].copy()
            if not df_derco_mes.empty:
                cds_derco = sorted(
                    {normalizar_texto(cd) for cd in df_derco_mes["Centro_Norm"].dropna() if normalizar_texto(cd)}
                )
                for cd_derco in cds_derco:
                    df_derco_cd = df_derco_mes[df_derco_mes["Centro_Norm"] == cd_derco]
                    for canal, grupo in df_derco_cd.groupby("Canal_Principal", dropna=False):
                        canal_str = str(canal).strip() if canal and str(canal).strip() else ""
                        if not canal_str:
                            continue
                        derco_canales_mensual.append({
                            "anio": year,
                            "mes": month,
                            "cd": cd_derco,
                            "canal": canal_str,
                            "lineas": int(len(grupo)),
                            "unidades": round_safe(grupo["Salida"].sum(), 2),
                            "pedidos": contar_pedidos_validos(grupo["Nro. de Doc. Externo"]),
                            "disponible": True,
                        })

    otif_ytd = construir_otif_ytd_desde_mensual(
        otif_mensual=otif_mensual,
        year=year,
        hasta_mes=hasta_mes,
    )
    fillrate_ytd = construir_fillrate_ytd_desde_mensual(
        fillrate_mensual=fillrate_mensual,
        year=year,
        hasta_mes=hasta_mes,
    )
    productividad_ytd_cliente = []
    for cliente_prod, cd_prod in sorted(universo_productividad):
        frames_ytd = []
        meses_incluidos = []
        meses_sin_datos = []
        for month in range(1, hasta_mes + 1):
            subset_prod = filtrar_subset_productividad_cliente_cd(
                registros_por_mes[month].get("df_productividad"),
                cliente=cliente_prod,
                cd=cd_prod,
            )
            if subset_prod is None or subset_prod.empty:
                meses_sin_datos.append(month)
                continue
            frames_ytd.append(subset_prod)
            meses_incluidos.append(month)
        df_subset_ytd = pd.concat(frames_ytd, ignore_index=True) if frames_ytd else None
        productividad_ytd_cliente.append(
            construir_fila_productividad_ytd(
                year=year,
                hasta_mes=hasta_mes,
                cliente=cliente_prod,
                cd=cd_prod,
                df_subset_ytd=df_subset_ytd,
                meses_incluidos=meses_incluidos,
                meses_sin_datos=meses_sin_datos,
            )
        )

    derco_ap_ytd = []
    for segmento_ap, cd_derco in sorted(universo_derco_ap):
        frames_ytd = []
        meses_incluidos = []
        meses_sin_datos = []
        for month in range(1, hasta_mes + 1):
            subset_derco_ap = filtrar_subset_derco_ap_productividad(
                registros_por_mes[month].get("df_productividad"),
                segmento_ap=segmento_ap,
                cd=cd_derco,
            )
            if subset_derco_ap is None or subset_derco_ap.empty:
                meses_sin_datos.append(month)
                continue
            frames_ytd.append(subset_derco_ap)
            meses_incluidos.append(month)
        df_subset_ytd = pd.concat(frames_ytd, ignore_index=True) if frames_ytd else None
        derco_ap_ytd.append(
            construir_fila_derco_ap_ytd(
                year=year,
                hasta_mes=hasta_mes,
                segmento_ap=segmento_ap,
                cd=cd_derco,
                df_subset_ytd=df_subset_ytd,
                meses_incluidos=meses_incluidos,
                meses_sin_datos=meses_sin_datos,
            )
        )

    # por_usuario_mensual: un registro por (cd, usuario, mes) para todos los meses disponibles
    por_usuario_mensual: list[dict] = []
    # por_usuario_canal_mensual: un registro por (cd, usuario, canal, mes) — solo DERCO
    por_usuario_canal_mensual: list[dict] = []
    # por_usuario_cliente_mensual: un registro por (cd, cliente, usuario, mes) — excluye DERCO
    # (DERCO ya tiene desglose más granular por canal en por_usuario_canal_mensual)
    por_usuario_cliente_mensual: list[dict] = []
    for _month in range(1, hasta_mes + 1):
        _df = registros_por_mes.get(_month, {}).get("df_productividad")
        if _df is not None and not _df.empty:
            por_usuario_mensual.extend(calcular_por_usuario(_df, year, _month))
            por_usuario_canal_mensual.extend(
                calcular_por_usuario_canal(_df, year, _month, cliente_filtro="DERCO")
            )
            por_usuario_cliente_mensual.extend(
                calcular_por_usuario_cliente(
                    _df, year, _month,
                    excluir_clientes={"DERCO", "GRUPO PLANET"},
                )
            )

    # Alias: mes más reciente como atajo rápido para consultas del periodo actual
    por_usuario = [f for f in por_usuario_mensual if f.get("mes") == hasta_mes]
    por_usuario_canal = [f for f in por_usuario_canal_mensual if f.get("mes") == hasta_mes]
    por_usuario_cliente = [f for f in por_usuario_cliente_mensual if f.get("mes") == hasta_mes]

    # lineas_no_asignadas_por_canal_mes: gap entre derco_canales_mensual (total del canal)
    # y la suma de por_usuario_canal_mensual (operadores que sí clasifican). El gap suele
    # ser por líneas con Fecha_Turno fuera del mes del archivo (spillover) o sin Registro
    # válido. Sirve para que el bot explique al usuario por qué suma operador != total canal.
    from collections import defaultdict as _dd
    _agg_puc = _dd(lambda: {"lineas": 0, "unidades": 0.0})
    for r in por_usuario_canal_mensual:
        k = (r["cd"], r["canal"], r["mes"])
        _agg_puc[k]["lineas"] += r["lineas"]
        _agg_puc[k]["unidades"] += r["unidades"]
    lineas_no_asignadas_por_canal_mes: list[dict] = []
    for r in derco_canales_mensual:
        k = ("CD " + str(r["cd"]).upper(), r["canal"], r["mes"])
        # CD en derco_canales_mensual viene ya normalizado a "QUILICURA"/"PUDAHUEL" etc;
        # _agg_puc usa cd_display ("CD QUILICURA"). Probamos ambas formas.
        suma = _agg_puc.get(k, {"lineas": 0, "unidades": 0.0})
        if suma["lineas"] == 0:
            k_alt = (str(r["cd"]), r["canal"], r["mes"])
            suma = _agg_puc.get(k_alt, suma)
        diff_lineas = r["lineas"] - suma["lineas"]
        diff_unidades = r.get("unidades", 0) - suma["unidades"]
        if diff_lineas > 0:
            lineas_no_asignadas_por_canal_mes.append({
                "cd": r["cd"],
                "anio": r.get("anio", year),
                "mes": r["mes"],
                "canal": r["canal"],
                "lineas_no_asignadas": int(diff_lineas),
                "unidades_no_asignadas": round_safe(diff_unidades, 2),
                "total_lineas_canal": int(r["lineas"]),
                "lineas_asignadas_a_operador": int(suma["lineas"]),
                "motivo": "Líneas con Fecha_Turno fuera del mes del archivo (spillover) o sin operador WMS registrado",
            })

    return {
        "disponible": bool(otif_mensual or productividad_mensual_cliente),
        "criterio_historico": "cierre_recalculado",
        "origen_historico": "recalculado_desde_fuente_viva",
        "fecha_generacion": fecha_generacion,
        "advertencia": "Los valores historicos pueden variar si la fuente NNSS cambia despues del cierre operativo.",
        "corte_operativo_disponible": False,
        "periodo_cobertura": {
            "anio": year,
            "desde_mes": 1,
            "hasta_mes": hasta_mes,
            "tipo": "mensual",
        },
        "nnss": {
            "otif_mensual": otif_mensual,
            "otif_ytd": otif_ytd,
            "fillrate_mensual": fillrate_mensual,
            "fillrate_ytd": fillrate_ytd,
        },
        "productividad": {
            "mensual_cliente": productividad_mensual_cliente,
            "ytd_cliente": productividad_ytd_cliente,
            "derco_ap_mensual": derco_ap_mensual,
            "derco_ap_ytd": derco_ap_ytd,
            "derco_canales_mensual": derco_canales_mensual,
            "por_usuario": por_usuario,
            "por_usuario_mensual": por_usuario_mensual,
            "por_usuario_canal": por_usuario_canal,
            "por_usuario_canal_mensual": por_usuario_canal_mensual,
            "por_usuario_cliente": por_usuario_cliente,
            "por_usuario_cliente_mensual": por_usuario_cliente_mensual,
            "lineas_no_asignadas_por_canal_mes": lineas_no_asignadas_por_canal_mes,
        },
    }


def leer_metadata_productividad(path: Path) -> tuple[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        centro = normalizar_texto(ws["A1"].value)
        cliente = normalizar_texto(ws["A4"].value)
    finally:
        wb.close()
    return centro, cliente


def descubrir_fuentes_productividad(year: int, month: int, verbose: bool = False) -> list[FuenteDetectada]:
    fuentes: list[FuenteDetectada] = []
    token_year = f"\\{year}\\"
    token_month = f"\\{month:02d}."

    if not PRODUCTIVIDAD_ROOT_OFICIAL.exists():
        return fuentes

    for path in PRODUCTIVIDAD_ROOT_OFICIAL.rglob("Mov*.xlsx"):
        if es_archivo_respaldo(path):
            continue
        texto_path = str(path)
        if token_year not in texto_path or token_month not in texto_path:
            continue
        try:
            centro, cliente = leer_metadata_productividad(path)
        except Exception:
            centro, cliente = "", ""
        fuentes.append(FuenteDetectada(ruta=path, cliente=cliente, centro=centro))
        if verbose:
            print(f"[PROD] Fuente detectada: {path} | centro={centro} | cliente={cliente}")
    return sorted(fuentes, key=lambda item: (normalizar_texto(item.centro), normalizar_texto(item.cliente), str(item.ruta)))


def limpiar_productividad_excel(fuente: FuenteDetectada) -> tuple[pd.DataFrame, dict[str, Any]]:
    if pd is None:
        raise RuntimeError("pandas no esta disponible para leer productividad.")

    meta = pd.read_excel(fuente.ruta, sheet_name=0, header=None, nrows=8, engine="openpyxl")
    centro = normalizar_texto(meta.iloc[0, 0] if len(meta.index) >= 1 else fuente.centro)
    cliente = normalizar_texto(meta.iloc[3, 0] if len(meta.index) >= 4 else fuente.cliente)
    df = pd.read_excel(fuente.ruta, sheet_name=0, header=8, engine="openpyxl")
    df.columns = [quitar_acentos(normalizar_texto(col)) if col is not None else "" for col in df.columns]
    if "" in df.columns:
        df = df.drop(columns=[""], errors="ignore")
    if "Articulo.1" in df.columns and "SKU" not in df.columns:
        df = df.rename(columns={"Articulo.1": "SKU"})
    if "Articulo" in df.columns and "Descripcion" not in df.columns:
        pass
    for columna in COLUMNAS_PRODUCTIVIDAD_BASE:
        if columna not in df.columns:
            df[columna] = None
    df = df[COLUMNAS_PRODUCTIVIDAD_BASE].copy()
    df = df.dropna(how="all")
    df["Centro"] = centro or normalizar_texto(fuente.centro)
    df["Cliente"] = cliente or normalizar_texto(fuente.cliente)
    df["Fecha"] = df["Fecha"].map(parse_fecha)
    df["Hora"] = df["Hora"].map(parse_fecha_hora_o_hora)
    df["Salida"] = df["Salida"].map(parse_numero)
    df["Entrada"] = df["Entrada"].map(parse_numero)
    df["Nro. de Doc. Externo"] = df["Nro. de Doc. Externo"].map(normalizar_numero_pedido)
    df["Comprobante externo"] = df["Comprobante externo"].map(normalizar_clave)
    df["Destino"] = df["Destino"].map(normalizar_texto)
    df["Registro"] = df["Registro"].map(normalizar_texto)
    df["Ubicacion"] = df["Ubicacion"].map(normalizar_texto)
    df["Tipo de operacion"] = df["Tipo de operacion"].map(normalizar_texto)
    df["Trabajo"] = df["Trabajo"].map(normalizar_texto)
    df["timestamp_operacion"] = df.apply(construir_timestamp_operacion, axis=1)
    df = df[df["timestamp_operacion"].notna()].copy()
    df = df[df["Salida"] > 0].copy()
    df["__archivo_fuente"] = str(fuente.ruta)
    metadata = {
        "path": str(fuente.ruta),
        "name": fuente.ruta.name,
        "modified": iso_mtime(fuente.ruta),
        "rows": int(len(df)),
        "cliente_detectado": cliente,
        "cd_detectado": centro,
    }
    return df, metadata


def parse_fecha_hora_o_hora(valor: Any) -> time | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, time):
        return valor
    if isinstance(valor, str):
        texto = normalizar_texto(valor)
        for formato in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(texto, formato).time()
            except ValueError:
                continue
    dt = parse_fecha(valor)
    return dt.time() if dt else None


def construir_timestamp_operacion(row: pd.Series) -> datetime | None:
    fecha = row.get("Fecha")
    hora = row.get("Hora")
    if not es_fecha_valida(fecha):
        return None
    if pd is not None and isinstance(fecha, pd.Timestamp):
        fecha = fecha.to_pydatetime()
    if not isinstance(fecha, datetime):
        return None
    if isinstance(hora, time):
        return datetime.combine(fecha.date(), hora)
    return datetime.combine(fecha.date(), time.min)


def calcular_fecha_turno(timestamp_operacion: datetime | None) -> date | None:
    if not es_fecha_valida(timestamp_operacion):
        return None
    if pd is not None and isinstance(timestamp_operacion, pd.Timestamp):
        timestamp_operacion = timestamp_operacion.to_pydatetime()
    if not isinstance(timestamp_operacion, datetime):
        return None
    if timestamp_operacion.time() <= time(6, 0, 0):
        return (timestamp_operacion.date() - timedelta(days=1))
    return timestamp_operacion.date()


def es_fin_de_semana(fecha_turno: date) -> bool:
    return fecha_turno.weekday() >= 5


def calcular_turno(timestamp_operacion: datetime | None) -> str:
    if not es_fecha_valida(timestamp_operacion):
        return "SIN_TURNO"
    if pd is not None and isinstance(timestamp_operacion, pd.Timestamp):
        timestamp_operacion = timestamp_operacion.to_pydatetime()
    if not isinstance(timestamp_operacion, datetime):
        return "SIN_TURNO"
    fecha_turno = calcular_fecha_turno(timestamp_operacion)
    if fecha_turno is None:
        return "SIN_TURNO"
    if es_fin_de_semana(fecha_turno):
        return "H.E"

    hora = timestamp_operacion.time()
    weekday = fecha_turno.weekday()
    if weekday <= 3:  # lunes a jueves
        if time(8, 0) <= hora < time(18, 0):
            return "a.m"
        if hora >= time(20, 0) or hora <= time(6, 0):
            return "p.m"
        return "H.E"
    if weekday == 4:  # viernes
        if time(8, 0) <= hora < time(17, 0):
            return "a.m"
        if hora >= time(21, 0) or hora <= time(6, 0):
            return "p.m"
        return "H.E"
    return "H.E"


def calcular_hora_operativa(timestamp_operacion: datetime | None) -> float | None:
    if not es_fecha_valida(timestamp_operacion):
        return None
    if pd is not None and isinstance(timestamp_operacion, pd.Timestamp):
        timestamp_operacion = timestamp_operacion.to_pydatetime()
    if not isinstance(timestamp_operacion, datetime):
        return None
    valor = timestamp_operacion.hour + (timestamp_operacion.minute / 60.0) + (timestamp_operacion.second / 3600.0)
    if valor < 6:
        return valor + 24
    return valor


def calcular_canal_derco(comprobante_externo: Any, destino: Any) -> tuple[str, str]:
    comprobante = normalizar_clave(comprobante_externo)
    destino_txt = normalizar_mayusculas(destino)
    clave = f"{comprobante[:2]}{destino_txt[:4]}"

    if clave == "46AP00":
        return "AP", "AP"
    if clave in {"31SODI", "31WALM", "31EASY", "31REND", "31HIPE"}:
        return "GT", "GT"
    if clave == "55LO B":
        return "LB", "LB"
    if clave == "46SG00":
        return "SG", "SG"
    if clave in {"91SG00", "91AP00", "91CORO"}:
        return "CAP", "CAP"
    return "MY", "MY"


def normalizar_ubicacion_lookup(valor: Any) -> str:
    texto = normalizar_etiqueta(valor)
    if not texto:
        return ""
    partes = texto.split("-")
    if len(partes) == 4 and partes[-1].isdigit():
        return "-".join(partes[:-1])
    if len(partes) >= 5 and partes[0] in {"P", "M"} and partes[1] == "EST":
        return "-".join(partes[:3])
    return texto


def normalizar_locacion_dim(valor: Any) -> str:
    texto = normalizar_etiqueta(valor)
    if texto == "RACK":
        return "RACK"
    if texto == "ESTANTERIA":
        return "ESTANTERIA"
    if texto == "PISO":
        return "PISO"
    if texto == "SISTEMICA":
        return "SISTEMICA"
    return texto or "SIN_DIM"


def tipo_ubicacion_corregida(locacion: Any) -> str:
    loc = normalizar_locacion_dim(locacion)
    if loc in {"PISO", "SISTEMICA"}:
        return "RACK"
    if loc:
        return loc
    return "SIN_DIM"


def cliente_meta_productividad(cliente: Any, centro: Any) -> str:
    cliente_txt = normalizar_etiqueta(cliente)
    centro_txt = normalizar_etiqueta(centro)
    if cliente_txt == "MASCOTAS LATINAS" and centro_txt == "PUDAHUEL":
        return "MASCOTAS LATINAS_P"
    return cliente_txt


def ajustar_canal_detalle_derco(row: pd.Series) -> tuple[str, str, str, str, str]:
    cliente = normalizar_mayusculas(row.get("Cliente"))
    tipo_dim = normalizar_locacion_dim(row.get("Tipo_Ubicacion_Dim"))
    tipo_corregida = tipo_ubicacion_corregida(tipo_dim)
    if cliente not in {"DERCO", "GRUPO PLANET"}:
        return "UNICO", "UNICO", tipo_dim or "SIN_DIM", tipo_corregida, "no_aplica"

    canal_principal, _ = calcular_canal_derco(row.get("Comprobante externo"), row.get("Destino"))
    tipo_meta = tipo_corregida if tipo_corregida != "SIN_DIM" else "SIN_DIM"

    # Fase 2 CES: MY con destino en Base CES se reclasifica como CES (mismo criterio
    # que canal_derco_auto.py para FillRate). La columna es_ces_destino la inyecta
    # _aplicar_es_ces_destino() antes del apply.
    es_ces = bool(row.get("es_ces_destino", False))
    if canal_principal == "MY" and es_ces:
        return "CES", "CES", tipo_meta, tipo_corregida, "ces_destino"

    if canal_principal == "AP":
        if tipo_corregida == "RACK":
            canal_detalle = "AP Rack"
        elif tipo_corregida == "ESTANTERIA":
            canal_detalle = "AP Estanteria"
        else:
            canal_detalle = "AP Sin clasificar"
    elif canal_principal in {"CAP", "MY", "SG"}:
        canal_detalle = canal_principal
    else:
        canal_detalle = canal_principal
    # Fase 2: la clasificacion ahora viene de canal_derco_utils.clasificar_ubicacion_dim
    # (regla de prefijos), no de la tabla DimUbicaciones.
    metodo = "regla_prefijos" if tipo_dim != "SIN_DIM" else "no_disponible"
    return canal_principal, canal_detalle, tipo_meta, tipo_corregida, metodo


# ----------------------------------------------------------------------------
# Fase 2 CES: detector de destinos CES en MovDerco (productividad)
# ----------------------------------------------------------------------------

# Cache modulo: Base CES se carga una sola vez por proceso.
_CES_MATCHER: Any = None  # callable(destino) -> nombre_ces | None
_CES_CARGADA = False
_CES_DISPONIBLE = False


def _cargar_ces_lazy() -> bool:
    """Carga Base CES una sola vez. Retorna True si esta disponible."""
    global _CES_MATCHER, _CES_CARGADA, _CES_DISPONIBLE
    if _CES_CARGADA:
        return _CES_DISPONIBLE
    _CES_CARGADA = True
    if not BASE_CES_PATH.exists():
        print(f"[WARN] Base CES no encontrada en {BASE_CES_PATH}; CES no se clasificara en productividad.")
        return False
    try:
        _ces_set, _CES_MATCHER = cargar_base_ces(str(BASE_CES_PATH))
        _CES_DISPONIBLE = True
        return True
    except Exception as exc:  # noqa: BLE001
        print(f"[WARN] No se pudo cargar Base CES ({exc}); CES no se clasificara en productividad.")
        return False


def _aplicar_es_ces_destino(df: pd.DataFrame) -> pd.DataFrame:
    """Inyecta columna booleana 'es_ces_destino' en el df.

    Solo evalua filas de DERCO/GRUPO PLANET (las unicas que aplican canal CES).
    Si Base CES no esta disponible, todas quedan en False (no se reclasifica nada,
    comportamiento equivalente al previo).
    """
    if "es_ces_destino" in df.columns:
        return df
    disponible = _cargar_ces_lazy()
    if not disponible or "Destino" not in df.columns:
        df["es_ces_destino"] = False
        return df
    # Solo evaluar filas DERCO/GRUPO PLANET para no gastar CPU en otros clientes.
    mask_derco = df["Cliente"].map(normalizar_mayusculas).isin({"DERCO", "GRUPO PLANET"})
    if not mask_derco.any():
        df["es_ces_destino"] = False
        return df
    destinos_unicos = df.loc[mask_derco, "Destino"].dropna().astype(str).unique()
    mapa_ces = {d: (_CES_MATCHER(d) is not None) for d in destinos_unicos}
    es_ces = df["Destino"].astype(str).map(mapa_ces).fillna(False).astype(bool)
    df["es_ces_destino"] = es_ces & mask_derco
    return df


def campos_pedidos_unicos(pedidos: int, campo_unicos: str) -> dict[str, int]:
    pedidos = int(pedidos)
    return {
        "pedidos": pedidos,
        campo_unicos: pedidos,
    }


def agregar_metricas(
    df: pd.DataFrame,
    campos_groupby: list[str],
    total_lineas: int,
    total_unidades: float,
    campo_pedidos_unicos: str | None = None,
) -> list[dict[str, Any]]:
    resultado = []
    for claves, grupo in df.groupby(campos_groupby):
        if not isinstance(claves, tuple):
            claves = (claves,)
        registro = {campo: normalizar_texto(valor) for campo, valor in zip(campos_groupby, claves)}
        lineas = int(len(grupo))
        unidades = round_safe(grupo["Salida"].sum(), 2)
        pedidos = int(grupo["Nro. de Doc. Externo"].replace("", pd.NA).dropna().nunique())
        registro.update({"lineas": lineas, "unidades": unidades})
        if campo_pedidos_unicos:
            registro.update(campos_pedidos_unicos(pedidos, campo_pedidos_unicos))
        else:
            registro["pedidos"] = pedidos
        registro.update(
            {
                "participacion_lineas_pct": porcentaje_safe(lineas, total_lineas),
                "participacion_unidades_pct": porcentaje_safe(unidades, total_unidades),
            }
        )
        resultado.append(registro)
    return sorted(resultado, key=lambda item: (-item["lineas"], -item["unidades"], str(item)))


def contar_pedidos_validos(serie: pd.Series) -> int:
    if pd is not None:
        return int(serie.replace("", pd.NA).dropna().nunique())
    valores = {normalizar_texto(v) for v in serie.tolist() if normalizar_texto(v)}
    return int(len(valores))


def serializar_fecha_productividad(valor: Any) -> str:
    if isinstance(valor, datetime):
        valor = valor.date()
    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")
    return normalizar_texto(valor)


def encontrar_archivo_dimensiones() -> Path | None:
    if not DIMENSIONES_ROOT.exists():
        return None
    for path in DIMENSIONES_ROOT.rglob("*.xlsx"):
        if path.name.lower() == DIMENSIONES_FILENAME.lower():
            return path
    return None


def parse_turno_meta(valor: Any) -> str:
    texto = normalizar_texto(valor).lower()
    if texto.endswith("a.m"):
        return "a.m"
    if texto.endswith("p.m"):
        return "p.m"
    return ""


def parse_canal_tipo_meta(valor: Any) -> tuple[str, str]:
    texto = normalizar_etiqueta(valor)
    if not texto:
        return "", ""
    turno = parse_turno_meta(valor)
    if turno:
        texto = texto[: -len(turno)].strip()
    tokens = texto.split()
    tipo = ""
    if tokens and tokens[-1] in {"RACK", "ESTANTERIA", "PISO", "SISTEMICA"}:
        tipo = tokens[-1]
        tokens = tokens[:-1]
    canal = " ".join(tokens).strip()
    if canal == "UNICO":
        canal = "UNICO"
    return canal, tipo


def cargar_dimensiones(
    verbose: bool = False,
) -> tuple[
    dict[str, Any],
    dict[str, Any],
    dict[tuple[str, str, str, str], float],
    dict[str, dict[str, str]],
]:
    archivo = encontrar_archivo_dimensiones()
    dimensiones_info = {
        "disponible": False,
        "archivo": None,
        "hojas_detectadas": [],
        "dimubicaciones": {
            "disponible": False,
        },
    }
    metas_info = {
        "disponible": False,
        "archivo": None,
        "filas": 0,
        "hoja": None,
        "clientes": [],
    }
    metas_map: dict[tuple[str, str, str, str], float] = {}

    if not archivo:
        return dimensiones_info, metas_info, metas_map, {}

    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        dimensiones_info["disponible"] = True
        dimensiones_info["archivo"] = str(archivo)
        dimensiones_info["hojas_detectadas"] = list(wb.sheetnames)
        if verbose:
            print(f"[DIM] Archivo detectado: {archivo}")

        ubicaciones_map: dict[str, dict[str, str]] = {}
        hoja_ubic = next((name for name in wb.sheetnames if normalizar_etiqueta(name) in {"UBICACIONES CDS", "DIMUBICACIONES", "UBICACIONES"}), None)
        if hoja_ubic:
            ws_ubic = wb[hoja_ubic]
            filas_ubic = list(ws_ubic.iter_rows(values_only=True))
            if filas_ubic:
                headers = [quitar_acentos(normalizar_texto(v)) for v in filas_ubic[0]]
                idx_ubic = {h: i for i, h in enumerate(headers)}
                for row in filas_ubic[1:]:
                    ubic = normalizar_ubicacion_lookup(row[idx_ubic.get("Ubicacion", -1)] if "Ubicacion" in idx_ubic else None)
                    if not ubic:
                        continue
                    locacion = row[idx_ubic.get("Locacion", -1)] if "Locacion" in idx_ubic else None
                    cd = row[idx_ubic.get("CD", -1)] if "CD" in idx_ubic else None
                    ubicaciones_map[ubic] = {
                        "locacion": normalizar_locacion_dim(locacion),
                        "cd": normalizar_etiqueta(cd),
                    }
            dimensiones_info["dimubicaciones"] = {
                "disponible": bool(ubicaciones_map),
                "hoja": hoja_ubic,
                "filas": len(ubicaciones_map),
            }

        hoja_meta = None
        for candidata in ("Meta actual", "Meta actual (2)", "Meta antigua"):
            if candidata in wb.sheetnames:
                hoja_meta = candidata
                break

        if hoja_meta:
            ws_meta = wb[hoja_meta]
            filas_meta = list(ws_meta.iter_rows(values_only=True))
            if filas_meta:
                headers = [quitar_acentos(normalizar_texto(v)) for v in filas_meta[0]]
                idx_meta = {h: i for i, h in enumerate(headers)}
                clientes_con_meta = set()
                filas_validas = 0
                for row in filas_meta[1:]:
                    cliente = normalizar_etiqueta(row[idx_meta.get("Cliente", idx_meta.get("Empresa", -1))] if ("Cliente" in idx_meta or "Empresa" in idx_meta) else None)
                    canal_turno = row[idx_meta.get("Canal x turno", -1)] if "Canal x turno" in idx_meta else None
                    meta_raw = row[idx_meta.get("Lineas x canal x turno", idx_meta.get("Líneas x canal x turno", -1))] if ("Lineas x canal x turno" in idx_meta or "Líneas x canal x turno" in idx_meta) else None
                    turno = parse_turno_meta(canal_turno)
                    canal, tipo = parse_canal_tipo_meta(canal_turno)
                    meta = parse_numero(meta_raw)
                    if not cliente or not canal or not tipo or not turno:
                        continue
                    metas_map[(cliente, canal, tipo, turno)] = meta
                    clientes_con_meta.add(cliente)
                    filas_validas += 1
                metas_info = {
                    "disponible": bool(metas_map),
                    "archivo": str(archivo),
                    "filas": filas_validas,
                    "hoja": hoja_meta,
                    "clientes": sorted(clientes_con_meta),
                }

        wb.close()
        return dimensiones_info, metas_info, metas_map, ubicaciones_map
    except Exception as exc:
        dimensiones_info["mensaje"] = f"No se pudo leer archivo de dimensiones: {exc}"
        metas_info["archivo"] = str(archivo)
        metas_info["mensaje"] = f"No se pudo leer metas: {exc}"
        return dimensiones_info, metas_info, metas_map, {}


def calcular_productividad(
    fuentes: list[FuenteDetectada],
    year: int,
    month: int,
    metas_map: dict[tuple[str, str, str, str], float],
    metas_info: dict[str, Any],
    ubicaciones_map: dict[str, dict[str, str]],
) -> tuple[dict[str, Any], dict[str, Any], list[str], list[str]]:
    if pd is None:
        return {
            "disponible": False,
            "mensaje": "pandas no esta disponible en este entorno.",
        }, {
            "disponible": False,
            "ruta_oficial": str(PRODUCTIVIDAD_ROOT_OFICIAL),
            "archivos_leidos": 0,
            "archivos": [],
            "mensaje": "pandas no esta disponible en este entorno.",
        }, [], []

    if not PRODUCTIVIDAD_ROOT_OFICIAL.exists():
        mensaje = "No se encontro la carpeta oficial de Productividad sincronizada localmente."
        return {
            "disponible": False,
            "mensaje": mensaje,
        }, {
            "disponible": False,
            "ruta_oficial": str(PRODUCTIVIDAD_ROOT_OFICIAL),
            "archivos_leidos": 0,
            "archivos": [],
            "mensaje": mensaje,
        }, [], []

    frames = []
    errores = []
    archivos_leidos = []
    for fuente in fuentes:
        try:
            df_fuente, info_fuente = limpiar_productividad_excel(fuente)
            if not df_fuente.empty:
                frames.append(df_fuente)
            archivos_leidos.append(info_fuente)
        except Exception as exc:
            errores.append(f"{fuente.ruta.name}: {exc}")

    if not frames:
        return {
            "disponible": False,
            "mensaje": "No se encontro una fuente local de productividad para el periodo solicitado.",
        }, {
            "disponible": False,
            "ruta_oficial": str(PRODUCTIVIDAD_ROOT_OFICIAL),
            "archivos_leidos": len(archivos_leidos),
            "archivos": archivos_leidos,
            "errores": errores[:5],
        }, [], []

    df = pd.concat(frames, ignore_index=True)
    df["Ubicacion_Clave"] = df["Ubicacion"].map(normalizar_ubicacion_lookup)
    # Fase 2: clasificacion Rack/Est unificada con canal_derco_auto.py via canal_derco_utils.
    # Antes: lookup contra Tabla Ubicaciones CDs.xlsx (DimUbicaciones) -> SIN_DIM si no estaba.
    # Ahora: regla de prefijos sobre Ubicacion (mismo criterio que data Derco).
    df["Tipo_Ubicacion_Dim"] = df["Ubicacion"].map(clasificar_ubicacion_dim)
    df["Cliente_Meta"] = df.apply(lambda row: cliente_meta_productividad(row.get("Cliente"), row.get("Centro")), axis=1)
    df["Fecha_Turno"] = df["timestamp_operacion"].map(calcular_fecha_turno)
    df["Turno"] = df["timestamp_operacion"].map(calcular_turno)
    df["Hora_Operativa"] = df["timestamp_operacion"].map(calcular_hora_operativa)

    # Fase 2 CES: idem productividad historica — clasificar destinos CES antes del canal.
    df = _aplicar_es_ces_destino(df)
    canales = df.apply(ajustar_canal_detalle_derco, axis=1, result_type="expand")
    canales.columns = [
        "Canal_Principal",
        "Canal_Detalle",
        "Tipo_Ubicacion_Meta",
        "Tipo_Ubicacion_Corregida",
        "Canal_Detalle_Metodo",
    ]
    df = pd.concat([df, canales], axis=1)
    # Canal mayorista = CAP + MY + SG + CES (CES son concesionarios, parte del mayorista
    # según regla de negocio confirmada por usuario 2026-05-15). En canales_originales CES
    # sigue apareciendo separado.
    df["Canal_Agrupado"] = df["Canal_Principal"].map(lambda c: "CAP-MY-SG-CES" if c in {"CAP", "MY", "SG", "CES"} else c)

    total_lineas = int(len(df))
    total_unidades = float(df["Salida"].sum())
    pedidos_total = int(df["Nro. de Doc. Externo"].replace("", pd.NA).dropna().nunique())
    dias_trabajados = int(df["Fecha_Turno"].dropna().nunique())
    fechas_turno_unicas = [f for f in df["Fecha_Turno"].dropna().tolist() if isinstance(f, date)]
    dias_laborables = len({f for f in fechas_turno_unicas if f.weekday() < 5})

    horas_grupo = (
        df.groupby(["Fecha_Turno", "Registro", "Turno"])["timestamp_operacion"]
        .agg(["min", "max"])
        .reset_index()
    )
    horas_total = 0.0
    for _, row in horas_grupo.iterrows():
        delta_horas = (row["max"] - row["min"]).total_seconds() / 3600.0
        if delta_horas > 1:
            delta_horas -= 1
        horas_total += max(delta_horas, 0.0)

    por_cd = agregar_metricas(df, ["Centro"], total_lineas, total_unidades)
    por_cliente = agregar_metricas(df, ["Cliente"], total_lineas, total_unidades, campo_pedidos_unicos="pedidos_unicos_periodo")
    for item in por_cliente:
        item["cliente"] = item.pop("Cliente", "")
    por_cliente_canal = agregar_metricas(df, ["Cliente", "Canal_Agrupado"], total_lineas, total_unidades)

    por_turno = []
    for turno, grupo in df.groupby("Turno"):
        por_turno.append(
            {
                "turno": turno,
                "lineas": int(len(grupo)),
                "unidades": round_safe(grupo["Salida"].sum(), 2),
                "pedidos": int(grupo["Nro. de Doc. Externo"].replace("", pd.NA).dropna().nunique()),
                "operadores_unicos": int(grupo["Registro"].replace("", pd.NA).dropna().nunique()),
            }
        )
    por_turno.sort(key=lambda item: (-item["lineas"], item["turno"]))

    derco_df = df[df["Cliente"].map(normalizar_mayusculas).isin({"DERCO", "GRUPO PLANET"})].copy()
    if not derco_df.empty:
        derco_df["Fecha_Turno_Texto"] = derco_df["Fecha_Turno"].map(serializar_fecha_productividad)
    derco_payload = {"disponible": not derco_df.empty, "ap_total": {}, "ap_detalle": [], "canales": [], "canales_originales": []}
    if not derco_df.empty:
        ap_total = derco_df[derco_df["Canal_Principal"] == "AP"]
        if not ap_total.empty:
            derco_payload["ap_total"] = {
                "lineas": int(len(ap_total)),
                "unidades": round_safe(ap_total["Salida"].sum(), 2),
                "pedidos": int(ap_total["Nro. de Doc. Externo"].replace("", pd.NA).dropna().nunique()),
            }
        ap_clasificadas = ap_total[ap_total["Canal_Detalle"].isin({"AP Rack", "AP Estanteria"})]
        for detalle, grupo in ap_clasificadas.groupby("Canal_Detalle"):
            derco_payload["ap_detalle"].append(
                {
                    "canal_detalle": detalle,
                    "lineas": int(len(grupo)),
                    "unidades": round_safe(grupo["Salida"].sum(), 2),
                    "pedidos": int(grupo["Nro. de Doc. Externo"].replace("", pd.NA).dropna().nunique()),
                }
            )
        ap_no_clasificadas = ap_total[~ap_total["Canal_Detalle"].isin({"AP Rack", "AP Estanteria"})]
        if not ap_no_clasificadas.empty:
            derco_payload["ap_no_clasificadas"] = {
                "lineas": int(len(ap_no_clasificadas)),
                "unidades": round_safe(ap_no_clasificadas["Salida"].sum(), 2),
                "pedidos": int(ap_no_clasificadas["Nro. de Doc. Externo"].replace("", pd.NA).dropna().nunique()),
            }
        for canal, grupo in derco_df.groupby("Canal_Agrupado"):
            derco_payload["canales"].append(
                {
                    "canal": canal,
                    "lineas": int(len(grupo)),
                    "unidades": round_safe(grupo["Salida"].sum(), 2),
                    "pedidos": int(grupo["Nro. de Doc. Externo"].replace("", pd.NA).dropna().nunique()),
                }
            )
        for canal, grupo in derco_df.groupby("Canal_Principal"):
            derco_payload["canales_originales"].append(
                {
                    "canal": canal,
                    "lineas": int(len(grupo)),
                    "unidades": round_safe(grupo["Salida"].sum(), 2),
                    "pedidos": int(grupo["Nro. de Doc. Externo"].replace("", pd.NA).dropna().nunique()),
                }
            )
        derco_payload["ap_detalle"].sort(key=lambda item: (-item["lineas"], item["canal_detalle"]))
        derco_payload["canales"].sort(key=lambda item: (-item["lineas"], item["canal"]))
        derco_payload["canales_originales"].sort(key=lambda item: (-item["lineas"], item["canal"]))
        if derco_payload["canales"]:
            top = derco_payload["canales"][0]
            derco_payload["top_canal_por_lineas"] = top["canal"]
            derco_payload["top_canal_por_unidades"] = max(derco_payload["canales"], key=lambda item: item["unidades"])["canal"]
        # Fase 2: detalle Rack/Est siempre viene de canal_derco_utils (regla de prefijos),
        # no depende de DimUbicaciones. Si Tipo_Ubicacion_Dim es SIN_DIM aqui significa
        # ubicacion vacia/nula, no falta de la tabla.
        match_ap = int(ap_total["Tipo_Ubicacion_Dim"].ne("SIN_DIM").sum()) if not ap_total.empty else 0
        if not ap_total.empty and match_ap == len(ap_total):
            derco_payload["ap_detalle_metodo"] = "regla_prefijos"
        elif not ap_total.empty and match_ap > 0:
            derco_payload["ap_detalle_metodo"] = "regla_prefijos_parcial"
            derco_payload["advertencia"] = "Algunas lineas AP no tienen Ubicacion registrada en MovDerco."
        else:
            derco_payload["ap_detalle_metodo"] = "no_disponible"
            derco_payload["advertencia"] = "Sin Ubicacion en MovDerco para detalle AP Rack / AP Estanteria."

    df["Fecha_Turno_Texto"] = df["Fecha_Turno"].map(serializar_fecha_productividad)
    fechas_validas = df[df["Fecha_Turno_Texto"] != ""].copy()
    fechas_validas_ordenadas = sorted({f for f in fechas_validas["Fecha_Turno_Texto"].tolist() if f})

    alertas_diarias_contadores = {
        "registros_sin_fecha_turno": int(df["Fecha_Turno_Texto"].eq("").sum()),
        "registros_sin_cliente": int(df["Cliente"].map(normalizar_texto_seguro).eq("").sum()),
        "registros_sin_cd": int(df["Centro"].map(normalizar_texto_seguro).eq("").sum()),
        "registros_sin_salida": int((df["Salida"] <= 0).sum()),
        "registros_sin_pedido": int(df["Nro. de Doc. Externo"].map(normalizar_texto_seguro).eq("").sum()),
        "registros_sin_turno": int(df["Turno"].map(normalizar_texto_seguro).isin({"", "SIN_TURNO"}).sum()),
        "registros_sin_canal": int(df["Canal_Agrupado"].map(normalizar_texto_seguro).eq("").sum()),
        "registros_derco_sin_canal": int(derco_df["Canal_Agrupado"].map(normalizar_texto_seguro).eq("").sum()) if not derco_df.empty else 0,
        "registros_derco_ap_sin_tipo_ubicacion": int(
            derco_df[
                derco_df["Canal_Principal"].eq("AP")
                & derco_df["Tipo_Ubicacion_Corregida"].map(normalizar_texto_seguro).isin({"", "SIN_DIM"})
            ].shape[0]
        ) if not derco_df.empty else 0,
        "fechas_fuera_periodo": int(
            fechas_validas[
                ~fechas_validas["Fecha_Turno"].map(lambda f: isinstance(f, date) and f.year == year and f.month == month)
            ].shape[0]
        ),
    }

    por_fecha = []
    for fecha, grupo in fechas_validas.groupby("Fecha_Turno_Texto", dropna=False):
        por_fecha.append(
            {
                "fecha": fecha,
                "lineas": int(len(grupo)),
                "unidades": round_safe(grupo["Salida"].sum(), 2),
                **campos_pedidos_unicos(
                    contar_pedidos_validos(grupo["Nro. de Doc. Externo"]),
                    "pedidos_unicos_fecha",
                ),
                "clientes": int(grupo["Cliente"].replace("", pd.NA).dropna().nunique()),
                "cds": int(grupo["Centro"].replace("", pd.NA).dropna().nunique()),
            }
        )
    por_fecha.sort(key=lambda item: item["fecha"])

    por_fecha_cliente = []
    for (fecha, cliente, cd), grupo in fechas_validas.groupby(["Fecha_Turno_Texto", "Cliente", "Centro"], dropna=False):
        por_fecha_cliente.append(
            {
                "fecha": fecha,
                "cliente": normalizar_texto(cliente),
                "cd": normalizar_texto(cd),
                "lineas": int(len(grupo)),
                "unidades": round_safe(grupo["Salida"].sum(), 2),
                **campos_pedidos_unicos(
                    contar_pedidos_validos(grupo["Nro. de Doc. Externo"]),
                    "pedidos_unicos_fecha",
                ),
            }
        )
    por_fecha_cliente.sort(key=lambda item: (item["fecha"], item["cliente"], item["cd"]))

    por_fecha_cliente_canal = []
    for (fecha, cliente, cd, canal, tipo_ubicacion, canal_detalle), grupo in fechas_validas.groupby(
        ["Fecha_Turno_Texto", "Cliente", "Centro", "Canal_Agrupado", "Tipo_Ubicacion_Corregida", "Canal_Detalle"],
        dropna=False,
    ):
        por_fecha_cliente_canal.append(
            {
                "fecha": fecha,
                "cliente": normalizar_texto(cliente),
                "cd": normalizar_texto(cd),
                "canal": normalizar_texto(canal),
                "tipo_ubicacion": normalizar_texto(tipo_ubicacion),
                "canal_detalle": normalizar_texto(canal_detalle),
                "lineas": int(len(grupo)),
                "unidades": round_safe(grupo["Salida"].sum(), 2),
                "pedidos": contar_pedidos_validos(grupo["Nro. de Doc. Externo"]),
            }
        )
    por_fecha_cliente_canal.sort(
        key=lambda item: (item["fecha"], item["cliente"], item["cd"], item["canal"], item["tipo_ubicacion"], item["canal_detalle"])
    )

    por_fecha_cliente_turno = []
    for (fecha, cliente, cd, turno), grupo in fechas_validas.groupby(
        ["Fecha_Turno_Texto", "Cliente", "Centro", "Turno"],
        dropna=False,
    ):
        por_fecha_cliente_turno.append(
            {
                "fecha": fecha,
                "cliente": normalizar_texto(cliente),
                "cd": normalizar_texto(cd),
                "turno": normalizar_texto(turno),
                "lineas": int(len(grupo)),
                "unidades": round_safe(grupo["Salida"].sum(), 2),
                "pedidos": contar_pedidos_validos(grupo["Nro. de Doc. Externo"]),
            }
        )
    por_fecha_cliente_turno.sort(key=lambda item: (item["fecha"], item["cliente"], item["cd"], item["turno"]))

    derco_por_fecha = []
    derco_ap_por_fecha = []
    derco_canal_por_fecha = []
    if not derco_df.empty:
        derco_fechas = derco_df[derco_df["Fecha_Turno_Texto"] != ""].copy()
        for fecha, grupo in derco_fechas.groupby("Fecha_Turno_Texto", dropna=False):
            derco_por_fecha.append(
                {
                    "fecha": fecha,
                    "lineas": int(len(grupo)),
                    "unidades": round_safe(grupo["Salida"].sum(), 2),
                    **campos_pedidos_unicos(
                        contar_pedidos_validos(grupo["Nro. de Doc. Externo"]),
                        "pedidos_unicos_fecha",
                    ),
                }
            )
            grupo_ap = grupo[grupo["Canal_Principal"] == "AP"]
            grupo_ap_rack = grupo_ap[grupo_ap["Canal_Detalle"] == "AP Rack"]
            grupo_ap_est = grupo_ap[grupo_ap["Canal_Detalle"] == "AP Estanteria"]
            derco_ap_por_fecha.append(
                {
                    "fecha": fecha,
                    "ap_total": {
                        "lineas": int(len(grupo_ap)),
                        "unidades": round_safe(grupo_ap["Salida"].sum(), 2),
                        **campos_pedidos_unicos(
                            contar_pedidos_validos(grupo_ap["Nro. de Doc. Externo"]),
                            "pedidos_unicos_fecha",
                        ),
                    },
                    "ap_rack": {
                        "lineas": int(len(grupo_ap_rack)),
                        "unidades": round_safe(grupo_ap_rack["Salida"].sum(), 2),
                        **campos_pedidos_unicos(
                            contar_pedidos_validos(grupo_ap_rack["Nro. de Doc. Externo"]),
                            "pedidos_unicos_fecha",
                        ),
                    },
                    "ap_estanteria": {
                        "lineas": int(len(grupo_ap_est)),
                        "unidades": round_safe(grupo_ap_est["Salida"].sum(), 2),
                        **campos_pedidos_unicos(
                            contar_pedidos_validos(grupo_ap_est["Nro. de Doc. Externo"]),
                            "pedidos_unicos_fecha",
                        ),
                    },
                }
            )
        for (fecha, canal), grupo in derco_fechas.groupby(["Fecha_Turno_Texto", "Canal_Agrupado"], dropna=False):
            derco_canal_por_fecha.append(
                {
                    "fecha": fecha,
                    "canal": normalizar_texto(canal),
                    "lineas": int(len(grupo)),
                    "unidades": round_safe(grupo["Salida"].sum(), 2),
                    "pedidos": contar_pedidos_validos(grupo["Nro. de Doc. Externo"]),
                }
            )
        derco_por_fecha.sort(key=lambda item: item["fecha"])
        derco_ap_por_fecha.sort(key=lambda item: item["fecha"])
        derco_canal_por_fecha.sort(key=lambda item: (item["fecha"], item["canal"]))

    sum_diario_lineas = int(sum(item["lineas"] for item in por_fecha))
    sum_diario_unidades = round_safe(sum(item["unidades"] for item in por_fecha), 2)
    sum_fecha_cliente_lineas = int(sum(item["lineas"] for item in por_fecha_cliente))
    sum_fecha_cliente_unidades = round_safe(sum(item["unidades"] for item in por_fecha_cliente), 2)
    sum_derco_fecha_lineas = int(sum(item["lineas"] for item in derco_por_fecha))
    sum_derco_ap_fecha_lineas = int(sum(item["ap_total"]["lineas"] for item in derco_ap_por_fecha))
    totales_diarios_no_cuadran_global = 0
    if sum_diario_lineas != total_lineas:
        totales_diarios_no_cuadran_global += 1
    if sum_diario_unidades != round_safe(total_unidades, 2):
        totales_diarios_no_cuadran_global += 1
    if sum_fecha_cliente_lineas != total_lineas:
        totales_diarios_no_cuadran_global += 1
    if sum_fecha_cliente_unidades != round_safe(total_unidades, 2):
        totales_diarios_no_cuadran_global += 1
    alertas_diarias_contadores["totales_diarios_no_cuadran_global"] = totales_diarios_no_cuadran_global

    diario_payload = {
        "disponible": True,
        "periodo": {"anio": year, "mes": month},
        "por_fecha": por_fecha,
        "fecha_min": fechas_validas_ordenadas[0] if fechas_validas_ordenadas else None,
        "fecha_max": fechas_validas_ordenadas[-1] if fechas_validas_ordenadas else None,
        "alertas": resumir_alertas_contadores(alertas_diarias_contadores),
    }

    metas_detalle = []
    meta_total = 0.0
    lineas_con_meta = 0
    meta_total_derco = 0.0
    lineas_con_meta_derco = 0
    if metas_map:
        agrupado_meta = (
            df.groupby(["Cliente_Meta", "Canal_Agrupado", "Tipo_Ubicacion_Meta", "Turno"])
            .size()
            .reset_index(name="lineas")
        )
        for _, row in agrupado_meta.iterrows():
            cliente = normalizar_etiqueta(row["Cliente_Meta"])
            canal = normalizar_etiqueta(row["Canal_Agrupado"])
            tipo_ubicacion = normalizar_etiqueta(row["Tipo_Ubicacion_Meta"])
            turno = normalizar_texto(row["Turno"])
            meta = metas_map.get((cliente, canal, tipo_ubicacion, turno), None)
            if meta is not None:
                meta_periodo = meta * max(dias_laborables, 1)
                meta_total += meta_periodo
                lineas_con_meta += int(row["lineas"])
                if cliente == "DERCO":
                    meta_total_derco += meta_periodo
                    lineas_con_meta_derco += int(row["lineas"])
                metas_detalle.append(
                    {
                        "cliente": cliente,
                        "canal": canal,
                        "tipo_ubicacion": tipo_ubicacion,
                        "turno": turno,
                        "lineas": int(row["lineas"]),
                        "meta_lineas": round_safe(meta, 2),
                        "meta_periodo": round_safe(meta_periodo, 2),
                        "cumplimiento_pct": porcentaje_safe(row["lineas"], meta_periodo),
                    }
                )

    alertas = []
    recomendaciones = []
    if horas_total == 0:
        recomendaciones.append("No se pudo estimar horas trabajadas para productividad.")
    if derco_payload["disponible"] and derco_payload.get("canales"):
        canal_riesgo = max(derco_payload["canales"], key=lambda item: item["lineas"])
        alertas.append(
            f"DERCO concentra mayor carga operativa en {canal_riesgo['canal']} segun lineas del periodo."
        )
    if totales_diarios_no_cuadran_global:
        alertas.append("Los agregados diarios de productividad no cuadran con los totales globales del periodo.")
    if derco_df.empty:
        recomendaciones.append("No se detectaron registros DERCO en productividad para el periodo.")
    elif sum_derco_fecha_lineas != int(len(derco_df)):
        alertas.append("La serie diaria DERCO no cuadra con el total mensual DERCO.")
    if derco_payload.get("ap_total") and sum_derco_ap_fecha_lineas != int(derco_payload["ap_total"].get("lineas", 0)):
        alertas.append("La serie diaria AP de DERCO no cuadra con el total AP mensual.")

    clientes_mov = sorted({normalizar_etiqueta(c) for c in df["Cliente_Meta"].dropna().tolist() if normalizar_texto(c)})
    clientes_meta = sorted(set(metas_info.get("clientes", [])))
    cobertura_total = bool(clientes_mov) and set(clientes_mov).issubset(set(clientes_meta))

    productividad = {
        "disponible": True,
        "periodo": {"anio": year, "mes": month},
        "global": {
            "lineas": total_lineas,
            "unidades": round_safe(total_unidades, 2),
            "pedidos": pedidos_total,
            "dias_trabajados": dias_trabajados,
            "productividad_lineas_dia": porcentaje_lineal(total_lineas, dias_trabajados),
            "productividad_unidades_dia": porcentaje_lineal(total_unidades, dias_trabajados),
            "horas_trabajadas_total": round_safe(horas_total, 2),
            "productividad_lineas_hora": porcentaje_lineal(total_lineas, horas_total),
            "productividad_unidades_hora": porcentaje_lineal(total_unidades, horas_total),
            "horas_trabajadas_metodo": "aproximado",
        },
        "por_cd": por_cd,
        "por_cliente": por_cliente,
        "por_cliente_canal": por_cliente_canal,
        "por_turno": por_turno,
        "diario": diario_payload,
        "por_fecha_cliente": por_fecha_cliente,
        "por_fecha_cliente_canal": por_fecha_cliente_canal,
        "por_fecha_cliente_turno": por_fecha_cliente_turno,
        "derco": derco_payload,
        "metas": {
            "disponible": bool(metas_map),
            "dias_laborables": dias_laborables,
            "cumplimiento_disponible": bool(metas_detalle),
            "alcance": "global" if cobertura_total else "segun_dimmetas_disponible",
            "cumplimiento_global_disponible": cobertura_total,
            "cumplimiento_total_metas_disponibles": porcentaje_safe(lineas_con_meta, meta_total),
            "cumplimiento_derco": porcentaje_safe(lineas_con_meta_derco, meta_total_derco),
            "clientes_con_meta": clientes_meta,
            "clientes_con_movimiento_sin_meta": sorted(set(clientes_mov) - set(clientes_meta)),
            "por_cliente_canal_turno": sorted(
                metas_detalle,
                key=lambda item: (-item["lineas"], item["cliente"], item["canal"], item["turno"]),
            ),
        },
    }
    productividad["derco"]["por_fecha"] = derco_por_fecha
    productividad["derco"]["ap_por_fecha"] = derco_ap_por_fecha
    productividad["derco"]["canal_por_fecha"] = derco_canal_por_fecha
    fuente_info = {
        "disponible": True,
        "ruta_oficial": str(PRODUCTIVIDAD_ROOT_OFICIAL),
        "archivos_leidos": len([item for item in archivos_leidos if item["rows"] > 0]),
        "archivos": archivos_leidos,
        "errores": errores[:5],
    }
    return productividad, fuente_info, alertas, recomendaciones


def porcentaje_lineal(numerador: float, denominador: float) -> float:
    if not denominador:
        return 0.0
    return round_safe(numerador / denominador, 2)


def normalizar_texto_seguro(valor: Any) -> str:
    if valor is None:
        return ""
    if pd is not None:
        try:
            if pd.isna(valor):
                return ""
        except Exception:
            pass
    texto = normalizar_texto(valor)
    if normalizar_etiqueta(texto) == "NAN":
        return ""
    return texto


def parse_fecha_desde_nombre_archivo(path: Path) -> datetime | None:
    token = path.stem[:14]
    if len(token) == 14 and token.isdigit():
        try:
            return datetime.strptime(token, "%Y%m%d%H%M%S")
        except ValueError:
            return None
    return None


def serializar_o_mtime_fecha(valor: datetime | None, path: Path | None = None) -> str | None:
    if valor is not None:
        return valor.strftime("%Y-%m-%d %H:%M:%S")
    if path is not None:
        return iso_mtime(path)
    return None


def normalizar_empresa_inventario(empresa: Any, cd: Any) -> str:
    empresa_txt = normalizar_mayusculas(empresa)
    cd_txt = normalizar_etiqueta(cd)
    if empresa_txt == "CERVECERIA ABI":
        return "ABINBEV"
    if empresa_txt == "MASCOTA LATINA":
        empresa_txt = "MASCOTAS LATINAS"
    if empresa_txt == "MASCOTAS LATINAS" and cd_txt == "PUDAHUEL":
        return "MASCOTAS LATINAS_P"
    return empresa_txt


def normalizar_cd_inventario(valor: Any) -> str:
    texto = normalizar_etiqueta(valor)
    if texto == "PUDAHUEL REFRIGERADO":
        return "PUDAHUEL"
    return texto


def construir_ubicacion_desde_componentes(cara: Any, columna: Any, nivel: Any) -> str:
    cara_txt = normalizar_texto(cara)
    columna_txt = normalizar_texto(columna)
    nivel_txt = normalizar_texto(nivel)
    if columna_txt.isdigit():
        columna_txt = columna_txt.zfill(3)
    if nivel_txt.isdigit():
        nivel_txt = nivel_txt.zfill(2)
    partes = [cara_txt, columna_txt, nivel_txt]
    if not any(partes):
        return ""
    return "-".join(parte for parte in partes if parte)


def es_texto_total_resumen(valor: Any) -> bool:
    texto = normalizar_etiqueta(valor)
    if not texto:
        return False
    return any(token in texto for token in ("TOTAL", "TOTAL KILOS", "TOTAL M3", "TOTAL PALLETS"))


def limpiar_filas_stock_wms(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int]]:
    descartes = {
        "filas_descartadas_no_operativas": 0,
        "filas_descartadas_totales_excel": 0,
    }
    if df.empty:
        return df, descartes

    empresa = df["Empresa"].map(normalizar_texto_seguro)
    cd = df["CD"].map(normalizar_texto_seguro)
    articulo = df["Articulo"].map(normalizar_texto_seguro)
    contenedor = df["Contenedor"].map(normalizar_texto_seguro)
    ubicacion = df["Ubicacion"].map(normalizar_texto_seguro)
    unids = df["Unids"].map(normalizar_texto_seguro)
    bultos = df["Bultos"].map(normalizar_texto_seguro)

    mask_total = empresa.map(es_texto_total_resumen) | cd.map(es_texto_total_resumen)
    mask_no_operativa = (
        empresa.eq("")
        | cd.eq("")
        | (articulo.eq("") & contenedor.eq(""))
        | (articulo.eq("") & contenedor.eq("") & unids.eq("") & bultos.eq("") & ubicacion.eq(""))
    ) & ~mask_total

    descartes["filas_descartadas_totales_excel"] = int(mask_total.sum())
    descartes["filas_descartadas_no_operativas"] = int(mask_no_operativa.sum())

    df_limpio = df.loc[~(mask_total | mask_no_operativa)].copy()
    return df_limpio, descartes


def parse_fecha_staging_desde_nombre_archivo(path: Path) -> datetime | None:
    match = re.search(r"(\d{14})$", path.stem)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%d%m%Y%H%M%S")
    except ValueError:
        return None


def normalizar_cliente_staging(valor: Any, cd: Any = None) -> str:
    texto = normalizar_etiqueta(valor)
    cd_txt = normalizar_etiqueta(cd)
    if not texto:
        return ""
    equivalencias = {
        "ABINBEV": "ABINBEV",
        "ABINVEB": "ABINBEV",
        "CERVECERIA ABI": "ABINBEV",
        "BARENTZ": "BARENTZ",
        "DERCO": "DERCO",
        "UNILEVER": "UNILEVER",
        "RUNO": "RUNO SPA",
        "RUNO SPA": "RUNO SPA",
        "POCHTECA": "POCHTECA",
        "DAIKIN": "DAIKIN",
        "DAIKIN CLIENTES": "DAIKIN CLIENTES",
        "MASCOTAS LATINAS": "MASCOTAS LATINAS",
        "MASCOTA LATINA": "MASCOTAS LATINAS",
        "CEPAS CHILE": "CEPAS CHILE",
        "COLLICO": "COLLICO",
        "DELIBEST": "DELIBEST",
        "INTIME": "INTIME",
        "NATIVOS DRINK": "NATIVO DRINKS SPA",
        "NATIVO DRINKS": "NATIVO DRINKS SPA",
        "NATIVO DRINKS SPA": "NATIVO DRINKS SPA",
        "TRES MONTE": "TRESMONTES",
        "TRES MONTES": "TRESMONTES",
        "BURASCHI": "BURASCHI",
        "SANTA ROSA": "SANTA ROSA",
    }
    cliente = equivalencias.get(texto, texto)
    if cliente == "MASCOTAS LATINAS" and cd_txt == "PUDAHUEL":
        return "MASCOTAS LATINAS_P"
    return cliente


def normalizar_cd_staging(valor: Any, fallback: Any = None) -> str:
    texto = normalizar_etiqueta(valor)
    if texto in {"1", "QUILICURA"}:
        return "QUILICURA"
    if texto in {"2", "PUDAHUEL"}:
        return "PUDAHUEL"
    if texto in {"3", "PUDAHUEL UNITARIO"}:
        return "PUDAHUEL UNITARIO"
    if texto:
        return normalizar_cd_inventario(texto)
    return normalizar_cd_inventario(fallback)


def normalizar_estado_disponibilidad_staging(valor: Any) -> str:
    texto = normalizar_etiqueta(valor)
    texto = texto.replace("ADISPONIBLE", "A DISPONIBLE")
    return texto


def normalizar_estado_staging(lugar: Any, ubicacion: Any) -> str:
    lugar_txt = normalizar_estado_disponibilidad_staging(lugar)
    ubicacion_txt = normalizar_estado_disponibilidad_staging(ubicacion)
    if "STAGING IN" in lugar_txt or "STAGINGIN" in lugar_txt:
        return "STAGING IN"
    if "STAGING OUT" in lugar_txt or "STAGINGOUT" in lugar_txt:
        return "STAGING OUT"
    if "STAGING IN" in ubicacion_txt or "STAGINGIN" in ubicacion_txt:
        return "STAGING IN"
    if "STAGING OUT" in ubicacion_txt or "STAGINGOUT" in ubicacion_txt:
        return "STAGING OUT"
    if "ALMACEN" in lugar_txt:
        return "ALMACEN"
    if "ALMACEN" in ubicacion_txt:
        return "ALMACEN"
    return "SIN CLASIFICAR"


def grupo_motivo_bloqueo_staging(valor: Any) -> str:
    texto = normalizar_estado_disponibilidad_staging(valor)
    if any(token in texto for token in ("MERMA", "SCRAP", "MERCADERIA VENCIDA")):
        return "MERMA"
    if "CONTROL DE CALIDAD" in texto:
        return "CONTROL DE CALIDAD"
    if "CUARENTENA" in texto:
        return "CUARENTENA"
    if "A DISPONIBLE" in texto:
        return "DISPONIBLE"
    return "OTRO"


def bool_bloqueado_staging(valor: Any) -> bool:
    texto = normalizar_etiqueta(valor)
    return texto in {"S", "SI", "TRUE", "1"}


def balde_antiguedad(dias: int | None) -> tuple[str, int]:
    if dias is None or dias < 0:
        return "SIN DATO", 99
    if dias <= 7:
        return "0-7 Días (W1)", 1
    if dias <= 14:
        return "8-14 Días (W2)", 2
    if dias <= 21:
        return "15-21 Días (W3)", 3
    return "> 21 Días (W+)", 4


def first_non_empty(serie: pd.Series) -> str:
    for valor in serie.tolist():
        texto = normalizar_texto_seguro(valor)
        if texto:
            return texto
    return ""


def descubrir_fuentes_staging(verbose: bool = False) -> list[FuenteDetectada]:
    fuentes: list[FuenteDetectada] = []
    if not STAGING_ROOT.exists():
        return fuentes

    for path in STAGING_ROOT.rglob("*.csv"):
        if es_archivo_respaldo(path):
            continue
        cliente = path.parent.name if path.parent != STAGING_ROOT else ""
        centro = path.parent.parent.name if path.parent.parent != STAGING_ROOT.parent else ""
        fuentes.append(
            FuenteDetectada(
                ruta=path,
                cliente=normalizar_cliente_staging(cliente, centro),
                centro=normalizar_cd_staging(centro),
            )
        )
    fuentes = sorted(
        fuentes,
        key=lambda item: (
            parse_fecha_staging_desde_nombre_archivo(item.ruta) or datetime.fromtimestamp(item.ruta.stat().st_mtime),
            normalizar_texto(item.centro),
            normalizar_texto(item.cliente),
            str(item.ruta),
        ),
        reverse=True,
    )
    if verbose:
        for fuente in fuentes[:20]:
            print(f"[INV-STG] Fuente detectada: {fuente.ruta} | cd={fuente.centro} | cliente={fuente.cliente}")
    return fuentes


def seleccionar_ultimas_fuentes_staging(fuentes: list[FuenteDetectada]) -> list[FuenteDetectada]:
    seleccion: dict[tuple[str, str], FuenteDetectada] = {}
    for fuente in fuentes:
        clave = (normalizar_texto(fuente.centro), normalizar_texto(fuente.cliente))
        actual = seleccion.get(clave)
        fecha_fuente = parse_fecha_staging_desde_nombre_archivo(fuente.ruta) or datetime.fromtimestamp(fuente.ruta.stat().st_mtime)
        fecha_actual = (
            parse_fecha_staging_desde_nombre_archivo(actual.ruta) or datetime.fromtimestamp(actual.ruta.stat().st_mtime)
            if actual else None
        )
        if actual is None or (fecha_actual and fecha_fuente > fecha_actual):
            seleccion[clave] = fuente
    return sorted(seleccion.values(), key=lambda item: (normalizar_texto(item.centro), normalizar_texto(item.cliente), str(item.ruta)))


def cargar_staging_csv(
    fuente: FuenteDetectada,
    columnas_necesarias: list[str] | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    if pd is None:
        raise RuntimeError("pandas no esta disponible para leer staging.")

    columnas_necesarias = columnas_necesarias or []
    ultimo_error = None
    for encoding in ("cp1252", "utf-8-sig", "utf-8"):
        try:
            df = pd.read_csv(
                fuente.ruta,
                sep=";",
                dtype=str,
                encoding=encoding,
                keep_default_na=False,
            )
            break
        except Exception as exc:
            ultimo_error = exc
            df = None
    if df is None:
        raise RuntimeError(f"No se pudo leer CSV staging: {ultimo_error}")

    renombres = {}
    for columna in df.columns:
        clave = quitar_acentos(normalizar_texto(columna))
        clave_norm = normalizar_etiqueta(columna)
        if clave == "Deposito":
            renombres[columna] = "deposito_origen"
        elif clave == "Empresa":
            renombres[columna] = "cliente_origen"
        elif clave == "Codigo Articulo":
            renombres[columna] = "Articulo"
        elif "ARTICULO DESCRIPCI" in clave_norm:
            renombres[columna] = "Descripcion"
        elif clave == "Pallet":
            renombres[columna] = "Pallet"
        elif clave == "Lote":
            renombres[columna] = "Lote"
        elif clave == "Bloqueado":
            renombres[columna] = "Bloqueado"
        elif clave == "Descripcion Estado":
            renombres[columna] = "Descripcion Estado"
        elif clave == "Cantidad":
            renombres[columna] = "Cantidad"
        elif clave == "Unidad":
            renombres[columna] = "Unidad"
        elif clave == "Bultos":
            renombres[columna] = "Bultos"
        elif clave == "Ubicacion":
            renombres[columna] = "Ubicacion"
        elif clave == "Lugar":
            renombres[columna] = "Lugar"
        elif clave == "Fecha de Alta":
            renombres[columna] = "Fecha de Alta"
        elif clave == "Fecha de Vencimiento":
            renombres[columna] = "Fecha de Vencimiento"
        else:
            renombres[columna] = columna
    df = df.rename(columns=renombres)

    base_cols = [
        "deposito_origen", "cliente_origen", "Articulo", "Descripcion", "Pallet", "Lote", "Bloqueado",
        "Descripcion Estado", "Cantidad", "Unidad", "Bultos", "Ubicacion", "Lugar", "Fecha de Alta",
        "Fecha de Vencimiento",
    ]
    for columna in base_cols:
        if columna not in df.columns:
            df[columna] = None
    df = df[base_cols].copy()
    df = df.dropna(how="all")

    fecha_descarga = parse_fecha_staging_desde_nombre_archivo(fuente.ruta)
    df["cliente_origen"] = df["cliente_origen"].map(normalizar_texto_seguro)
    df["deposito_origen"] = df["deposito_origen"].map(normalizar_texto_seguro)
    df["cliente"] = [
        normalizar_cliente_staging(fuente.cliente or cliente, deposito or fuente.centro)
        for cliente, deposito in zip(df["cliente_origen"], df["deposito_origen"])
    ]
    df["cd"] = [
        normalizar_cd_staging(deposito, fuente.centro)
        for deposito in df["deposito_origen"]
    ]
    df["Articulo"] = df["Articulo"].map(normalizar_clave)
    df["Descripcion"] = df["Descripcion"].map(normalizar_texto)
    df["Pallet"] = df["Pallet"].map(normalizar_clave)
    df["Lote"] = df["Lote"].map(normalizar_texto)
    df["Bloqueado"] = df["Bloqueado"].map(normalizar_texto)
    df["Descripcion Estado"] = df["Descripcion Estado"].map(normalizar_texto)
    df["Descripcion Estado Normalizada"] = df["Descripcion Estado"].map(normalizar_estado_disponibilidad_staging)
    df["Cantidad_num"] = df["Cantidad"].map(parse_numero)
    df["Bultos_num"] = df["Bultos"].map(parse_numero)
    df["Ubicacion"] = df["Ubicacion"].map(normalizar_texto)
    df["Lugar"] = df["Lugar"].map(normalizar_texto)
    df["estado_staging_normalizado"] = [
        normalizar_estado_staging(lugar, ubicacion)
        for lugar, ubicacion in zip(df["Lugar"], df["Ubicacion"])
    ]
    df["fecha_descarga"] = fecha_descarga or datetime.fromtimestamp(fuente.ruta.stat().st_mtime)
    df["archivo_fuente"] = fuente.ruta.name

    mask_vacia = (
        df["Pallet"].map(normalizar_texto_seguro).eq("")
        & df["Articulo"].map(normalizar_texto_seguro).eq("")
        & df["Ubicacion"].map(normalizar_texto_seguro).eq("")
        & df["Cantidad"].map(normalizar_texto_seguro).eq("")
    )
    df = df.loc[~mask_vacia].copy()

    info = {
        "path": str(fuente.ruta),
        "name": fuente.ruta.name,
        "modified": iso_mtime(fuente.ruta),
        "cliente_detectado": fuente.cliente,
        "cd_detectado": fuente.centro,
        "rows": int(len(df)),
        "fecha_descarga": serializar_o_mtime_fecha(fecha_descarga, fuente.ruta),
    }
    if columnas_necesarias:
        columnas_keep = [col for col in columnas_necesarias if col in df.columns]
        df = df[columnas_keep].copy()
    return df, info


def resumir_staging_alertas(df_snapshot: pd.DataFrame, extra: dict[str, int] | None = None) -> list[dict[str, Any]]:
    extra = extra or {}
    alertas_contadores = {
        "registros_sin_cliente": int(df_snapshot["cliente"].eq("").sum()),
        "registros_sin_cd": int(df_snapshot["cd"].eq("").sum()),
        "registros_sin_pallet": int(df_snapshot["Pallet"].eq("").sum()),
        "registros_sin_articulo": int(df_snapshot["Articulo"].eq("").sum()),
        "registros_sin_ubicacion": int(df_snapshot["Ubicacion"].eq("").sum()),
        "registros_sin_estado_staging": int(df_snapshot["estado_staging_normalizado"].eq("SIN CLASIFICAR").sum()),
        "registros_sin_cantidad": int(df_snapshot["Cantidad"].map(normalizar_texto_seguro).eq("").sum()),
        "registros_con_cantidad_negativa": int((df_snapshot["Cantidad_num"] < 0).sum()),
    }
    alertas_contadores.update({k: int(v) for k, v in extra.items()})
    return resumir_alertas_contadores(alertas_contadores)


def calcular_staging_snapshot(df_snapshot: pd.DataFrame, fecha_referencia: str | None) -> tuple[dict[str, Any], list[str]]:
    if df_snapshot is None or df_snapshot.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontraron datos de snapshot de staging.",
        }, []

    df_staging = df_snapshot[df_snapshot["estado_staging_normalizado"].isin(ESTADOS_STAGING_VALIDOS)].copy()
    alertas = resumir_staging_alertas(df_snapshot)
    mensajes = []
    if df_snapshot["estado_staging_normalizado"].eq("SIN CLASIFICAR").any():
        mensajes.append("Existen registros de staging sin clasificacion clara por Lugar/Ubicacion.")

    if df_staging.empty:
        return {
            "disponible": True,
            "fecha_referencia": fecha_referencia,
            "total_plts": 0,
            "total_skus": 0,
            "total_unidades": 0.0,
            "total_bultos": 0.0,
            "por_estado": [],
            "por_cd": [],
            "por_cliente": [],
            "top_clientes": [],
            "alertas": alertas,
        }, mensajes

    por_estado = []
    for estado, grupo in df_staging.groupby("estado_staging_normalizado", dropna=False):
        por_estado.append(
            {
                "estado_staging": estado,
                "plts": int(grupo.loc[grupo["Pallet"] != "", "Pallet"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
                "unidades": round_safe(grupo["Cantidad_num"].sum(), 2),
                "bultos": round_safe(grupo["Bultos_num"].sum(), 2),
            }
        )
    por_estado.sort(key=lambda item: (-item["plts"], item["estado_staging"]))

    por_cd = []
    for cd, grupo in df_staging.groupby("cd", dropna=False):
        por_cd.append(
            {
                "cd": cd,
                "plts": int(grupo.loc[grupo["Pallet"] != "", "Pallet"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
                "unidades": round_safe(grupo["Cantidad_num"].sum(), 2),
                "bultos": round_safe(grupo["Bultos_num"].sum(), 2),
            }
        )
    por_cd.sort(key=lambda item: item["cd"])

    por_cliente = []
    for (cliente, cd), grupo in df_staging.groupby(["cliente", "cd"], dropna=False):
        estados = (
            grupo.groupby("estado_staging_normalizado")["Cantidad_num"]
            .sum()
            .sort_values(ascending=False)
            .head(3)
            .reset_index()
        )
        por_cliente.append(
            {
                "cliente": cliente,
                "cd": cd,
                "plts": int(grupo.loc[grupo["Pallet"] != "", "Pallet"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
                "unidades": round_safe(grupo["Cantidad_num"].sum(), 2),
                "bultos": round_safe(grupo["Bultos_num"].sum(), 2),
                "estados_principales": [
                    {"estado_staging": row["estado_staging_normalizado"], "unidades": round_safe(row["Cantidad_num"], 2)}
                    for _, row in estados.iterrows()
                ],
            }
        )
    por_cliente.sort(key=lambda item: (-item["unidades"], item["cliente"], item["cd"]))

    top_clientes = []
    for item in por_cliente[:10]:
        estado_principal = item["estados_principales"][0]["estado_staging"] if item["estados_principales"] else ""
        top_clientes.append(
            {
                "cliente": item["cliente"],
                "cd": item["cd"],
                "plts": item["plts"],
                "unidades": item["unidades"],
                "estado_principal": estado_principal,
            }
        )

    payload = {
        "disponible": True,
        "fecha_referencia": fecha_referencia,
        "total_plts": int(df_staging.loc[df_staging["Pallet"] != "", "Pallet"].nunique()),
        "total_skus": int(df_staging.loc[df_staging["Articulo"] != "", "Articulo"].nunique()),
        "total_unidades": round_safe(df_staging["Cantidad_num"].sum(), 2),
        "total_bultos": round_safe(df_staging["Bultos_num"].sum(), 2),
        "por_estado": por_estado,
        "por_cd": por_cd,
        "por_cliente": por_cliente,
        "top_clientes": top_clientes,
        "alertas": alertas,
    }
    return payload, mensajes


def calcular_stock_bloqueado_staging(df_snapshot: pd.DataFrame) -> tuple[dict[str, Any], list[str]]:
    if df_snapshot is None or df_snapshot.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontraron datos de staging para bloqueos.",
        }, []

    df_staging = df_snapshot[df_snapshot["estado_staging_normalizado"].isin(ESTADOS_STAGING_VALIDOS)].copy()
    if df_staging.empty:
        return {
            "disponible": True,
            "unidades": 0.0,
            "plts": 0,
            "skus": 0,
            "por_cliente": [],
            "por_motivo": [],
            "alertas": resumir_alertas_contadores({"motivos_vacios": 0, "bloqueado_vacio": 0}),
        }, []

    desc_norm = df_staging["Descripcion Estado Normalizada"]
    mask = (
        df_staging["Bloqueado"].map(bool_bloqueado_staging)
        | desc_norm.ne("A DISPONIBLE")
        | desc_norm.str.contains("MERMA|CUARENTENA|CONTROL DE CALIDAD|MERCADERIA VENCIDA", na=False)
    )
    df_bloqueado = df_staging.loc[mask].copy()
    if df_bloqueado.empty:
        return {
            "disponible": True,
            "unidades": 0.0,
            "plts": 0,
            "skus": 0,
            "por_cliente": [],
            "por_motivo": [],
            "alertas": resumir_alertas_contadores({"motivos_vacios": 0, "bloqueado_vacio": 0}),
        }, []

    df_bloqueado["grupo_motivo"] = df_bloqueado["Descripcion Estado"].map(grupo_motivo_bloqueo_staging)
    alertas_contadores = {
        "motivos_vacios": int(df_bloqueado["Descripcion Estado"].map(normalizar_texto_seguro).eq("").sum()),
        "bloqueado_vacio": int(df_bloqueado["Bloqueado"].map(normalizar_texto_seguro).eq("").sum()),
    }
    mensajes = []

    por_cliente = []
    for (cliente, cd), grupo in df_bloqueado.groupby(["cliente", "cd"], dropna=False):
        top_motivos = grupo.groupby("Descripcion Estado")["Cantidad_num"].sum().sort_values(ascending=False).head(3).reset_index()
        por_cliente.append(
            {
                "cliente": cliente,
                "cd": cd,
                "unidades": round_safe(grupo["Cantidad_num"].sum(), 2),
                "plts": int(grupo.loc[grupo["Pallet"] != "", "Pallet"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
                "principales_motivos": [
                    {"motivo": row["Descripcion Estado"], "unidades": round_safe(row["Cantidad_num"], 2)}
                    for _, row in top_motivos.iterrows()
                ],
            }
        )
    por_cliente.sort(key=lambda item: (-item["unidades"], item["cliente"], item["cd"]))

    por_motivo = []
    for (motivo, grupo_motivo), grupo in df_bloqueado.groupby(["Descripcion Estado", "grupo_motivo"], dropna=False):
        por_motivo.append(
            {
                "motivo": motivo,
                "grupo_motivo": grupo_motivo,
                "unidades": round_safe(grupo["Cantidad_num"].sum(), 2),
                "plts": int(grupo.loc[grupo["Pallet"] != "", "Pallet"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
            }
        )
    por_motivo.sort(key=lambda item: (-item["unidades"], item["grupo_motivo"], item["motivo"]))

    payload = {
        "disponible": True,
        "unidades": round_safe(df_bloqueado["Cantidad_num"].sum(), 2),
        "plts": int(df_bloqueado.loc[df_bloqueado["Pallet"] != "", "Pallet"].nunique()),
        "skus": int(df_bloqueado.loc[df_bloqueado["Articulo"] != "", "Articulo"].nunique()),
        "por_cliente": por_cliente,
        "por_motivo": por_motivo,
        "alertas": resumir_alertas_contadores(alertas_contadores),
    }
    return payload, mensajes


def construir_pallets_actuales(df_snapshot: pd.DataFrame) -> pd.DataFrame:
    if df_snapshot is None or df_snapshot.empty:
        return pd.DataFrame()
    df_staging = df_snapshot[df_snapshot["estado_staging_normalizado"].isin(ESTADOS_STAGING_VALIDOS)].copy()
    if df_staging.empty:
        return pd.DataFrame()
    agrupado = (
        df_staging.groupby(["cliente", "cd", "Pallet"], dropna=False)
        .agg(
            unidades=("Cantidad_num", "sum"),
            bultos=("Bultos_num", "sum"),
            skus=("Articulo", lambda s: int(pd.Series([v for v in s if normalizar_texto(v)]).nunique())),
            articulo=("Articulo", first_non_empty),
            descripcion=("Descripcion", first_non_empty),
            ubicacion=("Ubicacion", first_non_empty),
            estado_staging=("estado_staging_normalizado", first_non_empty),
        )
        .reset_index()
    )
    return agrupado


def procesar_historico_staging(
    fuentes_historicas: list[FuenteDetectada],
    pallets_objetivo: set[str],
    pallets_out_objetivo: set[str],
) -> tuple[dict[str, datetime], dict[str, datetime]]:
    primeras_apariciones: dict[str, datetime] = {}
    primeras_apariciones_out: dict[str, datetime] = {}
    if not pallets_objetivo:
        return primeras_apariciones, primeras_apariciones_out

    columnas = ["Pallet", "estado_staging_normalizado"]
    for fuente in fuentes_historicas:
        fecha_archivo = parse_fecha_staging_desde_nombre_archivo(fuente.ruta) or datetime.fromtimestamp(fuente.ruta.stat().st_mtime)
        try:
            df_hist, _ = cargar_staging_csv(fuente, columnas_necesarias=columnas)
        except Exception:
            continue
        if df_hist.empty:
            continue
        df_hist = df_hist[df_hist["Pallet"].isin(pallets_objetivo)].copy()
        if df_hist.empty:
            continue
        for pallet in df_hist["Pallet"].dropna().unique().tolist():
            if pallet not in primeras_apariciones or fecha_archivo < primeras_apariciones[pallet]:
                primeras_apariciones[pallet] = fecha_archivo
        if pallets_out_objetivo:
            df_out = df_hist[
                df_hist["Pallet"].isin(pallets_out_objetivo)
                & df_hist["estado_staging_normalizado"].eq("STAGING OUT")
            ].copy()
            for pallet in df_out["Pallet"].dropna().unique().tolist():
                if pallet not in primeras_apariciones_out or fecha_archivo < primeras_apariciones_out[pallet]:
                    primeras_apariciones_out[pallet] = fecha_archivo
    return primeras_apariciones, primeras_apariciones_out


def calcular_staging_antiguedad(
    pallets_actuales: pd.DataFrame,
    primeras_apariciones: dict[str, datetime],
    fecha_referencia: datetime | None,
) -> tuple[dict[str, Any], list[str]]:
    if pallets_actuales is None or pallets_actuales.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontraron pallets actuales en staging para antiguedad.",
        }, []
    if fecha_referencia is None:
        return {
            "disponible": False,
            "mensaje": "No se pudo determinar fecha de referencia para antiguedad de staging.",
        }, []

    df = pallets_actuales.copy()
    df["fecha_ingreso_original"] = df["Pallet"].map(primeras_apariciones)
    df["dias"] = df["fecha_ingreso_original"].map(
        lambda valor: (fecha_referencia.date() - valor.date()).days if isinstance(valor, datetime) else None
    )
    df["balde_info"] = df["dias"].map(balde_antiguedad)
    df["balde_antiguedad"] = df["balde_info"].map(lambda v: v[0])
    df["orden_balde"] = df["balde_info"].map(lambda v: v[1])

    por_balde = []
    for (balde, orden), grupo in df.groupby(["balde_antiguedad", "orden_balde"], dropna=False):
        por_balde.append(
            {
                "balde_antiguedad": balde,
                "orden": int(orden),
                "plts": int(len(grupo)),
                "unidades": round_safe(grupo["unidades"].sum(), 2),
                "skus": int(grupo["articulo"].replace("", pd.NA).dropna().nunique()),
            }
        )
    por_balde.sort(key=lambda item: item["orden"])

    por_cliente = []
    for (cliente, cd), grupo in df.groupby(["cliente", "cd"], dropna=False):
        baldes = grupo.groupby("balde_antiguedad").size().sort_values(ascending=False)
        por_cliente.append(
            {
                "cliente": cliente,
                "cd": cd,
                "plts": int(len(grupo)),
                "unidades": round_safe(grupo["unidades"].sum(), 2),
                "balde_principal": baldes.index[0] if not baldes.empty else "",
                "plts_mayor_21_dias": int(grupo["dias"].fillna(-1).gt(21).sum()),
            }
        )
    por_cliente.sort(key=lambda item: (-item["plts_mayor_21_dias"], -item["plts"], item["cliente"], item["cd"]))

    mayores = df[df["dias"].fillna(-1).gt(21)].copy()
    mayores = mayores.sort_values(["dias", "unidades"], ascending=[False, False])
    mayores_21_dias = []
    for _, row in mayores.head(20).iterrows():
        mayores_21_dias.append(
            {
                "cliente": row["cliente"],
                "cd": row["cd"],
                "pallet": row["Pallet"],
                "dias": int(row["dias"]) if row["dias"] is not None else None,
                "articulo": row["articulo"],
                "descripcion": row["descripcion"],
                "unidades": round_safe(row["unidades"], 2),
                "ubicacion": row["ubicacion"],
                "estado_staging": row["estado_staging"],
                "fecha_ingreso_original": serializar_fecha(row["fecha_ingreso_original"]),
            }
        )

    payload = {
        "disponible": True,
        "fecha_referencia": serializar_fecha(fecha_referencia),
        "metodo": "primera_aparicion_historica_pallet",
        "nota": "La antigüedad usa primera aparición histórica del pallet. Si un pallet sale y vuelve a entrar, puede sobreestimar antigüedad; una mejora futura sería calcular por racha continua.",
        "por_balde": por_balde,
        "por_cliente": por_cliente,
        "mayores_21_dias": mayores_21_dias,
        "alertas": resumir_alertas_contadores(
            {
                "pallets_sin_primera_aparicion_historica": int(df["fecha_ingreso_original"].isna().sum()),
            }
        ),
    }
    return payload, []


def calcular_staging_out_permanencia(
    pallets_actuales: pd.DataFrame,
    primeras_apariciones_out: dict[str, datetime],
    fecha_referencia: datetime | None,
) -> tuple[dict[str, Any], list[str]]:
    if pallets_actuales is None or pallets_actuales.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontraron pallets actuales para permanencia en STAGING OUT.",
        }, []
    if fecha_referencia is None:
        return {
            "disponible": False,
            "mensaje": "No se pudo determinar fecha de referencia para permanencia en STAGING OUT.",
        }, []

    df = pallets_actuales[pallets_actuales["estado_staging"].eq("STAGING OUT")].copy()
    if df.empty:
        return {
            "disponible": True,
            "fecha_referencia": serializar_fecha(fecha_referencia),
            "metodo": "primera_aparicion_staging_out_pallet",
            "nota": "La permanencia considera primera aparición del pallet en STAGING OUT. Una mejora futura sería medir por racha continua.",
            "total_plts_staging_out": 0,
            "por_balde": [],
            "por_cliente": [],
            "mayores_21_dias": [],
            "alertas": [],
        }, []

    df["fecha_ingreso_original"] = df["Pallet"].map(primeras_apariciones_out)
    df["dias"] = df["fecha_ingreso_original"].map(
        lambda valor: (fecha_referencia.date() - valor.date()).days if isinstance(valor, datetime) else None
    )
    df["balde_info"] = df["dias"].map(balde_antiguedad)
    df["balde_antiguedad"] = df["balde_info"].map(lambda v: v[0])
    df["orden_balde"] = df["balde_info"].map(lambda v: v[1])

    por_balde = []
    for (balde, orden), grupo in df.groupby(["balde_antiguedad", "orden_balde"], dropna=False):
        por_balde.append(
            {
                "balde_antiguedad": balde,
                "orden": int(orden),
                "plts": int(len(grupo)),
                "unidades": round_safe(grupo["unidades"].sum(), 2),
                "skus": int(grupo["articulo"].replace("", pd.NA).dropna().nunique()),
            }
        )
    por_balde.sort(key=lambda item: item["orden"])

    por_cliente = []
    for (cliente, cd), grupo in df.groupby(["cliente", "cd"], dropna=False):
        baldes = grupo.groupby("balde_antiguedad").size().sort_values(ascending=False)
        por_cliente.append(
            {
                "cliente": cliente,
                "cd": cd,
                "plts": int(len(grupo)),
                "unidades": round_safe(grupo["unidades"].sum(), 2),
                "balde_principal": baldes.index[0] if not baldes.empty else "",
                "plts_mayor_21_dias": int(grupo["dias"].fillna(-1).gt(21).sum()),
            }
        )
    por_cliente.sort(key=lambda item: (-item["plts_mayor_21_dias"], -item["plts"], item["cliente"], item["cd"]))

    mayores = df[df["dias"].fillna(-1).gt(21)].copy()
    mayores = mayores.sort_values(["dias", "unidades"], ascending=[False, False])
    mayores_21_dias = []
    for _, row in mayores.head(20).iterrows():
        mayores_21_dias.append(
            {
                "cliente": row["cliente"],
                "cd": row["cd"],
                "pallet": row["Pallet"],
                "dias": int(row["dias"]) if row["dias"] is not None else None,
                "articulo": row["articulo"],
                "descripcion": row["descripcion"],
                "unidades": round_safe(row["unidades"], 2),
                "ubicacion": row["ubicacion"],
                "estado_staging": row["estado_staging"],
                "fecha_ingreso_original": serializar_fecha(row["fecha_ingreso_original"]),
            }
        )

    payload = {
        "disponible": True,
        "fecha_referencia": serializar_fecha(fecha_referencia),
        "metodo": "primera_aparicion_staging_out_pallet",
        "nota": "La permanencia considera primera aparición del pallet en STAGING OUT. Una mejora futura sería medir por racha continua.",
        "total_plts_staging_out": int(len(df)),
        "por_balde": por_balde,
        "por_cliente": por_cliente,
        "mayores_21_dias": mayores_21_dias,
        "alertas": resumir_alertas_contadores(
            {
                "pallets_staging_out_sin_primera_aparicion": int(df["fecha_ingreso_original"].isna().sum()),
            }
        ),
    }
    return payload, []


def descubrir_fuentes_stock_wms(verbose: bool = False) -> list[FuenteDetectada]:
    fuentes: list[FuenteDetectada] = []
    if not STOCK_WMS_ROOT.exists():
        return fuentes

    quilicura_dir = STOCK_WMS_ROOT / "Quilicura"
    pudahuel_dir = STOCK_WMS_ROOT / "Pudahuel"

    if quilicura_dir.exists():
        archivos = sorted(
            [p for p in quilicura_dir.glob("*.xlsx") if not es_archivo_respaldo(p)],
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if archivos:
            fuentes.append(FuenteDetectada(ruta=archivos[0], centro="QUILICURA"))

    if pudahuel_dir.exists():
        archivos = sorted(
            [p for p in pudahuel_dir.glob("*.xlsx") if not es_archivo_respaldo(p)],
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        archivo_pudahuel = next((p for p in archivos if "UNITARIO" not in normalizar_etiqueta(p.name)), None)
        archivo_unitario = next((p for p in archivos if "UNITARIO" in normalizar_etiqueta(p.name)), None)
        if archivo_pudahuel:
            fuentes.append(FuenteDetectada(ruta=archivo_pudahuel, centro="PUDAHUEL"))
        if archivo_unitario:
            fuentes.append(FuenteDetectada(ruta=archivo_unitario, centro="PUDAHUEL UNITARIO"))

    if verbose:
        for fuente in fuentes:
            print(f"[INV-STOCK] Fuente detectada: {fuente.ruta} | cd={fuente.centro}")
    return fuentes


def descubrir_fuentes_posiciones(verbose: bool = False) -> list[FuenteDetectada]:
    fuentes: list[FuenteDetectada] = []
    if not POSICIONES_ROOT.exists():
        return fuentes

    archivos = [
        POSICIONES_ROOT / "Quilicura" / "Posiciones Ocupadas.xlsx",
        POSICIONES_ROOT / "Quilicura" / "Posiciones Libres.xlsx",
        POSICIONES_ROOT / "Pudahuel" / "Posiciones Ocupadas Moderno.xlsx",
        POSICIONES_ROOT / "Pudahuel" / "Posiciones Libres Moderno.xlsx",
        POSICIONES_ROOT / "Pudahuel" / "Posiciones Ocupadas Unitario.xlsx",
        POSICIONES_ROOT / "Pudahuel" / "Posiciones Libres Unitario.xlsx",
        POSICIONES_ROOT / "Pudahuel" / "Posiciones Ocupadas Refrigerado.xlsx",
        POSICIONES_ROOT / "Pudahuel" / "Posiciones Libres Refrigerado.xlsx",
    ]
    for archivo in archivos:
        if archivo.exists() and not es_archivo_respaldo(archivo):
            fuentes.append(FuenteDetectada(ruta=archivo))

    if verbose:
        for fuente in fuentes:
            print(f"[INV-POS] Fuente detectada: {fuente.ruta}")
    return fuentes


def encontrar_archivo_dim_inventario() -> Path | None:
    directo = INVENTARIO_DIM_ROOT / INVENTARIO_DIM_FILENAME
    if directo.exists():
        return directo
    if not INVENTARIO_DIM_ROOT.exists():
        return None
    candidatos = sorted(
        [
            p for p in INVENTARIO_DIM_ROOT.glob("*.xlsx")
            if normalizar_etiqueta(p.name) == normalizar_etiqueta(INVENTARIO_DIM_FILENAME)
        ],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidatos[0] if candidatos else None


def detectar_header_stock(path: Path) -> int:
    if pd is None:
        return 5
    muestra = pd.read_excel(path, sheet_name=0, header=None, nrows=10, engine="openpyxl")
    esperadas = {"DEPOSITO", "EMPRESA", "ARTICULO", "CONTENEDOR", "DESCRIPCION"}
    mejor_fila = 4
    mejor_score = -1
    for idx, row in muestra.iterrows():
        headers = {normalizar_etiqueta(v) for v in row.tolist() if normalizar_texto(v)}
        score = len(headers & esperadas)
        if score > mejor_score:
            mejor_score = score
            mejor_fila = idx
    return mejor_fila + 1


def cargar_stock_wms_excel(fuente: FuenteDetectada) -> tuple[pd.DataFrame, dict[str, Any]]:
    if pd is None:
        raise RuntimeError("pandas no esta disponible para leer stock WMS.")

    header_row = detectar_header_stock(fuente.ruta)
    df = pd.read_excel(fuente.ruta, sheet_name=0, header=header_row - 1, engine="openpyxl")
    df.columns = [normalizar_texto(c) for c in df.columns]
    df = df.loc[:, [col for col in df.columns if col]]
    df = df.dropna(how="all")

    renombres = {}
    for columna in df.columns:
        clave = quitar_acentos(normalizar_texto(columna))
        if clave == "Deposito":
            renombres[columna] = "CD"
        elif clave == "Empresa Dsc.":
            renombres[columna] = "Empresa"
        elif clave == "Empresa":
            renombres[columna] = "empresa_origen"
        elif clave == "Articulo":
            renombres[columna] = "Articulo"
        elif clave == "Descripcion":
            renombres[columna] = "Descripcion"
        elif clave == "Sub-Deposito":
            renombres[columna] = "Sub-Deposito"
        elif clave == "Codigo de Referencia":
            renombres[columna] = "Codigo de Referencia"
        else:
            renombres[columna] = columna
    df = df.rename(columns=renombres)

    columnas_necesarias = [
        "CD", "empresa_origen", "Empresa", "Articulo", "Descripcion", "Contenedor", "Apto", "Motivo",
        "Sub-Deposito", "Ingreso", "Vencimiento", "Lote", "Bultos", "Unids", "Area", "Cara", "Columna",
        "Nivel", "Nro Despacho", "Codigo de Referencia",
    ]
    for columna in columnas_necesarias:
        if columna not in df.columns:
            df[columna] = None

    fecha_descarga = parse_fecha_desde_nombre_archivo(fuente.ruta)
    df["deposito_origen"] = df["CD"]
    df["CD"] = df["CD"].map(normalizar_cd_inventario)
    df["Empresa"] = [
        normalizar_empresa_inventario(emp, cd)
        for emp, cd in zip(df["Empresa"], df["CD"])
    ]
    df["empresa_origen"] = df["empresa_origen"].map(normalizar_texto)
    df["Articulo"] = df["Articulo"].map(normalizar_clave)
    df["Descripcion"] = df["Descripcion"].map(normalizar_texto)
    df["Contenedor"] = df["Contenedor"].map(normalizar_clave)
    df["Apto"] = df["Apto"].map(normalizar_texto)
    df["Motivo"] = df["Motivo"].map(normalizar_texto)
    df["Ubicacion"] = [
        construir_ubicacion_desde_componentes(cara, columna, nivel)
        for cara, columna, nivel in zip(df["Cara"], df["Columna"], df["Nivel"])
    ]
    df["Unids_num"] = df["Unids"].map(parse_numero)
    df["Bultos_num"] = df["Bultos"].map(parse_numero)
    df["fecha_descarga"] = fecha_descarga
    df["archivo_fuente"] = fuente.ruta.name
    df, descartes = limpiar_filas_stock_wms(df)

    info = {
        "path": str(fuente.ruta),
        "name": fuente.ruta.name,
        "modified": iso_mtime(fuente.ruta),
        "cd_detectado": fuente.centro,
        "rows": int(len(df)),
        "header_row": header_row,
        "fecha_descarga": serializar_o_mtime_fecha(fecha_descarga, fuente.ruta),
        **descartes,
    }
    return df, info
def grupo_motivo_bloqueo_wms(motivo: Any) -> str:
    texto = normalizar_etiqueta(motivo)
    if any(token in texto for token in ("MERMA", "SCRAP", "MERCADERIA VENCIDA")):
        return "Merma"
    if "CONTROL DE CALIDAD" in texto:
        return "Control de Calidad"
    if "CUARENTENA" in texto:
        return "Cuarentena"
    return "Otro"


def es_cliente_santa_rosa(cliente: Any) -> bool:
    texto = normalizar_etiqueta(cliente).replace("_", " ").replace(".", " ")
    texto = " ".join(texto.split())
    tokens = set(texto.split())
    return "ROSA" in tokens and ("SANTA" in tokens or "STA" in tokens)


def resumir_stock_bloqueado_subset(df: pd.DataFrame) -> dict[str, Any]:
    if df is None or df.empty:
        return {
            "unidades": 0.0,
            "plts": 0,
            "skus": 0,
            "por_cliente": [],
            "por_motivo": [],
        }

    por_cliente = []
    for (cliente, cd), grupo in df.groupby(["Empresa", "CD"], dropna=False):
        top_motivos = grupo.groupby("Motivo")["Unids_num"].sum().sort_values(ascending=False).head(3).reset_index()
        por_cliente.append(
            {
                "cliente": cliente,
                "cd": cd,
                "unidades": round_safe(grupo["Unids_num"].sum(), 2),
                "plts": int(grupo.loc[grupo["Contenedor"] != "", "Contenedor"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
                "principales_motivos": [
                    {"motivo": row["Motivo"], "unidades": round_safe(row["Unids_num"], 2)}
                    for _, row in top_motivos.iterrows()
                ],
            }
        )
    por_cliente.sort(key=lambda item: (-item["unidades"], item["cliente"], item["cd"]))

    por_motivo = []
    for (motivo, grupo_motivo), grupo in df.groupby(["Motivo", "grupo_motivo"], dropna=False):
        por_motivo.append(
            {
                "motivo": motivo,
                "grupo_motivo": grupo_motivo,
                "unidades": round_safe(grupo["Unids_num"].sum(), 2),
                "plts": int(grupo.loc[grupo["Contenedor"] != "", "Contenedor"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
            }
        )
    por_motivo.sort(key=lambda item: (-item["unidades"], item["grupo_motivo"], item["motivo"]))

    return {
        "unidades": round_safe(df["Unids_num"].sum(), 2),
        "plts": int(df.loc[df["Contenedor"] != "", "Contenedor"].nunique()),
        "skus": int(df.loc[df["Articulo"] != "", "Articulo"].nunique()),
        "por_cliente": por_cliente,
        "por_motivo": por_motivo,
    }


def cargar_dim_layout_inventario(verbose: bool = False) -> tuple[dict[str, Any], pd.DataFrame | None]:
    info = {
        "disponible": False,
        "archivo": None,
        "tabla": "Tabla14",
    }
    if pd is None:
        info["mensaje"] = "pandas no esta disponible para leer dimension de inventario."
        return info, None

    archivo = encontrar_archivo_dim_inventario()
    if not archivo:
        info["mensaje"] = "No se encontro Tabla Ubicaciones CDs.xlsx en la ruta oficial de inventario."
        return info, None

    try:
        df = pd.read_excel(archivo, sheet_name="Ubicaciones CDs", engine="openpyxl")
        df.columns = [normalizar_texto(c) for c in df.columns]
        renombres = {}
        for columna in df.columns:
            clave = quitar_acentos(normalizar_texto(columna))
            if clave == "Ubicacion":
                renombres[columna] = "Ubicacion"
            else:
                renombres[columna] = columna
        df = df.rename(columns=renombres)
        for columna in ["Ubicacion", "Locacion", "CD", "Estado", "Profundidad", "Altura", "Tipo"]:
            if columna not in df.columns:
                df[columna] = None
        df = df[["Ubicacion", "Locacion", "CD", "Estado", "Profundidad", "Altura", "Tipo"]].copy()
        df = df.dropna(how="all")
        df["Ubicacion"] = df["Ubicacion"].map(normalizar_texto)
        df = df[df["Ubicacion"] != ""].copy()
        df["Locacion"] = df["Locacion"].map(normalizar_texto)
        df["CD"] = df["CD"].map(normalizar_cd_inventario)
        df["Estado"] = df["Estado"].map(normalizar_texto)
        df["Tipo"] = df["Tipo"].map(normalizar_texto)
        df["Profundidad_num"] = df["Profundidad"].map(parse_numero)
        info.update({
            "disponible": True,
            "archivo": str(archivo),
            "filas": int(len(df)),
        })
        if verbose:
            print(f"[INV-DIM] Archivo detectado: {archivo}")
        return info, df
    except Exception as exc:
        info["archivo"] = str(archivo)
        info["mensaje"] = f"No se pudo leer dimension de inventario: {exc}"
        return info, None


def es_archivo_conteos_valido(path: Path) -> bool:
    nombre = normalizar_etiqueta(path.name)
    return nombre in {normalizar_etiqueta(v) for v in CONTEOS_INVENTARIO_VARIANTES} or (
        "REGISTROS DE CONTEO CICLICO" in nombre
    )


def encontrar_archivo_conteos_en_root(root: Path) -> Path | None:
    if not root.exists():
        return None
    directo = root / CONTEOS_INVENTARIO_FILENAME
    if directo.exists():
        return directo
    candidatos = sorted(
        [p for p in root.glob("*.xlsx") if es_archivo_conteos_valido(p)],
        key=lambda p: (0 if normalizar_etiqueta(p.name) == normalizar_etiqueta(CONTEOS_INVENTARIO_FILENAME) else 1, -p.stat().st_mtime),
    )
    return candidatos[0] if candidatos else None


def encontrar_archivo_conteos_inventario() -> tuple[Path | None, str | None]:
    oficial = encontrar_archivo_conteos_en_root(CONTEOS_OFICIAL_ROOT)
    if oficial:
        return oficial, "oficial"
    legado = encontrar_archivo_conteos_en_root(INVENTARIO_DIM_ROOT)
    if legado:
        return legado, "fallback_legado"
    return None, None


def normalizar_columna_conteos(columna: Any) -> str:
    clave = normalizar_etiqueta(columna).replace(".", "").replace("_", " ")
    clave = " ".join(clave.split())
    if clave == "CD":
        return "CD"
    if clave == "USUARIO":
        return "Usuario"
    if clave == "UBICACION":
        return "ubicacion"
    if clave in {"UBIC CONTADAS", "UBICACION CONTADAS"}:
        return "Ubic. Contadas"
    if clave in {"NRO INTENTO", "NRO DE INTENTO"}:
        return "Nro. Intento"
    if clave == "EMPRESA":
        return "Empresa"
    if clave in {"EMPRESA CAPTURA", "EMPRESACAPTURA"}:
        return "Empresa Captura"
    if clave in {"COD ARTICULO FOTO", "CODIGO ARTICULO FOTO"}:
        return "Cod. Articulo Foto"
    if clave == "CODIGO CAPTURA":
        return "Codigo Captura"
    if clave == "CANTIDAD FOTO":
        return "Cantidad Foto"
    if clave == "CANTIDAD CAPTURA":
        return "Cantidad Captura"
    if clave == "PALLET FOTO":
        return "Pallet Foto"
    if clave == "PALLET CAPTURA":
        return "Pallet Captura"
    if clave == "LOTE FOTO":
        return "Lote Foto"
    if clave == "LOTE CAPTURA":
        return "Lote Captura"
    if clave == "VTO FOTO":
        return "Vto. Foto"
    if clave == "VTO CAPTURA":
        return "Vto. Captura"
    if clave == "DIFERENCIA":
        return "Diferencia"
    if clave == "VALIDAR SKU-LOTE-FVCTO" or clave == "VALIDAR SKU LOTE FVCTO":
        return "Validar Sku-lote-Fvcto"
    if clave == "FECHA CONTEO":
        return "Fecha Conteo"
    if clave == "% CUMPLIMIENTO IRA":
        return "% Cumplimiento IRA"
    if clave == "% CUMPLIMIENTO ILA":
        return "% Cumplimiento ILA"
    if clave == "FACTOR DE AJUSTE":
        return "Factor de ajuste"
    if clave in {"ESTADO UBIC", "ESTADO UBIC."}:
        return "Estado Ubic."
    if clave in {"TIPO DE UBIC", "TIPO DE UBIC."}:
        return "Tipo de Ubic."
    if clave == "MES":
        return "MES"
    if clave == "AÑO" or clave == "ANO":
        return "AÑO"
    if clave == "BODEGA":
        return "BODEGA"
    return normalizar_texto(columna)


def parse_porcentaje_cumplimiento(valor: Any) -> float | None:
    if valor is None:
        return None
    if pd is not None:
        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass
    numero = parse_numero(valor)
    texto = normalizar_texto(valor)
    if texto == "":
        return None
    if numero > 1:
        numero = numero / 100.0
    if numero < 0:
        numero = 0.0
    if numero > 1:
        numero = 1.0
    return float(numero)


def normalizar_valor_comparable_conteo(valor: Any) -> str:
    texto = normalizar_texto_seguro(valor)
    if not texto:
        return ""
    etiqueta = normalizar_etiqueta(texto)
    if etiqueta in {"NAN", "NONE", "NULL", "NAT"}:
        return ""
    compacto = etiqueta.replace(" ", "")
    if compacto in {"", "/", "//", "///", "/-/", "-"}:
        return ""
    if re.fullmatch(r"[+-]?\d+(?:[.,]\d+)?", compacto):
        numero = parse_numero(compacto)
        if float(numero).is_integer():
            return str(int(numero))
        return f"{numero:.6f}".rstrip("0").rstrip(".")
    return etiqueta


def normalizar_fecha_comparable_conteo(valor: Any) -> str:
    fecha = parse_fecha(valor)
    if fecha is not None:
        return fecha.strftime("%Y-%m-%d")
    return normalizar_valor_comparable_conteo(valor)


def normalizar_numero_comparable_conteo(valor: Any) -> str:
    texto = normalizar_texto_seguro(valor)
    if not texto:
        return ""
    numero = parse_numero(valor)
    if float(numero).is_integer():
        return str(int(numero))
    return f"{numero:.6f}".rstrip("0").rstrip(".")


def llave_sku_lote_vto_conteo(codigo: Any, lote: Any, vto: Any) -> tuple[str, str, str]:
    return (
        normalizar_valor_comparable_conteo(codigo),
        normalizar_valor_comparable_conteo(lote),
        normalizar_fecha_comparable_conteo(vto),
    )


def calcular_ira_fila_conteo(row: pd.Series) -> float:
    llave_foto = llave_sku_lote_vto_conteo(row.get("Cod. Articulo Foto"), row.get("Lote Foto"), row.get("Vto. Foto"))
    llave_captura = llave_sku_lote_vto_conteo(row.get("Codigo Captura"), row.get("Lote Captura"), row.get("Vto. Captura"))
    if llave_foto != llave_captura or not any(llave_foto) or not any(llave_captura):
        return 0.0

    cantidad_foto = row.get("Cantidad Foto_num", 0.0) or 0.0
    cantidad_captura = row.get("Cantidad Captura_num", 0.0) or 0.0
    if cantidad_foto == 0 and cantidad_captura == 0:
        return 1.0

    denominador = max(cantidad_foto, cantidad_captura)
    if denominador <= 0:
        return 0.0

    diferencia_raw = normalizar_texto_seguro(row.get("Diferencia"))
    diferencia = row.get("Diferencia_num", 0.0) or 0.0
    if not diferencia_raw and (cantidad_foto or cantidad_captura):
        diferencia = cantidad_captura - cantidad_foto

    try:
        valor = 1.0 - (abs(float(diferencia)) / float(denominador))
    except Exception:
        return 0.0
    return max(0.0, min(1.0, float(valor)))


def calcular_ila_fila_conteo(row: pd.Series) -> float:
    foto = (
        normalizar_valor_comparable_conteo(row.get("ubicacion")),
        normalizar_valor_comparable_conteo(row.get("Cod. Articulo Foto")),
        normalizar_numero_comparable_conteo(row.get("Cantidad Foto")),
        normalizar_valor_comparable_conteo(row.get("Pallet Foto")),
        normalizar_valor_comparable_conteo(row.get("Lote Foto")),
        normalizar_fecha_comparable_conteo(row.get("Vto. Foto")),
    )
    captura = (
        normalizar_valor_comparable_conteo(row.get("ubicacion")),
        normalizar_valor_comparable_conteo(row.get("Codigo Captura")),
        normalizar_numero_comparable_conteo(row.get("Cantidad Captura")),
        normalizar_valor_comparable_conteo(row.get("Pallet Captura")),
        normalizar_valor_comparable_conteo(row.get("Lote Captura")),
        normalizar_fecha_comparable_conteo(row.get("Vto. Captura")),
    )
    if not any(foto[1:]) and not any(captura[1:]):
        return 0.0
    return 1.0 if foto == captura else 0.0


def auditoria_promedio_vs_excel(df: pd.DataFrame, columna_recalculada: str, columna_excel: str) -> dict[str, Any]:
    if columna_recalculada not in df.columns or columna_excel not in df.columns:
        return {"disponible": False, "mensaje": "No existen columnas para auditoria de referencia Excel."}

    df_valid = df[df[columna_excel].notna()].copy()
    if df_valid.empty:
        return {"disponible": False, "mensaje": "No existen valores de referencia Excel para auditoria."}

    diferencia_pp = round_safe((df_valid[columna_recalculada].mean() - df_valid[columna_excel].mean()) * 100.0, 2)
    diferencia_abs_pp = round_safe((df_valid[columna_recalculada] - df_valid[columna_excel]).abs().mean() * 100.0, 2)
    return {
        "disponible": True,
        "muestras": int(len(df_valid)),
        "promedio_recalculado_pct": round_safe(df_valid[columna_recalculada].mean() * 100.0, 2),
        "promedio_columna_excel_pct": round_safe(df_valid[columna_excel].mean() * 100.0, 2),
        "diferencia_promedio_pp": diferencia_pp,
        "diferencia_abs_promedio_pp": diferencia_abs_pp,
    }


def cargar_conteos_sheet(path: Path, hoja: str, cd_default: str) -> pd.DataFrame:
    if pd is None:
        raise RuntimeError("pandas no esta disponible para leer conteos cíclicos.")
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if hoja not in wb.sheetnames:
            raise KeyError(f"No existe la hoja {hoja}")
        ws = wb[hoja]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            df = pd.DataFrame()
        else:
            columnas = [normalizar_columna_conteos(c) for c in headers]
            df = pd.DataFrame(list(rows), columns=columnas)
    finally:
        if wb is not None:
            wb.close()

    columnas_faltantes = []
    for columna in [
        "CD", "Usuario", "ubicacion", "Ubic. Contadas", "Nro. Intento", "Empresa", "Empresa Captura",
        "Cod. Articulo Foto", "Codigo Captura", "Cantidad Foto", "Cantidad Captura", "Pallet Foto", "Pallet Captura",
        "Lote Foto", "Lote Captura", "Vto. Foto", "Vto. Captura", "Diferencia", "Validar Sku-lote-Fvcto",
        "Fecha Conteo", "% Cumplimiento IRA", "% Cumplimiento ILA", "Factor de ajuste", "Estado Ubic.",
        "% Cumplimiento ILA", "Factor de ajuste", "Estado Ubic.", "Tipo de Ubic.", "MES", "AÑO", "BODEGA",
    ]:
        if columna not in df.columns:
            df[columna] = None
            columnas_faltantes.append(columna)
    columnas_faltantes = [col for col in columnas_faltantes if col not in {"MES", "AÃ‘O", "BODEGA"}]
    df = df.dropna(how="all")
    df = df[
        [
            "CD", "Usuario", "ubicacion", "Ubic. Contadas", "Nro. Intento", "Empresa", "Empresa Captura",
            "Cod. Articulo Foto", "Codigo Captura", "Cantidad Foto", "Cantidad Captura", "Pallet Foto",
            "Pallet Captura", "Lote Foto", "Lote Captura", "Vto. Foto", "Vto. Captura", "Diferencia",
            "Validar Sku-lote-Fvcto", "Fecha Conteo", "% Cumplimiento IRA", "% Cumplimiento ILA",
            "% Cumplimiento ILA", "Factor de ajuste", "Estado Ubic.", "Tipo de Ubic.", "MES", "AÑO", "BODEGA",
        ]
    ].copy()
    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["CD"] = df["CD"].map(lambda v: normalizar_cd_inventario(v) or cd_default)
    df["Empresa"] = [normalizar_empresa_inventario(emp, cd) for emp, cd in zip(df["Empresa"], df["CD"])]
    df["Empresa Captura"] = df["Empresa Captura"].map(normalizar_texto_seguro)
    df["ubicacion"] = df["ubicacion"].map(normalizar_texto_seguro)
    df["Ubic. Contadas"] = df["Ubic. Contadas"].map(normalizar_texto_seguro)
    df["Usuario"] = df["Usuario"].map(normalizar_texto_seguro)
    df["Estado Ubic."] = df["Estado Ubic."].map(normalizar_texto_seguro)
    df["Tipo de Ubic."] = df["Tipo de Ubic."].map(lambda v: normalizar_locacion_dim(v) or "SIN_TIPO")
    df["Fecha Conteo_dt"] = df["Fecha Conteo"].map(parse_fecha)
    df["Cantidad Foto_num"] = df["Cantidad Foto"].map(parse_numero)
    df["Cantidad Captura_num"] = df["Cantidad Captura"].map(parse_numero)
    df["Diferencia_num"] = df["Diferencia"].map(parse_numero)
    df["IRA_excel_norm"] = df["% Cumplimiento IRA"].map(parse_porcentaje_cumplimiento)
    df["ILA_excel_norm"] = df["% Cumplimiento ILA"].map(parse_porcentaje_cumplimiento)
    df["IRA_calc_norm"] = df.apply(calcular_ira_fila_conteo, axis=1)
    df["ILA_calc_norm"] = df.apply(calcular_ila_fila_conteo, axis=1)
    df["fuente_hoja"] = hoja
    df.attrs["columnas_faltantes"] = columnas_faltantes
    return df


def cargar_conteos_inventario(
    year: int,
    month: int,
    verbose: bool = False,
) -> tuple[dict[str, Any], pd.DataFrame | None, dict[str, Any], list[str]]:
    info = {
        "disponible": False,
        "archivo": None,
        "hojas_usadas": [],
        "fecha_min_conteo": None,
        "fecha_max_conteo": None,
        "ruta_base": None,
        "origen": None,
    }
    archivo, origen = encontrar_archivo_conteos_inventario()
    alertas: list[str] = []
    if not archivo:
        info["mensaje"] = "No se encontró archivo de conteos cíclicos en la ruta oficial de registros de conteos ni en el fallback legado."
        return info, None, {"santa_rosa": {"disponible": False, "mensaje": "No se encontró fuente Santa Rosa en esta fase."}}, alertas

    hojas_plan = [("BD Quilicura", "QUILICURA"), ("BD Pudahuel", "PUDAHUEL")]
    frames = []
    hojas_usadas = []
    hojas_no_encontradas = []
    columnas_faltantes = []
    for hoja, cd_default in hojas_plan:
        try:
            df_hoja = cargar_conteos_sheet(archivo, hoja, cd_default)
            frames.append(df_hoja)
            hojas_usadas.append(hoja)
            columnas_faltantes.extend(df_hoja.attrs.get("columnas_faltantes", []))
            if verbose:
                print(f"[INV-CONTEOS] Hoja leida: {hoja} | filas={len(df_hoja)}")
        except Exception as exc:
            hojas_no_encontradas.append(hoja)
            alertas.append(f"No se pudo leer hoja de conteos {hoja}: {exc}")

    if not frames:
        info.update({
            "archivo": str(archivo),
            "mensaje": "No se pudieron leer hojas WMS de conteos cíclicos.",
            "hojas_usadas": [],
        })
        return info, None, {"santa_rosa": {"disponible": False, "mensaje": "No se encontró fuente Santa Rosa en esta fase."}}, alertas

    df_all = pd.concat(frames, ignore_index=True)
    fechas = df_all["Fecha Conteo_dt"].dropna()
    info.update({
        "disponible": True,
        "archivo": str(archivo),
        "hojas_usadas": hojas_usadas,
        "fecha_min_conteo": serializar_fecha(fechas.min()) if not fechas.empty else None,
        "fecha_max_conteo": serializar_fecha(fechas.max()) if not fechas.empty else None,
        "ruta_base": str(archivo.parent),
        "origen": origen,
    })
    if origen == "fallback_legado":
        alertas.append("Conteos cíclicos usó la ruta legada de inventario porque no se encontró la fuente oficial de registros de conteos.")

    if hojas_no_encontradas:
        alertas.append(f"Hojas de conteos no encontradas o no legibles: {', '.join(hojas_no_encontradas)}.")

    mask_periodo = df_all["Fecha Conteo_dt"].map(lambda d: isinstance(d, datetime) and d.year == year and d.month == month)
    df_periodo = df_all.loc[mask_periodo].copy()
    df_periodo.attrs["columnas_faltantes"] = columnas_faltantes
    df_periodo.attrs["hojas_no_encontradas"] = hojas_no_encontradas
    santa_rosa = {"disponible": False, "mensaje": "No se encontró fuente Santa Rosa en esta fase."}
    return info, df_periodo, {"santa_rosa": santa_rosa}, alertas


def calcular_ira_ponderado_subset(df: pd.DataFrame, value_col: str = "IRA_calc_norm") -> float | None:
    if value_col not in df.columns:
        return None
    df_valid = df[df[value_col].notna()].copy()
    if df_valid.empty:
        return None
    agrupado = (
        df_valid.groupby(["Fecha Conteo_dt", "ubicacion", "Tipo de Ubic."], dropna=False)
        .agg(
            ira_promedio=(value_col, "mean"),
            cantidad_sistema=("Cantidad Foto_num", "sum"),
        )
        .reset_index()
    )
    total = agrupado["cantidad_sistema"].sum()
    if total <= 0:
        return None
    valor = float((agrupado["ira_promedio"] * agrupado["cantidad_sistema"]).sum() / total)
    return round_safe(valor * 100.0, 2)


def calcular_ila_ponderado_subset(df: pd.DataFrame, value_col: str = "ILA_calc_norm") -> float | None:
    if value_col not in df.columns:
        return None
    df_valid = df[df["Ubic. Contadas"].map(normalizar_texto_seguro).ne("")].copy()
    df_valid = df_valid[df_valid[value_col].notna()].copy()
    if df_valid.empty:
        return None
    agrupado = (
        df_valid.groupby(["CD", "Fecha Conteo_dt", "Tipo de Ubic.", "Ubic. Contadas"], dropna=False)
        .agg(
            ila_ubicacion=(value_col, "min"),
        )
        .reset_index()
    )
    total = len(agrupado)
    if not total:
        return None
    valor = float(agrupado["ila_ubicacion"].sum() / total)
    return round_safe(valor * 100.0, 2)


def calcular_conteos_ciclicos(
    df_conteos: pd.DataFrame | None,
    fuente_info: dict[str, Any],
    year: int,
    month: int,
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any], list[str]]:
    periodo = {"anio": year, "mes": month}
    if df_conteos is None:
        mensaje = fuente_info.get("mensaje", "No se encontró fuente de conteos cíclicos.")
        return (
            {"disponible": False, "periodo": periodo, "fuente": fuente_info, "mensaje": mensaje},
            {
                "disponible": False,
                "periodo": periodo,
                "metodo": "replica_powerbi_desde_columnas_base",
                "requiere_validacion_powerbi": True,
                "mensaje": mensaje,
                "wms": {"disponible": False, "mensaje": mensaje},
                "santa_rosa": {"disponible": False, "mensaje": "No se encontró fuente Santa Rosa en esta fase."},
            },
            {
                "disponible": False,
                "periodo": periodo,
                "mensaje": mensaje,
                "avance_vs_stock_inicial": {"disponible": False, "mensaje": "No calculado en esta fase."},
            },
            [],
        )

    if df_conteos.empty:
        mensaje = "No se encontraron registros de conteo cíclico para el período solicitado."
        return (
            {
                "disponible": False,
                "periodo": periodo,
                "fuente": fuente_info,
                "mensaje": mensaje,
                "total_registros": 0,
                "ubicaciones_unicas_contadas": 0,
                "por_cd": [],
                "por_cliente": [],
                "por_tipo_ubicacion": [],
                "diferencias": {
                    "registros_con_diferencia": 0,
                    "registros_sin_diferencia": 0,
                    "diferencia_neta": 0.0,
                    "diferencia_absoluta": 0.0,
                },
                "alertas": [],
            },
            {
                "disponible": False,
                "periodo": periodo,
                "metodo": "replica_powerbi_desde_columnas_base",
                "requiere_validacion_powerbi": True,
                "mensaje": mensaje,
                "wms": {"disponible": False, "mensaje": mensaje, "alertas": []},
                "santa_rosa": {"disponible": False, "mensaje": "No se encontró fuente Santa Rosa en esta fase."},
            },
            {
                "disponible": False,
                "periodo": periodo,
                "mensaje": mensaje,
                "avance_vs_stock_inicial": {"disponible": False, "mensaje": "No calculado en esta fase."},
                "alertas": [],
            },
            [],
        )

    alertas_contadores = {
        "registros_sin_fecha_conteo": int(df_conteos["Fecha Conteo_dt"].isna().sum()),
        "registros_sin_ubicacion": int(df_conteos["ubicacion"].eq("").sum()),
        "registros_sin_ubicacion_contada": int(df_conteos["Ubic. Contadas"].eq("").sum()),
        "registros_sin_cliente": int(df_conteos["Empresa"].eq("").sum()),
        "registros_sin_cd": int(df_conteos["CD"].eq("").sum()),
        "registros_sin_cantidad_foto": int(df_conteos["Cantidad Foto"].map(normalizar_texto_seguro).eq("").sum()),
        "registros_sin_cumplimiento_ira": int(df_conteos["IRA_excel_norm"].isna().sum()),
        "registros_sin_cumplimiento_ila": int(df_conteos["ILA_excel_norm"].isna().sum()),
        "registros_con_cantidad_foto_negativa": int((df_conteos["Cantidad Foto_num"] < 0).sum()),
        "registros_con_diferencia_no_numerica": int((
            (df_conteos["Diferencia"].map(normalizar_texto_seguro).ne(""))
            & (df_conteos["Diferencia_num"].eq(0))
            & (~df_conteos["Diferencia"].astype(str).str.contains("0", na=False))
        ).sum()),
    }
    hojas_no_encontradas = df_conteos.attrs.get("hojas_no_encontradas", [])
    if hojas_no_encontradas:
        alertas_contadores["hojas_no_encontradas"] = len(hojas_no_encontradas)
    columnas_faltantes = df_conteos.attrs.get("columnas_faltantes", [])
    if columnas_faltantes:
        alertas_contadores["columnas_faltantes"] = len(columnas_faltantes)

    total_registros = int(len(df_conteos))
    ubicaciones_unicas_total = int(df_conteos.loc[df_conteos["Ubic. Contadas"] != "", "Ubic. Contadas"].nunique())
    registros_con_diferencia = int(df_conteos["Diferencia_num"].ne(0).sum())
    registros_sin_diferencia = int(df_conteos["Diferencia_num"].eq(0).sum())
    diferencia_neta = round_safe(df_conteos["Diferencia_num"].sum(), 2)
    diferencia_absoluta = round_safe(df_conteos["Diferencia_num"].abs().sum(), 2)

    por_cd = []
    for cd, grupo in df_conteos.groupby("CD", dropna=False):
        por_cd.append(
            {
                "cd": cd,
                "registros": int(len(grupo)),
                "ubicaciones_unicas": int(grupo.loc[grupo["Ubic. Contadas"] != "", "Ubic. Contadas"].nunique()),
                "clientes": int(grupo.loc[grupo["Empresa"] != "", "Empresa"].nunique()),
                "registros_con_diferencia": int(grupo["Diferencia_num"].ne(0).sum()),
                "diferencia_neta": round_safe(grupo["Diferencia_num"].sum(), 2),
                "diferencia_absoluta": round_safe(grupo["Diferencia_num"].abs().sum(), 2),
            }
        )
    por_cd.sort(key=lambda item: item["cd"])

    por_cliente = []
    for (cliente, cd), grupo in df_conteos.groupby(["Empresa", "CD"], dropna=False):
        por_cliente.append(
            {
                "cliente": cliente,
                "cd": cd,
                "registros": int(len(grupo)),
                "ubicaciones_unicas": int(grupo.loc[grupo["Ubic. Contadas"] != "", "Ubic. Contadas"].nunique()),
                "ira_ponderado_pct": calcular_ira_ponderado_subset(grupo),
                "ila_ponderado_pct": calcular_ila_ponderado_subset(grupo),
                "registros_con_diferencia": int(grupo["Diferencia_num"].ne(0).sum()),
                "diferencia_neta": round_safe(grupo["Diferencia_num"].sum(), 2),
            }
        )
    por_cliente.sort(key=lambda item: (-item["registros"], item["cliente"], item["cd"]))

    por_tipo = []
    for (cd, tipo), grupo in df_conteos.groupby(["CD", "Tipo de Ubic."], dropna=False):
        por_tipo.append(
            {
                "cd": cd,
                "tipo_ubicacion": tipo,
                "registros": int(len(grupo)),
                "ubicaciones_unicas": int(grupo.loc[grupo["Ubic. Contadas"] != "", "Ubic. Contadas"].nunique()),
                "ira_ponderado_pct": calcular_ira_ponderado_subset(grupo),
                "ila_ponderado_pct": calcular_ila_ponderado_subset(grupo),
            }
        )
    por_tipo.sort(key=lambda item: (item["cd"], -item["registros"], item["tipo_ubicacion"]))

    ira_total = calcular_ira_ponderado_subset(df_conteos)
    ila_total = calcular_ila_ponderado_subset(df_conteos)
    auditoria_ira = auditoria_promedio_vs_excel(df_conteos, "IRA_calc_norm", "IRA_excel_norm")
    auditoria_ila = auditoria_promedio_vs_excel(df_conteos, "ILA_calc_norm", "ILA_excel_norm")
    metodo = "replica_powerbi_desde_columnas_base"
    requiere_validacion = True

    conteos_payload = {
        "disponible": True,
        "periodo": periodo,
        "fuente": fuente_info,
        "total_registros": total_registros,
        "ubicaciones_unicas_contadas": ubicaciones_unicas_total,
        "por_cd": por_cd,
        "por_cliente": por_cliente,
        "por_tipo_ubicacion": por_tipo,
        "diferencias": {
            "registros_con_diferencia": registros_con_diferencia,
            "registros_sin_diferencia": registros_sin_diferencia,
            "diferencia_neta": diferencia_neta,
            "diferencia_absoluta": diferencia_absoluta,
        },
        "alertas": resumir_alertas_contadores(alertas_contadores),
    }

    ira_ila_payload = {
        "disponible": True,
        "periodo": periodo,
        "metodo": metodo,
        "requiere_validacion_powerbi": requiere_validacion,
        "nota": "IRA e ILA se recalculan desde columnas base. Las columnas calculadas del Excel se usan solo como referencia/auditoria.",
        "wms": {
            "disponible": True,
            "ira_ponderado_pct": ira_total,
            "ila_ponderado_pct": ila_total,
            "por_cd": [
                {
                    "cd": item["cd"],
                    "ira_ponderado_pct": calcular_ira_ponderado_subset(df_conteos[df_conteos["CD"].eq(item["cd"])]),
                    "ila_ponderado_pct": calcular_ila_ponderado_subset(df_conteos[df_conteos["CD"].eq(item["cd"])]),
                }
                for item in por_cd
            ],
            "por_tipo_ubicacion": por_tipo,
            "auditoria_excel_referencia": {
                "ira": auditoria_ira,
                "ila": auditoria_ila,
            },
            "alertas": resumir_alertas_contadores(alertas_contadores),
        },
        "santa_rosa": {"disponible": False, "mensaje": "No se encontró fuente Santa Rosa en esta fase."},
    }

    avance_payload = {
        "disponible": False,
        "periodo": periodo,
        "ubicaciones_contadas_periodo": ubicaciones_unicas_total,
        "total_ubicaciones_layout": 0.0,
        "avance_pct": None,
        "por_cd": [],
        "avance_vs_stock_inicial": {"disponible": False, "mensaje": "No calculado en esta fase."},
        "alertas": resumir_alertas_contadores(alertas_contadores),
    }

    mensajes = []
    if hojas_no_encontradas:
        mensajes.append(f"Conteos cíclicos: faltan hojas esperadas {', '.join(hojas_no_encontradas)}.")
    return conteos_payload, ira_ila_payload, avance_payload, mensajes


def calcular_avance_conteo(
    df_conteos: pd.DataFrame | None,
    df_layout: pd.DataFrame | None,
    year: int,
    month: int,
) -> tuple[dict[str, Any], list[str]]:
    periodo = {"anio": year, "mes": month}
    objetivo_mes = obtener_objetivo_mensual_conteo(year, month)
    avance_vs_objetivo_mes = {
        "disponible": objetivo_mes is not None and objetivo_mes > 0,
        "objetivo_mes_ubicaciones": int(objetivo_mes) if objetivo_mes is not None else 0,
        "ubicaciones_contadas_mes": 0,
        "avance_pct": None,
    }
    if df_conteos is None:
        return {
            "disponible": False,
            "periodo": periodo,
            "mensaje": "No se encontró fuente de conteos para calcular avance.",
            "avance_vs_objetivo_mes": avance_vs_objetivo_mes,
            "avance_vs_layout": {"disponible": False, "mensaje": "No calculado en esta fase."},
            "avance_vs_stock_inicial": {"disponible": False, "mensaje": "No calculado en esta fase."},
            "alertas": [],
        }, []
    if df_conteos.empty:
        return {
            "disponible": False,
            "periodo": periodo,
            "mensaje": "No se encontraron registros de conteo cíclico para el período solicitado.",
            "avance_vs_objetivo_mes": avance_vs_objetivo_mes,
            "avance_vs_layout": {"disponible": False, "mensaje": "No calculado en esta fase."},
            "avance_vs_stock_inicial": {"disponible": False, "mensaje": "No calculado en esta fase."},
            "alertas": [],
        }, []
    if df_layout is None or df_layout.empty:
        return {
            "disponible": False,
            "periodo": periodo,
            "mensaje": "No se encontró layout de ubicaciones para avance de conteo.",
            "avance_vs_objetivo_mes": avance_vs_objetivo_mes,
            "avance_vs_layout": {"disponible": False, "mensaje": "No se encontró layout de ubicaciones para avance de conteo."},
            "avance_vs_stock_inicial": {"disponible": False, "mensaje": "No calculado en esta fase."},
            "alertas": [],
        }, []

    layout_base = df_layout[
        ~df_layout["Locacion"].map(normalizar_etiqueta).isin(LOCACIONES_OCUPACION_EXCLUIDAS)
    ].copy()
    total_layout = round_safe(layout_base["Profundidad_num"].sum(), 2)
    ubicaciones_contadas = int(df_conteos.loc[df_conteos["Ubic. Contadas"] != "", "Ubic. Contadas"].nunique())
    avance_vs_objetivo_mes["ubicaciones_contadas_mes"] = ubicaciones_contadas
    if avance_vs_objetivo_mes["disponible"]:
        avance_vs_objetivo_mes["avance_pct"] = porcentaje_safe(
            ubicaciones_contadas,
            avance_vs_objetivo_mes["objetivo_mes_ubicaciones"],
        )
    por_cd = []
    for cd, grupo in df_conteos.groupby("CD", dropna=False):
        total_cd = round_safe(layout_base.loc[layout_base["CD"].eq(cd), "Profundidad_num"].sum(), 2)
        ubic_cd = int(grupo.loc[grupo["Ubic. Contadas"] != "", "Ubic. Contadas"].nunique())
        por_cd.append(
            {
                "cd": cd,
                "ubicaciones_contadas": ubic_cd,
                "total_ubicaciones_layout": total_cd,
                "avance_pct": porcentaje_safe(ubic_cd, total_cd),
            }
        )
    por_cd.sort(key=lambda item: item["cd"])
    avance_vs_layout = {
        "disponible": True,
        "total_ubicaciones_layout": total_layout,
        "ubicaciones_contadas_periodo": ubicaciones_contadas,
        "avance_pct": porcentaje_safe(ubicaciones_contadas, total_layout),
    }

    payload = {
        "disponible": True,
        "periodo": {"anio": year, "mes": month},
        "ubicaciones_contadas_periodo": ubicaciones_contadas,
        "total_ubicaciones_layout": total_layout,
        "avance_pct": (
            avance_vs_objetivo_mes["avance_pct"]
            if avance_vs_objetivo_mes["disponible"]
            else avance_vs_layout["avance_pct"]
        ),
        "por_cd": por_cd,
        "avance_vs_objetivo_mes": avance_vs_objetivo_mes,
        "avance_vs_layout": avance_vs_layout,
        "avance_vs_stock_inicial": {"disponible": False, "mensaje": "No calculado en esta fase."},
        "alertas": [],
    }
    return payload, []
def cargar_posiciones_excel(fuente: FuenteDetectada) -> tuple[pd.DataFrame, dict[str, Any]]:
    if pd is None:
        raise RuntimeError("pandas no esta disponible para leer posiciones.")
    df = pd.read_excel(fuente.ruta, sheet_name=0, header=3, engine="openpyxl")
    df.columns = [normalizar_texto(c) for c in df.columns]
    for columna in [
        "CD", "Area", "Cara", "Columna", "Nivel", "Estado", "Prof. Totales", "Prof. Ocupadas",
        "Prof. Libres", "Inhibida", "Mezcla",
    ]:
        if columna not in df.columns:
            df[columna] = None
    df = df.dropna(how="all")
    df["CD"] = df["CD"].map(normalizar_cd_inventario)
    df["Estado"] = df["Estado"].map(normalizar_texto)
    df["Inhibida"] = df["Inhibida"].map(normalizar_texto)
    df["Tipo_Archivo"] = fuente.ruta.stem
    df["Ubicacion"] = [
        construir_ubicacion_desde_componentes(cara, columna, nivel)
        for cara, columna, nivel in zip(df["Cara"], df["Columna"], df["Nivel"])
    ]
    info = {
        "path": str(fuente.ruta),
        "name": fuente.ruta.name,
        "modified": iso_mtime(fuente.ruta),
        "rows": int(len(df)),
    }
    return df, info
def resumir_alertas_contadores(contadores: dict[str, int]) -> list[dict[str, Any]]:
    return [
        {"tipo": clave, "cantidad": int(valor)}
        for clave, valor in contadores.items()
    ]


def obtener_objetivo_mensual_conteo(year: int, month: int) -> int | None:
    return OBJETIVOS_MENSUALES_CONTEO.get((year, month))


def calcular_stock_desde_wms(
    df_stock: pd.DataFrame,
    fecha_referencia: str | None,
    descartes: dict[str, int] | None = None,
) -> tuple[dict[str, Any], list[str]]:
    if df_stock is None or df_stock.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontraron datos de stock WMS.",
        }, []

    descartes = descartes or {}
    alertas_contadores = {
        "registros_sin_cliente": int(df_stock["Empresa"].eq("").sum()),
        "registros_sin_articulo": int(df_stock["Articulo"].eq("").sum()),
        "registros_sin_ubicacion": int(df_stock["Ubicacion"].eq("").sum()),
        "registros_con_unidades_negativas": int((df_stock["Unids_num"] < 0).sum()),
        "registros_sin_contenedor": int(df_stock["Contenedor"].eq("").sum()),
        "filas_descartadas_no_operativas": int(descartes.get("filas_descartadas_no_operativas", 0)),
        "filas_descartadas_totales_excel": int(descartes.get("filas_descartadas_totales_excel", 0)),
    }

    por_cd = []
    for cd, grupo in df_stock.groupby("CD", dropna=False):
        por_cd.append(
            {
                "cd": cd,
                "unidades": round_safe(grupo["Unids_num"].sum(), 2),
                "plts": int(grupo.loc[grupo["Contenedor"] != "", "Contenedor"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
                "ubicaciones": int(grupo.loc[grupo["Ubicacion"] != "", "Ubicacion"].nunique()),
            }
        )

    por_cliente = []
    for (cliente, cd), grupo in df_stock.groupby(["Empresa", "CD"], dropna=False):
        por_cliente.append(
            {
                "cliente": cliente,
                "cd": cd,
                "unidades": round_safe(grupo["Unids_num"].sum(), 2),
                "plts": int(grupo.loc[grupo["Contenedor"] != "", "Contenedor"].nunique()),
                "skus": int(grupo.loc[grupo["Articulo"] != "", "Articulo"].nunique()),
                "ubicaciones": int(grupo.loc[grupo["Ubicacion"] != "", "Ubicacion"].nunique()),
            }
        )
    por_cliente.sort(key=lambda item: (-item["unidades"], item["cliente"], item["cd"]))

    top_skus = []
    agrupado_sku = (
        df_stock.groupby(["Empresa", "CD", "Articulo", "Descripcion"], dropna=False)
        .agg(
            unidades=("Unids_num", "sum"),
            plts=("Contenedor", lambda s: int(pd.Series([v for v in s if normalizar_texto(v)]).nunique())),
        )
        .reset_index()
    )
    agrupado_sku = agrupado_sku.sort_values(["unidades", "Empresa", "CD", "Articulo"], ascending=[False, True, True, True])
    for _, row in agrupado_sku.head(10).iterrows():
        top_skus.append(
            {
                "cliente": row["Empresa"],
                "cd": row["CD"],
                "articulo": row["Articulo"],
                "descripcion": row["Descripcion"],
                "unidades": round_safe(row["unidades"], 2),
                "plts": int(row["plts"]),
            }
        )

    alertas = resumir_alertas_contadores(alertas_contadores)
    mensajes = []
    if alertas_contadores["registros_sin_ubicacion"]:
        mensajes.append("Stock WMS contiene registros sin ubicacion derivable.")
    if alertas_contadores["registros_con_unidades_negativas"]:
        mensajes.append("Stock WMS contiene unidades negativas que requieren revision.")
    if alertas_contadores["filas_descartadas_totales_excel"]:
        mensajes.append("Stock WMS descarto filas totales o de resumen del Excel antes de calcular KPIs.")

    stock_payload = {
        "disponible": True,
        "fecha_referencia": fecha_referencia,
        "total_unidades": round_safe(df_stock["Unids_num"].sum(), 2),
        "total_plts": int(df_stock.loc[df_stock["Contenedor"] != "", "Contenedor"].nunique()),
        "total_skus": int(df_stock.loc[df_stock["Articulo"] != "", "Articulo"].nunique()),
        "por_cd": sorted(por_cd, key=lambda item: item["cd"]),
        "por_cliente": por_cliente,
        "top_skus": top_skus,
        "alertas": alertas,
    }
    return stock_payload, mensajes
def calcular_stock_bloqueado_wms(
    df_stock: pd.DataFrame,
    descartes: dict[str, int] | None = None,
) -> tuple[dict[str, Any], list[str]]:
    if df_stock is None or df_stock.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontraron datos de stock WMS para bloqueos.",
        }, []

    descartes = descartes or {}
    apto_norm = df_stock["Apto"].map(normalizar_etiqueta)
    motivo_norm = df_stock["Motivo"].map(normalizar_etiqueta)
    mask = (apto_norm != "SI") | (motivo_norm != "A DISPONIBLE")
    df_bloqueado = df_stock.loc[mask].copy()

    if df_bloqueado.empty:
        return {
            "disponible": True,
            "kpi_principal": "operativo_powerbi",
            "unidades": 0.0,
            "plts": 0,
            "skus": 0,
            "por_cliente": [],
            "por_motivo": [],
            "total_fuente": {
                "unidades": 0.0,
                "plts": 0,
                "skus": 0,
                "por_cliente": [],
                "por_motivo": [],
            },
            "operativo_powerbi": {
                "disponible": True,
                "criterio": "Excluye Santa Rosa por tratamiento especial y para calzar con dashboard ejecutivo Power BI.",
                "clientes_excluidos": ["SANTA ROSA"],
                "unidades": 0.0,
                "plts": 0,
                "skus": 0,
                "por_cliente": [],
                "por_motivo": [],
            },
            "santa_rosa": {
                "disponible": False,
                "criterio": "Cliente con tratamiento especial.",
                "unidades": 0.0,
                "plts": 0,
                "skus": 0,
                "por_cliente": [],
                "por_motivo": [],
            },
            "alertas": resumir_alertas_contadores({
                "motivos_vacios": 0,
                "apto_vacio": 0,
                "bloqueos_sin_motivo": 0,
                "filas_descartadas_no_operativas": int(descartes.get("filas_descartadas_no_operativas", 0)),
                "filas_descartadas_totales_excel": int(descartes.get("filas_descartadas_totales_excel", 0)),
            }),
        }, []

    df_bloqueado["grupo_motivo"] = df_bloqueado["Motivo"].map(grupo_motivo_bloqueo_wms)
    mask_santa_rosa = df_bloqueado["Empresa"].map(es_cliente_santa_rosa)
    df_santa_rosa = df_bloqueado.loc[mask_santa_rosa].copy()
    df_operativo = df_bloqueado.loc[~mask_santa_rosa].copy()
    alertas_contadores = {
        "motivos_vacios": int(df_bloqueado["Motivo"].eq("").sum()),
        "apto_vacio": int(df_bloqueado["Apto"].eq("").sum()),
        "bloqueos_sin_motivo": int(df_bloqueado[df_bloqueado["Motivo"].eq("") & df_bloqueado["Apto"].map(normalizar_etiqueta).ne("SI")].shape[0]),
        "filas_descartadas_no_operativas": int(descartes.get("filas_descartadas_no_operativas", 0)),
        "filas_descartadas_totales_excel": int(descartes.get("filas_descartadas_totales_excel", 0)),
    }
    total_fuente = resumir_stock_bloqueado_subset(df_bloqueado)
    operativo_powerbi = resumir_stock_bloqueado_subset(df_operativo)
    santa_rosa = resumir_stock_bloqueado_subset(df_santa_rosa)

    mensajes = []
    if alertas_contadores["bloqueos_sin_motivo"]:
        mensajes.append("Hay bloqueos WMS sin motivo informado.")
    if alertas_contadores["filas_descartadas_totales_excel"]:
        mensajes.append("Stock bloqueado WMS excluyo filas totales o de resumen del Excel.")
    if not df_santa_rosa.empty:
        mensajes.append("Stock bloqueado WMS separo Santa Rosa del KPI ejecutivo por tratamiento especial.")

    payload = {
        "disponible": True,
        "kpi_principal": "operativo_powerbi",
        "unidades": operativo_powerbi["unidades"],
        "plts": operativo_powerbi["plts"],
        "skus": operativo_powerbi["skus"],
        "por_cliente": operativo_powerbi["por_cliente"],
        "por_motivo": operativo_powerbi["por_motivo"],
        "total_fuente": total_fuente,
        "operativo_powerbi": {
            "disponible": True,
            "criterio": "Excluye Santa Rosa por tratamiento especial y para calzar con dashboard ejecutivo Power BI.",
            "clientes_excluidos": ["SANTA ROSA"],
            **operativo_powerbi,
        },
        "santa_rosa": {
            "disponible": not df_santa_rosa.empty,
            "criterio": "Cliente con tratamiento especial.",
            **santa_rosa,
        },
        "alertas": resumir_alertas_contadores(alertas_contadores),
    }
    return payload, mensajes
def calcular_ocupacion(
    df_layout: pd.DataFrame | None,
    df_posiciones: pd.DataFrame | None,
    fecha_referencia: str | None,
) -> tuple[dict[str, Any], list[str]]:
    if df_layout is None or df_layout.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontro layout de ubicaciones para ocupacion.",
        }, []
    if df_posiciones is None or df_posiciones.empty:
        return {
            "disponible": False,
            "mensaje": "No se encontraron datos de posiciones para ocupacion.",
        }, []

    df_layout = df_layout.copy()
    df_posiciones = df_posiciones.copy()
    df_posiciones["Ubicacion"] = df_posiciones["Ubicacion"].map(normalizar_texto)
    df_posiciones["Estado"] = df_posiciones["Estado"].map(normalizar_texto)
    df_posiciones["Inhibida"] = df_posiciones["Inhibida"].map(normalizar_texto)

    pos_group = (
        df_posiciones.groupby("Ubicacion", dropna=False)
        .agg(
            estado_dinamico=("Estado", lambda s: next((normalizar_texto(v) for v in s if normalizar_texto(v)), "")),
            inhibida=("Inhibida", lambda s: next((normalizar_texto(v) for v in s if normalizar_texto(v)), "")),
        )
        .reset_index()
    )

    df = df_layout.merge(pos_group, on="Ubicacion", how="left")
    df["estado_dinamico"] = df["estado_dinamico"].fillna("").map(normalizar_texto)
    df["inhibida"] = df["inhibida"].fillna("").map(normalizar_texto)
    df["locacion_norm"] = df["Locacion"].map(normalizar_etiqueta)
    df["grupo_estado"] = df["estado_dinamico"].map(normalizar_etiqueta).replace(
        {"COMPLETO": "Ocupado", "OCUPADO": "Ocupado", "LIBRE": "Libre", "": "SD"}
    )
    df.loc[~df["grupo_estado"].isin({"Ocupado", "Libre"}), "grupo_estado"] = "SD"
    df["estado_operativo"] = "Activo"
    df.loc[df["grupo_estado"].eq("SD") | df["inhibida"].map(normalizar_etiqueta).eq("SI"), "estado_operativo"] = "Inactivo"

    layout_base = df[
        ~df["locacion_norm"].isin(LOCACIONES_OCUPACION_EXCLUIDAS)
        & df["inhibida"].map(normalizar_etiqueta).ne("SI")
    ].copy()
    layout_operativo = layout_base[
        layout_base["grupo_estado"].ne("SD") & layout_base["estado_operativo"].eq("Activo")
    ].copy()

    def resumen_ocupacion(grupo: pd.DataFrame) -> dict[str, float]:
        total = grupo["Profundidad_num"].sum()
        ocupadas = grupo.loc[grupo["grupo_estado"].eq("Ocupado"), "Profundidad_num"].sum()
        libres = grupo.loc[grupo["grupo_estado"].eq("Libre"), "Profundidad_num"].sum()
        sd = grupo.loc[grupo["grupo_estado"].eq("SD"), "Profundidad_num"].sum()
        return {
            "total": round_safe(total, 2),
            "ocupadas": round_safe(ocupadas, 2),
            "libres": round_safe(libres, 2),
            "sd": round_safe(sd, 2),
            "ocupacion_pct": porcentaje_safe(ocupadas, total),
        }

    base_total = resumen_ocupacion(layout_base)
    operativo_total = resumen_ocupacion(layout_operativo)

    por_cd = []
    for cd, grupo in layout_base.groupby("CD", dropna=False):
        base_cd = resumen_ocupacion(grupo)
        operativo_cd = resumen_ocupacion(layout_operativo.loc[layout_operativo["CD"].eq(cd)])
        por_cd.append(
            {
                "cd": cd,
                "total_ubicaciones_layout": base_cd["total"],
                "ocupadas": base_cd["ocupadas"],
                "libres": base_cd["libres"],
                "sd": base_cd["sd"],
                "ocupacion_pct": operativo_cd["ocupacion_pct"],
                "ocupacion_tecnica_pct": base_cd["ocupacion_pct"],
                "ocupacion_operativa_pct": operativo_cd["ocupacion_pct"],
            }
        )
    por_cd.sort(key=lambda item: item["cd"])

    por_locacion = []
    for (cd, locacion), grupo in layout_base.groupby(["CD", "Locacion"], dropna=False):
        resumen = resumen_ocupacion(grupo)
        por_locacion.append({"cd": cd, "locacion": locacion, **resumen})
    por_locacion.sort(key=lambda item: (item["cd"], -item["ocupadas"], item["locacion"]))

    por_tipo = []
    for (cd, tipo), grupo in layout_base.groupby(["CD", "Tipo"], dropna=False):
        resumen = resumen_ocupacion(grupo)
        por_tipo.append({"cd": cd, "tipo": tipo, **resumen})
    por_tipo.sort(key=lambda item: (item["cd"], -item["ocupadas"], item["tipo"]))

    pos_ubicaciones = set(df_posiciones.loc[df_posiciones["Ubicacion"] != "", "Ubicacion"].tolist())
    layout_ubicaciones = set(df_layout.loc[df_layout["Ubicacion"] != "", "Ubicacion"].tolist())
    alertas_contadores = {
        "ubicaciones_layout_sin_estado": int(df["estado_dinamico"].eq("").sum()),
        "ubicaciones_posiciones_no_en_layout": int(len(pos_ubicaciones - layout_ubicaciones)),
        "ubicaciones_inhibidas_excluidas": int(df["inhibida"].map(normalizar_etiqueta).eq("SI").sum()),
    }
    mensajes = []
    if alertas_contadores["ubicaciones_posiciones_no_en_layout"]:
        mensajes.append("Existen ubicaciones en posiciones que no aparecen en el layout base.")

    payload = {
        "disponible": True,
        "fecha_referencia": fecha_referencia,
        "kpi_principal": "ocupacion_operativa",
        "sd_definicion": "Ubicaciones existentes en layout/base, pero eliminadas del WMS actual por retiro fisico o desarme de racks.",
        "total_ubicaciones_layout": base_total["total"],
        "ocupadas": base_total["ocupadas"],
        "libres": base_total["libres"],
        "sd": base_total["sd"],
        "ocupacion_pct": operativo_total["ocupacion_pct"],
        "ocupacion_tecnica": {
            "total": base_total["total"],
            "ocupadas": base_total["ocupadas"],
            "ocupacion_pct": base_total["ocupacion_pct"],
            "uso": "referencial_layout_historico",
        },
        "ocupacion_operativa": {
            "total": operativo_total["total"],
            "ocupadas": operativo_total["ocupadas"],
            "ocupacion_pct": operativo_total["ocupacion_pct"],
            "requiere_validacion": True,
            "uso": "kpi_principal_operativo",
            "nota": "La ocupacion operativa excluye ubicaciones SD/inactivas y debe validarse contra criterio operacional.",
        },
        "por_cd": por_cd,
        "por_locacion": por_locacion,
        "por_tipo_ubicacion": por_tipo,
        "alertas": resumir_alertas_contadores(alertas_contadores),
        "observaciones": [
            "ubicaciones_layout_sin_estado representa ubicaciones SD o fuera del WMS actual y debe revisarse como reconciliacion entre layout historico y WMS."
        ],
    }
    return payload, mensajes
def calcular_inventario_inicial(year: int, month: int, verbose: bool = False) -> tuple[dict[str, Any], list[str], list[str]]:
    if pd is None:
        return {
            "disponible": False,
            "mensaje": "pandas no esta disponible en este entorno para inventario.",
        }, [], []

    stock_fuentes = descubrir_fuentes_stock_wms(verbose=verbose)
    staging_fuentes = descubrir_fuentes_staging(verbose=verbose)
    staging_snapshot_fuentes = seleccionar_ultimas_fuentes_staging(staging_fuentes)
    posiciones_fuentes = descubrir_fuentes_posiciones(verbose=verbose)
    dim_info, df_layout = cargar_dim_layout_inventario(verbose=verbose)
    conteos_fuente_info, df_conteos_periodo, conteos_extra, alertas_conteos_fuente = cargar_conteos_inventario(
        year=year,
        month=month,
        verbose=verbose,
    )

    fuentes_payload = {
        "stock": {
            "disponible": bool(stock_fuentes),
            "ruta": str(STOCK_WMS_ROOT),
            "archivos_usados": [str(f.ruta) for f in stock_fuentes],
            "fecha_referencia": None,
        },
        "staging": {
            "disponible": bool(staging_fuentes),
            "ruta": str(STAGING_ROOT),
            "archivos_usados": [str(f.ruta) for f in staging_snapshot_fuentes],
            "fecha_referencia": None,
            "historico_desde": None,
            "historico_hasta": None,
            "archivos_historicos_procesados": len(staging_fuentes),
        },
        "posiciones": {
            "disponible": bool(posiciones_fuentes),
            "ruta": str(POSICIONES_ROOT),
            "archivos_usados": [str(f.ruta) for f in posiciones_fuentes],
            "fecha_referencia": None,
        },
        "dim_ubicaciones": dim_info,
        "conteos": conteos_fuente_info,
    }

    stock_frames = []
    stock_archivos = []
    stock_fechas = []
    stock_descartes = {
        "filas_descartadas_no_operativas": 0,
        "filas_descartadas_totales_excel": 0,
    }
    for fuente in stock_fuentes:
        try:
            df_fuente, info_fuente = cargar_stock_wms_excel(fuente)
            stock_frames.append(df_fuente)
            stock_archivos.append(info_fuente)
            stock_descartes["filas_descartadas_no_operativas"] += int(info_fuente.get("filas_descartadas_no_operativas", 0))
            stock_descartes["filas_descartadas_totales_excel"] += int(info_fuente.get("filas_descartadas_totales_excel", 0))
            fecha = parse_fecha_desde_nombre_archivo(fuente.ruta)
            if fecha:
                stock_fechas.append(fecha)
        except Exception as exc:
            if verbose:
                print(f"[INV-STOCK] Error leyendo {fuente.ruta.name}: {exc}")

    staging_frames = []
    staging_archivos = []
    staging_fechas = []
    for fuente in staging_snapshot_fuentes:
        try:
            df_fuente, info_fuente = cargar_staging_csv(fuente)
            if not df_fuente.empty:
                staging_frames.append(df_fuente)
            staging_archivos.append(info_fuente)
            staging_fechas.append(df_fuente["fecha_descarga"].max() if not df_fuente.empty else (parse_fecha_staging_desde_nombre_archivo(fuente.ruta) or datetime.fromtimestamp(fuente.ruta.stat().st_mtime)))
        except Exception as exc:
            if verbose:
                print(f"[INV-STG] Error leyendo {fuente.ruta.name}: {exc}")

    posiciones_frames = []
    posiciones_archivos = []
    posiciones_fechas = []
    for fuente in posiciones_fuentes:
        try:
            df_fuente, info_fuente = cargar_posiciones_excel(fuente)
            posiciones_frames.append(df_fuente)
            posiciones_archivos.append(info_fuente)
            posiciones_fechas.append(datetime.fromtimestamp(fuente.ruta.stat().st_mtime))
        except Exception as exc:
            if verbose:
                print(f"[INV-POS] Error leyendo {fuente.ruta.name}: {exc}")

    fuentes_payload["stock"]["archivos_usados"] = stock_archivos
    fuentes_payload["staging"]["archivos_usados"] = staging_archivos
    fuentes_payload["posiciones"]["archivos_usados"] = posiciones_archivos
    fuentes_payload["stock"]["fecha_referencia"] = serializar_o_mtime_fecha(max(stock_fechas) if stock_fechas else None)
    fuentes_payload["staging"]["fecha_referencia"] = serializar_o_mtime_fecha(max(staging_fechas) if staging_fechas else None)
    fuentes_payload["posiciones"]["fecha_referencia"] = serializar_o_mtime_fecha(max(posiciones_fechas) if posiciones_fechas else None)
    if staging_fuentes:
        fechas_historicas = [
            parse_fecha_staging_desde_nombre_archivo(f.ruta) or datetime.fromtimestamp(f.ruta.stat().st_mtime)
            for f in staging_fuentes
        ]
        fuentes_payload["staging"]["historico_desde"] = serializar_fecha(min(fechas_historicas))
        fuentes_payload["staging"]["historico_hasta"] = serializar_fecha(max(fechas_historicas))
    elif not STAGING_ROOT.exists():
        fuentes_payload["staging"]["mensaje"] = "No se encontro la carpeta oficial de Staging IN-OUT sincronizada localmente."
    else:
        fuentes_payload["staging"]["mensaje"] = "No se detectaron archivos CSV de staging en la ruta oficial."

    df_stock = pd.concat(stock_frames, ignore_index=True) if stock_frames else None
    df_staging_snapshot = pd.concat(staging_frames, ignore_index=True) if staging_frames else None
    df_posiciones = pd.concat(posiciones_frames, ignore_index=True) if posiciones_frames else None
    fecha_referencia = max(stock_fechas + staging_fechas + posiciones_fechas) if (stock_fechas or staging_fechas or posiciones_fechas) else None
    fecha_referencia_txt = serializar_o_mtime_fecha(fecha_referencia)

    stock_payload, alertas_stock = calcular_stock_desde_wms(df_stock, fecha_referencia_txt, stock_descartes)
    bloqueado_payload, alertas_bloqueado = calcular_stock_bloqueado_wms(df_stock, stock_descartes)
    staging_payload, alertas_staging = calcular_staging_snapshot(df_staging_snapshot, fecha_referencia_txt)
    bloqueado_staging_payload, alertas_bloqueado_staging = calcular_stock_bloqueado_staging(df_staging_snapshot)
    pallets_actuales = construir_pallets_actuales(df_staging_snapshot) if df_staging_snapshot is not None else pd.DataFrame()
    pallets_objetivo = set(pallets_actuales["Pallet"].dropna().tolist()) if not pallets_actuales.empty else set()
    pallets_out_objetivo = set(
        pallets_actuales.loc[pallets_actuales["estado_staging"].eq("STAGING OUT"), "Pallet"].dropna().tolist()
    ) if not pallets_actuales.empty else set()
    primeras_apariciones, primeras_apariciones_out = procesar_historico_staging(
        staging_fuentes,
        pallets_objetivo,
        pallets_out_objetivo,
    )
    staging_antiguedad_payload, alertas_antiguedad = calcular_staging_antiguedad(
        pallets_actuales,
        primeras_apariciones,
        fecha_referencia,
    )
    staging_out_payload, alertas_staging_out = calcular_staging_out_permanencia(
        pallets_actuales,
        primeras_apariciones_out,
        fecha_referencia,
    )
    conteos_payload, ira_ila_payload, _avance_payload_base, alertas_conteos = calcular_conteos_ciclicos(
        df_conteos_periodo,
        conteos_fuente_info,
        year=year,
        month=month,
    )
    avance_payload, alertas_avance = calcular_avance_conteo(
        df_conteos_periodo,
        df_layout,
        year=year,
        month=month,
    )
    ocupacion_payload, alertas_ocupacion = calcular_ocupacion(df_layout, df_posiciones, fecha_referencia_txt)

    inventario = {
        "disponible": bool(
            stock_payload.get("disponible")
            or staging_payload.get("disponible")
            or ocupacion_payload.get("disponible")
        ),
        "fecha_referencia": fecha_referencia_txt,
        "fuentes": fuentes_payload,
        "stock": stock_payload,
        "stock_bloqueado_wms": bloqueado_payload,
        "staging": staging_payload,
        "stock_bloqueado_staging": bloqueado_staging_payload,
        "staging_antiguedad": staging_antiguedad_payload,
        "staging_out_permanencia": staging_out_payload,
        "conteos_ciclicos": conteos_payload,
        "ira_ila": ira_ila_payload,
        "avance_conteo": avance_payload,
        "ocupacion": ocupacion_payload,
        "pendientes": {
            "santa_rosa_conteos": "Pendiente si se confirma fuente externa",
        },
    }
    alertas = deduplicar_lista(
        alertas_stock
        + alertas_bloqueado
        + alertas_staging
        + alertas_bloqueado_staging
        + alertas_antiguedad
        + alertas_staging_out
        + alertas_conteos_fuente
        + alertas_conteos
        + alertas_avance
        + alertas_ocupacion
    )
    recomendaciones = []
    if inventario["disponible"] and not ocupacion_payload.get("disponible"):
        recomendaciones.append("Validar cruce entre posiciones y layout antes de publicar ocupacion.")
    if staging_antiguedad_payload.get("disponible"):
        recomendaciones.append("La antigüedad de staging usa primera aparición histórica del pallet y puede sobreestimar reingresos.")
    if staging_out_payload.get("disponible"):
        recomendaciones.append("La permanencia en STAGING OUT usa primera aparición histórica y conviene validar casos de reingreso.")
    return inventario, alertas, recomendaciones


def construir_alertas_recomendaciones(payload: dict[str, Any]) -> tuple[list[str], list[str]]:
    alertas = list(payload.get("alertas", []))
    recomendaciones = list(payload.get("recomendaciones", []))

    nnss = payload.get("nnss", {}) or {}
    pendientes = nnss.get("pendientes", {}) or {}
    if pendientes.get("mayores_7_dias"):
        alertas.append("Existen pedidos pendientes con mas de 7 dias que requieren seguimiento.")

    productividad = payload.get("productividad", {}) or {}
    global_prod = productividad.get("global", {}) or {}
    if productividad.get("disponible") and global_prod.get("dias_trabajados", 0) == 0:
        recomendaciones.append("Validar fechas de productividad: no se detectaron dias trabajados en el periodo.")

    inventario = payload.get("inventario", {}) or {}
    ocupacion = inventario.get("ocupacion", {}) or {}
    ocupacion_operativa = ocupacion.get("ocupacion_operativa", {}) or {}
    if inventario.get("disponible") and ocupacion_operativa.get("requiere_validacion"):
        recomendaciones.append("La ocupacion operativa de inventario requiere validacion contra criterio operacional.")

    return deduplicar_lista(alertas), deduplicar_lista(recomendaciones)


def deduplicar_lista(items: list[str]) -> list[str]:
    vistos = set()
    salida = []
    for item in items:
        texto = normalizar_texto(item)
        if not texto or texto in vistos:
            continue
        vistos.add(texto)
        salida.append(texto)
    return salida


def escribir_json(payload: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return output_path


def fecha_corte_periodo(year: int, month: int) -> datetime:
    now = datetime.now()
    if now.year == year and now.month == month:
        return now
    ultimo_dia = calendar.monthrange(year, month)[1]
    return datetime(year, month, ultimo_dia, 23, 59, 59)


def build_output_path(year: int, month: int, output: str | None) -> Path:
    if output:
        return Path(output)
    fecha = datetime.now().strftime("%Y%m%d")
    return LOGDIR / f"resumen_kpi_ops_{fecha}.json"


def resumen_inspeccion_nnss(fuentes: list[FuenteDetectada]) -> list[dict[str, Any]]:
    salida = []
    for fuente in fuentes:
        salida.append({"ruta": str(fuente.ruta), "hoja": fuente.hoja})
    return salida


def resumen_inspeccion_productividad(fuentes: list[FuenteDetectada]) -> dict[str, Any]:
    salida = {
        "disponible": PRODUCTIVIDAD_ROOT_OFICIAL.exists(),
        "ruta_oficial": str(PRODUCTIVIDAD_ROOT_OFICIAL),
        "archivos_detectados": [],
    }
    if not PRODUCTIVIDAD_ROOT_OFICIAL.exists():
        salida["mensaje"] = "No se encontro la carpeta oficial de Productividad sincronizada localmente."
        return salida

    for fuente in fuentes:
        salida["archivos_detectados"].append(
            {
                "path": str(fuente.ruta),
                "name": fuente.ruta.name,
                "modified": iso_mtime(fuente.ruta),
                "cliente_detectado": fuente.cliente,
                "cd_detectado": fuente.centro,
            }
        )
    return salida


def resumen_inspeccion_inventario(
    stock_fuentes: list[FuenteDetectada],
    staging_fuentes: list[FuenteDetectada],
    posiciones_fuentes: list[FuenteDetectada],
) -> dict[str, Any]:
    dim_archivo = encontrar_archivo_dim_inventario()
    staging_snapshot_fuentes = seleccionar_ultimas_fuentes_staging(staging_fuentes)
    fechas_historicas = [
        parse_fecha_staging_desde_nombre_archivo(fuente.ruta) or datetime.fromtimestamp(fuente.ruta.stat().st_mtime)
        for fuente in staging_fuentes
    ]
    return {
        "stock": {
            "disponible": STOCK_WMS_ROOT.exists(),
            "ruta": str(STOCK_WMS_ROOT),
            "archivos_detectados": [
                {
                    "path": str(fuente.ruta),
                    "name": fuente.ruta.name,
                    "modified": iso_mtime(fuente.ruta),
                    "cd_detectado": fuente.centro,
                }
                for fuente in stock_fuentes
            ],
        },
        "staging": {
            "disponible": STAGING_ROOT.exists(),
            "ruta": str(STAGING_ROOT),
            "archivos_detectados": [
                {
                    "path": str(fuente.ruta),
                    "name": fuente.ruta.name,
                    "modified": iso_mtime(fuente.ruta),
                    "cliente_detectado": fuente.cliente,
                    "cd_detectado": fuente.centro,
                }
                for fuente in staging_snapshot_fuentes
            ],
            "archivos_historicos": len(staging_fuentes),
            "historico_desde": serializar_fecha(min(fechas_historicas)) if fechas_historicas else None,
            "historico_hasta": serializar_fecha(max(fechas_historicas)) if fechas_historicas else None,
        },
        "posiciones": {
            "disponible": POSICIONES_ROOT.exists(),
            "ruta": str(POSICIONES_ROOT),
            "archivos_detectados": [
                {
                    "path": str(fuente.ruta),
                    "name": fuente.ruta.name,
                    "modified": iso_mtime(fuente.ruta),
                }
                for fuente in posiciones_fuentes
            ],
        },
        "dim_ubicaciones": {
            "disponible": dim_archivo is not None,
            "archivo": str(dim_archivo) if dim_archivo else None,
            "tabla": "Tabla14",
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Genera resumen JSON de KPI operativos para EgakatOpsBot.")
    parser.add_argument("--year", type=int, default=datetime.now().year)
    parser.add_argument("--month", type=int, default=datetime.now().month)
    parser.add_argument("--output", type=str, default=None)
    parser.add_argument("--inspect-only", action="store_true")
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    if args.month < 1 or args.month > 12:
        print("Mes invalido. Debe estar entre 1 y 12.")
        return 1

    nnss_fuentes = descubrir_fuentes_nnss(verbose=args.verbose)
    productividad_fuentes = descubrir_fuentes_productividad(args.year, args.month, verbose=args.verbose)
    stock_fuentes = descubrir_fuentes_stock_wms(verbose=args.verbose)
    staging_fuentes = descubrir_fuentes_staging(verbose=args.verbose)
    posiciones_fuentes = descubrir_fuentes_posiciones(verbose=args.verbose)
    dimensiones_info, metas_info, metas_map, ubicaciones_map = cargar_dimensiones(verbose=args.verbose)

    if args.inspect_only:
        salida = {
            "periodo": {"anio": args.year, "mes": args.month},
            "nnss": resumen_inspeccion_nnss(nnss_fuentes),
            "productividad": resumen_inspeccion_productividad(productividad_fuentes),
            "inventario": resumen_inspeccion_inventario(stock_fuentes, staging_fuentes, posiciones_fuentes),
            "dimensiones": dimensiones_info,
            "metas": metas_info,
        }
        print(json.dumps(salida, ensure_ascii=False, indent=2))
        return 0

    fecha_generacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fecha_consulta = fecha_corte_periodo(args.year, args.month)
    df_nnss, fuente_nnss = leer_consulta_fr(nnss_fuentes, args.year, args.month, verbose=args.verbose)
    nnss_payload, alertas_nnss, recomendaciones_nnss = calcular_nnss(
        df_nnss,
        year=args.year,
        month=args.month,
        fecha_consulta=fecha_consulta,
    )
    historico_payload = construir_historico_otif_mensual(
        nnss_fuentes=nnss_fuentes,
        year=args.year,
        hasta_mes=args.month,
        fecha_generacion=fecha_generacion,
        ubicaciones_map=ubicaciones_map,
    )
    recepciones_payload = construir_payload_recepciones(
        raiz_onedrive=RECEPCIONES_ROOT_OFICIAL,
        year=args.year,
        hasta_mes=args.month,
    )
    historico_payload["recepciones"] = recepciones_payload
    productividad_payload, fuente_productividad, alertas_prod, recomendaciones_prod = calcular_productividad(
        productividad_fuentes,
        year=args.year,
        month=args.month,
        metas_map=metas_map,
        metas_info=metas_info,
        ubicaciones_map=ubicaciones_map,
    )
    inventario_payload, alertas_inv, recomendaciones_inv = calcular_inventario_inicial(
        year=args.year,
        month=args.month,
        verbose=args.verbose,
    )

    payload = {
        "disponible": bool(
            nnss_payload.get("disponible")
            or productividad_payload.get("disponible")
            or inventario_payload.get("disponible")
        ),
        "fecha_generacion": fecha_generacion,
        "fuentes": {
            "nnss": fuente_nnss,
            "productividad": fuente_productividad,
            "dimensiones": dimensiones_info,
            "metas": metas_info,
        },
        "nnss": nnss_payload,
        "historico": historico_payload,
        "productividad": productividad_payload,
        "inventario": inventario_payload,
        "alertas": deduplicar_lista(alertas_nnss + alertas_prod + alertas_inv),
        "recomendaciones": deduplicar_lista(recomendaciones_nnss + recomendaciones_prod + recomendaciones_inv),
    }
    payload["alertas"], payload["recomendaciones"] = construir_alertas_recomendaciones(payload)

    output_path = build_output_path(args.year, args.month, args.output)
    escribir_json(payload, output_path)
    print(f"Resumen KPI operativo guardado en: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
