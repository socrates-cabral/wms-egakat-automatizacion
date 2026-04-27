import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from datetime import datetime, timedelta
import unicodedata
import re
import json
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


LOG_DIR = Path(r"C:\ClaudeWork\logs\validaciones_negocio")
LOG_DIR.mkdir(parents=True, exist_ok=True)

RUTA_STOCK_WMS = Path(
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA\Datos para Dashboard - Stock WMS Semanal"
)
RUTA_STAGING = Path(
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA\Datos para Dashboard - Stagin IN- OUT"
)
RUTA_CLIENTES_EK = Path(
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA\Datos para Dashboard - Clientes EK"
)


# =====================================================================================
# Normalizacion / lectura
# =====================================================================================

def strip_accents(text: str) -> str:
    text = str(text)
    return ''.join(
        c for c in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(c)
    )


def fix_common_visual_encoding(text: str) -> str:
    if text is None:
        return ""
    s = str(text)
    replacements = {
        "descripci¢n": "descripcion",
        "descripciã³n": "descripcion",
        "descripciã“n": "descripcion",
        "ubicaci¢n": "ubicacion",
        "ubicaciã³n": "ubicacion",
        "recepci¢n": "recepcion",
        "recepciã³n": "recepcion",
        "preparaci¢n": "preparacion",
        "preparaciã³n": "preparacion",
        "generaci¢n": "generacion",
        "generaciã³n": "generacion",
        "reubicaci¢n": "reubicacion",
        "reubicaciã³n": "reubicacion",
        "despachad¢": "despachado",
        "vencimient¢": "vencimiento",
        "articul¢": "articulo",
        "deposit¢": "deposito",
    }
    s_low = s.lower()
    for bad, good in replacements.items():
        s_low = s_low.replace(bad, good)
    return s_low


def norm_name(text: Any) -> str:
    text = "" if text is None else str(text)
    text = fix_common_visual_encoding(text)
    text = re.sub(r"(?i)\bn[°º]\.?(\s*)", "n", text)
    text = text.replace("№", "n")
    text = strip_accents(text)
    text = text.replace("°", "")
    text = text.replace("º", "")
    text = text.replace("fh.", "fh ")
    text = text.replace("descripcin", "descripcion")
    text = text.replace(".", " ")
    text = text.replace("/", " ")
    text = text.replace("-", " ")
    text = text.replace("_", " ")
    text = text.lower()
    text = re.sub(r"\s+", " ", text).strip()
    return text


def read_csv_flexible(path: Path, delimiter: str, encoding_candidates: List[str]) -> Tuple[pd.DataFrame, str]:
    last_error = None
    for enc in encoding_candidates:
        try:
            df = pd.read_csv(
                path,
                sep=delimiter,
                dtype=str,
                encoding=enc,
                engine="python",
                keep_default_na=False,
                on_bad_lines="error"
            )
            return df, enc
        except Exception as e:
            last_error = e
    raise RuntimeError(f"No se pudo leer CSV con los encodings probados: {last_error}")


def detect_header_row_excel(path: Path, search_window_rows: int, header_contains_any: List[str]) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    targets = set(header_contains_any)
    best_row = 1
    best_score = -1

    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=search_window_rows, values_only=True),
        start=1
    ):
        normalized = [norm_name(v) for v in row if v is not None and str(v).strip() != ""]
        score = sum(1 for x in normalized if x in targets)
        if score > best_score:
            best_score = score
            best_row = idx

    return best_row


def read_excel_schema(path: Path, header_row_1_based: int) -> pd.DataFrame:
    return pd.read_excel(path, header=header_row_1_based - 1, dtype=str)


DATASET_SPECS: Dict[str, Dict[str, Any]] = {
    "stock_wms": {
        "dataset_group": "stock_wms",
        "kind": "excel",
        "header_mode": "detect_in_top_rows",
        "search_window_rows": 15,
        "header_contains_any": [
            "deposito", "empresa", "articulo", "descripcion",
            "contenedor", "sub deposito"
        ],
    },
    "staging_estandar": {
        "dataset_group": "staging",
        "kind": "csv",
        "delimiter": ";",
        "encoding_candidates": ["utf-8", "latin1", "cp1252"],
    },
    "staging_unilever": {
        "dataset_group": "staging",
        "kind": "csv",
        "delimiter": ";",
        "encoding_candidates": ["utf-8", "latin1", "cp1252"],
    },
    "pedidos_preparados": {
        "dataset_group": "pedidos_preparados",
        "kind": "excel",
        "header_mode": "fixed_row",
        "header_row_1_based": 1,
    },
    "recepciones_recibidas": {
        "dataset_group": "recepciones",
        "kind": "excel",
        "header_mode": "fixed_row",
        "header_row_1_based": 1,
    },
}


# =====================================================================================
# Utilidades de validacion de negocio
# =====================================================================================

def is_blank(v: Any) -> bool:
    return v is None or str(v).strip() == "" or str(v).strip().lower() in {"nan", "nat", "none"}


def safe_numeric_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
        errors="coerce"
    )


def safe_datetime_series(series: pd.Series) -> pd.Series:
    """
    Parsea fechas mixtas sin romper formatos ISO tipo YYYY-MM-DD.
    - Si el valor parte con año de 4 dígitos, se interpreta como year-first.
    - En el resto de los casos, se interpreta como day-first.
    """
    s = series.astype(str).str.strip()
    s = s.replace({"": pd.NA, "nan": pd.NA, "NaT": pd.NA, "None": pd.NA})

    iso_mask = s.str.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}(?:\s+.*)?$", na=False)

    out = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    if iso_mask.any():
        iso_values = s.loc[iso_mask].str.replace("/", "-", regex=False)
        out.loc[iso_mask] = pd.to_datetime(iso_values, errors="coerce", yearfirst=True)
    if (~iso_mask).any():
        out.loc[~iso_mask] = pd.to_datetime(s.loc[~iso_mask], errors="coerce", dayfirst=True)
    return out


def make_col_map(df: pd.DataFrame) -> Dict[str, str]:
    return {norm_name(c): c for c in df.columns}


def col_exists(df: pd.DataFrame, col_norm: str) -> bool:
    return col_norm in make_col_map(df)


def get_col(df: pd.DataFrame, col_norm: str) -> Optional[str]:
    return make_col_map(df).get(col_norm)


def excel_row_from_index(df: pd.DataFrame, idx: int, header_row_1_based: int) -> int:
    return idx + header_row_1_based + 1

def row_nonblank_count(df: pd.DataFrame) -> pd.Series:
    meaningful_cols = [c for c in df.columns if norm_name(c) != ""]
    if not meaningful_cols:
        meaningful_cols = list(df.columns)
    if not meaningful_cols:
        return pd.Series(dtype="int64")

    work = df[meaningful_cols].copy()
    for c in meaningful_cols:
        work[c] = work[c].apply(lambda v: None if is_blank(v) else str(v).strip())
    return work.notna().sum(axis=1)


def drop_fully_blank_rows(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    if df.empty:
        return df, 0
    counts = row_nonblank_count(df)
    mask_keep = counts > 0
    removed = int((~mask_keep).sum())
    return df.loc[mask_keep].copy(), removed


def trim_trailing_noise_rows(
    df: pd.DataFrame,
    key_cols_norm: List[str],
    numeric_cols_norm: Optional[List[str]] = None,
    date_cols_norm: Optional[List[str]] = None,
) -> Tuple[pd.DataFrame, int]:
    if df.empty:
        return df, 0

    numeric_cols_norm = numeric_cols_norm or []
    date_cols_norm = date_cols_norm or []

    col_map = make_col_map(df)
    key_cols = [col_map[c] for c in key_cols_norm if c in col_map]
    numeric_cols = [col_map[c] for c in numeric_cols_norm if c in col_map]
    date_cols = [col_map[c] for c in date_cols_norm if c in col_map]

    if not key_cols and not numeric_cols and not date_cols:
        return df, 0

    trim_count = 0
    while len(df) > 0:
        row = df.iloc[-1]

        keys_blank = all(is_blank(row[c]) for c in key_cols) if key_cols else True

        nums_blank = True
        for c in numeric_cols:
            val = safe_numeric_series(pd.Series([row[c]])).iloc[0]
            if pd.notna(val) and val != 0:
                nums_blank = False
                break

        dates_blank = all(pd.isna(safe_datetime_series(pd.Series([row[c]])).iloc[0]) for c in date_cols) if date_cols else True
        nonblank_count = int(row_nonblank_count(df.tail(1)).iloc[0]) if len(df) else 0

        if keys_blank and nums_blank and dates_blank and nonblank_count <= 1:
            df = df.iloc[:-1].copy()
            trim_count += 1
        else:
            break

    return df, trim_count


def trim_stock_tail_residual_rows(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    """
    Recorta solo la cola residual típica de stock WMS:
    filas al final del archivo donde ya no existe dato operativo real,
    aunque puedan arrastrar textos accesorios como Depósito o Empresa.
    """
    if df.empty:
        return df, 0

    col_map = make_col_map(df)

    soft_text_cols = [
        col_map[c] for c in ["deposito", "empresa"]
        if c in col_map
    ]
    hard_blank_cols = [
        col_map[c] for c in ["articulo", "descripcion", "contenedor", "sub deposito"]
        if c in col_map
    ]
    numeric_cols = [
        col_map[c] for c in ["bultos", "unids", "kilos", "peso recep(kg)", "vh unitario"]
        if c in col_map
    ]
    date_cols = [
        col_map[c] for c in ["ingreso", "vencimiento", "fecha de elaboracion"]
        if c in col_map
    ]

    if not hard_blank_cols:
        return df, 0

    trim_count = 0
    while len(df) > 0:
        row = df.iloc[-1]

        hard_blank = all(is_blank(row[c]) for c in hard_blank_cols)

        nums_blank = True
        for c in numeric_cols:
            val = safe_numeric_series(pd.Series([row[c]])).iloc[0]
            if pd.notna(val) and val != 0:
                nums_blank = False
                break

        dates_blank = True
        for c in date_cols:
            val = safe_datetime_series(pd.Series([row[c]])).iloc[0]
            if pd.notna(val):
                dates_blank = False
                break

        allowed_nonblank = set(soft_text_cols)
        actual_nonblank = {
            c for c in df.columns
            if not is_blank(row[c])
        }

        nonblank_count = len(actual_nonblank)
        only_soft_nonblank = actual_nonblank.issubset(allowed_nonblank)

        if hard_blank and nums_blank and dates_blank and (only_soft_nonblank or nonblank_count <= len(allowed_nonblank)):
            df = df.iloc[:-1].copy()
            trim_count += 1
        else:
            break

    return df, trim_count


def clean_dataset_business_df(df: pd.DataFrame, dataset_group: str) -> Tuple[pd.DataFrame, List[str]]:
    notes: List[str] = []
    original_rows = len(df)

    df, removed_blank = drop_fully_blank_rows(df)
    if removed_blank:
        notes.append(f"Se eliminaron {removed_blank} filas completamente vacías antes de validar negocio.")

    trailing_rules = {
        "stock_wms": {
            "key_cols_norm": ["deposito", "articulo", "descripcion", "contenedor", "sub deposito"],
            "numeric_cols_norm": ["bultos", "unids", "kilos", "peso recep(kg)"],
            "date_cols_norm": ["ingreso", "vencimiento", "fecha de elaboracion"],
        },
        "staging": {
            "key_cols_norm": ["pallet", "codigo articulo", "articulo descripcion", "lote"],
            "numeric_cols_norm": ["cantidad"],
            "date_cols_norm": ["fecha de alta", "fecha de vencimiento"],
        },
        "pedidos_preparados": {
            "key_cols_norm": ["npedido", "articulo", "lote", "noc"],
            "numeric_cols_norm": ["cant prep"],
            "date_cols_norm": ["fh generacion", "fh preparacion", "fh despachado", "fh salida"],
        },
        "recepciones": {
            "key_cols_norm": ["n recepcion", "articulo", "lote"],
            "numeric_cols_norm": ["cantidad recibida"],
            "date_cols_norm": ["fh generacion", "fh inicio de recepcion", "fh fin de recepcion", "fh inicio de guardado", "fh fin de guardado"],
        },
    }

    cfg = trailing_rules.get(dataset_group)
    if cfg:
        df, trimmed = trim_trailing_noise_rows(df, **cfg)
        if trimmed:
            notes.append(f"Se recortaron {trimmed} filas residuales al final del archivo antes de validar negocio.")

    if dataset_group == "stock_wms":
        df, trimmed_stock_tail = trim_stock_tail_residual_rows(df)
        if trimmed_stock_tail:
            notes.append(
                f"Se recortaron {trimmed_stock_tail} filas residuales de cola en stock WMS "
                f"(sin dato operativo real) antes de validar negocio."
            )

    if len(df) != original_rows:
        df = df.copy()

    return df, notes


def build_sample_rows(
    df: pd.DataFrame,
    mask: pd.Series,
    header_row_1_based: int,
    cols_to_show: List[str],
    limit: int = 5
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    subset = df.loc[mask, cols_to_show].head(limit)
    for idx, (_, row) in enumerate(subset.iterrows()):
        original_idx = subset.index[idx]
        item: Dict[str, Any] = {"fila_excel": excel_row_from_index(df, int(original_idx), header_row_1_based)}
        for c in cols_to_show:
            item[c] = row[c]
        rows.append(item)
    return rows


def severity_by_count(
    count: int,
    total_rows: int,
    critical: bool = False,
    moderate: bool = False
) -> str:
    if count <= 0:
        return "OK"

    pct = (count / total_rows) if total_rows else 0

    if critical:
        if count >= 5 or pct >= 0.02:
            return "ERROR"
        return "PARCIAL"

    if moderate:
        if count >= 5 or pct >= 0.02:
            return "PARCIAL"
        return "WARNING"

    if count >= 5 or pct >= 0.02:
        return "PARCIAL"
    return "WARNING"


def add_finding(
    findings: List[Dict[str, Any]],
    severity: str,
    rule: str,
    dataset_group: str,
    detail: str,
    count: int,
    column: Optional[str] = None,
    sample: Optional[List[Dict[str, Any]]] = None,
) -> None:
    if severity == "OK":
        return
    findings.append({
        "severidad": severity,
        "regla": rule,
        "dataset": dataset_group,
        "columna": column,
        "detalle": detail,
        "cantidad_afectada": int(count),
        "muestra": sample or [],
    })


def final_state_from_findings(findings: List[Dict[str, Any]]) -> str:
    severities = [f["severidad"] for f in findings]
    if "ERROR" in severities:
        return "ERROR"
    if "PARCIAL" in severities:
        return "PARCIAL"
    if "WARNING" in severities:
        return "WARNING"
    return "OK"


# =====================================================================================
# Reglas genericas
# =====================================================================================

def rule_required_not_blank(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
    col_norm: str,
    rule_name: str,
    critical: bool = True,
) -> Optional[Dict[str, Any]]:
    actual_col = get_col(df, col_norm)
    if not actual_col:
        return None

    mask = df[actual_col].apply(is_blank)
    count = int(mask.sum())
    if count == 0:
        return None

    severity = severity_by_count(count, len(df), critical=critical, moderate=not critical)
    sample = build_sample_rows(df, mask, header_row_1_based, [actual_col])
    return {
        "severidad": severity,
        "regla": rule_name,
        "dataset": dataset_group,
        "columna": actual_col,
        "detalle": f"Se detectaron {count} registros con '{actual_col}' vacío.",
        "cantidad_afectada": count,
        "muestra": sample,
    }


def rule_negative_values(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
    col_norm: str,
    rule_name: str,
    allow_zero: bool = True,
    critical: bool = True,
) -> Optional[Dict[str, Any]]:
    actual_col = get_col(df, col_norm)
    if not actual_col:
        return None

    serie = safe_numeric_series(df[actual_col])
    mask = serie < 0 if allow_zero else serie <= 0
    mask = mask.fillna(False)
    count = int(mask.sum())
    if count == 0:
        return None

    severity = severity_by_count(count, len(df), critical=critical, moderate=not critical)
    sample = build_sample_rows(df, mask, header_row_1_based, [actual_col])
    return {
        "severidad": severity,
        "regla": rule_name,
        "dataset": dataset_group,
        "columna": actual_col,
        "detalle": f"Se detectaron {count} valores {'negativos' if allow_zero else '<= 0'} en '{actual_col}'.",
        "cantidad_afectada": count,
        "muestra": sample,
    }


def rule_date_order(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
    start_col_norm: str,
    end_col_norm: str,
    rule_name: str,
    critical: bool = True,
) -> Optional[Dict[str, Any]]:
    start_col = get_col(df, start_col_norm)
    end_col = get_col(df, end_col_norm)
    if not start_col or not end_col:
        return None

    s1 = safe_datetime_series(df[start_col])
    s2 = safe_datetime_series(df[end_col])
    mask = s1.notna() & s2.notna() & (s2 < s1)
    count = int(mask.sum())
    if count == 0:
        return None

    severity = severity_by_count(count, len(df), critical=critical, moderate=not critical)
    sample = build_sample_rows(df, mask, header_row_1_based, [start_col, end_col])
    return {
        "severidad": severity,
        "regla": rule_name,
        "dataset": dataset_group,
        "columna": f"{start_col} -> {end_col}",
        "detalle": f"Se detectaron {count} registros donde '{end_col}' es menor que '{start_col}'.",
        "cantidad_afectada": count,
        "muestra": sample,
    }


def rule_future_dates(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
    col_norm: str,
    rule_name: str,
    tolerance_days: int = 2,
    critical: bool = False,
    execution_ts: Optional[pd.Timestamp] = None,
) -> Optional[Dict[str, Any]]:
    actual_col = get_col(df, col_norm)
    if not actual_col:
        return None

    serie = safe_datetime_series(df[actual_col])
    base_ts = execution_ts if execution_ts is not None else pd.Timestamp.now()
    limit_date = pd.Timestamp(base_ts).normalize() + pd.Timedelta(days=tolerance_days)
    mask = serie.notna() & (serie > limit_date)
    count = int(mask.sum())
    if count == 0:
        return None

    severity = severity_by_count(count, len(df), critical=critical, moderate=not critical)
    sample = build_sample_rows(df, mask, header_row_1_based, [actual_col])
    return {
        "severidad": severity,
        "regla": rule_name,
        "dataset": dataset_group,
        "columna": actual_col,
        "detalle": f"Se detectaron {count} fechas futuras fuera de tolerancia en '{actual_col}' (límite: {limit_date.date()}).",
        "cantidad_afectada": count,
        "muestra": sample,
    }


def rule_duplicates_on_keys(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
    key_cols_norm: List[str],
    rule_name: str,
    moderate: bool = True,
) -> Optional[Dict[str, Any]]:
    col_map = make_col_map(df)
    actual_cols = [col_map[c] for c in key_cols_norm if c in col_map]
    if len(actual_cols) != len(key_cols_norm):
        return None

    key_df = df[actual_cols].copy()
    for c in actual_cols:
        key_df[c] = key_df[c].astype(str).str.strip()

    mask_blank_key = pd.Series(False, index=df.index)
    for c in actual_cols:
        mask_blank_key = mask_blank_key | key_df[c].apply(is_blank)

    work = key_df.loc[~mask_blank_key].copy()
    if work.empty:
        return None

    dup_mask_work = work.duplicated(subset=actual_cols, keep=False)
    dup_indices = work.index[dup_mask_work]
    mask = pd.Series(False, index=df.index)
    mask.loc[dup_indices] = True

    count = int(mask.sum())
    if count == 0:
        return None

    severity = severity_by_count(count, len(df), critical=False, moderate=moderate)
    sample = build_sample_rows(df, mask, header_row_1_based, actual_cols)
    return {
        "severidad": severity,
        "regla": rule_name,
        "dataset": dataset_group,
        "columna": ", ".join(actual_cols),
        "detalle": f"Se detectaron {count} registros duplicados usando la clave: {', '.join(actual_cols)}.",
        "cantidad_afectada": count,
        "muestra": sample,
    }



def rule_exact_line_duplicates_staging(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
) -> Optional[Dict[str, Any]]:
    """
    Reemplaza la antigua heurística de duplicado sospechoso por pallet/código/lote.
    En staging, esas combinaciones pueden repetirse válidamente para muchos clientes.
    Por eso aquí solo se busca duplicado prácticamente exacto de línea operativa,
    usando la mayor granularidad disponible sin amarrarse a reglas por cliente.
    """
    required_base = ["pallet", "codigo articulo"]
    context_candidates = [
        "lote", "contenedor", "cantidad", "area", "cara", "columna", "nivel",
        "ubicacion", "motivo", "destino", "sub deposito", "fecha de alta",
        "fecha de vencimiento", "unidad", "bultos", "unids", "kilos",
    ]

    col_map = make_col_map(df)
    if not all(c in col_map for c in required_base):
        return None

    actual_cols: List[str] = [col_map[c] for c in required_base]
    for c in context_candidates:
        if c in col_map and col_map[c] not in actual_cols:
            actual_cols.append(col_map[c])

    # Para que la regla sea sana, exigimos al menos algo más que pallet + código.
    if len(actual_cols) < 5:
        return None

    work = df[actual_cols].copy()
    for c in actual_cols:
        work[c] = work[c].apply(lambda v: "" if is_blank(v) else str(v).strip())

    # Ignorar filas donde la clave base venga vacía.
    base_blank = pd.Series(False, index=df.index)
    for c in [col_map[x] for x in required_base]:
        base_blank = base_blank | work[c].eq("")
    work = work.loc[~base_blank].copy()
    if work.empty:
        return None

    dup_mask_work = work.duplicated(subset=actual_cols, keep=False)
    dup_indices = work.index[dup_mask_work]
    mask = pd.Series(False, index=df.index)
    mask.loc[dup_indices] = True

    count = int(mask.sum())
    if count == 0:
        return None

    sample_cols = actual_cols[: min(len(actual_cols), 8)]
    sample = build_sample_rows(df, mask, header_row_1_based, sample_cols)
    severity = severity_by_count(count, len(df), critical=False, moderate=True)
    return {
        "severidad": severity,
        "regla": "DUPLICADO_EXACTO_LINEA_STAGING",
        "dataset": dataset_group,
        "columna": ", ".join(actual_cols),
        "detalle": (
            f"Se detectaron {count} registros duplicados casi exactos en staging, "
            f"usando granularidad operativa: {', '.join(actual_cols)}."
        ),
        "cantidad_afectada": count,
        "muestra": sample,
    }


def rule_exact_line_duplicates_pedidos(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
) -> Optional[Dict[str, Any]]:
    """
    Reemplaza la heurística rígida por NºPedido + Articulo + Lote.
    En pedidos preparados, esa combinación puede repetirse válidamente
    por fraccionamiento, múltiples pallets origen o múltiples ubicaciones.
    Aquí solo se marca duplicado cuando la línea es prácticamente exacta,
    usando la mayor granularidad operativa disponible.
    """
    required_base = ["npedido", "articulo"]
    context_candidates = [
        "lote",
        "pallet origen",
        "ubicacion origen",
        "pallet picking",
        "cant prep",
        "op",
        "viaje",
        "destino",
        "tipo",
        "fh generacion",
        "fh preparacion",
        "fh despachado",
        "fh salida",
    ]

    col_map = make_col_map(df)
    if not all(c in col_map for c in required_base):
        return None

    actual_cols: List[str] = [col_map[c] for c in required_base]
    for c in context_candidates:
        if c in col_map and col_map[c] not in actual_cols:
            actual_cols.append(col_map[c])

    # Exigir granularidad adicional para no volver a caer en una regla gruesa.
    if len(actual_cols) < 6:
        return None

    work = df[actual_cols].copy()
    for c in actual_cols:
        work[c] = work[c].apply(lambda v: "" if is_blank(v) else str(v).strip())

    base_blank = pd.Series(False, index=df.index)
    for c in [col_map[x] for x in required_base]:
        base_blank = base_blank | work[c].eq("")
    work = work.loc[~base_blank].copy()
    if work.empty:
        return None

    dup_mask_work = work.duplicated(subset=actual_cols, keep=False)
    dup_indices = work.index[dup_mask_work]
    mask = pd.Series(False, index=df.index)
    mask.loc[dup_indices] = True

    count = int(mask.sum())
    if count == 0:
        return None

    sample_cols = actual_cols[: min(len(actual_cols), 8)]
    sample = build_sample_rows(df, mask, header_row_1_based, sample_cols)
    severity = severity_by_count(count, len(df), critical=False, moderate=True)
    return {
        "severidad": severity,
        "regla": "DUPLICADO_EXACTO_LINEA_PEDIDOS",
        "dataset": dataset_group,
        "columna": ", ".join(actual_cols),
        "detalle": (
            f"Se detectaron {count} registros duplicados casi exactos en pedidos preparados, "
            f"usando granularidad operativa: {', '.join(actual_cols)}."
        ),
        "cantidad_afectada": count,
        "muestra": sample,
    }


def rule_outliers_iqr(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
    col_norm: str,
    rule_name: str,
    min_non_null: int = 12,
    max_alert_ratio: float = 0.005,
    max_alert_count: int = 50,
) -> Optional[Dict[str, Any]]:
    actual_col = get_col(df, col_norm)
    if not actual_col:
        return None

    serie_full = safe_numeric_series(df[actual_col])
    serie = serie_full.dropna()
    serie = serie[serie >= 0]
    if len(serie) < min_non_null:
        return None

    q1 = serie.quantile(0.25)
    q3 = serie.quantile(0.75)
    iqr = q3 - q1
    p995 = serie.quantile(0.995)

    if pd.isna(iqr) or iqr <= 0:
        if pd.isna(p995):
            return None
        upper = p995
    else:
        upper = max(q3 + (6 * iqr), p995)

    if pd.isna(upper) or upper <= 0:
        return None

    mask = serie_full > upper
    count = int(mask.fillna(False).sum())
    if count == 0:
        return None

    ratio = count / len(df) if len(df) else 0
    if count > max_alert_count or ratio > max_alert_ratio:
        return None

    sample = build_sample_rows(df, mask.fillna(False), header_row_1_based, [actual_col])
    return {
        "severidad": "WARNING",
        "regla": rule_name,
        "dataset": dataset_group,
        "columna": actual_col,
        "detalle": f"Se detectaron {count} outliers altos en '{actual_col}' (umbral robusto: {upper:.2f}).",
        "cantidad_afectada": count,
        "muestra": sample,
    }


# =====================================================================================
# Reglas especificas por dataset
# =====================================================================================


# =====================================================================================
# Reglas especificas por dataset
# =====================================================================================

def rule_stock_contenedor_vacio_con_stock(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
) -> Optional[Dict[str, Any]]:
    stock_col = get_col(df, "bultos") or get_col(df, "unids") or get_col(df, "kilos")
    cont_col = get_col(df, "contenedor")
    art_col = get_col(df, "articulo")
    if not stock_col or not cont_col:
        return None

    stock = safe_numeric_series(df[stock_col])
    mask = stock.gt(0) & df[cont_col].apply(is_blank)
    count = int(mask.fillna(False).sum())
    if count == 0:
        return None

    cols = [c for c in [art_col, cont_col, stock_col] if c]
    sample = build_sample_rows(df, mask.fillna(False), header_row_1_based, cols)
    severity = severity_by_count(count, len(df), critical=False, moderate=True)
    return {
        "severidad": severity,
        "regla": "STOCK_CONTENEDOR_VACIO",
        "dataset": dataset_group,
        "columna": cont_col,
        "detalle": f"Se detectaron {count} registros con stock positivo y contenedor vacío.",
        "cantidad_afectada": count,
        "muestra": sample,
    }


def rule_staging_pallet_vacio_con_cantidad(
    df: pd.DataFrame,
    dataset_group: str,
    header_row_1_based: int,
) -> Optional[Dict[str, Any]]:
    pallet_col = get_col(df, "pallet")
    qty_col = get_col(df, "cantidad")
    art_col = get_col(df, "codigo articulo") or get_col(df, "articulo descripcion")
    if not pallet_col or not qty_col:
        return None

    qty = safe_numeric_series(df[qty_col])
    mask = qty.gt(0) & df[pallet_col].apply(is_blank)
    count = int(mask.fillna(False).sum())
    if count == 0:
        return None

    cols = [c for c in [art_col, pallet_col, qty_col] if c]
    sample = build_sample_rows(df, mask.fillna(False), header_row_1_based, cols)
    severity = severity_by_count(count, len(df), critical=False, moderate=True)
    return {
        "severidad": severity,
        "regla": "STAGING_PALLET_VACIO",
        "dataset": dataset_group,
        "columna": pallet_col,
        "detalle": f"Se detectaron {count} registros con cantidad positiva y pallet vacío.",
        "cantidad_afectada": count,
        "muestra": sample,
    }


# =====================================================================================
# Motor por dataset
# =====================================================================================

def load_dataset(path: str, schema_name: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    file_path = Path(path)
    spec = DATASET_SPECS[schema_name]
    meta: Dict[str, Any] = {
        "archivo": str(file_path),
        "schema": schema_name,
        "dataset_group": spec["dataset_group"],
        "header_row_1_based": None,
        "sheet_used": None,
        "notas": [],
    }

    if spec["kind"] == "csv":
        df, enc = read_csv_flexible(
            file_path,
            spec["delimiter"],
            spec.get("encoding_candidates", ["utf-8", "latin1", "cp1252"])
        )
        meta["header_row_1_based"] = 1
        meta["sheet_used"] = None
        meta["notas"].append(f"CSV leído con encoding: {enc}")
        return df, meta

    if spec["header_mode"] == "fixed_row":
        header_row = spec["header_row_1_based"]
    else:
        header_row = detect_header_row_excel(
            file_path,
            spec["search_window_rows"],
            [norm_name(x) for x in spec["header_contains_any"]]
        )

    df = read_excel_schema(file_path, header_row)
    meta["header_row_1_based"] = header_row
    meta["sheet_used"] = "Sheet1"
    return df, meta


def execute_rules(df: pd.DataFrame, schema_name: str, header_row_1_based: int, execution_ts: Optional[pd.Timestamp] = None) -> List[Dict[str, Any]]:
    dataset_group = DATASET_SPECS[schema_name]["dataset_group"]
    findings: List[Dict[str, Any]] = []

    def push(found: Optional[Dict[str, Any]]) -> None:
        if found:
            findings.append(found)

    if dataset_group == "staging":
        push(rule_required_not_blank(df, dataset_group, header_row_1_based, "codigo articulo", "ARTICULO_VACIO", critical=True))
        push(rule_negative_values(df, dataset_group, header_row_1_based, "cantidad", "CANTIDAD_NEGATIVA", allow_zero=True, critical=True))
        push(rule_staging_pallet_vacio_con_cantidad(df, dataset_group, header_row_1_based))
        push(rule_date_order(df, dataset_group, header_row_1_based, "fecha de alta", "fecha de vencimiento", "VENCIMIENTO_MENOR_A_ALTA", critical=False))
        push(rule_future_dates(df, dataset_group, header_row_1_based, "fecha de alta", "FECHA_ALTA_FUTURA", tolerance_days=2, critical=False, execution_ts=execution_ts))
        push(rule_exact_line_duplicates_staging(df, dataset_group, header_row_1_based))
        push(rule_outliers_iqr(df, dataset_group, header_row_1_based, "cantidad", "OUTLIER_CANTIDAD"))

    elif dataset_group == "stock_wms":
        push(rule_required_not_blank(df, dataset_group, header_row_1_based, "articulo", "ARTICULO_VACIO", critical=True))
        push(rule_required_not_blank(df, dataset_group, header_row_1_based, "descripcion", "DESCRIPCION_VACIA", critical=False))
        push(rule_stock_contenedor_vacio_con_stock(df, dataset_group, header_row_1_based))
        for stock_candidate in ["bultos", "unids", "kilos", "peso recep(kg)"]:
            push(rule_negative_values(df, dataset_group, header_row_1_based, stock_candidate, f"NEGATIVO_{norm_name(stock_candidate).replace(' ', '_').upper()}", allow_zero=True, critical=True))
        push(rule_duplicates_on_keys(df, dataset_group, header_row_1_based, ["deposito", "articulo", "contenedor", "sub deposito"], "DUPLICADO_STOCK_EXACTO", moderate=True))
        for outlier_candidate in ["bultos", "unids", "kilos"]:
            push(rule_outliers_iqr(df, dataset_group, header_row_1_based, outlier_candidate, f"OUTLIER_{norm_name(outlier_candidate).replace(' ', '_').upper()}"))

    elif dataset_group == "pedidos_preparados":
        push(rule_required_not_blank(df, dataset_group, header_row_1_based, "npedido", "NPEDIDO_VACIO", critical=True))
        push(rule_required_not_blank(df, dataset_group, header_row_1_based, "articulo", "ARTICULO_VACIO", critical=True))
        push(rule_negative_values(df, dataset_group, header_row_1_based, "cant prep", "CANT_PREP_NO_VALIDA", allow_zero=False, critical=True))
        push(rule_date_order(df, dataset_group, header_row_1_based, "fh generacion", "fh preparacion", "PREPARACION_MENOR_A_GENERACION", critical=True))
        push(rule_date_order(df, dataset_group, header_row_1_based, "fh preparacion", "fh despachado", "DESPACHO_MENOR_A_PREPARACION", critical=True))
        push(rule_date_order(df, dataset_group, header_row_1_based, "fh despachado", "fh salida", "SALIDA_MENOR_A_DESPACHADO", critical=False))
        push(rule_future_dates(df, dataset_group, header_row_1_based, "fh salida", "FH_SALIDA_FUTURA", tolerance_days=2, critical=False, execution_ts=execution_ts))
        push(rule_exact_line_duplicates_pedidos(df, dataset_group, header_row_1_based))
        push(rule_outliers_iqr(df, dataset_group, header_row_1_based, "cant prep", "OUTLIER_CANT_PREP"))

    elif dataset_group == "recepciones":
        push(rule_required_not_blank(df, dataset_group, header_row_1_based, "n recepcion", "NRECEPCION_VACIO", critical=True))
        push(rule_required_not_blank(df, dataset_group, header_row_1_based, "articulo", "ARTICULO_VACIO", critical=True))
        push(rule_negative_values(df, dataset_group, header_row_1_based, "cantidad recibida", "CANTIDAD_RECIBIDA_NO_VALIDA", allow_zero=False, critical=True))
        push(rule_date_order(df, dataset_group, header_row_1_based, "fh generacion", "fh inicio de recepcion", "INICIO_RECEPCION_MENOR_A_GENERACION", critical=True))
        push(rule_date_order(df, dataset_group, header_row_1_based, "fh inicio de recepcion", "fh fin de recepcion", "FIN_RECEPCION_MENOR_A_INICIO", critical=True))
        push(rule_date_order(df, dataset_group, header_row_1_based, "fh inicio de guardado", "fh fin de guardado", "FIN_GUARDADO_MENOR_A_INICIO", critical=True))
        push(rule_future_dates(df, dataset_group, header_row_1_based, "fh fin de guardado", "FH_FIN_GUARDADO_FUTURA", tolerance_days=2, critical=False, execution_ts=execution_ts))
        push(rule_duplicates_on_keys(df, dataset_group, header_row_1_based, ["n recepcion", "articulo", "lote"], "DUPLICADO_RECEPCION_ARTICULO_LOTE", moderate=True))
        push(rule_outliers_iqr(df, dataset_group, header_row_1_based, "cantidad recibida", "OUTLIER_CANTIDAD_RECIBIDA"))

    return findings


def validate_business(path: str, schema_name: str) -> Dict[str, Any]:
    file_path = Path(path)
    result: Dict[str, Any] = {
        "archivo": str(file_path),
        "schema": schema_name,
        "dataset_group": DATASET_SPECS[schema_name]["dataset_group"],
        "estado": "OK",
        "sheet_used": None,
        "header_row_1_based": None,
        "row_count": None,
        "column_count": None,
        "errores": [],
        "warnings": [],
        "notas": [],
        "hallazgos": [],
        "actual_columns": [],
    }

    if not file_path.exists():
        result["estado"] = "ERROR"
        result["errores"].append("El archivo no existe.")
        return result

    try:
        df, meta = load_dataset(path, schema_name)
        result["sheet_used"] = meta["sheet_used"]
        result["header_row_1_based"] = meta["header_row_1_based"]
        result["notas"].extend(meta["notas"])
        result["row_count"] = len(df)
        result["column_count"] = len(df.columns)
        result["actual_columns"] = list(df.columns)

        if len(df) == 0:
            result["estado"] = "WARNING"
            result["warnings"].append("Archivo sin registros. No aplica validación de negocio.")
            return result

        df, cleaning_notes = clean_dataset_business_df(df, result["dataset_group"])
        result["notas"].extend(cleaning_notes)
        result["row_count"] = len(df)

        if len(df) == 0:
            result["estado"] = "WARNING"
            result["warnings"].append("Archivo sin registros utilizables luego de la limpieza operativa. No aplica validación de negocio.")
            return result

        findings = execute_rules(df, schema_name, result["header_row_1_based"] or 1, execution_ts=pd.Timestamp.now())
        result["hallazgos"] = findings

        for f in findings:
            msg = f"[{f['severidad']}] {f['regla']}: {f['detalle']}"
            if f["severidad"] == "ERROR":
                result["errores"].append(msg)
            else:
                result["warnings"].append(msg)

        result["estado"] = final_state_from_findings(findings)

    except Exception as e:
        result["estado"] = "ERROR"
        result["errores"].append(f"Error procesando archivo: {e}")

    return result


# =====================================================================================
# Integracion / resumen
# =====================================================================================

def merge_structure_business_states(structure_state: str, business_state: str) -> str:
    ranking = {"OK": 0, "WARNING": 1, "PARCIAL": 2, "ERROR": 3}
    reverse = {0: "OK", 1: "WARNING", 2: "PARCIAL", 3: "ERROR"}
    worst = max(ranking.get(structure_state, 0), ranking.get(business_state, 0))
    return reverse[worst]


def _latest_match(folder: Path, pattern: str, name_contains: Optional[str] = None) -> Optional[Path]:
    if not folder.exists():
        return None
    files = [p for p in folder.glob(pattern) if p.is_file()]
    if name_contains:
        files = [p for p in files if name_contains.lower() in p.name.lower()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def _latest_recursive(folder: Path, pattern: str, name_contains: Optional[str] = None) -> Optional[Path]:
    if not folder.exists():
        return None
    files = [p for p in folder.rglob(pattern) if p.is_file()]
    if name_contains:
        files = [p for p in files if name_contains.lower() in p.name.lower()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def encontrar_stock_wms() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []

    q = _latest_match(RUTA_STOCK_WMS / "Quilicura", "*.xlsx")
    if q:
        resultados.append((q, "stock_wms"))

    folder = RUTA_STOCK_WMS / "Pudahuel"
    if folder.exists():
        files = sorted([p for p in folder.glob("*.xlsx") if p.is_file()], key=lambda p: p.stat().st_mtime, reverse=True)
        usados = set()
        for p in files:
            if p.name not in usados:
                resultados.append((p, "stock_wms"))
                usados.add(p.name)
            if len(usados) >= 2:
                break

    return resultados


def encontrar_staging() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []

    clientes_quilicura = [
        "ABINBEV", "DAIKIN", "DAIKIN CLIENTES", "DERCO",
        "MASCOTAS LATINAS", "POCHTECA"
    ]
    clientes_pudahuel = [
        "BARENTZ", "BURASCHI", "CEPAS CHILE", "COLLICO", "DELIBEST",
        "INTIME", "NATIVOS DRINK", "TRES MONTE", "UNILEVER", "RUNO"
    ]

    for cliente in clientes_quilicura:
        f = _latest_match(RUTA_STAGING / "Quilicura" / cliente, "*.csv")
        if f:
            schema = "staging_unilever" if cliente.upper() == "UNILEVER" else "staging_estandar"
            resultados.append((f, schema))

    for cliente in clientes_pudahuel:
        f = _latest_match(RUTA_STAGING / "Pudahuel" / cliente, "*.csv")
        if f:
            schema = "staging_unilever" if cliente.upper() == "UNILEVER" else "staging_estandar"
            resultados.append((f, schema))

    return resultados


def encontrar_clientes_ek_preparacion() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []
    clientes = ["ABINBEV", "DAIKIN", "DERCO", "MASCOTAS LATINAS", "POCHTECA"]
    for cliente in clientes:
        carpeta = RUTA_CLIENTES_EK / cliente / "Preparación"
        f = _latest_recursive(carpeta, "*.xlsx", "Pedidos Preparados")
        if f:
            resultados.append((f, "pedidos_preparados"))
    return resultados


def encontrar_clientes_ek_recepciones() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []
    clientes = ["ABINBEV", "DAIKIN", "DERCO", "MASCOTAS LATINAS", "POCHTECA"]
    for cliente in clientes:
        carpeta = RUTA_CLIENTES_EK / cliente / "Recepciones"
        f = _latest_recursive(carpeta, "*.xlsx")
        if f:
            resultados.append((f, "recepciones_recibidas"))
    return resultados


def save_results_txt(results: List[Dict[str, Any]]) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = LOG_DIR / f"negocio_{ts}.txt"

    ok = sum(1 for r in results if r["estado"] == "OK")
    warn = sum(1 for r in results if r["estado"] == "WARNING")
    parcial = sum(1 for r in results if r["estado"] == "PARCIAL")
    err = sum(1 for r in results if r["estado"] == "ERROR")

    lines: List[str] = []
    lines.append("RESUMEN VALIDACION DE NEGOCIO")
    lines.append("=" * 90)
    lines.append(f"Fecha ejecucion : {datetime.now().isoformat()}")
    lines.append(f"Total revisados : {len(results)}")
    lines.append(f"OK              : {ok}")
    lines.append(f"WARNING         : {warn}")
    lines.append(f"PARCIAL         : {parcial}")
    lines.append(f"ERROR           : {err}")
    lines.append("")

    for r in results:
        lines.append(f"[{r['estado']}] {r['schema']} | {r['archivo']}")
        lines.append(f"   Header row : {r['header_row_1_based']}")
        lines.append(f"   Filas      : {r['row_count']}")
        lines.append(f"   Columnas   : {r['column_count']}")
        if r["warnings"]:
            for w in r["warnings"]:
                lines.append(f"   - WARNING: {w}")
        if r["errores"]:
            for e in r["errores"]:
                lines.append(f"   - ERROR: {e}")
        if r["notas"]:
            for n in r["notas"]:
                lines.append(f"   - NOTA: {n}")
        if r["hallazgos"]:
            lines.append("   - HALLAZGOS DETALLE:")
            for h in r["hallazgos"]:
                lines.append(f"      * {h['severidad']} | {h['regla']} | {h['detalle']}")
        lines.append("")

    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def save_results_json(results: List[Dict[str, Any]]) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = LOG_DIR / f"negocio_{ts}.json"
    payload = {
        "timestamp": datetime.now().isoformat(),
        "results": results
    }
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def _self_test_dates_v13() -> None:
    casos = pd.Series(["2026-03-10", "10/03/2026", "2026/03/10", "03/10/2026"])
    salida = safe_datetime_series(casos)
    assert str(salida.iloc[0].date()) == "2026-03-10"
    assert str(salida.iloc[1].date()) == "2026-03-10"
    assert str(salida.iloc[2].date()) == "2026-03-10"
    assert str(salida.iloc[3].date()) == "2026-10-03"


def _self_test_stock_tail_trim_v13() -> None:
    df = pd.DataFrame({
        "Depósito": ["PUDAHUEL", "PUDAHUEL", "PUDAHUEL", "PUDAHUEL"],
        "Empresa": ["EK", "EK", "EK", "EK"],
        "Artículo": ["SKU1", "", "", ""],
        "Descripción": ["Prod 1", "", "", ""],
        "Contenedor": ["CONT1", "", "", ""],
        "Sub-Depósito": ["A1", "", "", ""],
        "Bultos": ["10", "", "", ""],
        "Unids ": ["100", "", "", ""],
        "Kilos ": ["55", "", "", ""],
    })
    out, removed = trim_stock_tail_residual_rows(df)
    assert removed == 3
    assert len(out) == 1


def _self_test_staging_exact_dup_v13() -> None:
    df = pd.DataFrame({
        "Pallet": ["PLT1", "PLT1", "PLT1"],
        "Codigo Articulo": ["SKU1", "SKU1", "SKU1"],
        "Lote": ["L1", "L1", "L1"],
        "Contenedor": ["C1", "C1", "C1"],
        "Cantidad": ["10", "10", "12"],
        "Area": ["A", "A", "A"],
        "Cara": ["01", "01", "01"],
        "Columna": ["001", "001", "001"],
        "Nivel": ["01", "01", "01"],
    })
    found = rule_exact_line_duplicates_staging(df, "staging", 1)
    assert found is not None
    assert found["cantidad_afectada"] == 2


def run_auto_local_examples() -> None:
    _self_test_dates_v13()
    _self_test_stock_tail_trim_v13()
    _self_test_staging_exact_dup_v13()
    examples: List[Tuple[Path, str]] = []
    examples.extend(encontrar_stock_wms())
    examples.extend(encontrar_staging())
    examples.extend(encontrar_clientes_ek_preparacion())
    examples.extend(encontrar_clientes_ek_recepciones())

    if not examples:
        print("No se encontraron archivos reales para validar.")
        return

    print("=" * 90)
    print("VALIDADOR DE NEGOCIO WMS EGAKAT - AUTO LOCAL v1.3")
    print("=" * 90)

    results = []
    for path, schema_name in examples:
        print(f"Validando: {path} | schema={schema_name}")
        results.append(validate_business(str(path), schema_name))

    txt = save_results_txt(results)
    js = save_results_json(results)

    ok = sum(1 for r in results if r["estado"] == "OK")
    warn = sum(1 for r in results if r["estado"] == "WARNING")
    parcial = sum(1 for r in results if r["estado"] == "PARCIAL")
    err = sum(1 for r in results if r["estado"] == "ERROR")

    print("-" * 90)
    print(f"Total: {len(results)} | OK: {ok} | WARNING: {warn} | PARCIAL: {parcial} | ERROR: {err}")
    print(f"TXT :  {txt}")
    print(f"JSON:  {js}")


if __name__ == "__main__":
    run_auto_local_examples()
