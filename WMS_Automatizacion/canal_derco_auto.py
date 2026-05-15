# =============================================================================
# canal_derco_auto.py
# Automatiza la columna "Canal" en data Derco.xlsx
# Lee todos los MovDerco (anio/mes) + Base CES desde SharePoint local (OneDrive)
# NO modifica MovDerco ni Base CES -- solo reescribe la col Canal en data Derco
# Hace backup de data Derco antes de tocar nada
# =============================================================================
# DECISIONES DE NEGOCIO (confirmadas con usuario):
#   - Llave de cruce: Nro Aplica (OP) <-> MovDerco Comprobante (100% match verificado)
#   - Ubicacion -> tipo:
#       QE + digito, P-EST-, I-BBR-  -> Estanteria
#       Q + digito (no QE)            -> Rack
#       QP, MAQ, PISOD                -> Piso (cuenta como Rack)
#       resto                         -> Rack (default seguro)
#   - Pedidos AP mixtos Rack/Est: predominio por lineas; empate -> AP_R
#   - Canal CES: pedido MY cuyo Destino corresponde a un concesionario de Base CES
#   - Canales separados: AP_R, AP_E, CES, GT, SG, CAP, MY, LB -- NO unificar
#   - data Derco: solo la col "Canal" se reescribe; resto del archivo intacto
# =============================================================================

import os
import re
import csv
import sys
import glob
import time
import shutil
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from canal_derco_utils import (  # noqa: E402
    canal_principal_derco,
    cargar_base_ces,
    clasificar_ubicacion,
    clasificar_ubicacion_estricta,
    norm_str_canal,
    resolver_canal_con_ces,
)

warnings.filterwarnings("ignore")

# -- RUTAS --------------------------------------------------------------------
BASE = r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA"

DATA_DERCO = BASE + r"\Datos para Dashboard - NNSS Operacional\Quilicura\data Derco.xlsx"
BASE_CES = BASE + r"\Datos para Dashboard - Productividad\Archivos Soporte\Base CES.xlsx"
MOV_GLOB = BASE + r"\Datos para Dashboard - Productividad\CD QUILICURA\*\*\MovDerco.xlsx"

HOJA_DD = "Seguimiento de pedidos"
COL_CANAL = "Canal"
COL_OP = "Nro Aplica"
COL_ESTADO = "Estado Pedido"
ESTADO_FINAL = "Con Salida"  # una OP es definitiva al llegar a este estado

BACKUP_DIR = Path(__file__).parent / "_backups_data_derco"
LOG_DIR = Path(r"C:\ClaudeWork\logs")
METRICAS_CSV = LOG_DIR / "canal_derco_metricas.csv"
MAX_BACKUPS = 5

# Acumula el reporte para volcarlo a log al final
_REPORTE: list[str] = []


def log(msg: str = "") -> None:
    print(msg)
    _REPORTE.append(msg)


# -- HELPERS ------------------------------------------------------------------

# norm_str_canal y canal_principal_derco viven en canal_derco_utils.py (Fase 2:
# compartidos con generar_resumen_kpi_ops.py para que productividad clasifique
# CES igual que FillRate). Alias local para retrocompatibilidad de este modulo.
norm_str = norm_str_canal


def fix_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Corrige nombres de columna mal codificados (latin1) por las tildes."""
    df.columns = [str(c).encode("latin1", "ignore").decode("latin1") for c in df.columns]
    return df


def find_col(df: pd.DataFrame, pattern: str):
    matches = [c for c in df.columns if re.search(pattern, c, re.IGNORECASE)]
    return matches[0] if matches else None


# canal_principal_derco y cargar_base_ces viven en canal_derco_utils.py (Fase 2).
# Wrapper local cargar_ces para preservar la firma usada en main() de este modulo:
# antes no aceptaba log_callback, ahora delegamos al log local del modulo.

def cargar_ces(path: str):
    """Wrapper de cargar_base_ces (canal_derco_utils) que enchufa el log local."""
    return cargar_base_ces(path, log_callback=log)


# -- 2. MOVDERCO --------------------------------------------------------------

def cargar_movderco(patron_glob: str) -> pd.DataFrame:
    """Lee y consolida todos los MovDerco.xlsx. Retorna cols: OP, Comprobante_ext, Destino, Tipo_ubic."""
    archivos = sorted(glob.glob(patron_glob))
    if not archivos:
        raise FileNotFoundError(f"No se encontraron MovDerco en: {patron_glob}")

    log(f"  Archivos MovDerco encontrados: {len(archivos)}")
    partes = []

    for ruta in archivos:
        mes_dir = os.path.basename(os.path.dirname(ruta))
        try:
            xl = pd.ExcelFile(ruta)
            raw = pd.read_excel(ruta, sheet_name=xl.sheet_names[0], header=8)
            raw = fix_cols(raw)

            col_op = find_col(raw, r"^Comprobante$")
            col_ext = find_col(raw, r"Comprobante externo")
            col_dest = find_col(raw, r"^Destino$")
            col_ubic = find_col(raw, r"bica")
            if not all([col_op, col_ext, col_dest, col_ubic]):
                log(f"    {mes_dir}: columnas faltantes -- omitido")
                continue

            raw = raw[
                raw[col_op].notna()
                & (raw[col_op].astype(str).str.strip() != "")
                & (raw[col_op].astype(str) != "Comprobante")
                & (~raw[col_op].astype(str).str.startswith("El reporte"))
            ].copy()
            if len(raw) == 0:
                log(f"    {mes_dir}: vacio -- omitido")
                continue

            sub = raw[[col_op, col_ext, col_dest, col_ubic]].copy()
            sub.columns = ["OP", "Comprobante_ext", "Destino", "Ubicacion"]
            sub["OP"] = (
                sub["OP"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            )
            sub = sub[sub["OP"].str.match(r"^\d+$")]
            sub["Tipo_ubic"] = sub["Ubicacion"].apply(clasificar_ubicacion)
            partes.append(sub)
            log(f"    {mes_dir}: {len(sub):,} lineas validas")
        except Exception as e:  # noqa: BLE001
            log(f"    {mes_dir}: ERROR -- {e}")

    if not partes:
        raise ValueError("No se pudo leer ningun MovDerco.")

    df = pd.concat(partes, ignore_index=True)
    log(f"  Total MovDerco consolidado: {len(df):,} lineas")
    return df


# -- 3. RESUMEN POR OP --------------------------------------------------------

def construir_resumen_op(mov: pd.DataFrame, matcher) -> dict:
    """Una entrada por OP: canal_ppal, rack_lines, est_lines, es_ces, ces_nombre."""
    mov = mov.copy()

    # Canal principal por linea (luego se toma el mas frecuente por OP)
    mov["canal_ppal"] = [
        canal_principal_derco(ce, de)
        for ce, de in zip(mov["Comprobante_ext"], mov["Destino"])
    ]

    # CES: resolver sobre destinos unicos (mas barato) y mapear
    destinos_unicos = mov["Destino"].dropna().astype(str).unique()
    mapa_ces = {d: matcher(d) for d in destinos_unicos}
    mov["ces_nombre"] = mov["Destino"].astype(str).map(mapa_ces)

    rack = mov[mov["Tipo_ubic"].isin(["RACK", "PISO"])].groupby("OP").size().rename("rack_lines")
    est = mov[mov["Tipo_ubic"] == "EST"].groupby("OP").size().rename("est_lines")
    canal_ppal = (
        mov.groupby("OP")["canal_ppal"].agg(lambda x: x.value_counts().idxmax()).rename("canal_ppal")
    )
    ces_nombre = (
        mov.groupby("OP")["ces_nombre"]
        .agg(lambda x: next((v for v in x if pd.notna(v)), None))
        .rename("ces_nombre")
    )

    resumen = pd.concat([canal_ppal, rack, est, ces_nombre], axis=1).reset_index()
    resumen["rack_lines"] = resumen["rack_lines"].fillna(0).astype(int)
    resumen["est_lines"] = resumen["est_lines"].fillna(0).astype(int)
    resumen["es_ces"] = resumen["ces_nombre"].notna()

    log(f"  OPs unicos en MovDerco: {len(resumen):,}")
    return resumen.set_index("OP").to_dict("index")


# resolver_canal vive en canal_derco_utils.resolver_canal_con_ces (Fase 2).
# Alias local para retrocompatibilidad.
resolver_canal = resolver_canal_con_ces


# -- 4. REESCRIBIR data Derco -------------------------------------------------

def hacer_backup(path_dd: str) -> Path:
    BACKUP_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = BACKUP_DIR / f"data Derco_BACKUP_{ts}.xlsx"
    shutil.copy2(path_dd, destino)
    # Conservar solo los ultimos MAX_BACKUPS
    backups = sorted(BACKUP_DIR.glob("data Derco_BACKUP_*.xlsx"))
    for viejo in backups[:-MAX_BACKUPS]:
        viejo.unlink(missing_ok=True)
    log(f"  Backup creado: {destino}")
    return destino


def reescribir_canal(path_dd: str, nuevos_canales: list) -> int:
    """Reescribe SOLO la columna Canal de la hoja HOJA_DD; el resto queda intacto."""
    wb = openpyxl.load_workbook(path_dd)
    ws = wb[HOJA_DD]

    canal_col_idx = None
    for cell in ws[1]:
        if str(cell.value).strip() == COL_CANAL:
            canal_col_idx = cell.column
            break
    if canal_col_idx is None:
        raise ValueError(f"Columna '{COL_CANAL}' no encontrada en la hoja '{HOJA_DD}'")

    for i, valor in enumerate(nuevos_canales):
        ws.cell(row=i + 2, column=canal_col_idx).value = valor

    wb.save(path_dd)
    return len(nuevos_canales)


def registrar_metricas(metricas: dict) -> None:
    """Anexa una fila al CSV de metricas para seguir la tendencia de tiempos."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    nuevo = not METRICAS_CSV.exists()
    with METRICAS_CSV.open("a", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        if nuevo:
            w.writerow(list(metricas.keys()))
        w.writerow(list(metricas.values()))


# -- MAIN ---------------------------------------------------------------------

def main() -> None:
    dry_run = "--dry-run" in sys.argv
    inicio = datetime.now()
    t0 = time.perf_counter()
    tiempos: dict[str, float] = {}
    log("=" * 70)
    log(f"CANAL DERCO AUTO -- inicio {inicio:%Y-%m-%d %H:%M:%S}"
        + ("  [DRY-RUN: no se escribe nada]" if dry_run else ""))
    log("=" * 70)

    log("\n[1/4] Cargando Base CES...")
    t = time.perf_counter()
    ces_set, matcher = cargar_ces(BASE_CES)
    tiempos["base_ces"] = time.perf_counter() - t

    log("\n[2/4] Cargando MovDerco (todos los meses)...")
    t = time.perf_counter()
    mov = cargar_movderco(MOV_GLOB)
    tiempos["movderco"] = time.perf_counter() - t

    log("\n[3/4] Calculando canales por OP...")
    t = time.perf_counter()
    resumen = construir_resumen_op(mov, matcher)
    tiempos["resumen_op"] = time.perf_counter() - t

    log("\n[4/4] Actualizando data Derco...")
    t = time.perf_counter()
    dd = fix_cols(pd.read_excel(DATA_DERCO, sheet_name=HOJA_DD))
    if COL_OP not in dd.columns or COL_CANAL not in dd.columns:
        raise ValueError(f"data Derco no tiene las columnas '{COL_OP}' / '{COL_CANAL}'")

    def op_str(v):
        try:
            return str(int(float(str(v).strip())))
        except (ValueError, TypeError):
            return str(v).strip()

    nuevos = []
    sin_match = 0
    for _, fila in dd.iterrows():
        op = op_str(fila[COL_OP])
        info = resumen.get(op)
        if info is None:
            nuevos.append(fila[COL_CANAL])  # OP sin picking en MovDerco -> se conserva
            sin_match += 1
            continue
        nuevos.append(
            resolver_canal(info["canal_ppal"], info["rack_lines"], info["est_lines"], info["es_ces"])
        )
    tiempos["lectura_calculo_dd"] = time.perf_counter() - t

    # Comparacion antes/despues
    antes = dd[COL_CANAL].astype(str).value_counts().to_dict()
    despues = pd.Series(nuevos).astype(str).value_counts().to_dict()
    log("\n  Comparacion Canal (antes -> despues):")
    for c in sorted(set(antes) | set(despues)):
        a, d = antes.get(c, 0), despues.get(c, 0)
        log(f"    {c:10s}: {a:7,} -> {d:7,}  ({d - a:+,})")
    log(f"\n  OPs sin picking en MovDerco (canal conservado): {sin_match:,}")

    # Diagnostico CES
    ces_asignados = [r["ces_nombre"] for r in resumen.values() if r.get("ces_nombre")]
    log(f"  Concesionarios CES detectados: {pd.Series(ces_asignados).nunique()} distintos, "
        f"{len(ces_asignados):,} OPs")
    ces_sin_uso = sorted(ces_set - set(ces_asignados))
    if ces_sin_uso:
        log(f"  [!] {len(ces_sin_uso)} concesionarios de Base CES sin pedidos en el periodo "
            f"(revisar truncado de Destino):")
        for n in ces_sin_uso[:30]:
            log(f"        - {n}")
        if len(ces_sin_uso) > 30:
            log(f"        ... y {len(ces_sin_uso) - 30} mas")

    # Backup + escritura
    t = time.perf_counter()
    if dry_run:
        log("\n  [DRY-RUN] data Derco NO fue modificado.")
    else:
        hacer_backup(DATA_DERCO)
        n = reescribir_canal(DATA_DERCO, nuevos)
        log(f"\n  Filas actualizadas en data Derco: {n:,}")
    tiempos["backup_escritura"] = time.perf_counter() - t

    total = time.perf_counter() - t0

    # Filas provisionales: OP aun no "Con Salida" -> su Canal puede cambiar en corridas futuras
    if COL_ESTADO in dd.columns:
        provisionales = int((dd[COL_ESTADO].astype(str).str.strip() != ESTADO_FINAL).sum())
    else:
        provisionales = -1  # columna ausente

    # Auditoria: ubicaciones que cayeron en el default catch-all (RACK por descarte).
    # Si aparece > 0 hay que agregar regla explicita en canal_derco_utils.py.
    ubic_unicas = mov["Ubicacion"].dropna().astype(str).unique()
    ubic_default = sorted({u for u in ubic_unicas if clasificar_ubicacion_estricta(u) is None})
    if ubic_default:
        log(f"  [!] {len(ubic_default)} ubicaciones nuevas SIN regla explicita (cayeron en default RACK):")
        for u in ubic_default[:10]:
            log(f"        - {u}")
        if len(ubic_default) > 10:
            log(f"        ... y {len(ubic_default) - 10} mas")

    # Metricas de ejecucion -- para observar como crece el tiempo con los registros
    log("\n  Metricas de ejecucion (seguimiento del crecimiento):")
    log(f"    MovDerco lineas leidas : {len(mov):>9,}")
    log(f"    OPs en MovDerco        : {len(resumen):>9,}")
    log(f"    Filas data Derco       : {len(dd):>9,}")
    log(f"    Filas provisionales    : {provisionales:>9,}  (Estado != '{ESTADO_FINAL}', Canal no definitivo)")
    log(f"    Ubic. sin regla explic.: {len(ubic_default):>9,}  (default catch-all -- agregar regla si crece)")
    log(f"    t Base CES             : {tiempos['base_ces']:>8.1f}s")
    log(f"    t MovDerco (carga)     : {tiempos['movderco']:>8.1f}s")
    log(f"    t Resumen por OP       : {tiempos['resumen_op']:>8.1f}s")
    log(f"    t Lectura+calculo DD   : {tiempos['lectura_calculo_dd']:>8.1f}s")
    log(f"    t Backup+escritura     : {tiempos['backup_escritura']:>8.1f}s")
    log(f"    t TOTAL                : {total:>8.1f}s")

    registrar_metricas({
        "fecha": inicio.strftime("%Y-%m-%d %H:%M:%S"),
        "modo": "dry-run" if dry_run else "real",
        "movderco_lineas": len(mov),
        "ops_movderco": len(resumen),
        "filas_data_derco": len(dd),
        "filas_provisionales": provisionales,
        "ubicaciones_sin_regla": len(ubic_default),
        "t_base_ces_s": round(tiempos["base_ces"], 1),
        "t_movderco_s": round(tiempos["movderco"], 1),
        "t_resumen_op_s": round(tiempos["resumen_op"], 1),
        "t_lectura_calculo_dd_s": round(tiempos["lectura_calculo_dd"], 1),
        "t_backup_escritura_s": round(tiempos["backup_escritura"], 1),
        "t_total_s": round(total, 1),
    })

    log("\n" + "=" * 70)
    log(f"COMPLETADO en {total:.1f}s")
    log("=" * 70)

    # Volcar log
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_path = LOG_DIR / f"canal_derco_{inicio:%Y-%m-%d_%H%M%S}.log"
    log_path.write_text("\n".join(_REPORTE), encoding="utf-8")
    print(f"\nLog guardado en: {log_path}")
    print(f"Metricas acumuladas en: {METRICAS_CSV}")


if __name__ == "__main__":
    main()
