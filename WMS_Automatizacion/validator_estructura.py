
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from datetime import datetime
import unicodedata
import re
import json
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


LOG_DIR = Path(r"C:\ClaudeWork\logs\validaciones_estructura")
LOG_DIR.mkdir(parents=True, exist_ok=True)

RUTA_STOCK_WMS = Path(
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA\Datos para Dashboard - Stock WMS Semanal"
)
RUTA_POSICIONES = Path(
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA\Datos para Dashboard - Consulta de Posiciones"
)
RUTA_STAGING = Path(
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA\Datos para Dashboard - Stagin IN- OUT"
)
RUTA_CLIENTES_EK = Path(
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA\Datos para Dashboard - Clientes EK"
)


def strip_accents(text: str) -> str:
    text = str(text)
    return ''.join(
        c for c in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(c)
    )


def fix_common_visual_encoding(text: str) -> str:
    if text is None:
        return ""
    s = str(text)
    replacements = {
        "descripci¢n": "descripcion",
        "descripciã³n": "descripcion",
        "descripciã“n": "descripcion",
        "ubicaci¢n": "ubicacion",
        "ubicaciã³n": "ubicacion",
        "recepci¢n": "recepcion",
        "recepciã³n": "recepcion",
        "preparaci¢n": "preparacion",
        "preparaciã³n": "preparacion",
        "generaci¢n": "generacion",
        "generaciã³n": "generacion",
        "reubicaci¢n": "reubicacion",
        "reubicaciã³n": "reubicacion",
        "despachad¢": "despachado",
        "vencimient¢": "vencimiento",
        "articul¢": "articulo",
        "deposit¢": "deposito",
    }
    s_low = s.lower()
    for bad, good in replacements.items():
        s_low = s_low.replace(bad, good)
    return s_low


def norm_name(text: Any) -> str:
    text = "" if text is None else str(text)

    # 1) reparar basura visual antes de cualquier otra cosa
    text = fix_common_visual_encoding(text)

    # 2) normalizar prefijos tipo N°, Nº, N°., Nº. ANTES de strip_accents
    #    para evitar que Nº -> No
    text = re.sub(r"(?i)\bn[°º]\.?\s*", "n", text)
    text = text.replace("№", "n")

    # 3) ahora sí quitar acentos
    text = strip_accents(text)

    # 4) limpiar símbolos sueltos que puedan quedar
    text = text.replace("°", "")
    text = text.replace("º", "")

    # 5) casos especiales ya usados en el sistema
    text = text.replace("fh.", "fh ")
    text = text.replace("descripcin", "descripcion")

    # 6) normalización general
    text = text.replace(".", " ")
    text = text.replace("/", " ")
    text = text.replace("-", " ")
    text = text.replace("_", " ")
    text = text.lower()
    text = re.sub(r"\s+", " ", text).strip()
    return text


SCHEMAS: Dict[str, Dict[str, Any]] = {
    "stock_wms": {
        "kind": "excel",
        "header_mode": "detect_in_top_rows",
        "search_window_rows": 15,
        "header_contains_any": [
            "deposito", "empresa", "articulo", "descripcion",
            "contenedor", "sub deposito"
        ],
        "required_columns": [
            "Deposíto", "Empresa", "Artículo", "Descripción",
            "Contenedor", "Sub-Depósito"
        ],
        "optional_columns": [
            "Empresa Dsc.", "Destino", "Apto", "Motivo",
            "Ingreso", "Vencimiento", "Lote", "Bultos",
            "Unids", "U/Bulto", "Kilos", "Peso Recep(Kg)",
            "M3", "Area", "Cara", "Columna", "Nivel",
            "Lote 2", "Lote de Inspección", "Fecha de Elaboración",
            "Nro Despacho", "Código de Referencia", "VH Unitario"
        ],
        "allow_extra_columns": True,
        "allow_zero_rows": True,
        "critical_type_rules": {
            "deposito": "text",
            "empresa": "text",
            "articulo": "text",
            "descripcion": "text",
            "contenedor": "text",
            "sub deposito": "text",
        }
    },

    "staging_estandar": {
        "kind": "csv",
        "delimiter": ";",
        "encoding_candidates": ["utf-8", "latin1", "cp1252"],
        "required_columns": [
            "Deposito", "Empresa", "Codigo Articulo", "Articulo Descripción",
            "Rubro", "Subrubro", "Pallet", "Lote", "Bloqueado",
            "Codigo Estado", "Descripcion Estado", "Fecha de Alta",
            "Fecha de Vencimiento", "Cantidad", "Unidad", "Bultos",
            "Litros", "KGS", "Lugar", "Ubicacion"
        ],
        "optional_columns": [],
        "allow_extra_columns": False,
        "allow_zero_rows": True,
        "critical_type_rules": {
            "deposito": "text",
            "empresa": "text",
            "codigo articulo": "text",
            "articulo descripcion": "text",
            "rubro": "text_or_blank",
            "subrubro": "text_or_blank",
            "pallet": "text_or_blank",
            "lote": "text_or_blank",
            "bloqueado": "text_or_blank",
            "codigo estado": "text_or_blank",
            "descripcion estado": "text_or_blank",
            "fecha de alta": "date_or_blank",
            "fecha de vencimiento": "date_or_blank",
            "cantidad": "numeric_or_blank",
            "unidad": "text_or_blank",
            "bultos": "numeric_or_blank",
            "litros": "numeric_or_blank",
            "kgs": "numeric_or_blank",
            "lugar": "text_or_blank",
            "ubicacion": "text_or_blank",
        }
    },

    "staging_unilever": {
        "kind": "csv",
        "delimiter": ";",
        "encoding_candidates": ["utf-8", "latin1", "cp1252"],
        "required_columns": [
            "Deposito", "Empresa", "Codigo Articulo", "Articulo Descripción",
            "Rubro", "Subrubro", "Pallet", "Lote", "Bloqueado",
            "Codigo Estado", "Descripcion Estado", "Fecha de Alta",
            "Fecha de Vencimiento", "Cantidad", "Unidad", "Bultos",
            "Litros", "KGS", "Lugar", "Ubicacion", "FhElab"
        ],
        "optional_columns": [],
        "allow_extra_columns": False,
        "allow_zero_rows": True,
        "critical_type_rules": {
            "deposito": "text",
            "empresa": "text",
            "codigo articulo": "text",
            "articulo descripcion": "text",
            "rubro": "text_or_blank",
            "subrubro": "text_or_blank",
            "pallet": "text_or_blank",
            "lote": "text_or_blank",
            "bloqueado": "text_or_blank",
            "codigo estado": "text_or_blank",
            "descripcion estado": "text_or_blank",
            # En Unilever estas fechas pueden venir en formato no estándar del export.
            # Se validan de forma más flexible en v4.
            "fecha de alta": "text_or_blank",
            "fecha de vencimiento": "text_or_blank",
            "cantidad": "numeric_or_blank",
            "unidad": "text_or_blank",
            "bultos": "numeric_or_blank",
            "litros": "numeric_or_blank",
            "kgs": "numeric_or_blank",
            "lugar": "text_or_blank",
            "ubicacion": "text_or_blank",
            "fhelab": "text_or_blank",
        }
    },

    "posiciones": {
        "kind": "excel",
        "header_mode": "fixed_row",
        "header_row_1_based": 4,
        "required_columns": [
            "CD", "Area", "Cara", "Columna", "Nivel", "Estado",
            "Prof. Totales", "Prof. Ocupadas", "Prof. Libres",
            "Inhibida", "Mezcla"
        ],
        "optional_columns": [],
        "allow_extra_columns": False,
        "allow_zero_rows": True,
        "critical_type_rules": {
            "cd": "text",
            "area": "text",
            "cara": "text",
            "columna": "text",
            "nivel": "text",
            "estado": "text",
            "prof totales": "numeric_or_blank",
            "prof ocupadas": "numeric_or_blank",
            "prof libres": "numeric_or_blank",
            "inhibida": "text_or_blank",
            "mezcla": "text_or_blank",
        }
    },

    "pedidos_preparados": {
        "kind": "excel",
        "header_mode": "fixed_row",
        "header_row_1_based": 1,
        "required_columns": [
            "Depósito", "Empresa", "Viaje", "Dock", "Destino", "Tipo",
            "OP", "N°Pedido", "N°OC", "Articulo", "Descripción", "Lote",
            "Vencimiento", "Cant.Prep", "Kilos", "M3", "Litros",
            "Pallet Picking", "Hubo Mov. de Reubicación", "Pallet Origen",
            "Ubicación Origen", "Fh. Generación", "Fh Preparación",
            "Fh Despachado", "Fh Salida", "Preparador", "Control Isla"
        ],
        "optional_columns": [],
        "allow_extra_columns": False,
        "allow_zero_rows": True,
        "critical_type_rules": {
            "deposito": "text",
            "empresa": "text",
            "viaje": "text_or_blank",
            "dock": "text_or_blank",
            "destino": "text_or_blank",
            "tipo": "text_or_blank",
            "op": "numeric_or_text",
            "npedido": "numeric_or_text",
            "noc": "text_or_blank",
            "articulo": "text",
            "descripcion": "text",
            "lote": "text_or_blank",
            "vencimiento": "date_or_blank",
            "cant prep": "numeric_or_blank",
            "kilos": "numeric_or_blank",
            "m3": "numeric_or_blank",
            "litros": "numeric_or_blank",
            "pallet picking": "text_or_blank",
            "hubo mov de reubicacion": "text_or_blank",
            "pallet origen": "text_or_blank",
            "ubicacion origen": "text_or_blank",
            "fh generacion": "date_or_blank",
            "fh preparacion": "date_or_blank",
            "fh despachado": "date_or_blank",
            "fh salida": "date_or_blank",
            "preparador": "text_or_blank",
            "control isla": "text_or_blank",
        }
    },

    "recepciones_recibidas": {
        "kind": "excel",
        "header_mode": "fixed_row",
        "header_row_1_based": 1,
        "required_columns": [
            "CD", "Empresa", "Viaje", "Dock", "Tipo de Origen", "Origen",
            "Tipo", "OP", "N° Recepción", "N° OC", "Articulo", "Descripción",
            "Lote", "Vencimiento", "Cantidad Recibida", "Kilos", "M3",
            "Litros", "Pallet", "Ubicación Guardado", "Fh. Generación",
            "Fh. Inicio de Recepción", "Fh. Fin de Recepción",
            "Fh. Inicio de Guardado", "Fh. Fin de Guardado"
        ],
        "optional_columns": [],
        "allow_extra_columns": False,
        "allow_zero_rows": True,
        "critical_type_rules": {
            "cd": "text",
            "empresa": "text",
            "viaje": "text_or_blank",
            "dock": "text_or_blank",
            "tipo de origen": "text_or_blank",
            "origen": "text_or_blank",
            "tipo": "text_or_blank",
            "op": "numeric_or_text",
            "n recepcion": "numeric_or_text",
            "n oc": "text_or_blank",
            "articulo": "text",
            "descripcion": "text",
            "lote": "text_or_blank",
            "vencimiento": "date_or_blank",
            "cantidad recibida": "numeric_or_blank",
            "kilos": "numeric_or_blank",
            "m3": "numeric_or_blank",
            "litros": "numeric_or_blank",
            "pallet": "text_or_blank",
            "ubicacion guardado": "text_or_blank",
            "fh generacion": "date_or_blank",
            "fh inicio de recepcion": "date_or_blank",
            "fh fin de recepcion": "date_or_blank",
            "fh inicio de guardado": "date_or_blank",
            "fh fin de guardado": "date_or_blank",
        }
    },
}


def read_csv_flexible(path: Path, delimiter: str, encoding_candidates: List[str]) -> Tuple[pd.DataFrame, str]:
    last_error = None
    for enc in encoding_candidates:
        try:
            df = pd.read_csv(
                path,
                sep=delimiter,
                dtype=str,
                encoding=enc,
                engine="python",
                keep_default_na=False,
                on_bad_lines="error"
            )
            return df, enc
        except Exception as e:
            last_error = e
    raise RuntimeError(f"No se pudo leer CSV con los encodings probados: {last_error}")


def detect_header_row_excel(path: Path, search_window_rows: int, header_contains_any: List[str]) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    targets = set(header_contains_any)
    best_row = 1
    best_score = -1

    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=search_window_rows, values_only=True),
        start=1
    ):
        normalized = [norm_name(v) for v in row if v is not None and str(v).strip() != ""]
        score = sum(1 for x in normalized if x in targets)
        if score > best_score:
            best_score = score
            best_row = idx

    return best_row


def read_excel_schema(path: Path, header_row_1_based: int) -> pd.DataFrame:
    return pd.read_excel(path, header=header_row_1_based - 1, dtype=str)


def header_analysis(df: pd.DataFrame, schema: Dict[str, Any]) -> Dict[str, Any]:
    actual_original = list(df.columns)
    actual_norm = [norm_name(c) for c in actual_original]
    required_norm = [norm_name(c) for c in schema["required_columns"]]
    optional_norm = [norm_name(c) for c in schema.get("optional_columns", [])]

    actual_set = set(actual_norm)
    required_set = set(required_norm)
    allowed_set = required_set | set(optional_norm)

    missing = [c for c in required_norm if c not in actual_set]
    extras = [c for c in actual_norm if c not in allowed_set] if not schema.get("allow_extra_columns", False) else []

    duplicates = []
    seen = set()
    for c in actual_norm:
        if c in seen and c not in duplicates:
            duplicates.append(c)
        seen.add(c)

    return {
        "actual_original": actual_original,
        "actual_norm": actual_norm,
        "missing_required": missing,
        "extra_columns": extras,
        "duplicate_columns": duplicates,
    }


DATE_PATTERNS = [
    re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}$"),
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$"),
]

def is_blank(v: Any) -> bool:
    return v is None or str(v).strip() == "" or str(v).strip().lower() in {"nan", "nat", "none"}

def looks_date(v: Any) -> bool:
    if is_blank(v):
        return True
    s = str(v).strip()
    if isinstance(v, (pd.Timestamp, datetime)):
        return True
    return any(p.match(s) for p in DATE_PATTERNS)

def looks_numeric(v: Any) -> bool:
    if is_blank(v):
        return True
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        float(s)
        return True
    except Exception:
        return False

def looks_text(v: Any) -> bool:
    return not is_blank(v)

def value_matches_rule(v: Any, rule: str) -> bool:
    if rule == "text":
        return looks_text(v)
    if rule == "text_or_blank":
        return True if is_blank(v) else looks_text(v)
    if rule == "date_or_blank":
        return looks_date(v)
    if rule == "numeric_or_blank":
        return looks_numeric(v)
    if rule == "numeric_or_text":
        return True if is_blank(v) else (looks_numeric(v) or looks_text(v))
    return True


def _is_total_row(row_values: List[Any]) -> bool:
    nonblank = [str(v).strip() for v in row_values if not is_blank(v)]
    if not nonblank:
        return False
    # Si solo tiene valor en la primera celda y comienza con "Total"
    first = nonblank[0].lower()
    return first.startswith("total")


def row_shift_suspicion(df: pd.DataFrame, schema: Dict[str, Any], sample_limit: int = 500) -> List[str]:
    issues: List[str] = []
    rules = schema.get("critical_type_rules", {})
    if not rules:
        return issues

    col_map = {norm_name(c): c for c in df.columns}
    inspect_cols = [c for c in rules.keys() if c in col_map]
    if not inspect_cols:
        return issues

    limit = min(len(df), sample_limit)
    if limit == 0:
        return issues

    schema_name = schema.get("_schema_name", "")
    min_bad = 3
    if schema_name == "stock_wms":
        min_bad = 4

    for idx in range(limit):
        row = df.iloc[idx]
        row_list = row.tolist()

        nonblank_count = sum(0 if is_blank(v) else 1 for v in row_list)
        if nonblank_count == 0:
            continue

        # v4: ignorar filas-resumen/totales al final del archivo
        if schema_name == "stock_wms" and _is_total_row(row_list):
            continue

        bad_cols = []
        for norm_col in inspect_cols:
            original_col = col_map[norm_col]
            v = row[original_col]
            if not value_matches_rule(v, rules[norm_col]):
                bad_cols.append(norm_col)

        if len(bad_cols) >= min_bad:
            issues.append(
                f"Fila {idx + 2}: posible corrimiento/desalineacion. Columnas sospechosas: {', '.join(bad_cols[:8])}"
            )

        if len(issues) >= 25:
            break

    return issues


def validate_structure(path: str, schema_name: str) -> Dict[str, Any]:
    file_path = Path(path)
    schema = dict(SCHEMAS[schema_name])
    schema["_schema_name"] = schema_name

    result: Dict[str, Any] = {
        "archivo": str(file_path),
        "schema": schema_name,
        "estado": "OK",
        "sheet_used": None,
        "header_row_1_based": None,
        "row_count": None,
        "column_count": None,
        "errores": [],
        "warnings": [],
        "notas": [],
        "actual_columns": [],
    }

    if not file_path.exists():
        result["estado"] = "ERROR"
        result["errores"].append("El archivo no existe.")
        return result

    try:
        if schema["kind"] == "csv":
            df, enc = read_csv_flexible(
                file_path,
                schema["delimiter"],
                schema.get("encoding_candidates", ["utf-8", "latin1", "cp1252"])
            )
            result["notas"].append(f"CSV leido con encoding: {enc}")
            result["header_row_1_based"] = 1
            result["sheet_used"] = None
        else:
            if schema["header_mode"] == "fixed_row":
                header_row = schema["header_row_1_based"]
            else:
                header_row = detect_header_row_excel(
                    file_path,
                    schema["search_window_rows"],
                    [norm_name(x) for x in schema["header_contains_any"]]
                )
            result["header_row_1_based"] = header_row
            result["sheet_used"] = "Sheet1"
            df = read_excel_schema(file_path, header_row)

        result["row_count"] = len(df)
        result["column_count"] = len(df.columns)

        head = header_analysis(df, schema)
        result["actual_columns"] = head["actual_original"]

        if head["missing_required"]:
            result["errores"].append(
                "Faltan columnas requeridas: " + ", ".join(head["missing_required"])
            )

        if head["duplicate_columns"]:
            result["errores"].append(
                "Columnas duplicadas detectadas: " + ", ".join(head["duplicate_columns"])
            )

        if head["extra_columns"]:
            result["warnings"].append(
                "Columnas extra no esperadas: " + ", ".join(head["extra_columns"])
            )

        if len(df) == 0:
            if schema.get("allow_zero_rows", False):
                result["warnings"].append("Archivo sin registros, pero con estructura valida.")
            else:
                result["errores"].append("Archivo sin registros.")

        shift_issues = row_shift_suspicion(df, schema)
        if shift_issues:
            result["errores"].extend(shift_issues)

        if result["errores"]:
            result["estado"] = "ERROR"
        elif result["warnings"]:
            result["estado"] = "WARNING"
        else:
            result["estado"] = "OK"

    except Exception as e:
        result["estado"] = "ERROR"
        result["errores"].append(f"Error procesando archivo: {e}")

    return result


def _latest_match(folder: Path, pattern: str, name_contains: Optional[str] = None) -> Optional[Path]:
    if not folder.exists():
        return None
    files = [p for p in folder.glob(pattern) if p.is_file()]
    if name_contains:
        files = [p for p in files if name_contains.lower() in p.name.lower()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def _latest_recursive(folder: Path, pattern: str, name_contains: Optional[str] = None) -> Optional[Path]:
    if not folder.exists():
        return None
    files = [p for p in folder.rglob(pattern) if p.is_file()]
    if name_contains:
        files = [p for p in files if name_contains.lower() in p.name.lower()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def encontrar_stock_wms() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []

    q = _latest_match(RUTA_STOCK_WMS / "Quilicura", "*.xlsx")
    if q:
        resultados.append((q, "stock_wms"))

    folder = RUTA_STOCK_WMS / "Pudahuel"
    if folder.exists():
        files = sorted([p for p in folder.glob("*.xlsx") if p.is_file()], key=lambda p: p.stat().st_mtime, reverse=True)
        usados = set()
        for p in files:
            if p.name not in usados:
                resultados.append((p, "stock_wms"))
                usados.add(p.name)
            if len(usados) >= 2:
                break

    return resultados


def encontrar_staging() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []

    clientes_quilicura = [
        "ABINBEV", "DAIKIN", "DAIKIN CLIENTES", "DERCO",
        "MASCOTAS LATINAS", "POCHTECA"
    ]
    clientes_pudahuel = [
        "BARENTZ", "BURASCHI", "CEPAS CHILE", "COLLICO", "DELIBEST",
        "INTIME", "NATIVOS DRINK", "TRES MONTE", "UNILEVER", "RUNO"
    ]

    for cliente in clientes_quilicura:
        f = _latest_match(RUTA_STAGING / "Quilicura" / cliente, "*.csv")
        if f:
            schema = "staging_unilever" if cliente.upper() == "UNILEVER" else "staging_estandar"
            resultados.append((f, schema))

    for cliente in clientes_pudahuel:
        f = _latest_match(RUTA_STAGING / "Pudahuel" / cliente, "*.csv")
        if f:
            schema = "staging_unilever" if cliente.upper() == "UNILEVER" else "staging_estandar"
            resultados.append((f, schema))

    return resultados


def encontrar_posiciones() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []

    archivos = [
        RUTA_POSICIONES / "Quilicura" / "Posiciones Ocupadas.xlsx",
        RUTA_POSICIONES / "Quilicura" / "Posiciones Libres.xlsx",
        RUTA_POSICIONES / "Pudahuel" / "Posiciones Ocupadas Moderno.xlsx",
        RUTA_POSICIONES / "Pudahuel" / "Posiciones Libres Moderno.xlsx",
        RUTA_POSICIONES / "Pudahuel" / "Posiciones Ocupadas Unitario.xlsx",
        RUTA_POSICIONES / "Pudahuel" / "Posiciones Libres Unitario.xlsx",
        RUTA_POSICIONES / "Pudahuel" / "Posiciones Ocupadas Refrigerado.xlsx",
        RUTA_POSICIONES / "Pudahuel" / "Posiciones Libres Refrigerado.xlsx",
    ]

    for f in archivos:
        if f.exists():
            resultados.append((f, "posiciones"))

    return resultados


def encontrar_clientes_ek_preparacion() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []
    clientes = ["ABINBEV", "DAIKIN", "DERCO", "MASCOTAS LATINAS", "POCHTECA"]
    for cliente in clientes:
        carpeta = RUTA_CLIENTES_EK / cliente / "Preparación"
        f = _latest_recursive(carpeta, "*.xlsx", "Pedidos Preparados")
        if f:
            resultados.append((f, "pedidos_preparados"))
    return resultados


def encontrar_clientes_ek_recepciones() -> List[Tuple[Path, str]]:
    resultados: List[Tuple[Path, str]] = []
    clientes = ["ABINBEV", "DAIKIN", "DERCO", "MASCOTAS LATINAS", "POCHTECA"]
    for cliente in clientes:
        carpeta = RUTA_CLIENTES_EK / cliente / "Recepciones"
        f = _latest_recursive(carpeta, "*.xlsx")
        if f:
            resultados.append((f, "recepciones_recibidas"))
    return resultados


def save_results_txt(results: List[Dict[str, Any]]) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = LOG_DIR / f"estructura_{ts}.txt"

    ok = sum(1 for r in results if r["estado"] == "OK")
    warn = sum(1 for r in results if r["estado"] == "WARNING")
    err = sum(1 for r in results if r["estado"] == "ERROR")

    lines: List[str] = []
    lines.append("RESUMEN VALIDACION DE ESTRUCTURA")
    lines.append("=" * 90)
    lines.append(f"Fecha ejecucion : {datetime.now().isoformat()}")
    lines.append(f"Total revisados : {len(results)}")
    lines.append(f"OK              : {ok}")
    lines.append(f"WARNING         : {warn}")
    lines.append(f"ERROR           : {err}")
    lines.append("")

    for r in results:
        lines.append(f"[{r['estado']}] {r['schema']} | {r['archivo']}")
        lines.append(f"   Header row : {r['header_row_1_based']}")
        lines.append(f"   Filas      : {r['row_count']}")
        lines.append(f"   Columnas   : {r['column_count']}")
        if r["warnings"]:
            for w in r["warnings"]:
                lines.append(f"   - WARNING: {w}")
        if r["errores"]:
            for e in r["errores"]:
                lines.append(f"   - ERROR: {e}")
        if r["notas"]:
            for n in r["notas"]:
                lines.append(f"   - NOTA: {n}")
        lines.append("")

    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def save_results_json(results: List[Dict[str, Any]]) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = LOG_DIR / f"estructura_{ts}.json"
    payload = {
        "timestamp": datetime.now().isoformat(),
        "results": results
    }
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def _self_test_norm_name() -> None:
    casos = {
        "N°Pedido": "npedido",
        "NºPedido": "npedido",
        "N°OC": "noc",
        "NºOC": "noc",
        "N° Recepción": "nrecepcion",
        "Nº Recepción": "nrecepcion",
    }
    for entrada, esperado in casos.items():
        salida = norm_name(entrada)
        assert salida == esperado, f"{entrada} -> {salida} != {esperado}"


def run_auto_local_examples() -> None:
    _self_test_norm_name()

    examples: List[Tuple[Path, str]] = []
    examples.extend(encontrar_stock_wms())
    examples.extend(encontrar_staging())
    examples.extend(encontrar_posiciones())
    examples.extend(encontrar_clientes_ek_preparacion())
    examples.extend(encontrar_clientes_ek_recepciones())

    if not examples:
        print("No se encontraron archivos reales para validar.")
        return

    print("=" * 90)
    print("VALIDADOR DE ESTRUCTURA WMS EGAKAT - AUTO LOCAL v5")
    print("=" * 90)

    results = []
    for path, schema_name in examples:
        print(f"Validando: {path} | schema={schema_name}")
        results.append(validate_structure(str(path), schema_name))

    txt = save_results_txt(results)
    js = save_results_json(results)

    ok = sum(1 for r in results if r["estado"] == "OK")
    warn = sum(1 for r in results if r["estado"] == "WARNING")
    err = sum(1 for r in results if r["estado"] == "ERROR")

    print("-" * 90)
    print(f"Total: {len(results)} | OK: {ok} | WARNING: {warn} | ERROR: {err}")
    print(f"TXT :  {txt}")
    print(f"JSON:  {js}")


if __name__ == "__main__":
    run_auto_local_examples()
