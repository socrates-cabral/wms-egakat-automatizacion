"""
vdr_comparador.py — v1.0
Comparador de archivos Base VDR — Derco Parts / EGA KAT
Detecta cambios en VDR SAP y VDR FISICO entre el archivo mas reciente
y el anterior, genera Excel de diferencias en OneDrive EGA KAT.

Uso: py vdr_comparador.py
Trigger: Task Scheduler cada 1 hora L-V

Flujo:
  1. Detecta carpeta del mes en curso (dinamica)
  2. Identifica archivo nuevo vs ultimo procesado (estado en .txt)
  3. Compara columnas VDR SAP y VDR FISICO por Material WMS
  4. Valida equivalencia Material WMS <-> Material SAP entre archivos
  5. Genera Reporte_VDR_DDMMYYYY_HHMMSS.xlsx solo si hay diferencias
  6. Power Automate detecta el archivo nuevo y envia correo con adjunto
"""

import sys
import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ─── CONFIGURACION ────────────────────────────────────────────────────────────

BASE_ORIGEN = r"C:\Users\Socrates Cabral\Grupo Planet SpA\José Caceres - Base VDR"
BASE_SALIDA = r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA\Reportes VDR"
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))          # C:\ClaudeWork\VDR_Comparador
ESTADO_FILE = os.path.join(_SCRIPT_DIR, "vdr_ultimo_procesado.txt")
LOGDIR      = os.path.join(os.path.dirname(_SCRIPT_DIR), "logs")  # C:\ClaudeWork\logs

# Columnas a cargar de cada archivo Base VDR
COLUMNAS_REQUERIDAS = [
    "Material WMS",
    "Material SAP",
    "Desc_Material",
    "Categoria",
    "VDR SAP",
    "VDR FISICO",
]

# Columnas VDR a comparar (nuevos vs anterior)
COLUMNAS_VDR = ["VDR SAP", "VDR FISICO"]

# Mapa de nombres de mes en espanol
MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# Estilos Excel
FILL_AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_HEADER   = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
FONT_HEADER   = Font(bold=True, color="FFFFFF")
FONT_BOLD     = Font(bold=True)

# ─── LOG ──────────────────────────────────────────────────────────────────────

os.makedirs(LOGDIR, exist_ok=True)
LOGFILE = os.path.join(LOGDIR, f"vdr_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")


def log(msg):
    """Escribe en consola y archivo de log — sin Unicode, solo ASCII."""
    # Reemplazar simbolos Unicode por equivalentes ASCII
    msg_ascii = (msg
        .replace("\u2192", "->")
        .replace("\u2714", "[OK]")
        .replace("\u2718", "[X]")
        .replace("\u25b6", ">>")
        .replace("\u2705", "[OK]")
        .replace("\u274c", "[ERROR]")
        .replace("\u26a0", "[AVISO]")
    )
    ts = f"[{datetime.now().strftime('%H:%M:%S')}]"
    linea = f"{ts} {msg_ascii}"
    print(linea, flush=True)
    with open(LOGFILE, "a", encoding="utf-8") as f:
        f.write(linea + "\n")


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def carpeta_mes_actual():
    """Retorna el nombre de la carpeta del mes en curso. Ej: '03. Marzo'"""
    hoy = datetime.now()
    return f"{hoy.month:02d}. {MESES_ES[hoy.month]}"


def extraer_fecha_nombre(nombre_archivo):
    """
    Extrae fecha del nombre 'Base VDR DD-MM-YYYY.xlsx'.
    Retorna objeto date o None si no coincide el patron.
    """
    patron = r"Base VDR (\d{2})-(\d{2})-(\d{4})\.xlsx"
    m = re.search(patron, nombre_archivo, re.IGNORECASE)
    if not m:
        return None
    try:
        return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
    except ValueError:
        return None


def listar_archivos_ordenados(carpeta_mes_path):
    """
    Lista archivos Base VDR en la carpeta, ordenados por fecha extraida del nombre.
    Retorna lista de (nombre_archivo, fecha).
    """
    archivos = []
    for f in Path(carpeta_mes_path).iterdir():
        if f.is_file() and f.suffix.lower() == ".xlsx":
            fecha = extraer_fecha_nombre(f.name)
            if fecha:
                archivos.append((f.name, fecha))
    archivos.sort(key=lambda x: x[1])
    return archivos


def leer_estado():
    """
    Lee el estado del ultimo archivo procesado.
    Formato guardado: 'nombre_carpeta_mes|nombre_archivo'
    Retorna (carpeta_mes, nombre_archivo) o (None, None).
    """
    if not os.path.exists(ESTADO_FILE):
        return None, None
    try:
        with open(ESTADO_FILE, "r", encoding="utf-8") as f:
            contenido = f.read().strip()
        if "|" in contenido:
            partes = contenido.split("|", 1)
            return partes[0], partes[1]
    except Exception:
        pass
    return None, None


def escribir_estado(carpeta_mes, nombre_archivo):
    """Guarda el ultimo archivo procesado en ESTADO_FILE."""
    os.makedirs(os.path.dirname(ESTADO_FILE), exist_ok=True)
    with open(ESTADO_FILE, "w", encoding="utf-8") as f:
        f.write(f"{carpeta_mes}|{nombre_archivo}")


def cargar_datos(ruta_xlsx):
    """
    Carga columnas requeridas de un archivo Base VDR.
    Retorna dict: {Material_WMS_str: {col: valor, ...}}
    """
    log(f"  Cargando: {os.path.basename(ruta_xlsx)}")
    wb = openpyxl.load_workbook(ruta_xlsx, read_only=True, data_only=True)
    ws = wb.active

    # Detectar indices de columnas por nombre en fila 1
    encabezados = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
        if cell.value and str(cell.value).strip() in COLUMNAS_REQUERIDAS:
            encabezados[str(cell.value).strip()] = col_idx

    # Verificar que esten todas las columnas requeridas
    faltantes = [c for c in COLUMNAS_REQUERIDAS if c not in encabezados]
    if faltantes:
        log(f"  [AVISO] Columnas no encontradas: {faltantes}")

    datos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat_wms_idx = encabezados.get("Material WMS")
        if not mat_wms_idx:
            continue
        mat_wms = row[mat_wms_idx - 1]
        if mat_wms is None:
            continue
        clave = str(mat_wms).strip()
        registro = {}
        for col_nombre, col_idx in encabezados.items():
            val = row[col_idx - 1]
            registro[col_nombre] = val if val is not None else ""
        datos[clave] = registro

    wb.close()
    log(f"  Registros cargados: {len(datos)}")
    return datos


# ─── EXCEL SALIDA ─────────────────────────────────────────────────────────────

def estilo_encabezado(ws, fila=1):
    """Aplica estilo a la fila de encabezado."""
    for cell in ws[fila]:
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws):
    """Ajusta ancho de columnas segun contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def escribir_hoja_diferencias(wb, datos_ant, datos_nuevo, nombre_ant, nombre_nuevo):
    """
    Crea hoja 'Diferencias_VDR' con SKUs donde cambio VDR SAP o VDR FISICO.
    Retorna cantidad de diferencias encontradas.
    """
    ws = wb.create_sheet("Diferencias_VDR")
    encabezados = [
        "Material WMS", "Material SAP", "Desc_Material", "Categoria",
        "VDR_SAP_anterior", "VDR_SAP_nuevo", "Delta_VDR_SAP",
        "VDR_FISICO_anterior", "VDR_FISICO_nuevo", "Delta_VDR_FISICO",
        "Archivo_anterior", "Archivo_nuevo",
    ]
    ws.append(encabezados)
    estilo_encabezado(ws)

    count = 0
    for mat_wms, reg_nuevo in datos_nuevo.items():
        if mat_wms not in datos_ant:
            continue
        reg_ant = datos_ant[mat_wms]

        vdr_sap_ant   = reg_ant.get("VDR SAP", "")
        vdr_sap_nuevo = reg_nuevo.get("VDR SAP", "")
        vdr_fis_ant   = reg_ant.get("VDR FISICO", "")
        vdr_fis_nuevo = reg_nuevo.get("VDR FISICO", "")

        delta_sap = "" if vdr_sap_ant == vdr_sap_nuevo else f"{vdr_sap_ant} -> {vdr_sap_nuevo}"
        delta_fis = "" if vdr_fis_ant == vdr_fis_nuevo else f"{vdr_fis_ant} -> {vdr_fis_nuevo}"

        if not delta_sap and not delta_fis:
            continue

        fila = [
            mat_wms,
            reg_nuevo.get("Material SAP", ""),
            reg_nuevo.get("Desc_Material", ""),
            reg_nuevo.get("Categoria", ""),
            vdr_sap_ant, vdr_sap_nuevo, delta_sap,
            vdr_fis_ant, vdr_fis_nuevo, delta_fis,
            nombre_ant, nombre_nuevo,
        ]
        ws.append(fila)
        # Resaltar fila en amarillo
        for cell in ws[ws.max_row]:
            cell.fill = FILL_AMARILLO
        count += 1

    autofit(ws)
    return count


def escribir_hoja_equivalencia(wb, datos_ant, datos_nuevo):
    """
    Crea hoja 'Cambios_Equivalencia' con SKUs donde cambio Material SAP.
    Retorna cantidad de cambios.
    """
    ws = wb.create_sheet("Cambios_Equivalencia")
    encabezados = [
        "Material WMS", "Material SAP (anterior)", "Material SAP (nuevo)", "Desc_Material",
    ]
    ws.append(encabezados)
    estilo_encabezado(ws)

    count = 0
    for mat_wms, reg_nuevo in datos_nuevo.items():
        if mat_wms not in datos_ant:
            continue
        sap_ant   = str(datos_ant[mat_wms].get("Material SAP", "")).strip()
        sap_nuevo = str(reg_nuevo.get("Material SAP", "")).strip()
        if sap_ant != sap_nuevo:
            ws.append([
                mat_wms, sap_ant, sap_nuevo,
                reg_nuevo.get("Desc_Material", ""),
            ])
            for cell in ws[ws.max_row]:
                cell.fill = FILL_AMARILLO
            count += 1

    autofit(ws)
    return count


def escribir_hoja_skus(wb, nombre_hoja, registros):
    """Crea hoja con lista de SKUs (nuevos o eliminados)."""
    ws = wb.create_sheet(nombre_hoja)
    encabezados = ["Material WMS", "Material SAP", "Desc_Material", "Categoria"]
    ws.append(encabezados)
    estilo_encabezado(ws)
    for mat_wms, reg in registros.items():
        ws.append([
            mat_wms,
            reg.get("Material SAP", ""),
            reg.get("Desc_Material", ""),
            reg.get("Categoria", ""),
        ])
    autofit(ws)
    return len(registros)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    inicio = datetime.now()
    log("=" * 60)
    log("  VDR Comparador v1.0 — EGA KAT / Derco Parts")
    log(f"  {inicio.strftime('%d/%m/%Y %H:%M:%S')}")
    log("=" * 60)

    # 1. Carpeta del mes en curso
    nombre_mes = carpeta_mes_actual()
    carpeta_mes_path = os.path.join(BASE_ORIGEN, nombre_mes)
    log(f"  Carpeta mes: {nombre_mes}")

    if not os.path.isdir(carpeta_mes_path):
        log(f"  [AVISO] Carpeta del mes no existe aun: {carpeta_mes_path}")
        log("  Sin novedad. Fin.")
        return

    # 2. Listar archivos ordenados por fecha del nombre
    archivos = listar_archivos_ordenados(carpeta_mes_path)
    log(f"  Archivos encontrados: {len(archivos)}")
    for nombre, fecha in archivos:
        log(f"    {fecha}  |  {nombre}")

    if len(archivos) < 2:
        log("  [AVISO] Se necesitan al menos 2 archivos para comparar. Sin novedad.")
        return

    # 3. Leer estado del ultimo procesado
    estado_mes, estado_archivo = leer_estado()
    archivo_mas_reciente = archivos[-1][0]
    log(f"  Ultimo procesado: {estado_mes} | {estado_archivo}")
    log(f"  Mas reciente:     {nombre_mes} | {archivo_mas_reciente}")

    # 4. Verificar si hay novedad
    if estado_mes == nombre_mes and estado_archivo == archivo_mas_reciente:
        log("  Sin archivos nuevos. Fin.")
        return

    # 5. Definir archivo nuevo y anterior
    archivo_nuevo = archivos[-1][0]
    archivo_ant   = archivos[-2][0]
    ruta_nuevo    = os.path.join(carpeta_mes_path, archivo_nuevo)
    ruta_ant      = os.path.join(carpeta_mes_path, archivo_ant)

    log(f"\n  Archivo anterior: {archivo_ant}")
    log(f"  Archivo nuevo:    {archivo_nuevo}")

    # 6. Cargar datos
    log("\n  Cargando datos...")
    try:
        datos_ant   = cargar_datos(ruta_ant)
        datos_nuevo = cargar_datos(ruta_nuevo)
    except Exception as e:
        log(f"  [ERROR] Fallo al leer archivos: {e}")
        return

    # 7. Identificar SKUs nuevos y eliminados
    set_ant   = set(datos_ant.keys())
    set_nuevo = set(datos_nuevo.keys())
    skus_nuevos     = {k: datos_nuevo[k] for k in set_nuevo - set_ant}
    skus_eliminados = {k: datos_ant[k]   for k in set_ant - set_nuevo}
    log(f"\n  SKUs en anterior:  {len(set_ant)}")
    log(f"  SKUs en nuevo:     {len(set_nuevo)}")
    log(f"  SKUs nuevos:       {len(skus_nuevos)}")
    log(f"  SKUs eliminados:   {len(skus_eliminados)}")

    # 8. Construir workbook de diferencias
    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    log("\n  Comparando VDR SAP y VDR FISICO...")
    n_dif   = escribir_hoja_diferencias(wb, datos_ant, datos_nuevo, archivo_ant, archivo_nuevo)
    log(f"  Diferencias VDR:      {n_dif}")

    log("  Comparando equivalencia Material WMS <-> Material SAP...")
    n_equiv = escribir_hoja_equivalencia(wb, datos_ant, datos_nuevo)
    log(f"  Cambios equivalencia: {n_equiv}")

    n_nuevos     = escribir_hoja_skus(wb, "SKUs_Nuevos", skus_nuevos)
    n_eliminados = escribir_hoja_skus(wb, "SKUs_Eliminados", skus_eliminados)
    log(f"  SKUs_Nuevos hoja:     {n_nuevos}")
    log(f"  SKUs_Eliminados hoja: {n_eliminados}")

    total_novedades = n_dif + n_equiv + n_nuevos + n_eliminados

    # 9. Guardar reporte solo si hay diferencias
    if total_novedades == 0:
        log("\n  Sin diferencias detectadas. No se genera reporte.")
    else:
        os.makedirs(BASE_SALIDA, exist_ok=True)
        ts_salida   = datetime.now().strftime("%d%m%Y_%H%M%S")
        nombre_rep  = f"Reporte_VDR_{ts_salida}.xlsx"
        ruta_rep    = os.path.join(BASE_SALIDA, nombre_rep)
        wb.save(ruta_rep)
        log(f"\n  [OK] Reporte generado: {ruta_rep}")
        log(f"       Diferencias VDR:      {n_dif}")
        log(f"       Cambios equivalencia: {n_equiv}")
        log(f"       SKUs nuevos:          {n_nuevos}")
        log(f"       SKUs eliminados:      {n_eliminados}")

    # 10. Actualizar estado
    escribir_estado(nombre_mes, archivo_nuevo)
    log(f"  Estado actualizado: {nombre_mes} | {archivo_nuevo}")

    dur = int((datetime.now() - inicio).total_seconds())
    log(f"\n  Duracion total: {dur // 60}m {dur % 60}s")
    log("=" * 60)


if __name__ == "__main__":
    main()


# ─── COMANDO TASK SCHEDULER ───────────────────────────────────────────────────
# Ejecutar UNA VEZ en PowerShell como Administrador para crear la tarea:
#
# schtasks /create /tn "VDR Comparador - EGA KAT" ^
#   /tr "\"C:\Users\Socrates Cabral\AppData\Local\Python\pythoncore-3.14-64\python.exe\" \"C:\ClaudeWork\vdr_comparador.py\"" ^
#   /sc HOURLY ^
#   /mo 1 ^
#   /d MON,TUE,WED,THU,FRI ^
#   /st 08:00 ^
#   /et 19:00 ^
#   /ru "Socrates Cabral" ^
#   /rp ^
#   /f
#
# Para verificar que se creo:
#   schtasks /query /tn "VDR Comparador - EGA KAT" /fo LIST /v
#
# Para ejecutar manualmente (prueba):
#   schtasks /run /tn "VDR Comparador - EGA KAT"
