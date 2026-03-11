"""
extractor.py — Agente Extractor Egakat
Lee datos desde Excel, PDF, SQL o carpeta OneDrive y los prepara para análisis.

Uso CLI:
    py AI_Agent/agentes/extractor.py excel   "ruta/archivo.xlsx"
    py AI_Agent/agentes/extractor.py pdf     "ruta/archivo.pdf"
    py AI_Agent/agentes/extractor.py sql     "SELECT * FROM tabla" --db "ruta.db"
    py AI_Agent/agentes/extractor.py onedrive "Stock WMS"
    py AI_Agent/agentes/extractor.py analizar "ruta/archivo.xlsx"   ← extrae + envía a Claude

Uso como módulo (otros agentes):
    from agentes.extractor import extraer_excel, extraer_pdf, extraer_sql
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import os
import re
import json
import argparse
from pathlib import Path
from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent.parent.parent
load_dotenv(dotenv_path=BASE_DIR / ".env")

# OneDrive base
ONEDRIVE = Path(os.getenv("ONEDRIVE_BASE",
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA"))

MAX_FILAS_PREVIEW  = 10    # filas visibles en el resumen
MAX_CHARS_CLAUDE   = 4000  # chars enviados a Claude
MAX_TEXTO_PDF      = 3000  # chars extraídos de PDF


# ══════════════════════════════════════════════════════════════════════════════
#  EXTRACTORES
# ══════════════════════════════════════════════════════════════════════════════

def extraer_excel(ruta: str | Path, hoja: str | int = 0) -> dict:
    """
    Lee un Excel y retorna metadatos + preview + datos crudos.
    Retorna dict con claves: archivo, hojas, hoja_activa, columnas, filas_total,
                             preview (list of dicts), resumen_texto
    """
    import openpyxl
    ruta = Path(ruta)
    if not ruta.exists():
        # buscar en OneDrive recursivamente
        encontrado = list(ONEDRIVE.rglob(ruta.name))
        if encontrado:
            ruta = encontrado[0]
        else:
            return {"error": f"Archivo no encontrado: {ruta.name}"}

    wb    = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hoja_nombre = wb.sheetnames[hoja] if isinstance(hoja, int) else hoja
    ws    = wb[hoja_nombre]

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return {"error": "Hoja vacía"}

    headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(filas[0])]
    datos   = [dict(zip(headers, fila)) for fila in filas[1:] if any(v is not None for v in fila)]

    preview = datos[:MAX_FILAS_PREVIEW]

    # Resumen en texto para Claude
    resumen = (
        f"Archivo: {ruta.name}\n"
        f"Hoja: {hoja_nombre} | Total hojas: {len(wb.sheetnames)}\n"
        f"Columnas ({len(headers)}): {', '.join(headers)}\n"
        f"Filas de datos: {len(datos)}\n\n"
        f"Primeras {min(MAX_FILAS_PREVIEW, len(preview))} filas:\n"
    )
    for i, fila in enumerate(preview, 1):
        resumen += f"  {i}. " + " | ".join(f"{k}: {v}" for k, v in fila.items() if v is not None) + "\n"

    wb.close()
    return {
        "archivo":      str(ruta),
        "hojas":        wb.sheetnames,
        "hoja_activa":  hoja_nombre,
        "columnas":     headers,
        "filas_total":  len(datos),
        "preview":      preview,
        "resumen_texto": resumen,
    }


def extraer_pdf(ruta: str | Path) -> dict:
    """
    Extrae texto de un PDF. Retorna dict con texto, paginas, resumen_texto.
    """
    import pdfplumber
    ruta = Path(ruta)
    if not ruta.exists():
        return {"error": f"Archivo no encontrado: {ruta}"}

    texto_total = ""
    n_paginas   = 0

    with pdfplumber.open(ruta) as pdf:
        n_paginas = len(pdf.pages)
        for pagina in pdf.pages:
            t = pagina.extract_text()
            if t:
                texto_total += t + "\n"

    if not texto_total.strip():
        return {"error": "PDF sin texto extraíble (puede ser imagen escaneada)"}

    texto_limitado = texto_total[:MAX_TEXTO_PDF]
    resumen = (
        f"Archivo: {ruta.name}\n"
        f"Páginas: {n_paginas}\n"
        f"Caracteres totales: {len(texto_total)}\n\n"
        f"Contenido (primeros {MAX_TEXTO_PDF} chars):\n{texto_limitado}"
    )

    return {
        "archivo":       str(ruta),
        "paginas":       n_paginas,
        "texto_total":   texto_total,
        "resumen_texto": resumen,
    }


def extraer_sql(query: str, db_path: str = None, conn_string: str = None) -> dict:
    """
    Ejecuta una query SQL.
    - db_path: ruta a archivo SQLite
    - conn_string: cadena ODBC para SQL Server / Access
    """
    try:
        if db_path:
            import sqlite3
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
        elif conn_string:
            import pyodbc
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
        else:
            return {"error": "Especifica --db (SQLite) o --conn (ODBC)"}

        cursor.execute(query)
        columnas = [desc[0] for desc in cursor.description] if cursor.description else []
        filas    = cursor.fetchall()
        conn.close()

        datos   = [dict(zip(columnas, fila)) for fila in filas]
        preview = datos[:MAX_FILAS_PREVIEW]

        resumen = (
            f"Query: {query[:200]}\n"
            f"Columnas: {', '.join(columnas)}\n"
            f"Filas resultado: {len(datos)}\n\n"
            f"Primeras {min(MAX_FILAS_PREVIEW, len(preview))} filas:\n"
        )
        for i, fila in enumerate(preview, 1):
            resumen += f"  {i}. " + " | ".join(f"{k}: {v}" for k, v in fila.items()) + "\n"

        return {
            "query":         query,
            "columnas":      columnas,
            "filas_total":   len(datos),
            "preview":       preview,
            "resumen_texto": resumen,
        }

    except Exception as e:
        return {"error": str(e)}


def buscar_onedrive(termino: str) -> dict:
    """
    Busca archivos Excel/PDF en OneDrive que coincidan con el término.
    Retorna lista de coincidencias.
    """
    if not ONEDRIVE.exists():
        return {"error": f"OneDrive no encontrado en: {ONEDRIVE}"}

    resultados = []
    for ext in ("*.xlsx", "*.xls", "*.pdf", "*.csv"):
        for p in ONEDRIVE.rglob(ext):
            if termino.lower() in p.name.lower() or termino.lower() in str(p.parent).lower():
                resultados.append({
                    "nombre": p.name,
                    "ruta":   str(p),
                    "tamaño": f"{p.stat().st_size / 1024:.1f} KB",
                })

    if not resultados:
        return {"error": f"Sin resultados para '{termino}' en OneDrive"}

    resumen = f"Búsqueda: '{termino}' — {len(resultados)} archivo(s) encontrado(s):\n"
    for r in resultados[:10]:
        resumen += f"  • {r['nombre']} ({r['tamaño']}) → {r['ruta']}\n"

    return {"resultados": resultados[:10], "resumen_texto": resumen}


# ══════════════════════════════════════════════════════════════════════════════
#  ENVIAR A CLAUDE
# ══════════════════════════════════════════════════════════════════════════════

def analizar_con_claude(resumen_texto: str, pregunta: str = "") -> str:
    """Envía el resumen extraído a Claude para análisis."""
    from anthropic import Anthropic

    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        return "ERROR: ANTHROPIC_API_KEY no está en .env"

    client = Anthropic(api_key=api_key)

    contexto = resumen_texto[:MAX_CHARS_CLAUDE]
    if len(resumen_texto) > MAX_CHARS_CLAUDE:
        contexto += f"\n... [truncado a {MAX_CHARS_CLAUDE} chars]"

    if not pregunta:
        pregunta = (
            "Analiza estos datos logísticos. Identifica: "
            "1) qué información contiene, "
            "2) KPIs o métricas relevantes que se pueden calcular, "
            "3) anomalías o puntos de atención, "
            "4) qué preguntas de negocio se pueden responder con estos datos."
        )

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1500,
        system=(
            "Eres un analista de operaciones logísticas experto. "
            "Trabajas para Egakat SPA, una empresa 3PL chilena. "
            "Respondes en español de forma concisa y orientada a acciones concretas."
        ),
        messages=[{
            "role": "user",
            "content": f"Datos extraídos:\n\n{contexto}\n\nAnálisis solicitado: {pregunta}"
        }]
    )
    return response.content[0].text


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Extractor de datos Egakat — Excel, PDF, SQL, OneDrive"
    )
    parser.add_argument("fuente", choices=["excel", "pdf", "sql", "onedrive", "analizar"],
                        help="Tipo de fuente de datos")
    parser.add_argument("objetivo", help="Ruta, query o término de búsqueda")
    parser.add_argument("--hoja",   default="0",  help="Nombre o índice de hoja Excel")
    parser.add_argument("--db",     default=None, help="Ruta archivo SQLite")
    parser.add_argument("--conn",   default=None, help="Cadena de conexión ODBC")
    parser.add_argument("--ask",    default="",   help="Pregunta específica para Claude")
    parser.add_argument("--json",   action="store_true", help="Output en JSON")

    args = parser.parse_args()

    # ── Extraer ──────────────────────────────────────────────────────────────
    resultado = {}

    if args.fuente == "excel":
        hoja = int(args.hoja) if args.hoja.isdigit() else args.hoja
        resultado = extraer_excel(args.objetivo, hoja)

    elif args.fuente == "pdf":
        resultado = extraer_pdf(args.objetivo)

    elif args.fuente == "sql":
        resultado = extraer_sql(args.objetivo, db_path=args.db, conn_string=args.conn)

    elif args.fuente == "onedrive":
        resultado = buscar_onedrive(args.objetivo)

    elif args.fuente == "analizar":
        # Detecta tipo por extensión y envía a Claude automáticamente
        ext = Path(args.objetivo).suffix.lower()
        if ext in (".xlsx", ".xls"):
            resultado = extraer_excel(args.objetivo)
        elif ext == ".pdf":
            resultado = extraer_pdf(args.objetivo)
        elif ext == ".csv":
            resultado = extraer_excel(args.objetivo)
        else:
            resultado = {"error": f"Extensión no soportada para análisis automático: {ext}"}

    # ── Error ────────────────────────────────────────────────────────────────
    if "error" in resultado:
        print(f"[ERROR] {resultado['error']}")
        sys.exit(1)

    # ── Output ───────────────────────────────────────────────────────────────
    if args.json:
        # Sin preview completo para no saturar
        salida = {k: v for k, v in resultado.items() if k not in ("preview",)}
        print(json.dumps(salida, ensure_ascii=False, indent=2, default=str))
        sys.exit(0)

    # Mostrar resumen
    print("\n" + "═" * 60)
    print(resultado.get("resumen_texto", "Sin resumen disponible"))
    print("═" * 60)

    # Si es analizar o hay --ask, enviar a Claude
    if args.fuente == "analizar" or args.ask:
        resumen = resultado.get("resumen_texto", "")
        if resumen:
            print("\n[Enviando a Claude para análisis...]\n")
            analisis = analizar_con_claude(resumen, args.ask)
            print(analisis)


if __name__ == "__main__":
    main()
