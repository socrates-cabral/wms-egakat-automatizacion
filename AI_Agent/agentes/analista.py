"""
analista.py — Agente Analista Logístico Egakat
Calcula KPIs, detecta anomalías y genera análisis narrativo con Claude.

Uso CLI:
    py AI_Agent/agentes/analista.py stock    "archivo.xlsx"
    py AI_Agent/agentes/analista.py staging  "archivo.csv"
    py AI_Agent/agentes/analista.py vdr      "archivo.xlsx"
    py AI_Agent/agentes/analista.py nps      "archivo.xlsx"
    py AI_Agent/agentes/analista.py comparar "archivo1.xlsx" "archivo2.xlsx"
    py AI_Agent/agentes/analista.py informe  "stock WMS"  --guardar

Uso como módulo:
    from agentes.analista import analizar_stock, analizar_staging
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import os
import json
import argparse
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent.parent.parent
load_dotenv(dotenv_path=BASE_DIR / ".env")

OD = Path(os.getenv(
    "ONEDRIVE_BASE",
    r"C:\Users\Socrates Cabral\OneDrive - EGA KAT LOGISTICA SPA"
))

MAX_CHARS_CLAUDE = 5000


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _buscar_archivo(nombre: str) -> Path | None:
    """Busca archivo en ClaudeWork y OneDrive."""
    p = Path(nombre)
    if p.exists():
        return p
    for base in (BASE_DIR, OD):
        encontrados = list(base.rglob(Path(nombre).name))
        if encontrados:
            return encontrados[0]
    return None


def _leer_excel_pandas(ruta: Path, hoja=0) -> "pd.DataFrame":
    import pandas as pd
    import openpyxl
    # Intentar con header=0 primero, luego auto-detectar
    for sheet in ([hoja] if hoja != 0 else [0, 1]):
        for header_row in (0, None):
            try:
                df = pd.read_excel(ruta, sheet_name=sheet, header=header_row, engine="openpyxl")
                if header_row is None:
                    # Auto-detectar fila de headers
                    for i, row in df.iterrows():
                        if row.notna().sum() >= 3 and row.astype(str).str.len().mean() > 2:
                            df.columns = [str(c).strip() if str(c) != "nan" else f"Col{j}"
                                          for j, c in enumerate(row)]
                            df = df.iloc[i+1:].reset_index(drop=True)
                            break
                df = df.dropna(how="all")
                if len(df) > 2:
                    for col in df.columns:
                        converted = pd.to_numeric(df[col], errors="coerce")
                        if converted.notna().sum() > len(df) * 0.3:
                            df[col] = converted
                    return df
            except Exception:
                continue
    return None


def _leer_csv_pandas(ruta: Path) -> "pd.DataFrame":
    import pandas as pd
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            df = pd.read_csv(ruta, encoding=enc, sep=None, engine="python")
            for col in df.columns:
                converted = pd.to_numeric(df[col], errors="coerce")
                if converted.notna().sum() > len(df) * 0.3:
                    df[col] = converted
            return df.dropna(how="all")
        except Exception:
            continue
    return None


def _claude(system: str, user: str, max_tokens: int = 1800) -> str:
    from anthropic import Anthropic
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        return "ERROR: ANTHROPIC_API_KEY no está en .env"
    client = Anthropic(api_key=api_key)
    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=max_tokens,
        system=system,
        messages=[{"role": "user", "content": user[:MAX_CHARS_CLAUDE]}]
    )
    return resp.content[0].text


SYSTEM_LOGISTICA = (
    "Eres un analista senior de operaciones logísticas de Egakat SPA, empresa 3PL chilena. "
    "Tu análisis es conciso, orientado a acción y usa terminología logística precisa. "
    "Respondes en español. Prioriza anomalías críticas y KPIs accionables."
)


# ══════════════════════════════════════════════════════════════════════════════
#  ANALIZADORES
# ══════════════════════════════════════════════════════════════════════════════

def analizar_stock(ruta: str) -> dict:
    """Analiza reporte de stock WMS: KPIs de inventario, % bloqueo, top SKUs."""
    import pandas as pd

    archivo = _buscar_archivo(ruta)
    if not archivo:
        return {"error": f"Archivo no encontrado: {ruta}"}

    # Elegir hoja con más columnas (datos más ricos)
    import openpyxl, pandas as pd
    wb = openpyxl.load_workbook(archivo, read_only=True)
    n_hojas = len(wb.sheetnames)
    wb.close()

    df = None
    for idx in range(n_hojas):
        candidato = _leer_excel_pandas(archivo, hoja=idx)
        if candidato is not None and not candidato.empty:
            if df is None or len(candidato.columns) > len(df.columns):
                df = candidato

    if df is None or df.empty:
        return {"error": "No se pudo leer el archivo Excel"}

    # Detectar columnas clave
    cols = [c.lower() for c in df.columns]
    col_mapa = {}
    for i, c in enumerate(cols):
        if any(k in c for k in ["no bloq", "n", "disponib"]) and "bloq" not in c[:2]:
            col_mapa.setdefault("disponible", df.columns[i])
        if any(k in c for k in ["bloq", " s", "s "]):
            col_mapa.setdefault("bloqueado", df.columns[i])
        if any(k in c for k in ["total", "general"]):
            col_mapa.setdefault("total", df.columns[i])
        if any(k in c for k in ["rubro", "subrub", "categ"]):
            col_mapa.setdefault("rubro", df.columns[i])
        if any(k in c for k in ["descrip", "articu", "nombre"]):
            col_mapa.setdefault("descripcion", df.columns[i])

    # KPIs numéricos
    kpis = {}
    kpis["total_skus"] = len(df)
    kpis["columnas_detectadas"] = list(df.columns)

    resumen = f"Archivo: {archivo.name}\nFilas: {len(df)} | Columnas: {len(df.columns)}\n"
    resumen += f"Columnas: {', '.join(df.columns.tolist())}\n\n"

    if "total" in col_mapa:
        col_t = col_mapa["total"]
        serie = pd.to_numeric(df[col_t], errors="coerce").dropna()
        if not serie.empty:
            kpis["unidades_total"] = int(serie.sum())
            kpis["promedio_por_sku"] = round(float(serie.mean()), 1)
            kpis["sku_max"] = int(serie.max())
            resumen += f"Unidades totales: {kpis['unidades_total']:,}\n"
            resumen += f"Promedio por SKU: {kpis['promedio_por_sku']}\n"

    if "bloqueado" in col_mapa and "total" in col_mapa:
        col_b = col_mapa["bloqueado"]
        col_t = col_mapa["total"]
        serie_b = pd.to_numeric(df[col_b], errors="coerce").fillna(0)
        serie_t = pd.to_numeric(df[col_t], errors="coerce").fillna(0)
        total_bloq = int(serie_b.sum())
        total_all  = int(serie_t.sum())
        pct_bloq   = round(total_bloq / total_all * 100, 2) if total_all > 0 else 0
        kpis["unidades_bloqueadas"] = total_bloq
        kpis["pct_bloqueado"]       = pct_bloq
        skus_100_bloq = int((serie_b > 0) & (pd.to_numeric(df.get(col_mapa.get("disponible",""), 0), errors="coerce").fillna(0) == 0) if "disponible" in col_mapa else serie_b > 0)
        resumen += f"Unidades bloqueadas: {total_bloq:,} ({pct_bloq}% del total)\n"

    if "rubro" in col_mapa and "total" in col_mapa:
        col_r = col_mapa["rubro"]
        col_t = col_mapa["total"]
        df["_total_num"] = pd.to_numeric(df[col_t], errors="coerce")
        top_rubros = (
            df.groupby(col_r)["_total_num"]
            .sum().sort_values(ascending=False).head(5)
        )
        resumen += f"\nTop 5 rubros por unidades:\n"
        for rubro, cant in top_rubros.items():
            resumen += f"  • {rubro}: {int(cant):,} unidades\n"

    # Preview primeras filas para Claude
    resumen += f"\nMuestra datos (primeras 8 filas):\n"
    resumen += df.head(8).to_string(index=False, max_colwidth=30) + "\n"

    analisis = _claude(
        SYSTEM_LOGISTICA,
        f"Analiza este reporte de stock WMS de Egakat SPA:\n\n{resumen}\n\n"
        "Entrega: 1) Estado del inventario, 2) Alertas críticas, "
        "3) KPIs más importantes, 4) Acciones recomendadas."
    )

    return {"kpis": kpis, "resumen_texto": resumen, "analisis": analisis}


def analizar_staging(ruta: str) -> dict:
    """Analiza reporte staging IN/OUT: pallets por cliente, flujo entrada/salida."""
    import pandas as pd

    archivo = _buscar_archivo(ruta)
    if not archivo:
        return {"error": f"Archivo no encontrado: {ruta}"}

    ext = archivo.suffix.lower()
    df  = _leer_csv_pandas(archivo) if ext == ".csv" else _leer_excel_pandas(archivo)
    if df is None or df.empty:
        return {"error": "No se pudo leer el archivo"}

    kpis   = {"total_registros": len(df), "columnas": list(df.columns)}
    resumen = f"Archivo: {archivo.name}\nRegistros: {len(df)}\n"
    resumen += f"Columnas: {', '.join(df.columns.tolist())}\n\n"

    # Detectar columnas de cantidad
    cols_num = df.select_dtypes(include="number").columns.tolist()
    if cols_num:
        for col in cols_num[:3]:
            total = df[col].sum()
            resumen += f"Total {col}: {total:,.0f}\n"
            kpis[f"total_{col}"] = float(total)

    # Detección cliente/empresa
    for col in df.columns:
        if any(k in col.lower() for k in ["empresa", "cliente", "compañia"]):
            conteo = df[col].value_counts().head(10)
            resumen += f"\nRegistros por {col}:\n"
            for nombre, cnt in conteo.items():
                resumen += f"  • {nombre}: {cnt}\n"
            kpis["clientes_top"] = conteo.to_dict()
            break

    resumen += f"\nMuestra (primeras 6 filas):\n"
    resumen += df.head(6).to_string(index=False, max_colwidth=25) + "\n"

    analisis = _claude(
        SYSTEM_LOGISTICA,
        f"Analiza este reporte de Staging IN/OUT de Egakat SPA:\n\n{resumen}\n\n"
        "Entrega: 1) Flujo de pallets, 2) Clientes con mayor movimiento, "
        "3) Anomalías detectadas, 4) Recomendaciones operativas."
    )

    return {"kpis": kpis, "resumen_texto": resumen, "analisis": analisis}


def analizar_nps(ruta: str) -> dict:
    """Analiza reporte NPS/CSAT: score, distribución, comentarios clave."""
    import pandas as pd

    archivo = _buscar_archivo(ruta)
    if not archivo:
        return {"error": f"Archivo no encontrado: {ruta}"}

    # Intentar todas las hojas
    import openpyxl
    wb     = openpyxl.load_workbook(archivo, read_only=True)
    hojas  = wb.sheetnames
    wb.close()

    resumen  = f"Archivo: {archivo.name}\nHojas: {', '.join(hojas)}\n\n"
    analisis_data = []

    for hoja in hojas[:3]:
        df = _leer_excel_pandas(archivo, hoja=hoja)
        if df is None or df.empty:
            continue
        resumen += f"--- Hoja: {hoja} ({len(df)} filas) ---\n"
        resumen += f"Columnas: {', '.join(df.columns.tolist())}\n"

        # Detectar columna de score NPS
        for col in df.columns:
            if any(k in col.lower() for k in ["score", "nps", "nota", "puntaj", "calific"]):
                serie = pd.to_numeric(df[col], errors="coerce").dropna()
                if not serie.empty:
                    promotores  = int((serie >= 9).sum())
                    detractores = int((serie <= 6).sum())
                    pasivos     = int(((serie >= 7) & (serie <= 8)).sum())
                    total       = len(serie)
                    nps_score   = round((promotores - detractores) / total * 100, 1)
                    resumen += (
                        f"NPS Score: {nps_score}\n"
                        f"Promotores (9-10): {promotores} ({round(promotores/total*100,1)}%)\n"
                        f"Pasivos (7-8): {pasivos} ({round(pasivos/total*100,1)}%)\n"
                        f"Detractores (0-6): {detractores} ({round(detractores/total*100,1)}%)\n"
                    )
                    analisis_data.append({"hoja": hoja, "nps": nps_score, "total": total})

        resumen += df.head(5).to_string(index=False, max_colwidth=30) + "\n\n"

    analisis = _claude(
        SYSTEM_LOGISTICA,
        f"Analiza estos resultados de encuesta NPS/CSAT de Egakat SPA:\n\n{resumen}\n\n"
        "Entrega: 1) Score NPS y benchmark, 2) Perfil de respondentes, "
        "3) Temas recurrentes en comentarios, 4) Acciones prioritarias para mejorar."
    )

    return {"hojas": hojas, "resumen_texto": resumen, "analisis": analisis}


def comparar_archivos(ruta1: str, ruta2: str) -> dict:
    """Compara dos Excel/CSV y detecta diferencias: filas nuevas, eliminadas, cambios."""
    import pandas as pd

    a1 = _buscar_archivo(ruta1)
    a2 = _buscar_archivo(ruta2)
    if not a1:
        return {"error": f"Archivo no encontrado: {ruta1}"}
    if not a2:
        return {"error": f"Archivo no encontrado: {ruta2}"}

    def leer(path):
        return (_leer_csv_pandas(path) if path.suffix == ".csv"
                else _leer_excel_pandas(path))

    df1 = leer(a1)
    df2 = leer(a2)
    if df1 is None or df2 is None:
        return {"error": "No se pudieron leer los archivos"}

    resumen = (
        f"Comparación:\n"
        f"  Archivo A: {a1.name} — {len(df1)} filas, {len(df1.columns)} cols\n"
        f"  Archivo B: {a2.name} — {len(df2)} filas, {len(df2.columns)} cols\n\n"
    )

    diff_filas = len(df2) - len(df1)
    resumen += f"Diferencia de filas: {'+' if diff_filas >= 0 else ''}{diff_filas}\n"

    # Columnas comunes
    cols_comunes = list(set(df1.columns) & set(df2.columns))
    cols_solo_a  = list(set(df1.columns) - set(df2.columns))
    cols_solo_b  = list(set(df2.columns) - set(df1.columns))
    resumen += f"Columnas comunes: {len(cols_comunes)}\n"
    if cols_solo_a:
        resumen += f"Solo en A: {', '.join(cols_solo_a)}\n"
    if cols_solo_b:
        resumen += f"Solo en B: {', '.join(cols_solo_b)}\n"

    # Diferencias numéricas en columnas comunes
    resumen += "\nDiferencias en columnas numéricas:\n"
    for col in cols_comunes:
        s1 = pd.to_numeric(df1[col], errors="coerce").dropna()
        s2 = pd.to_numeric(df2[col], errors="coerce").dropna()
        if not s1.empty and not s2.empty:
            d_sum = s2.sum() - s1.sum()
            if abs(d_sum) > 0.01:
                resumen += f"  • {col}: {s1.sum():,.1f} → {s2.sum():,.1f} (Δ {'+' if d_sum >= 0 else ''}{d_sum:,.1f})\n"

    analisis = _claude(
        SYSTEM_LOGISTICA,
        f"Compara estos dos reportes logísticos y explica las diferencias relevantes:\n\n{resumen}\n\n"
        "Entrega: 1) Resumen de cambios, 2) Diferencias críticas, "
        "3) Posibles causas, 4) Acciones recomendadas."
    )

    return {"resumen_texto": resumen, "analisis": analisis}


def informe_completo(termino: str, guardar: bool = False) -> dict:
    """
    Busca los archivos más recientes del término en OneDrive y genera informe completo.
    Si guardar=True, lo escribe en OneDrive/Reportes NPS o carpeta correspondiente.
    """
    from agentes.m365 import listar_carpeta, subir_archivo

    resultado = listar_carpeta(termino, dias=7)
    if "error" in resultado:
        return resultado

    recientes = resultado.get("recientes", [])
    if not recientes:
        return {"error": f"Sin archivos recientes para '{termino}'"}

    # Tomar el más reciente no-ini
    archivo = next(
        (r["ruta"] for r in recientes if not r["nombre"].endswith(".ini")),
        None
    )
    if not archivo:
        return {"error": "Sin archivos válidos recientes"}

    ext = Path(archivo).suffix.lower()
    if "nps" in termino.lower():
        resultado_analisis = analizar_nps(archivo)
    elif "staging" in termino.lower():
        resultado_analisis = analizar_staging(archivo)
    else:
        resultado_analisis = analizar_stock(archivo)

    if "error" in resultado_analisis:
        return resultado_analisis

    # Guardar en OneDrive si se pide
    if guardar:
        ts       = datetime.now().strftime("%Y%m%d_%H%M")
        nombre   = f"Analisis_{Path(archivo).stem}_{ts}.txt"
        tmp      = BASE_DIR / "logs" / nombre
        tmp.parent.mkdir(exist_ok=True)
        tmp.write_text(
            resultado_analisis.get("analisis", ""),
            encoding="utf-8"
        )
        subir_archivo(str(tmp), "nps" if "nps" in termino.lower() else "vdr", nombre)
        resultado_analisis["guardado_en"] = nombre

    return resultado_analisis


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Agente Analista Logístico Egakat")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_st  = sub.add_parser("stock",    help="Analizar stock WMS")
    p_st.add_argument("archivo")

    p_sg  = sub.add_parser("staging",  help="Analizar staging IN/OUT")
    p_sg.add_argument("archivo")

    p_np  = sub.add_parser("nps",      help="Analizar encuesta NPS/CSAT")
    p_np.add_argument("archivo")

    p_cm  = sub.add_parser("comparar", help="Comparar dos archivos")
    p_cm.add_argument("archivo1")
    p_cm.add_argument("archivo2")

    p_inf = sub.add_parser("informe",  help="Informe completo desde OneDrive")
    p_inf.add_argument("termino")
    p_inf.add_argument("--guardar", action="store_true")

    args = parser.parse_args()

    if args.cmd == "stock":
        r = analizar_stock(args.archivo)
    elif args.cmd == "staging":
        r = analizar_staging(args.archivo)
    elif args.cmd == "nps":
        r = analizar_nps(args.archivo)
    elif args.cmd == "comparar":
        r = comparar_archivos(args.archivo1, args.archivo2)
    elif args.cmd == "informe":
        r = informe_completo(args.termino, args.guardar)

    if "error" in r:
        print(f"[ERROR] {r['error']}")
        sys.exit(1)

    print("\n" + "═" * 60)
    print(r.get("resumen_texto", ""))
    print("═" * 60)
    print("\n📊 ANÁLISIS CLAUDE:\n")
    print(r.get("analisis", ""))

    if "guardado_en" in r:
        print(f"\n✅ Informe guardado: {r['guardado_en']}")


if __name__ == "__main__":
    main()
