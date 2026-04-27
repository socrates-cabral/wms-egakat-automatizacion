import sys
sys.stdout.reconfigure(encoding="utf-8")
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

OUT = Path("C:/ClaudeWork/analisis_estado_sp.txt")
BASE = Path("C:/Users/Socrates Cabral/OneDrive - EGA KAT LOGISTICA SPA/Datos para Dashboard - Productividad")

ARCHIVOS = {
    "CD QUILICURA/2026/04. Abril/MovABInbev.xlsx":       ("ABInbev/CERVECERIA ABI", "Fecha"),
    "CD QUILICURA/2026/04. Abril/MovBha.xlsx":            ("BHA",                    "Fecha"),
    "CD QUILICURA/2026/04. Abril/MovDaikin.xlsx":         ("DAIKIN",                 "Fecha"),
    "CD QUILICURA/2026/04. Abril/MovDerco.xlsx":          ("DERCO",                  "Fecha"),
    "CD QUILICURA/2026/04. Abril/MovMascota.xlsx":        ("MASCOTAS LATINAS",       "Fecha"),
    "CD QUILICURA/2026/04. Abril/MovPochteca.xlsx":       ("POCHTECA",               "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/MovBarentz.xlsx":         ("BARENTZ",                "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/MovBuraschi.xlsx":        ("BURASCHI",               "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/MovCepas Chile.xlsx":     ("CEPAS CHILE",            "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/MovCollico.xlsx":         ("COLLICO",                "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/MovDelibest.xlsx":        ("DELIBEST",               "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/Movintime.xlsx":          ("INTIME",                 "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/MovMascota Latina.xlsx":  ("MASCOTAS LATINA PUD",    "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/MovRuno.xlsx":            ("RUNO SPA",               "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/Movtresmontes.xlsx":      ("TRES MONTES",            "Fecha"),
    "CD PUDAHUEL/2026/04. Abril/MovUnilever.xlsx":        ("UNILEVER",               "Fecha"),
}

lines = [f"{'Cliente':<22} {'Filas':>6}  {'Desde':<12} {'Hasta':<12}  {'Dias_unicos':>11}  Arch_bytes"]

for rel, (nombre, col_fecha) in ARCHIVOS.items():
    f = BASE / rel
    if not f.exists():
        lines.append(f"{nombre:<22} {'NO EXISTE':>6}")
        continue
    tam = f.stat().st_size
    try:
        # detectar header row buscando "Fecha" en las primeras 12 filas
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        header_row = None
        for i, row in enumerate(ws.iter_rows(max_row=12, values_only=True)):
            if any(str(v).strip() == col_fecha for v in row if v):
                header_row = i
                break
        wb.close()

        if header_row is None:
            lines.append(f"{nombre:<22} {'SIN HEADER':>6}  {tam:>10} bytes")
            continue

        df = pd.read_excel(f, header=header_row, engine="openpyxl")
        # limpiar col_fecha
        df[col_fecha] = pd.to_datetime(df[col_fecha], errors="coerce")
        df = df.dropna(subset=[col_fecha])

        nf = len(df)
        desde = df[col_fecha].min().date() if nf > 0 else "—"
        hasta = df[col_fecha].max().date() if nf > 0 else "—"
        dias = df[col_fecha].dt.date.nunique()
        lines.append(f"{nombre:<22} {nf:>6}  {str(desde):<12} {str(hasta):<12}  {dias:>11}  {tam:>10}")
    except Exception as e:
        lines.append(f"{nombre:<22} ERROR: {e}")

OUT.write_text("\n".join(lines), encoding="utf-8")
print("OK")
