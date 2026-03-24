import sys
sys.stdout.reconfigure(encoding="utf-8")

import os
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
import streamlit as st
from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).parent.parent.parent / ".env")

MESES_MAP = {
    "01 Enero": 1, "02 Febrero": 2, "03 Marzo": 3, "04 Abril": 4,
    "05 Mayo": 5, "06 Junio": 6, "07 Julio": 7, "08 Agosto": 8,
    "09 Septiembre": 9, "10 Octubre": 10, "11 Noviembre": 11, "12 Diciembre": 12,
}
NOMBRES_MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

def _get_excel_path() -> Path:
    ruta = os.getenv("EXCEL_FP_PATH", "")
    if ruta and Path(ruta).exists():
        return Path(ruta)
    # fallback
    fallback = Path(r"C:\ClaudeWork\Plantilla-para-controlar-gastos.xlsm")
    if fallback.exists():
        return fallback
    raise FileNotFoundError(
        "No se encontró el Excel. Configura EXCEL_FP_PATH en .env o en Ajustes."
    )


def _load_workbook(ruta: Path = None):
    if ruta is None:
        ruta = _get_excel_path()
    return openpyxl.load_workbook(str(ruta), data_only=True, keep_vba=True)


def _detectar_formato(wb) -> str:
    """Detecta formato: 'nuevo' (hoja Transacciones) o 'antiguo' (hojas mensuales)."""
    if "Transacciones" in wb.sheetnames:
        return "nuevo"
    for hoja in wb.sheetnames:
        if hoja in MESES_MAP:
            return "antiguo"
    return "desconocido"


def _cargar_tx_nuevo(wb) -> pd.DataFrame:
    """Lee formato nuevo: hoja Transacciones — headers en fila 1, Tipo en col B."""
    if "Transacciones" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["Transacciones"]
    headers_raw = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip().lower() if h else "" for h in headers_raw]

    # Find column indices flexibly
    def _col(names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return None

    idx_fecha    = _col(["fecha"])
    idx_tipo     = _col(["tipo"])
    idx_grupo    = _col(["grupo"])
    idx_concepto = _col(["concepto"])
    idx_detalle  = _col(["detalle"])
    idx_importe  = _col(["importe", "monto", "valor"])
    idx_cuenta   = _col(["cuenta"])

    if idx_importe is None:
        return pd.DataFrame()

    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        def _get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        importe = _get(idx_importe)
        if not isinstance(importe, (int, float)) or importe <= 0:
            continue

        tipo = str(_get(idx_tipo) or "Gasto").strip()
        fecha_raw = _get(idx_fecha)
        fecha = fecha_raw if isinstance(fecha_raw, datetime) else None
        mes_num = fecha.month if fecha else 1

        filas.append({
            "mes": mes_num,
            "mes_nombre": NOMBRES_MESES.get(mes_num, ""),
            "tipo_tx": tipo,
            "grupo": str(_get(idx_grupo) or "Varios y Otros").strip(),
            "concepto": str(_get(idx_concepto) or "").strip(),
            "fecha": fecha,
            "detalle": str(_get(idx_detalle) or "").strip(),
            "importe": float(importe),
            "cuenta": str(_get(idx_cuenta) or "").strip(),
        })

    if not filas:
        return pd.DataFrame(columns=["mes", "mes_nombre", "tipo_tx", "grupo", "concepto",
                                      "fecha", "detalle", "importe", "cuenta"])
    df = pd.DataFrame(filas)
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    return df


@st.cache_data(ttl=300)
def cargar_transacciones(ruta_str: str = None) -> pd.DataFrame:
    """Carga todas las transacciones del año. Detecta formato automáticamente:
    - 'nuevo': hoja Transacciones (Plantilla_FinanzasPersonales.xlsx)
    - 'antiguo': hojas mensuales (Plantilla-para-controlar-gastos.xlsm)
    Retorna DataFrame unificado con columna tipo_tx.
    """
    ruta = Path(ruta_str) if ruta_str else _get_excel_path()
    wb = _load_workbook(ruta)
    fmt = _detectar_formato(wb)

    if fmt == "nuevo":
        df = _cargar_tx_nuevo(wb)
        if "tipo_tx" not in df.columns:
            df["tipo_tx"] = "Gasto"
        return df

    # Formato antiguo: hojas mensuales
    filas = []
    for hoja, num_mes in MESES_MAP.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        for row in ws.iter_rows(min_row=8, values_only=True):
            # Col A is empty (index 0), B=GRUPO(1), C=CONCEPTO(2), D=FECHA(3), E=DETALLE(4), F=IMPORTE(5)
            if len(row) < 6:
                continue
            grupo, concepto, fecha, detalle, importe = (
                row[1], row[2], row[3], row[4], row[5]
            )
            # Algunos meses tienen cols B y C intercambiadas — intentar swap antes de descartar
            if grupo is None and concepto is not None:
                grupo, concepto = concepto, grupo
            if grupo is None:
                continue
            if not isinstance(importe, (int, float)):
                continue
            if importe == 0:
                continue
            # Convención: IMPORTE negativo = Ingreso (resta del SUMA → aumenta saldo)
            tipo_tx = "Ingreso" if importe < 0 else "Gasto"
            filas.append({
                "mes": num_mes,
                "mes_nombre": NOMBRES_MESES[num_mes],
                "tipo_tx": tipo_tx,
                "grupo": str(grupo).strip(),
                "concepto": str(concepto).strip() if concepto else "",
                "fecha": fecha if isinstance(fecha, datetime) else None,
                "detalle": str(detalle).strip() if detalle else "",
                "importe": abs(float(importe)),
                "cuenta": "",
            })
    if not filas:
        return pd.DataFrame(columns=["mes", "mes_nombre", "tipo_tx", "grupo", "concepto",
                                      "fecha", "detalle", "importe", "cuenta"])
    df = pd.DataFrame(filas)
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    return df


@st.cache_data(ttl=300)
def cargar_saldos_mensuales(ruta_str: str = None) -> dict:
    """Retorna dict {num_mes: {saldo_inicial, saldo_actual}}."""
    ruta = Path(ruta_str) if ruta_str else _get_excel_path()
    wb = _load_workbook(ruta)
    saldos = {}
    for hoja, num_mes in MESES_MAP.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        saldo_actual = ws["F4"].value
        saldo_inicial = ws["F5"].value
        if saldo_actual is not None or saldo_inicial is not None:
            saldos[num_mes] = {
                "saldo_inicial": float(saldo_inicial) if saldo_inicial else 0.0,
                "saldo_actual": float(saldo_actual) if saldo_actual else 0.0,
            }
    return saldos


@st.cache_data(ttl=300)
def cargar_categorias(ruta_str: str = None) -> pd.DataFrame:
    """Carga tabla maestra grupos/conceptos/tipo desde hoja Categorias."""
    ruta = Path(ruta_str) if ruta_str else _get_excel_path()
    wb = _load_workbook(ruta)
    if "Categorias" not in wb.sheetnames:
        return pd.DataFrame(columns=["grupo", "concepto", "tipo"])
    ws = wb["Categorias"]
    # Col A=GRUPOS, B=CONCEPTOS, D=Prioridad, E=GRUPOS_TIPO, F=Tipo
    tipo_map = {}
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        grupo_col, concepto_col = row[0], row[1]
        grupo_tipo, tipo = (row[4] if len(row) > 4 else None), (row[5] if len(row) > 5 else None)
        if grupo_tipo and tipo:
            tipo_map[str(grupo_tipo).strip()] = str(tipo).strip()
        if grupo_col and concepto_col:
            filas.append({
                "grupo": str(grupo_col).strip(),
                "concepto": str(concepto_col).strip(),
            })
    df = pd.DataFrame(filas) if filas else pd.DataFrame(columns=["grupo", "concepto"])
    df["tipo"] = df["grupo"].map(tipo_map).fillna("Variable")
    return df


@st.cache_data(ttl=300)
def cargar_resumen_anual(ruta_str: str = None) -> pd.DataFrame:
    """Carga hoja Resumen → DataFrame con grupos como índice y meses como columnas."""
    ruta = Path(ruta_str) if ruta_str else _get_excel_path()
    wb = _load_workbook(ruta)
    if "Resumen" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["Resumen"]
    meses_cols = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
                  "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
    filas = []
    header_row = None
    col_indices = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        for j, val in enumerate(row):
            if val and str(val).upper() in meses_cols:
                header_row = i
                col_indices[str(val).upper()] = j
    if not col_indices:
        return pd.DataFrame()
    concepto_col_idx = 1  # col B = index 1
    for row in ws.iter_rows(min_row=(header_row or 3) + 1, values_only=True):
        concepto = row[concepto_col_idx] if len(row) > concepto_col_idx else None
        if concepto is None:
            continue
        fila = {"Grupo": str(concepto).strip()}
        for mes, idx in col_indices.items():
            val = row[idx] if len(row) > idx else 0
            fila[mes] = float(val) if isinstance(val, (int, float)) and val else 0.0
        filas.append(fila)
    if not filas:
        return pd.DataFrame()
    return pd.DataFrame(filas)  # Grupo queda como columna normal


_MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def _parse_clp_ultima(linea: str):
    """Extrae el ÚLTIMO monto CLP de una línea: '$ X.XXX.XXX' → float."""
    matches = re.findall(r'\$\s*([\d.]+)', linea)
    if not matches:
        return None
    raw = matches[-1].replace(".", "")
    try:
        return float(raw)
    except ValueError:
        return None


def _parse_clp_primera(linea: str):
    """Extrae el PRIMER monto CLP de una línea."""
    m = re.search(r'\$\s*([\d.]+)', linea)
    if not m:
        return None
    try:
        return float(m.group(1).replace(".", ""))
    except ValueError:
        return None


def parsear_liquidacion(archivo_pdf_bytes: bytes) -> dict:
    """Parsea liquidación de sueldo PDF → dict con campos estructurados.

    Formato real (Egakat / REX):
        Liquidación de sueldo Febrero 2026
        Sueldo Base $ 1.700.000
        Cotizacion AFP: en AFP Provida 11.45% sobre: 2.262.985 $ 259.112
        Salud: en Consalud UF6.081 (...) $ 241.967
        Impuesto: 4.0% sobre: (...) menos ... $ 32.343
        Líquido a pagar: $ 1.722.668
    """
    import pdfplumber
    import io

    campos = {
        "periodo": None,
        "empresa": None,
        "nombre_trabajador": None,
        "cargo": None,
        "dias_trabajados": None,
        "sueldo_base": None,
        "bono": None,
        "gratificacion": None,
        "colacion": None,
        "movilizacion": None,
        "total_haberes_afectos": None,
        "total_haberes_exentos": None,
        "afp": None,
        "salud": None,
        "cesantia": None,
        "impuesto": None,
        "total_descuentos_legales": None,
        "anticipo": None,
        "seguro_complementario": None,
        "total_otros_descuentos": None,
        "liquido": None,
    }

    try:
        with pdfplumber.open(io.BytesIO(archivo_pdf_bytes)) as pdf:
            lineas = []
            for page in pdf.pages:
                texto = page.extract_text() or ""
                lineas.extend(texto.splitlines())
    except Exception:
        return campos

    for linea in lineas:
        l = linea.strip()
        ll = l.lower()

        # Período: "Liquidación de sueldo Febrero 2026"
        if campos["periodo"] is None:
            m = re.search(r'liquidaci[oó]n de sueldo\s+(\w+)\s+(\d{4})', l, re.IGNORECASE)
            if m:
                campos["periodo"] = f"{m.group(1).capitalize()} {m.group(2)}"

        # Empresa (primera línea no vacía con letras)
        if campos["empresa"] is None and re.match(r'^[A-Z][A-Z\s]+(?:SPA|SA|LTDA|S\.A\.)', l):
            campos["empresa"] = l.strip()

        # Nombre trabajador
        if campos["nombre_trabajador"] is None:
            m = re.search(r'Nombre:\s+(.+?)\s{2,}', l)
            if m:
                campos["nombre_trabajador"] = m.group(1).strip()

        # Cargo
        if campos["cargo"] is None:
            m = re.search(r'Cargo:\s+(.+?)\s{2,}', l)
            if m:
                campos["cargo"] = m.group(1).strip()

        # Días trabajados
        if campos["dias_trabajados"] is None:
            m = re.search(r'D[ií]as trabajados:\s*(\d+)', l, re.IGNORECASE)
            if m:
                campos["dias_trabajados"] = int(m.group(1))

        # Sueldo Base (haber, no el del header)
        if campos["sueldo_base"] is None and re.match(r'sueldo base\s*\$', ll):
            campos["sueldo_base"] = _parse_clp_ultima(l)

        # Bono (cualquier línea que empiece con "Bono")
        if campos["bono"] is None and ll.startswith("bono"):
            v = _parse_clp_ultima(l)
            if v and v >= 1000:
                campos["bono"] = v

        # Gratificación
        if campos["gratificacion"] is None and re.match(r'gratificaci[oó]n\s*\$', ll):
            campos["gratificacion"] = _parse_clp_ultima(l)

        # Colación
        if campos["colacion"] is None and re.match(r'colaci[oó]n\s*\$', ll):
            campos["colacion"] = _parse_clp_ultima(l)

        # Movilización
        if campos["movilizacion"] is None and re.match(r'movilizaci[oó]n\s*\$', ll):
            campos["movilizacion"] = _parse_clp_ultima(l)

        # Total Haberes Afectos
        if campos["total_haberes_afectos"] is None and "total haberes afectos" in ll:
            campos["total_haberes_afectos"] = _parse_clp_ultima(l)

        # Total Haberes Exentos
        if campos["total_haberes_exentos"] is None and "total haberes exentos" in ll:
            campos["total_haberes_exentos"] = _parse_clp_ultima(l)

        # AFP: "Cotizacion AFP: en AFP Provida 11.45% sobre: 2.262.985 $ 259.112"
        if campos["afp"] is None and re.match(r'cotizaci[oó]n afp:', ll):
            campos["afp"] = _parse_clp_ultima(l)

        # Salud: "Salud: en Consalud UF6.081 (...) $ 241.967"
        if campos["salud"] is None and re.match(r'salud:', ll):
            campos["salud"] = _parse_clp_ultima(l)

        # Seguro de Cesantía
        if campos["cesantia"] is None and "seguro de cesantia" in ll:
            campos["cesantia"] = _parse_clp_ultima(l)

        # Impuesto: "Impuesto: 4.0% sobre: (...) menos ... $ 32.343"
        if campos["impuesto"] is None and re.match(r'impuesto:', ll):
            campos["impuesto"] = _parse_clp_ultima(l)

        # Total Descuentos Legales
        if campos["total_descuentos_legales"] is None and "total descuentos legales" in ll:
            campos["total_descuentos_legales"] = _parse_clp_ultima(l)

        # Anticipo (línea simple: "Anticipo $ 380.000")
        if campos["anticipo"] is None and re.match(r'anticipo\s*\$', ll):
            campos["anticipo"] = _parse_clp_ultima(l)

        # Seguro Complementario
        if campos["seguro_complementario"] is None and "seguro complementario" in ll:
            campos["seguro_complementario"] = _parse_clp_ultima(l)

        # Total Otros Descuentos
        if campos["total_otros_descuentos"] is None and "total otros descuentos" in ll:
            campos["total_otros_descuentos"] = _parse_clp_ultima(l)

        # Líquido a pagar
        if campos["liquido"] is None and re.match(r'l[ií]quido a pagar:', ll):
            campos["liquido"] = _parse_clp_ultima(l)

    return campos


def cargar_liquidaciones_carpeta(carpeta: str) -> list:
    """Carga todas las liquidaciones PDF de una carpeta → lista de dicts ordenada por período.

    Cada dict contiene todos los campos de parsear_liquidacion() más:
      - archivo: nombre del archivo
      - fecha_periodo: datetime object (ej. datetime(2026, 2, 1))
    Los datos se retornan en memoria — nunca se persisten en disco.
    """
    carpeta_path = Path(carpeta)
    if not carpeta_path.exists():
        return []

    pdfs = sorted(carpeta_path.glob("Liquidacion_contrato_*.pdf"))
    resultados = []

    for pdf_path in pdfs:
        try:
            datos = parsear_liquidacion(pdf_path.read_bytes())
            datos["archivo"] = pdf_path.name

            # Intentar parsear fecha desde nombre: Liquidacion_contrato_1_2026-02_1.pdf
            fecha_periodo = None
            m_nombre = re.search(r'(\d{4})-(\d{2})', pdf_path.name)
            if m_nombre:
                from datetime import datetime as _dt
                try:
                    fecha_periodo = _dt(int(m_nombre.group(1)), int(m_nombre.group(2)), 1)
                except ValueError:
                    pass

            # Fallback: parsear desde campo periodo "Febrero 2026"
            if fecha_periodo is None and datos.get("periodo"):
                partes = str(datos["periodo"]).split()
                if len(partes) == 2:
                    mes_str = partes[0].lower()
                    anio_str = partes[1]
                    mes_num = _MESES_ES.get(mes_str)
                    if mes_num and anio_str.isdigit():
                        from datetime import datetime as _dt
                        fecha_periodo = _dt(int(anio_str), mes_num, 1)

            datos["fecha_periodo"] = fecha_periodo
            resultados.append(datos)
        except Exception:
            continue

    resultados.sort(key=lambda x: x["fecha_periodo"] or datetime.min)
    return resultados


def parsear_amipass_archivo(archivo_bytes: bytes, nombre_archivo: str) -> float:
    """Extrae monto Amipass de distintos formatos de archivo → float o None.

    Soporta: PDF, XLSX, XLS, TXT, CSV.
    Retorna el primer monto ≥ 1.000 encontrado. Nunca persiste datos en disco.
    """
    ext = Path(nombre_archivo).suffix.lower()

    try:
        if ext == ".pdf":
            import pdfplumber, io
            with pdfplumber.open(io.BytesIO(archivo_bytes)) as pdf:
                texto = " ".join(p.extract_text() or "" for p in pdf.pages)
            # Buscar montos CLP: "$ 58.000" o "58.000" o "58000"
            matches = re.findall(r'\$?\s*([\d.]{4,})', texto)
            for raw in matches:
                try:
                    val = float(raw.replace(".", ""))
                    if val >= 1000:
                        return val
                except ValueError:
                    continue

        elif ext in (".xlsx", ".xls"):
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), read_only=True, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, (int, float)) and cell >= 1000:
                            return float(cell)

        elif ext in (".txt", ".csv"):
            texto = archivo_bytes.decode("utf-8", errors="replace")
            matches = re.findall(r'[\d.]{4,}', texto)
            for raw in matches:
                try:
                    val = float(raw.replace(".", ""))
                    if val >= 1000:
                        return val
                except ValueError:
                    continue

    except Exception:
        pass

    return None


def cargar_afp_movimientos(ruta_xlsx: str) -> pd.DataFrame:
    """Carga Excel AFP ProVida → DataFrame con FECHA, GIROS, APORTES, DESCRIPCION."""
    try:
        df = pd.read_excel(ruta_xlsx, header=0)
        df.columns = [str(c).strip().upper() for c in df.columns]
        rename = {}
        for col in df.columns:
            if "fecha" in col.lower():
                rename[col] = "FECHA"
            elif "giro" in col.lower():
                rename[col] = "GIROS"
            elif "aporte" in col.lower() or "abono" in col.lower():
                rename[col] = "APORTES"
            elif "descrip" in col.lower() or "concepto" in col.lower():
                rename[col] = "DESCRIPCION"
        df = df.rename(columns=rename)
        for col in ["FECHA", "GIROS", "APORTES", "DESCRIPCION"]:
            if col not in df.columns:
                df[col] = None
        df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce")
        df["GIROS"] = pd.to_numeric(df["GIROS"], errors="coerce").fillna(0)
        df["APORTES"] = pd.to_numeric(df["APORTES"], errors="coerce").fillna(0)
        return df[["FECHA", "GIROS", "APORTES", "DESCRIPCION"]].dropna(subset=["FECHA"])
    except Exception as e:
        return pd.DataFrame(columns=["FECHA", "GIROS", "APORTES", "DESCRIPCION"])


def cargar_gastos_compartidos(ruta_str: str = None) -> dict:
    """Carga hoja Gastos Compartidos → dict con fecha, items, total, por_persona."""
    ruta = Path(ruta_str) if ruta_str else _get_excel_path()
    wb = _load_workbook(ruta)
    if "Gastos Compartidos" not in wb.sheetnames:
        return {}
    ws = wb["Gastos Compartidos"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    fecha = rows[0][0] if rows else None
    items = []
    total_total = 0
    total_persona = 0
    for row in rows[2:]:
        concepto, total, por_persona = row[0], row[1], row[2]
        if concepto is None:
            continue
        if str(concepto).lower() == "total":
            total_total = float(total) if isinstance(total, (int, float)) else 0
            total_persona = float(por_persona) if isinstance(por_persona, (int, float)) else 0
            continue
        items.append({
            "concepto": str(concepto).strip(),
            "total": float(total) if isinstance(total, (int, float)) else 0,
            "por_persona": float(por_persona) if isinstance(por_persona, (int, float)) else 0,
        })
    return {
        "fecha": str(fecha) if fecha else "",
        "items": items,
        "total": total_total,
        "por_persona": total_persona,
    }


@st.cache_data(ttl=300)
def cargar_patrimonio_mensual(ruta_str: str = None) -> pd.DataFrame:
    """Lee hoja Patrimonio del formato nuevo → DataFrame con snapshots mensuales."""
    ruta = Path(ruta_str) if ruta_str else _get_excel_path()
    try:
        wb = _load_workbook(ruta)
    except Exception:
        return pd.DataFrame()
    if "Patrimonio" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["Patrimonio"]
    headers_raw = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers_raw)]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        fila = {}
        for i, h in enumerate(headers):
            fila[h] = row[i] if i < len(row) else None
        filas.append(fila)
    return pd.DataFrame(filas) if filas else pd.DataFrame(columns=headers)


@st.cache_data(ttl=300)
def cargar_inversiones(ruta_str: str = None) -> pd.DataFrame:
    """
    Lee hoja 'Inversiones' del Excel.
    Columnas esperadas: Activo | Ticker_CG | Cantidad | Precio_Compra_CLP | Fecha_Compra
    Retorna DataFrame vacío si la hoja no existe.
    """
    ruta = Path(ruta_str) if ruta_str else _get_excel_path()
    try:
        wb = _load_workbook(ruta)
    except Exception:
        return pd.DataFrame()
    if "Inversiones" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["Inversiones"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        return pd.DataFrame()
    headers = [str(c).strip() if c else f"col{i}" for i, c in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        data.append(dict(zip(headers, row)))
    if not data:
        return pd.DataFrame()
    df = pd.DataFrame(data)
    # Normalizar columnas clave
    rename = {}
    for col in df.columns:
        cl = col.lower().replace(" ", "_")
        if "activo" in cl:            rename[col] = "activo"
        elif "ticker" in cl:          rename[col] = "ticker_cg"
        elif "cantidad" in cl:        rename[col] = "cantidad"
        elif "precio" in cl and "compra" in cl: rename[col] = "precio_compra_clp"
        elif "fecha" in cl:           rename[col] = "fecha_compra"
    df = df.rename(columns=rename)
    for col in ["activo", "ticker_cg", "cantidad", "precio_compra_clp", "fecha_compra"]:
        if col not in df.columns:
            df[col] = None
    df["cantidad"]          = pd.to_numeric(df["cantidad"], errors="coerce").fillna(0)
    df["precio_compra_clp"] = pd.to_numeric(df["precio_compra_clp"], errors="coerce").fillna(0)
    df = df[df["cantidad"] > 0].reset_index(drop=True)
    return df


@st.cache_data(ttl=300)
def cargar_config_excel(ruta_str: str = None) -> dict:
    """Lee hoja Config del formato nuevo → dict {parámetro: valor}."""
    ruta = Path(ruta_str) if ruta_str else _get_excel_path()
    try:
        wb = _load_workbook(ruta)
    except Exception:
        return {}
    if "Config" not in wb.sheetnames:
        return {}
    ws = wb["Config"]
    config = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        val = row[1] if len(row) > 1 else None
        if key and val is not None:
            config[key] = val
    return config
