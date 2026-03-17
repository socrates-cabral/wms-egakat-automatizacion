import sys
sys.stdout.reconfigure(encoding="utf-8")

"""
bank_scraper.py — Movimientos bancarios BancoEstado.

CUENTAS SOPORTADAS:
  - Cuenta Corriente
  - CuentaRUT
  - Línea de Crédito
  - Cartola (estado de cuenta histórico)

ESTRATEGIAS DISPONIBLES:
  A) Carga manual:        usuario descarga el Excel del banco y lo sube a la app (recomendado)
  B) Scraping automático: login Playwright → navega → descarga Excel → parsea

FORMATO EXCEL BancoEstado (igual para todas las cuentas):
  Fila header: Fecha | N° Operación | Descripción | Cheques/Cargos $ | Depósitos/Abonos $ | Saldo $
  Cargo  → columna 3, valor negativo
  Abono  → columna 4, valor positivo

CREDENCIALES EN .env (no mostrar a nadie, ni a Claude):
  BANCO_ESTADO_RUT=XX.XXX.XXX-X
  BANCO_ESTADO_CLAVE=tu_clave
"""

import os
import json
import time
import re
import io
from pathlib import Path
from datetime import datetime, date, timedelta
from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).parent.parent.parent / ".env")

_DATA_DIR  = Path(__file__).parent.parent / "data"
_MOVS_FILE = _DATA_DIR / "bank_movimientos.json"
_DL_DIR    = _DATA_DIR / "bank_downloads"

TIMEOUT_NAV  = 60_000
TIMEOUT_ELEM = 30_000
TIMEOUT_DOWN = 60_000

NOMBRES_MESES = {
    1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril",
    5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto",
    9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre",
}

# ── Tipos de cuenta BancoEstado ──────────────────────────────────────────────
TIPOS_CUENTA_BE = {
    "cuenta_corriente": {
        "label":       "Cuenta Corriente",
        "menu_padre":  "Cuentas",       # menú izquierdo nivel 1
        "menu_hijo":   "Cuenta Corriente",  # submenú
        "url_path":    "cuentas/cuenta-corriente",
        "prefijo":     "CC",
    },
    "cuentarut": {
        "label":       "CuentaRUT",
        "menu_padre":  "Cuentas",
        "menu_hijo":   "CuentaRUT",
        "url_path":    "cuentas/cuentarut",
        "prefijo":     "RUT",
    },
    "linea_credito": {
        "label":       "Línea de Crédito",
        "menu_padre":  "Cuentas",
        "menu_hijo":   "Línea de Crédito",
        "url_path":    "cuentas/linea-de-credito",
        "prefijo":     "LC",
    },
    "cartola": {
        "label":       "Cartola (histórico)",
        "menu_padre":  "Cuentas",
        "menu_hijo":   "Cartolas",
        "url_path":    "cuentas/cartolas",
        "prefijo":     "CART",
    },
    "visa": {
        "label":       "Visa Smartmas (Tarjeta Crédito)",
        "menu_padre":  "Tarjetas",      # menú principal — NO bajo Cuentas
        "menu_hijo":   None,
        "url_path":    "tarjetas",
        "prefijo":     "VISA",
    },
}

# ── Grupos automáticos por descripción ───────────────────────────────────────
_GRUPOS = [
    (r"PAGO HIPOTECARIO|DIVIDENDO",        "Vivienda"),
    (r"TRANSFERENCIA|TRASPASO|TEF ",       "Transferencias"),
    (r"COMPRA HIP LIDER|LIDER|JUMBO|WALMART|UNIMARC|SANTA ISABEL", "Supermercado"),
    (r"COMPRA.*FARMACIA|FARMACIAS",        "Salud"),
    (r"COMPRA.*REST|RESTAUR|BURGER|CAFE|GELATO|PIZZA|SUSHI", "Restaurantes"),
    (r"MERCADOPAGO|MERCADO PAGO",          "MercadoPago"),
    (r"INTERESES|IMPUESTO LINEA|COMISION MANTENCION", "Cargos bancarios"),
    (r"COMPRA.*EASY|SODIMAC|HOME DEPOT",   "Hogar"),
    (r"COMPRA.*PARIS|FALABELLA|RIPLEY",    "Retail"),
    (r"SUELDO|REMUNERACION",              "Ingresos"),
]

def _clasificar(descripcion: str) -> str:
    d = descripcion.upper()
    for patron, grupo in _GRUPOS:
        if re.search(patron, d):
            return grupo
    return "Otros"


# ── I/O local ─────────────────────────────────────────────────────────────────

def _cargar() -> list:
    if _MOVS_FILE.exists():
        try:
            return json.loads(_MOVS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []


def _guardar(movs: list):
    _DATA_DIR.mkdir(exist_ok=True)
    _MOVS_FILE.write_text(
        json.dumps(movs, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )


def obtener_movimientos_banco() -> list:
    return _cargar()


def resumen_banco(movs: list) -> dict:
    if not movs:
        return {"total": 0, "gastos": 0, "ingresos": 0, "bancos": [], "ultimo_sync": None}
    gastos   = sum(m["importe"] for m in movs if m.get("tipo") == "cargo")
    ingresos = sum(m["importe"] for m in movs if m.get("tipo") == "abono")
    bancos   = list({m.get("banco", "?") for m in movs})
    fechas   = [m.get("_sync") for m in movs if m.get("_sync")]
    return {
        "total":       len(movs),
        "gastos":      gastos,
        "ingresos":    ingresos,
        "bancos":      bancos,
        "ultimo_sync": max(fechas) if fechas else None,
    }


def _merge(nuevos: list) -> dict:
    existentes = _cargar()
    ids_exist  = {m["_id"] for m in existentes}
    agregados  = [m for m in nuevos if m["_id"] not in ids_exist]
    todos      = existentes + agregados
    if len(todos) > 5000:
        todos = sorted(todos, key=lambda x: x.get("fecha", ""), reverse=True)[:5000]
    _guardar(todos)
    return {"ok": True, "total": len(nuevos), "nuevos": len(agregados), "error": None}


# ═══════════════════════════════════════════════════════════════════════════════
# PARSER EXCEL — formato nativo BancoEstado
# ═══════════════════════════════════════════════════════════════════════════════

def _detectar_tipo_cuenta(nombre_hoja: str, ws=None) -> str:
    """
    Detecta el tipo de cuenta desde el nombre de la hoja o el contenido.
    BancoEstado usa nombres como 'ltimos Movimientos 37100103371' (sin acento).
    Busca también en las primeras filas del worksheet.
    """
    h = nombre_hoja.lower()
    if "cuentarut" in h or "cuenta rut" in h:
        return "CuentaRUT"
    if "corriente" in h:
        return "Cuenta Corriente"
    if "linea" in h or "línea" in h or "credito" in h or "crédito" in h:
        return "Línea de Crédito"
    if "cartola" in h:
        return "Cartola"

    # Buscar en primeras 5 filas del contenido
    if ws is not None:
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            txt = " ".join(str(c) for c in row if c is not None).lower()
            if "cuentarut" in txt or "cuenta rut" in txt:
                return "CuentaRUT"
            if "corriente" in txt:
                return "Cuenta Corriente"
            if "linea de credito" in txt or "línea de crédito" in txt:
                return "Línea de Crédito"
            if "visa" in txt or "tarjeta de credito" in txt or "smartmas" in txt:
                return "Visa Smartmas"
            if "cartola" in txt:
                return "Cartola"

    return "BancoEstado"


def parsear_excel_bancoestado(source, tipo_cuenta: str = None) -> list:
    """
    Parsea cualquier Excel de BancoEstado (Cuenta Corriente, CuentaRUT, Línea de Crédito, Cartola).
    El formato de todas las cuentas es idéntico.

    source: Path | bytes | BytesIO
    tipo_cuenta: etiqueta opcional — si no se da, se detecta desde el nombre de la hoja.
    Retorna lista de movimientos en formato interno.
    """
    import openpyxl

    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(source, data_only=True)
    else:
        if isinstance(source, bytes):
            source = io.BytesIO(source)
        wb = openpyxl.load_workbook(source, data_only=True)

    ws = wb.active
    # Detectar tipo desde nombre de hoja + contenido si no se proporcionó
    if not tipo_cuenta:
        tipo_cuenta = _detectar_tipo_cuenta(ws.title, ws)
    movimientos = []
    header_row  = None

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Buscar fila de encabezado de datos
        if header_row is None:
            if row and str(row[0]).strip().lower() == "fecha":
                header_row = i
            continue

        # Fila vacía o subtotales → terminar
        if not row or row[0] is None:
            break
        fecha_val = str(row[0]).strip()
        if not re.match(r"\d{2}/\d{2}/\d{4}", fecha_val):
            break

        num_op  = str(row[1]).strip() if row[1] is not None else ""
        desc    = str(row[2]).strip() if row[2] is not None else ""
        cargo_v = row[3]   # negativo o 0
        abono_v = row[4]   # positivo o 0

        # Convertir a float
        try:
            cargo = float(str(cargo_v).replace(".", "").replace(",", ".").replace("$", "")) if cargo_v is not None else 0
        except ValueError:
            cargo = 0
        try:
            abono = float(str(abono_v).replace(".", "").replace(",", ".").replace("$", "")) if abono_v is not None else 0
        except ValueError:
            abono = 0

        # Determinar tipo
        if cargo != 0:
            importe = abs(cargo)
            tipo    = "cargo"
        elif abono != 0:
            importe = abs(abono)
            tipo    = "abono"
        else:
            continue  # fila sin monto

        if importe == 0:
            continue

        # Parsear fecha
        fecha_dt = None
        try:
            fecha_dt = datetime.strptime(fecha_val, "%d/%m/%Y").date()
        except ValueError:
            pass

        mes_num = fecha_dt.month if fecha_dt else 0

        prefijo_id = "".join(c for c in tipo_cuenta.upper() if c.isalpha())[:4]
        mov = {
            "banco":        "BancoEstado",
            "tipo_cuenta":  tipo_cuenta,
            "fecha":        str(fecha_dt) if fecha_dt else fecha_val,
            "mes":          mes_num,
            "mes_nombre":   NOMBRES_MESES.get(mes_num, ""),
            "num_op":       num_op,
            "descripcion":  desc[:80],
            "importe":      importe,
            "tipo":         tipo,
            "grupo":        _clasificar(desc) if tipo == "cargo" else "Ingresos banco",
            "_sync":        datetime.now().isoformat(),
            "_id":          f"BE_{prefijo_id}_{fecha_val}_{num_op}_{int(importe)}_{tipo}",
        }
        movimientos.append(mov)

    return movimientos


def cargar_excel_manual(source, tipo_cuenta: str = None) -> dict:
    """
    Carga un Excel BancoEstado subido manualmente.
    source: Path | bytes | BytesIO
    tipo_cuenta: opcional — si no se da, se detecta desde el nombre de hoja.
    Retorna resumen {ok, total, nuevos, error}.
    """
    try:
        movs = parsear_excel_bancoestado(source, tipo_cuenta=tipo_cuenta)
        if not movs:
            return {"ok": False, "error": "No se encontraron movimientos en el archivo.", "total": 0, "nuevos": 0}
        return _merge(movs)
    except Exception as e:
        return {"ok": False, "error": str(e), "total": 0, "nuevos": 0}


# ═══════════════════════════════════════════════════════════════════════════════
# SCRAPER AUTOMÁTICO — Playwright → descarga Excel → parsea
# ═══════════════════════════════════════════════════════════════════════════════

def _descargar_todas_cartolas(page, destino: Path) -> list:
    """
    En la página de Cartolas ya abierta, itera años (2024→2026),
    busca cartolas disponibles y descarga cada una.
    Retorna lista de Path de archivos descargados.
    """
    from playwright.sync_api import TimeoutError as PWTimeout

    AÑOS = ["2026", "2025", "2024"]
    archivos = []

    for año in AÑOS:
        try:
            # Seleccionar año en el dropdown
            # BancoEstado usa dropdown custom — click en el contenedor → click en la opción
            year_dropdown = page.locator("text=Selecciona un año, text=2026, text=2025, text=2024").first
            # Abrir dropdown de año
            page.locator("div:has-text('Filtrar por año') + div, [class*='year'], [class*='anio']").first.click(timeout=8_000)
            time.sleep(0.4)
            page.locator(f"li:has-text('{año}'), option:has-text('{año}'), [role='option']:has-text('{año}')").first.click(timeout=8_000)
            time.sleep(0.3)
        except Exception:
            try:
                # Fallback: buscar directamente el texto del año en los items del dropdown
                page.locator(f"text={año}").nth(1).click(timeout=5_000)
                time.sleep(0.3)
            except Exception:
                continue

        # Click en Buscar
        try:
            page.locator("button:has-text('Buscar')").click(timeout=8_000)
            page.wait_for_load_state("networkidle", timeout=TIMEOUT_NAV)
            time.sleep(1)
        except Exception:
            continue

        # Contar botones Descargar en la tabla
        botones = page.locator("button:has-text('Descargar'), a:has-text('Descargar')").all()
        if not botones:
            continue

        for i in range(len(botones)):
            try:
                # Re-localizar porque el DOM puede refrescar
                btn = page.locator("button:has-text('Descargar'), a:has-text('Descargar')").nth(i)
                with page.expect_download(timeout=TIMEOUT_DOWN) as dl_info:
                    btn.click()
                dl = dl_info.value
                ruta = destino / f"cartola_{año}_{i+1:02d}_{dl.suggested_filename}"
                dl.save_as(str(ruta))
                archivos.append(ruta)
                time.sleep(0.8)
            except Exception:
                continue

    return archivos


def scrape_bancoestado(tipo_cuenta: str = "cuenta_corriente", headless: bool = True) -> dict:
    """
    Login en BancoEstado → navega al tipo de cuenta → descarga Excel → parsea.

    tipo_cuenta: 'cuenta_corriente' | 'cuentarut' | 'linea_credito' | 'cartola'
    Retorna {ok, total, nuevos, error}.
    """
    cfg   = TIPOS_CUENTA_BE.get(tipo_cuenta, TIPOS_CUENTA_BE["cuenta_corriente"])
    rut   = os.getenv("BANCO_ESTADO_RUT", "")
    clave = os.getenv("BANCO_ESTADO_CLAVE", "")
    if not rut or not clave:
        return {"ok": False, "error": "BANCO_ESTADO_RUT / BANCO_ESTADO_CLAVE no configurados en .env", "total": 0, "nuevos": 0}

    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

    _DL_DIR.mkdir(parents=True, exist_ok=True)
    error_msg   = None
    excel_path  = None

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        ctx = browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            accept_downloads=True,
        )
        page = ctx.new_page()

        try:
            # ── Login ────────────────────────────────────────────────────────
            # URL pública home: bancoestado.cl/content/bancoestado-public/.../home.html
            # Post-login redirige a: nwm.bancoestado.cl/content/bancoestado/.../home.html
            # Navegar directo a la URL con el modal de login abierto (#/login)
            page.goto(
                "https://www.bancoestado.cl/content/bancoestado-public/cl/es/home/home.html#/login",
                timeout=TIMEOUT_NAV,
                wait_until="domcontentloaded",
            )
            time.sleep(2)

            if headless:
                # Modo automático: llenar formulario (puede fallar si hay bot-detection)
                rut_norm = rut.replace(".", "").replace("-", "").lower()
                page.wait_for_selector(
                    "input[placeholder*='12345678'], input[placeholder*='RUT']",
                    timeout=TIMEOUT_ELEM,
                )
                rut_field = page.locator(
                    "input[placeholder*='12345678'], input[placeholder*='RUT']"
                ).first
                rut_field.click()
                # Teclear lento carácter a carácter para evitar detección
                for c in rut_norm:
                    rut_field.type(c, delay=80)
                time.sleep(0.5)
                for c in clave:
                    page.locator("input[type='password']").first.type(c, delay=80)
                time.sleep(0.4)
                page.locator("button:has-text('Ingresar')").first.click()
            else:
                # Modo visible (semi-manual): el usuario hace login normalmente
                # El scraper espera hasta que detecta el portal privado
                print("\n" + "="*55)
                print("  Ingresa a BancoEstado en el browser que se abrió.")
                print("  El scraper continuará automáticamente al detectar")
                print("  que estás logueado (hasta 2 minutos de espera).")
                print("="*55 + "\n")

            # Esperar redirect a nwm.bancoestado.cl (hasta 3 min en modo visible)
            timeout_login = 180_000 if not headless else TIMEOUT_NAV
            page.wait_for_url("**/nwm.bancoestado.cl/**", timeout=timeout_login)
            time.sleep(2)

            # Cerrar popup "Tienes un Avance en Cuotas" (aparece post-login)
            try:
                page.click("button:has-text('Cerrar')", timeout=6_000)
                time.sleep(0.5)
            except PWTimeout:
                pass

            # ── Navegar al tipo de cuenta seleccionado ────────────────────────
            # Post-login el dominio cambia a nwm.bancoestado.cl
            # URL confirmada: nwm.bancoestado.cl/content/bancoestado/cl/es/home/...
            menu_padre = cfg["menu_padre"]
            menu_hijo  = cfg["menu_hijo"]

            # Expandir menú padre
            page.click(f"a:has-text('{menu_padre}')", timeout=TIMEOUT_ELEM)
            time.sleep(0.6)

            if menu_hijo:
                page.click(f"a:has-text('{menu_hijo}')", timeout=TIMEOUT_ELEM)
                page.wait_for_load_state("networkidle", timeout=TIMEOUT_NAV)
                time.sleep(1.5)

                # Cartolas: tiene dos tabs — "Saldos y movimientos" | "Cartolas"
                # Para descargar cartolas históricas hay que estar en el tab "Cartolas"
                if tipo_cuenta == "cartola":
                    try:
                        page.click("button:has-text('Cartolas'), a:has-text('Cartolas')", timeout=8_000)
                        page.wait_for_load_state("networkidle", timeout=TIMEOUT_NAV)
                        time.sleep(1)
                    except Exception:
                        pass
            else:
                # Tarjetas: menú directo
                page.wait_for_load_state("networkidle", timeout=TIMEOUT_NAV)
                time.sleep(1)
                if tipo_cuenta == "visa":
                    try:
                        # Seleccionar Visa Smartmas
                        visa_card = page.locator("text=Visa Smartmas").first
                        visa_card.locator("..").locator("button:has-text('Seleccionar')").click()
                        page.wait_for_load_state("networkidle", timeout=TIMEOUT_NAV)
                        time.sleep(1)
                    except Exception:
                        pass

            # ── Descargar archivo ─────────────────────────────────────────────
            if tipo_cuenta == "cartola":
                # Cartolas: iterar por año y descargar todas las disponibles
                archivos_descargados = _descargar_todas_cartolas(page, _DL_DIR)
                if not archivos_descargados:
                    error_msg = "No se encontraron cartolas para descargar."
            else:
                # Saldos/movimientos: botón Excel en la vista
                with page.expect_download(timeout=TIMEOUT_DOWN) as dl_info:
                    page.click(
                        "button:has-text('Excel'), a:has-text('Excel'), "
                        "button:has-text('Descargar'), a[download], "
                        "button[title*='Excel'], img[alt*='Excel']",
                        timeout=TIMEOUT_ELEM,
                    )
                dl = dl_info.value
                excel_path = _DL_DIR / dl.suggested_filename
                dl.save_as(str(excel_path))
            dl = dl_info.value
            excel_path = _DL_DIR / dl.suggested_filename
            dl.save_as(str(excel_path))

        except PWTimeout as e:
            error_msg = f"Timeout: {e}"
        except Exception as e:
            error_msg = str(e)
        finally:
            browser.close()

    if error_msg:
        return {"ok": False, "error": error_msg, "total": 0, "nuevos": 0}

    if tipo_cuenta == "cartola":
        # Parsear y consolidar todas las cartolas descargadas
        if not archivos_descargados:
            return {"ok": False, "error": "No se descargaron cartolas.", "total": 0, "nuevos": 0}
        todos_movs = []
        for arch in archivos_descargados:
            try:
                movs = parsear_excel_bancoestado(arch, tipo_cuenta="Cartola")
                todos_movs.extend(movs)
            except Exception:
                continue
        if not todos_movs:
            return {"ok": False, "error": "No se pudieron parsear las cartolas.", "total": 0, "nuevos": 0}
        return _merge(todos_movs)
    else:
        if not excel_path or not excel_path.exists():
            return {"ok": False, "error": "No se descargó el archivo Excel.", "total": 0, "nuevos": 0}
        return cargar_excel_manual(excel_path, tipo_cuenta=cfg["label"])


# ── Modo semi-manual: browser visible para debug / captcha ────────────────────

def scrape_bancoestado_visible(tipo_cuenta: str = "cuenta_corriente") -> dict:
    return scrape_bancoestado(tipo_cuenta=tipo_cuenta, headless=False)


# ═══════════════════════════════════════════════════════════════════════════════
# BCI — placeholder (en desarrollo)
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_bci() -> dict:
    return {"ok": False, "error": "BCI scraper en desarrollo. Usa carga manual por ahora.", "total": 0, "nuevos": 0}


# ═══════════════════════════════════════════════════════════════════════════════
# Conversión al formato estándar de transacciones de la app
# ═══════════════════════════════════════════════════════════════════════════════

def movimientos_banco_a_transacciones(movs: list = None) -> list:
    if movs is None:
        movs = _cargar()
    return [
        {
            "mes":        m.get("mes", 0),
            "mes_nombre": m.get("mes_nombre", ""),
            "grupo":      m.get("grupo", "Banco"),
            "concepto":   m.get("descripcion", ""),
            "fecha":      m.get("fecha"),
            "detalle":    f"{m.get('banco','')} #{m.get('num_op','')}",
            "importe":    m.get("importe", 0),
            "fuente":     "bancoestado",
        }
        for m in movs
    ]


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("accion", choices=["auto", "visible", "excel"], help="auto=scraper headless | visible=con pantalla | excel=parsear archivo local")
    parser.add_argument("--archivo", help="Ruta al Excel (solo para accion=excel)")
    args = parser.parse_args()

    if args.accion == "excel":
        if not args.archivo:
            print("Especifica --archivo ruta_al_excel.xlsx")
        else:
            res = cargar_excel_manual(Path(args.archivo))
            print(f"{'✅' if res['ok'] else '❌'} nuevos={res['nuevos']} total={res['total']} error={res['error']}")
    elif args.accion == "visible":
        res = scrape_bancoestado_visible()
        print(f"{'✅' if res['ok'] else '❌'} nuevos={res['nuevos']} total={res['total']} error={res['error']}")
    else:
        res = scrape_bancoestado()
        print(f"{'✅' if res['ok'] else '❌'} nuevos={res['nuevos']} total={res['total']} error={res['error']}")
