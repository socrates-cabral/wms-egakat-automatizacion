import sys
sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from datetime import date

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Paleta de colores ───────────────────────────────────────────────────────
C_HEADER_BG    = "0C1422"
C_HEADER_FONT  = "94A3B8"
C_SECTION_BG   = "1E293B"
C_SECTION_FONT = "14B8A6"
C_ODD_BG       = "111D2E"
C_EVEN_BG      = "0F1826"
C_DATA_FONT    = "E2E8F0"
C_BORDER       = "1E2D45"

# Colores importe por tipo
C_GASTO    = "F43F5E"
C_INGRESO  = "34D399"
C_INVERSION= "F59E0B"

# ─── Helpers de estilo ───────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font(hex_color: str, bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", color=hex_color, bold=bold, size=size)

def _border() -> Border:
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")

def _apply_header(ws, row: int, headers: list, widths: list = None):
    """Aplica estilo de encabezado a la fila indicada."""
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = _fill(C_HEADER_BG)
        cell.font = _font(C_HEADER_FONT, bold=True, size=10)
        cell.border = _border()
        cell.alignment = _center()
    if widths:
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

def _apply_section_title(ws, row: int, text: str, n_cols: int):
    """Fila de título de sección (merge + estilo oscuro teal)."""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=n_cols
    )
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = _fill(C_SECTION_BG)
    cell.font = _font(C_SECTION_FONT, bold=True, size=10)
    cell.alignment = _center()
    cell.border = _border()

def _apply_data_row(ws, row: int, values: list, n_cols: int,
                    importe_col: int = None, tipo_tx: str = None):
    """Aplica estilo a una fila de datos (alterna odd/even)."""
    bg = C_ODD_BG if row % 2 != 0 else C_EVEN_BG
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        if col_idx <= len(values):
            cell.value = values[col_idx - 1]
        cell.fill = _fill(bg)
        cell.border = _border()
        cell.alignment = _left()
        # Color importe según tipo
        if importe_col and col_idx == importe_col and tipo_tx:
            color = {
                "Gasto": C_GASTO,
                "Ingreso": C_INGRESO,
                "Inversión": C_INVERSION,
            }.get(tipo_tx, C_DATA_FONT)
            cell.font = _font(color, size=10)
        else:
            cell.font = _font(C_DATA_FONT, size=10)


# ─── Sheet 1: Transacciones ──────────────────────────────────────────────────
def _crear_transacciones(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Transacciones")
    ws.freeze_panes = "A2"

    headers = ["Fecha", "Tipo", "Grupo", "Concepto", "Detalle", "Importe", "Cuenta"]
    widths  = [12, 13, 22, 20, 18, 14, 16]
    _apply_header(ws, 1, headers, widths)
    ws.row_dimensions[1].height = 20

    # Datos de muestra
    muestras = [
        # (fecha_obj, Tipo, Grupo, Concepto, Detalle, Importe, Cuenta)
        (date(2026, 3, 1),  "Gasto",     "Alimentación",    "Supermercado",  "Líder Macul",   45000,   "Cuenta Vista"),
        (date(2026, 3, 5),  "Ingreso",   "Sueldo Líquido",  "Sueldo Egakat", "Marzo 2026",    1700000, "Cuenta Vista"),
        (date(2026, 3, 10), "Inversión", "USDT / Cripto",   "Compra USDT",   "10 unidades",   370000,  "Cuenta Ahorro"),
    ]

    for row_idx, (fecha, tipo, grupo, concepto, detalle, importe, cuenta) in enumerate(muestras, start=2):
        valores = [fecha, tipo, grupo, concepto, detalle, importe, cuenta]
        _apply_data_row(ws, row_idx, valores, len(headers), importe_col=6, tipo_tx=tipo)
        # Fecha format
        ws.cell(row=row_idx, column=1).number_format = "DD/MM/YYYY"
        # Importe format
        ws.cell(row=row_idx, column=6).number_format = "#,##0"

    # Data Validation — Tipo (col B)
    dv_tipo = DataValidation(
        type="list",
        formula1='"Gasto,Ingreso,Inversión,Ahorro,Transferencia"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_tipo.sqref = "B2:B2000"
    ws.add_data_validation(dv_tipo)

    # Data Validation — Cuenta (col G)
    dv_cuenta = DataValidation(
        type="list",
        formula1='"Cuenta Vista,Cuenta Ahorro,Efectivo,Tarjeta Crédito,USDT,Otra"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_cuenta.sqref = "G2:G2000"
    ws.add_data_validation(dv_cuenta)

    # Nota: validación dinámica de Grupo basada en Tipo requiere Named Ranges o macros VBA.
    # Se omite la validación automática de Grupo. Ver hoja Categorias para referencia.

    # Formato fecha columna A (filas vacías)
    for r in range(2, 2001):
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=6).number_format = "#,##0"


# ─── Sheet 2: Categorias ─────────────────────────────────────────────────────
def _crear_categorias(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Categorias")
    ws.freeze_panes = "A2"

    headers = ["Grupo", "Tipo_Clasificacion", "Uso", "Concepto_Ejemplo", "Tipo_Tx"]
    widths  = [25, 20, 12, 28, 12]
    _apply_header(ws, 1, headers, widths)

    # Gastos
    gastos = [
        ("Hogar y Vivienda",         "Fijo",         "Gasto",    "Arriendo, Dividendo",         "Gasto"),
        ("Familia e Hijos",          "Fijo",         "Gasto",    "Colegio, Ropa niños",         "Gasto"),
        ("Financiero - Deudas",      "Fijo",         "Gasto",    "Crédito BCI, Tarjeta",        "Gasto"),
        ("Alimentación",             "Variable",     "Gasto",    "Supermercado, Restorán",      "Gasto"),
        ("Salud y Cuidado Personal", "Variable",     "Gasto",    "Médico, Farmacia",            "Gasto"),
        ("Transporte",               "Variable",     "Gasto",    "Metro, Bencina, Uber",        "Gasto"),
        ("Servicios Básicos",        "Fijo",         "Gasto",    "Luz, Agua, Internet",         "Gasto"),
        ("Educación y Formación",    "Variable",     "Gasto",    "Cursos, Libros",              "Gasto"),
        ("Ahorro e Inversión",       "Fijo",         "Gasto",    "DAP, Fondo Mutuo",            "Ahorro/Inversión"),
        ("Suscripciones Digitales",  "Prescindible", "Gasto",    "Netflix, Spotify",            "Gasto"),
        ("Ocio y Vida Social",       "Prescindible", "Gasto",    "Cine, Restaurante",           "Gasto"),
        ("Mascotas",                 "Variable",     "Gasto",    "Veterinario, Comida",         "Gasto"),
        ("Regalos y Donaciones",     "Prescindible", "Gasto",    "Regalo cumpleaños",           "Gasto"),
        ("Varios y Otros",           "Variable",     "Gasto",    "Imprevistos",                 "Gasto"),
        ("Seguros",                  "Fijo",         "Gasto",    "Seguro vida, auto",           "Gasto"),
    ]

    row_idx = 2
    for valores in gastos:
        _apply_data_row(ws, row_idx, list(valores), len(headers))
        row_idx += 1

    # Sección Ingresos (fila 17 = row_idx actual → título en fila 17)
    _apply_section_title(ws, row_idx, "--- INGRESOS ---", len(headers))
    row_idx += 1

    ingresos = [
        ("Sueldo Líquido",    "Fijo",     "Ingreso",    "Egakat mensual",        "Ingreso"),
        ("Anticipo",          "Variable", "Ingreso",    "Anticipo quincena",     "Ingreso"),
        ("Bono",              "Variable", "Ingreso",    "Bono anual",            "Ingreso"),
        ("Arriendo Recibido", "Fijo",     "Ingreso",    "Propiedad 1803",        "Ingreso"),
        ("Freelance",         "Variable", "Ingreso",    "Proyecto externo",      "Ingreso"),
        ("Otros Ingresos",    "Variable", "Ingreso",    "Varios",                "Ingreso"),
    ]

    for valores in ingresos:
        _apply_data_row(ws, row_idx, list(valores), len(headers))
        row_idx += 1

    # Sección Inversiones
    _apply_section_title(ws, row_idx, "--- INVERSIONES ---", len(headers))
    row_idx += 1

    inversiones = [
        ("USDT / Cripto",      "Variable", "Inversión",  "Compra USDT",           "Inversión"),
        ("Depósito a Plazo",   "Fijo",     "Inversión",  "DAP BancoEstado",       "Inversión"),
        ("Fondo Mutuo",        "Variable", "Inversión",  "Fondo conservador",     "Inversión"),
        ("APV",                "Fijo",     "Inversión",  "APV voluntario AFP",    "Inversión"),
    ]

    for valores in inversiones:
        _apply_data_row(ws, row_idx, list(valores), len(headers))
        row_idx += 1


# ─── Sheet 3: Patrimonio ─────────────────────────────────────────────────────
def _crear_patrimonio(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Patrimonio")
    ws.freeze_panes = "A2"

    headers = [
        "Período", "Cuenta Vista", "Cuenta Ahorro", "USDT Uni",
        "Precio USDT", "Valor USDT CLP", "AFP Saldo",
        "Propiedad 1", "Deuda Hipotecaria", "Otras Deudas", "Patrimonio Neto",
    ]
    widths = [10, 14, 14, 12, 13, 16, 14, 14, 18, 14, 16]
    _apply_header(ws, 1, headers, widths)

    # Fila de muestra
    row_idx = 2
    valores_muestra = [
        "Ene-2026", 1679673, 10349996, 909, 37000,
        None,  # fórmula
        8774527, 0, 0, 0,
        None,  # fórmula
    ]

    bg = C_ODD_BG if row_idx % 2 != 0 else C_EVEN_BG
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if col_idx <= len(valores_muestra) and valores_muestra[col_idx - 1] is not None:
            cell.value = valores_muestra[col_idx - 1]
        cell.fill = _fill(bg)
        cell.font = _font(C_DATA_FONT, size=10)
        cell.border = _border()
        cell.alignment = _left()

    # Fórmulas
    ws.cell(row=row_idx, column=6).value  = f"=D{row_idx}*E{row_idx}"
    ws.cell(row=row_idx, column=11).value = (
        f"=B{row_idx}+C{row_idx}+F{row_idx}+G{row_idx}+"
        f"H{row_idx}-I{row_idx}-J{row_idx}"
    )

    # Formatos numéricos
    fmt_clp = "#,##0"
    for col_idx in [2, 3, 5, 6, 7, 8, 9, 10, 11]:
        ws.cell(row=row_idx, column=col_idx).number_format = fmt_clp


# ─── Sheet 4: Config ─────────────────────────────────────────────────────────
def _crear_config(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Config")

    headers = ["Parámetro", "Valor", "Descripción"]
    widths  = [28, 16, 38]
    _apply_header(ws, 1, headers, widths)

    filas = [
        ("Ingresos Mensuales",    2160668, "Suma sueldo + arriendo + bonos"),
        ("USDT Precio CLP",       37000,   "Actualizar manualmente"),
        ("AFP Saldo",             8774527, "Saldo cuenta individual AFP"),
        ("Dividendo Mensual",     595821,  "Cuota hipoteca mensual"),
        ("ISAPRE Mensual",        241967,  "Descuento salud mensual"),
        ("Fondo Emergencia Meta", 6000000, "3 meses de gastos objetivo"),
    ]

    for row_idx, valores in enumerate(filas, start=2):
        _apply_data_row(ws, row_idx, list(valores), len(headers))
        # Formato CLP columna B
        cell_val = ws.cell(row=row_idx, column=2)
        if isinstance(valores[1], (int, float)):
            cell_val.number_format = "#,##0"


# ─── Función principal ───────────────────────────────────────────────────────
def crear_plantilla() -> openpyxl.Workbook:
    """Crea y retorna el Workbook completo. No guarda en disco."""
    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _crear_transacciones(wb)
    _crear_categorias(wb)
    _crear_patrimonio(wb)
    _crear_config(wb)

    return wb


def main():
    ruta = Path(__file__).parent.parent / "Plantilla_FinanzasPersonales.xlsx"
    wb = crear_plantilla()
    wb.save(str(ruta))
    print(f"Plantilla creada: {ruta}")


if __name__ == "__main__":
    main()
