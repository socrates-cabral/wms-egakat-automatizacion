"""
nps_descarga.py v3.0
Descarga CSAT (386641) y NPS (418429) desde LimeSurvey RemoteControl 2 API.
Genera Excel PowerBI y lo sube directamente a SharePoint usando Microsoft Graph API.
Egakat SPA — Control de Gestión y Mejora Continua
"""

import sys
if sys.stdout:
    sys.stdout.reconfigure(encoding="utf-8")

import os
import csv
import json
import base64
import logging
import requests
import openpyxl
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── IMPORTAR GRAPH API DESDE WMS_Automatizacion ────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent.parent / "WMS_Automatizacion"))
from azure_graph import get_token, get_drive_id

# ── CONFIGURACIÓN ──────────────────────────────────────────────────────────────
load_dotenv(dotenv_path=Path(__file__).parent.parent / ".env")

LIMESURVEY_URL      = os.getenv("LIMESURVEY_URL",      "https://egakat.limesurvey.net/index.php/admin/remotecontrol")
LIMESURVEY_USER     = os.getenv("LIMESURVEY_USER",     "")
LIMESURVEY_PASSWORD = os.getenv("LIMESURVEY_PASSWORD", "")
SURVEY_ID_CSAT      = int(os.getenv("LIMESURVEY_SURVEY_ID_CSAT", "386641"))
SURVEY_ID_NPS       = int(os.getenv("LIMESURVEY_SURVEY_ID_NPS",  "418429"))

DIR_NPS      = Path(__file__).parent
TOKENS_CSAT  = DIR_NPS / "tokens_csat.csv"   # Actualizar cada ronda mensual
TOKENS_NPS   = DIR_NPS / "tokens_nps.csv"    # Crear cuando lance NPS
CONTACTOS    = DIR_NPS / "Contactos_Clientes.xlsx"

ARCHIVO_LOCAL = DIR_NPS / "NPS_PBI_datos.xlsx"  # Se guarda local temporalmente
FOLDER_SP     = "NPS_EK/Reportes NPS"           # Ruta exacta en SharePoint
NOMBRE_SP     = "NPS_PBI_datos.xlsx"

LOG_DIR  = Path(__file__).parent.parent / "logs"
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"nps_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger(__name__)


# ── ESTRUCTURA DE ENCUESTAS ────────────────────────────────────────────────────
PREGUNTAS_CSAT = {
    "G01Q01": (1, "Satisfacción general del servicio"),
    "G01Q05": (5, "Cumplimiento de tiempos de entrega"),
    "G01Q06": (6, "Precisión de pedidos preparados"),
    "G01Q07": (7, "Información y seguimiento de pedidos"),
}

AREAS_CSAT = {
    "G01Q02[SQ001]": "Recepción",
    "G01Q02[SQ002]": "Preparación de pedidos",
    "G01Q02[SQ003]": "Despacho",
    "G01Q02[SQ004]": "Gestión de inventarios",
    "G01Q02[SQ005]": "Transporte",
    "G01Q02[SQ006]": "Calidad",
    "G01Q02[SQ007]": "Servicio al cliente",
}

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo",    4: "Abril",
    5: "Mayo",  6: "Junio",   7: "Julio",    8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# ── API LIMESURVEY ─────────────────────────────────────────────────────────────
class LimeSurveyAPI:
    def __init__(self, url: str):
        self.url = url
        self.session_key = None
        self._id = 0

    def _call(self, method: str, params: list):
        self._id += 1
        r = requests.post(
            self.url,
            json={"method": method, "params": params, "id": self._id},
            headers={"content-type": "application/json"},
            timeout=60,
        )
        r.raise_for_status()
        data = r.json()
        if data.get("error"):
            raise RuntimeError(f"API error: {data['error']}")
        return data.get("result")

    def open(self, user: str, password: str):
        result = self._call("get_session_key", [user, password])
        if not result or len(result) < 10:
            raise RuntimeError(f"Login LimeSurvey fallido: {result}")
        self.session_key = result
        log.info("Sesión LimeSurvey abierta OK")

    def close(self):
        if self.session_key:
            self._call("release_session_key", [self.session_key])
            self.session_key = None

    def export_responses(self, survey_id: int) -> list[dict]:
        result = self._call("export_responses",[
            self.session_key, survey_id, "json", "es", "complete", "code", "short",
        ])
        if not result or (isinstance(result, dict) and "No Data" in str(result.get("status", ""))):
            return[]
        decoded = base64.b64decode(result).decode("utf-8")
        return json.loads(decoded).get("responses",[])


# ── MAPEO TOKEN → CLIENTE ──────────────────────────────────────────────────────
def cargar_contactos() -> dict:
    if not CONTACTOS.exists():
        log.warning(f"Contactos no encontrado: {CONTACTOS} — mapeo limitado")
        return {}
    import io, zipfile
    with open(CONTACTOS, "rb") as fh:
        raw = fh.read()
    if not zipfile.is_zipfile(io.BytesIO(raw)):
        log.error("Contactos_Clientes.xlsx corrupto o inválido")
        return {}
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        cliente, nombre, cargo, _, email = row[0], row[1], row[2], row[3], row[4]
        if email:
            mapping[email.strip().lower()] = {
                "cliente":   str(cliente).strip() if cliente else "",
                "contacto":  str(nombre).strip()  if nombre  else "",
                "cargo":     str(cargo).strip()   if cargo   else "",
                "email":     email.strip(),
            }
    wb.close()
    log.info(f"Contactos cargados: {len(mapping)} emails")
    return mapping

def cargar_tokens(csv_path: Path, contactos: dict) -> dict:
    if not csv_path.exists():
        log.warning(f"Tokens CSV no encontrado: {csv_path}")
        return {}
    mapping = {}
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            token = row.get("token", "").strip()
            email = row.get("email", "").strip().lower()
            if not token:
                continue
            info = contactos.get(email, {
                "cliente":  email or token[:8],
                "contacto": "",
                "cargo":    "",
                "email":    email,
            })
            mapping[token] = info
    log.info(f"Tokens cargados desde {csv_path.name}: {len(mapping)} tokens")
    return mapping


# ── CLASIFICACIÓN Y PROCESAMIENTO ──────────────────────────────────────────────
def clasificar_csat(nota) -> str:
    try: n = float(nota)
    except: return ""
    if n >= 4: return "Promotor"
    if n >= 3: return "Pasivo"
    return "Detractor"

def clasificar_nps(nota) -> str:
    try: n = float(nota)
    except: return ""
    if n >= 9: return "Promotor"
    if n >= 7: return "Pasivo"
    return "Detractor"

def _fecha(fecha_str: str) -> tuple[str, int]:
    try:
        dt = datetime.fromisoformat(fecha_str)
        return MESES_ES[dt.month], dt.year
    except Exception:
        return "Desconocido", datetime.now().year

def _nota_int(valor) -> int | None:
    try: return int(float(valor))
    except: return None

def procesar_csat(responses: list, token_map: dict) -> tuple[list, list, list]:
    f_clientes, f_areas, f_clientes_mes = [], [],[]
    for resp in responses:
        token = resp.get("token", "")
        info  = token_map.get(token)
        if not info: continue
        cliente = info["cliente"]
        mes, anio = _fecha(resp.get("submitdate", ""))
        for campo, (num, desc) in PREGUNTAS_CSAT.items():
            nota = _nota_int(resp.get(campo, ""))
            if nota is not None:
                f_clientes.append({"Cliente": cliente, "N° Pregunta": num, "Descripción pregunta": desc, "Nota": nota, "Mes": mes, "Año": anio, "Tipo": "CSAT"})
        for campo, area in AREAS_CSAT.items():
            nota = _nota_int(resp.get(campo, ""))
            if nota is not None:
                f_areas.append({"Cliente": cliente, "Área": area, "Nota": nota, "Mes": mes, "Año": anio})
        nota_gen = _nota_int(resp.get("G01Q01", ""))
        comentarios = " | ".join(filter(None,[resp.get("G01Q04", "").strip(), resp.get("G01Q13", "").strip()]))
        f_clientes_mes.append({"Cliente": cliente, "Nota": nota_gen, "Comentarios": comentarios, "Mes": mes, "Año": anio, "Clasificacion": clasificar_csat(resp.get("G01Q01", "")), "Tipo": "CSAT"})
    return f_clientes, f_areas, f_clientes_mes

def procesar_nps(responses: list, token_map: dict) -> tuple[list, list]:
    f_clientes, f_clientes_mes = [],[]
    for resp in responses:
        token = resp.get("token", "")
        info  = token_map.get(token)
        if not info: continue
        cliente = info["cliente"]
        mes, anio = _fecha(resp.get("submitdate", ""))
        nota_raw = (resp.get("Q00") or resp.get("G01Q01") or resp.get("NPS1") or "")
        nota = _nota_int(nota_raw)
        if nota is not None:
            f_clientes.append({"Cliente": cliente, "N° Pregunta": 0, "Descripción pregunta": "¿Qué tan probable es que recomiende Egakat? (0-10)", "Nota": nota, "Mes": mes, "Año": anio, "Tipo": "NPS"})
        f_clientes_mes.append({"Cliente": cliente, "Nota": nota, "Comentarios": resp.get("G01Q03", "").strip(), "Mes": mes, "Año": anio, "Clasificacion": clasificar_nps(nota_raw), "Tipo": "NPS"})
    return f_clientes, f_clientes_mes


# ── EXCEL Y GRAPH API ──────────────────────────────────────────────────────────
def _celda_header(ws, row, col, val, width=18):
    c = ws.cell(row=row, column=col, value=val)
    c.font, c.fill, c.alignment = Font(name="Calibri", bold=True, color="FFFFFF", size=10), PatternFill("solid", fgColor="1F497D"), Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _celda_dato(ws, row, col, val, bg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=val)
    c.font, c.fill, c.alignment = Font(name="Calibri", size=10, color="404040"), PatternFill("solid", fgColor=bg), Alignment(vertical="center", wrap_text=True)
    return c

def _escribir_hoja(wb, nombre: str, headers: list, filas: list, widths: dict = None):
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines, ws.freeze_panes, ws.row_dimensions[1].height = False, "A2", 20
    widths = widths or {}
    for i, h in enumerate(headers, 1): _celda_header(ws, 1, i, h, width=widths.get(h, 18))
    for r_idx, fila in enumerate(filas, 2):
        bg = "F2F2F2" if r_idx % 2 == 0 else "FFFFFF"
        for c_idx, h in enumerate(headers, 1): _celda_dato(ws, r_idx, c_idx, fila.get(h, ""), bg=bg)
    if filas: ws.auto_filter.ref = ws.dimensions

def generar_excel(f_clientes, f_areas, f_mes_total, contactos, ruta):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _escribir_hoja(wb, "fClientes", ["Cliente", "N° Pregunta", "Descripción pregunta", "Nota", "Mes", "Año", "Tipo"], f_clientes, {"Cliente": 32, "Descripción pregunta": 42})
    _escribir_hoja(wb, "fÁreas",["Cliente", "Área", "Nota", "Mes", "Año"], f_areas, {"Cliente": 32, "Área": 30})
    _escribir_hoja(wb, "fClientes_mes",["Cliente", "Nota", "Comentarios", "Mes", "Año", "Clasificacion", "Tipo"], f_mes_total, {"Cliente": 32, "Comentarios": 65})
    
    c_vistos = {}
    for info in contactos.values():
        if info.get("cliente") and info["cliente"] not in c_vistos: c_vistos[info["cliente"]] = info
    d_clientes = sorted([{"Cliente": v["cliente"], "Sistema": "WMS", "Contacto": v["contacto"], "Email": v["email"], "Cargo": v["cargo"]} for v in c_vistos.values()], key=lambda x: x["Cliente"])
    _escribir_hoja(wb, "dClientes", ["Cliente", "Sistema", "Contacto", "Email", "Cargo"], d_clientes, {"Cliente": 32, "Email": 36, "Cargo": 36})
    
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    log.info(f"Excel generado localmente: {ruta.name}")

def subir_a_sharepoint(ruta_local: Path, folder_sp: str, nombre_sp: str):
    log.info("Iniciando subida a SharePoint vía Graph API...")
    try:
        token = get_token()
        drive_id = get_drive_id(token)
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{folder_sp}/{nombre_sp}:/content"
        with open(ruta_local, "rb") as f:
            data = f.read()
        resp = requests.put(url, data=data, headers={"Authorization": f"Bearer {token}", "Content-Type": "application/octet-stream"}, timeout=120)
        if resp.status_code in (200, 201):
            log.info(f"[OK] Archivo subido exitosamente a SharePoint: {folder_sp}/{nombre_sp}")
        else:
            log.error(f"[FALLO] Error API {resp.status_code}: {resp.text}")
    except Exception as e:
        log.error(f"[FALLO] Excepción al subir por Graph API: {e}")

# ── ENTRY POINT ────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("NPS Descarga v3.0 (Graph API Directo) — inicio")
    log.info("=" * 60)

    if not LIMESURVEY_USER: 
        log.error("Falta usuario de LimeSurvey en .env")
        sys.exit(1)

    contactos  = cargar_contactos()
    token_csat = cargar_tokens(TOKENS_CSAT, contactos)
    token_nps  = cargar_tokens(TOKENS_NPS,  contactos)

    api = LimeSurveyAPI(LIMESURVEY_URL)
    try:
        api.open(LIMESURVEY_USER, LIMESURVEY_PASSWORD)
        log.info(f"── CSAT ({SURVEY_ID_CSAT}) ──")
        resp_csat = api.export_responses(SURVEY_ID_CSAT)
        f_cli_csat, f_areas, f_mes_csat = procesar_csat(resp_csat, token_csat) if resp_csat else ([], [],[])
        
        log.info(f"── NPS ({SURVEY_ID_NPS}) ──")
        resp_nps = api.export_responses(SURVEY_ID_NPS)
        f_cli_nps, f_mes_nps = procesar_nps(resp_nps, token_nps) if resp_nps else ([],[])
        
    finally:
        api.close()

    # Combinar datos
    f_clientes_total = f_cli_csat + f_cli_nps
    f_mes_total      = f_mes_csat + f_mes_nps

    log.info("\n── Generando Excel y Subiendo a SharePoint ──")
    # Generar el Excel local temporalmente
    generar_excel(f_clientes_total, f_areas, f_mes_total, contactos, ARCHIVO_LOCAL)
    
    # Subir el archivo generado directo a SharePoint vía Graph API
    if ARCHIVO_LOCAL.exists():
        subir_a_sharepoint(ARCHIVO_LOCAL, FOLDER_SP, NOMBRE_SP)

    log.info("=" * 60)
    log.info("NPS Descarga v3.0 — fin OK")
    log.info("=" * 60)

if __name__ == "__main__":
    main()