"""
nps_descarga.py v2.1
Descarga CSAT (386641) y NPS (418429) desde LimeSurvey RemoteControl 2 API.
Genera Excel PowerBI con 4 hojas: fClientes, fÁreas, fClientes_mes, dClientes.
Egakat SPA — Control de Gestión y Mejora Continua

Nombres estandarizados: se usa el nombre comercial del WMS, no la razón social.
Actualizar NOMBRES_WMS cuando cambien clientes o se confirmen pendientes.
"""

import sys
if sys.stdout:
    sys.stdout.reconfigure(encoding="utf-8")

import os
import csv
import json
import base64
import logging
import requests
import openpyxl
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIGURACIÓN ──────────────────────────────────────────────────────────────
load_dotenv(dotenv_path=Path(__file__).parent.parent / ".env")

LIMESURVEY_URL      = os.getenv("LIMESURVEY_URL",      "https://egakat.limesurvey.net/index.php/admin/remotecontrol")
LIMESURVEY_USER     = os.getenv("LIMESURVEY_USER",     "")
LIMESURVEY_PASSWORD = os.getenv("LIMESURVEY_PASSWORD", "")
SURVEY_ID_CSAT      = int(os.getenv("LIMESURVEY_SURVEY_ID_CSAT", "386641"))
SURVEY_ID_NPS       = int(os.getenv("LIMESURVEY_SURVEY_ID_NPS",  "418429"))

DIR_NPS      = Path(__file__).parent
TOKENS_CSAT  = DIR_NPS / "tokens_csat.csv"   # Actualizar cada ronda mensual
TOKENS_NPS   = DIR_NPS / "tokens_nps.csv"    # Crear cuando lance NPS 25/03
CONTACTOS    = DIR_NPS / "Contactos_Clientes.xlsx"

ONEDRIVE_BASE = Path(os.getenv("ONEDRIVE_PATH", "")) / "Reportes NPS"
ARCHIVO_PBI   = ONEDRIVE_BASE / "NPS_PBI_datos.xlsx"  # Nombre fijo — PowerBI apunta aquí

LOG_DIR  = Path(__file__).parent.parent / "logs"
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"nps_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger(__name__)


# ── ESTRUCTURA DE ENCUESTAS ────────────────────────────────────────────────────

# CSAT — preguntas numéricas individuales → fClientes
PREGUNTAS_CSAT = {
    "G01Q01": (1, "Satisfacción general del servicio"),
    "G01Q05": (5, "Cumplimiento de tiempos de entrega"),
    "G01Q06": (6, "Precisión de pedidos preparados"),
    "G01Q07": (7, "Información y seguimiento de pedidos"),
}

# CSAT — sub-preguntas de áreas → fÁreas
AREAS_CSAT = {
    "G01Q02[SQ001]": "Recepción",
    "G01Q02[SQ002]": "Preparación de pedidos",
    "G01Q02[SQ003]": "Despacho",
    "G01Q02[SQ004]": "Gestión de inventarios",
    "G01Q02[SQ005]": "Transporte",
    "G01Q02[SQ006]": "Calidad",
    "G01Q02[SQ007]": "Servicio al cliente",
}

# ── ESTANDARIZACIÓN DE NOMBRES (Razón social → Nombre WMS) ────────────────────
# Norma: el nombre oficial para Power BI es el nombre del WMS (nombre comercial).
# Fuente: Reportes de Ubicación de Contenedor Quilicura + Pudahuel.
# Actualizar cuando se confirmen clientes pendientes (NATIVO, nuevos).
#
# Clave   = nombre en Contactos_Clientes.xlsx (razón social o nombre SAC)
# Valor   = (nombre estándar Power BI, sistema)
# Sistema: "WMS" | "Odoo" | "SAP"
NOMBRES_WMS = {
    # ── WMS Egakat (Quilicura + Pudahuel) ────────────────────────────────────
    "Barentz Chile SpA":                           ("BARENTZ",              "WMS"),  # Cód 33
    "Buraschi Trading SpA":                        ("BURASCHI",             "WMS"),  # Cód 23
    "Cepas Chile S.A":                             ("CEPAS CHILE",          "WMS"),  # Cód 11
    "Daikin Airconditioning Chile S.A":            ("DAIKIN",               "WMS"),  # Cód 3
    "Confecciones Nazal Ltda":                     ("INTIME",               "WMS"),  # Cód 16
    "Comercializadora RyA S.A":                    ("MASCOTAS LATINAS",     "WMS"),  # Cód 13
    "Pochteca Chile S.A":                          ("POCHTECA",             "WMS"),  # Cód 4
    "Runo SpA":                                    ("RUNO SPA",             "WMS"),  # Cód 31
    "Saeg International Group":                    ("SAEG",                 "WMS"),  # Cód 28
    "Santa Rosa Chile":                            ("SANTA ROSA",           "WMS"),  # Cód 17
    "Unilever Chile Ltda":                         ("UNILEVER",             "WMS"),  # Cód 18
    "Grupo Planet":                                ("DERCO",                "WMS"),  # Cód 2
    "AB-InBev Cervecería Chile":                   ("CERVECERÍA CHILE",     "WMS"),  # sin stock hoy
    "Comercializadora Gecorp Ltda":                ("GECORP",               "WMS"),  # sin stock hoy
    "Composites Express SpA":                      ("COMPOSITES",           "WMS"),  # sin stock hoy
    "Importadora y Comercializadora BBC Chile S.A": ("BBCH",                "WMS"),  # sin stock hoy
    "Tresmontes Lucchetti":                        ("TRESMONTES LUCCHETTI", "WMS"),  # sin stock hoy

    # ── Egakat Express (Pudahuel) — Odoo ─────────────────────────────────────
    # NOTA: NATIVO DRINKS SPA (Cód 34 WMS) es empresa relacionada a San Joaquín SpA
    # (mismo dueño: Juan Pablo Barahona). NATIVO opera en WMS; San Joaquín en Odoo.
    # Para próxima ronda encuesta: agregar contacto separado para NATIVO DRINKS SPA.
    "Sociedad Comercial San Joaquin SpA":          ("SAN JOAQUIN SPA",      "Odoo"),
    # "Syntheon Chile SpA": EXCLUIDO — es proveedor de infraestructura IMO (carga peligrosa).
    # Egakat no tiene permisos/infraestructura propia → opera con 1 persona en instalaciones Syntheon.
    # La mercancía es de POCHTECA (cliente Egakat, Cód 4) — Pochteca es quien evalúa el servicio.
    # Fabiola Segovia (Syntheon) NO debe ser participante de la encuesta → retirar próxima ronda.
    "Comercial Dicalla SpA":                       ("MARLEY COFFEE",        "Odoo"),

    # ── Pudahuel — SAP del cliente (Egakat opera con licencia) ───────────────
    "Comercial Mabe Limitada":                     ("MABE",                 "SAP"),
}


def estandarizar_cliente(nombre: str) -> tuple[str, str]:
    """
    Convierte razón social a (nombre estándar Power BI, sistema).
    Si no hay mapeo devuelve (nombre original, 'WMS') como default.
    """
    resultado = NOMBRES_WMS.get(nombre)
    if isinstance(resultado, tuple):
        return resultado
    return (nombre, "WMS")


MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo",    4: "Abril",
    5: "Mayo",  6: "Junio",   7: "Julio",    8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


# ── API LIMESURVEY ─────────────────────────────────────────────────────────────

class LimeSurveyAPI:
    """Cliente LimeSurvey RemoteControl 2 (JSON-RPC)."""

    def __init__(self, url: str):
        self.url = url
        self.session_key = None
        self._id = 0

    def _call(self, method: str, params: list):
        self._id += 1
        r = requests.post(
            self.url,
            json={"method": method, "params": params, "id": self._id},
            headers={"content-type": "application/json"},
            timeout=60,
        )
        r.raise_for_status()
        data = r.json()
        if data.get("error"):
            raise RuntimeError(f"API error: {data['error']}")
        return data.get("result")

    def open(self, user: str, password: str):
        result = self._call("get_session_key", [user, password])
        if not result or len(result) < 10:
            raise RuntimeError(f"Login LimeSurvey fallido: {result}")
        self.session_key = result
        log.info("Sesión LimeSurvey abierta OK")

    def close(self):
        if self.session_key:
            self._call("release_session_key", [self.session_key])
            self.session_key = None

    def list_participants(self, survey_id: int, start: int = 0, limit: int = 1000) -> list[dict]:
        """Lista participantes del survey. Retorna [] si no hay."""
        result = self._call("list_participants", [
            self.session_key, survey_id, start, limit, False,
            ["tid", "token", "email", "firstname", "lastname", "emailstatus"],
        ])
        if not result or isinstance(result, dict) and "status" in result:
            return []
        return result

    def delete_participants(self, survey_id: int, tids: list[str]) -> dict:
        """Elimina participantes por tid. Retorna resultado API."""
        return self._call("delete_participants", [self.session_key, survey_id, tids])

    def export_responses(self, survey_id: int) -> list[dict]:
        """Exporta respuestas completas. Retorna [] si no hay datos."""
        result = self._call("export_responses", [
            self.session_key,
            survey_id,
            "json",       # formato
            "es",         # idioma
            "complete",   # solo completas
            "code",       # encabezados: código de pregunta (G01Q01, etc.)
            "short",      # valores cortos (número, no etiqueta)
        ])
        if not result:
            return []
        if isinstance(result, dict) and "No Data" in str(result.get("status", "")):
            return []
        decoded = base64.b64decode(result).decode("utf-8")
        return json.loads(decoded).get("responses", [])


# ── MAPEO TOKEN → CLIENTE ──────────────────────────────────────────────────────

def cargar_contactos() -> dict:
    """
    Lee Contactos_Clientes.xlsx → dict {email_lower: {cliente, contacto, cargo, email}}.
    La columna 'Cliente' es el nombre oficial usado en Power BI.
    """
    if not CONTACTOS.exists():
        log.warning(f"Contactos no encontrado: {CONTACTOS} — mapeo limitado")
        return {}

    import io, zipfile
    with open(CONTACTOS, "rb") as fh:
        raw = fh.read()
    if not zipfile.is_zipfile(io.BytesIO(raw)):
        log.error("Contactos_Clientes.xlsx corrupto o inválido")
        return {}

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    ws = wb.active
    mapping = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        cliente, nombre, cargo, _, email = row[0], row[1], row[2], row[3], row[4]
        if email:
            mapping[email.strip().lower()] = {
                "cliente":   str(cliente).strip() if cliente else "",
                "contacto":  str(nombre).strip()  if nombre  else "",
                "cargo":     str(cargo).strip()   if cargo   else "",
                "email":     email.strip(),
            }

    wb.close()
    log.info(f"Contactos cargados: {len(mapping)} emails")
    return mapping


def cargar_tokens(csv_path: Path, contactos: dict) -> dict:
    """
    Lee tokens CSV exportado desde LimeSurvey → dict {token: info_cliente}.
    Requiere columnas: token, email.
    Actualizar este archivo al inicio de cada ronda de encuesta.
    """
    if not csv_path.exists():
        log.warning(f"Tokens CSV no encontrado: {csv_path}")
        return {}

    mapping = {}
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            token = row.get("token", "").strip()
            email = row.get("email", "").strip().lower()
            if not token:
                continue
            info = contactos.get(email, {
                "cliente":  email or token[:8],
                "contacto": "",
                "cargo":    "",
                "email":    email,
            })
            # Aplicar nombre estándar y sistema
            info = dict(info)
            nombre_std, sistema = estandarizar_cliente(info["cliente"])
            info["cliente"] = nombre_std
            info["sistema"] = sistema
            mapping[token] = info

    log.info(f"Tokens cargados desde {csv_path.name}: {len(mapping)} tokens")
    return mapping


# ── CLASIFICACIÓN ──────────────────────────────────────────────────────────────

def clasificar_csat(nota) -> str:
    """CSAT escala 1-5: 4-5=Promotor, 3=Pasivo, 1-2=Detractor."""
    try:
        n = float(nota)
    except (TypeError, ValueError):
        return ""
    if n >= 4:
        return "Promotor"
    if n >= 3:
        return "Pasivo"
    return "Detractor"


def clasificar_nps(nota) -> str:
    """NPS escala 0-10: 9-10=Promotor, 7-8=Pasivo, 0-6=Detractor."""
    try:
        n = float(nota)
    except (TypeError, ValueError):
        return ""
    if n >= 9:
        return "Promotor"
    if n >= 7:
        return "Pasivo"
    return "Detractor"


def _fecha(fecha_str: str) -> tuple[str, int]:
    """Parsea fecha LimeSurvey → (mes_español, año)."""
    try:
        dt = datetime.fromisoformat(fecha_str)
        return MESES_ES[dt.month], dt.year
    except Exception:
        return "Desconocido", datetime.now().year


def _nota_int(valor) -> int | None:
    """Convierte valor de respuesta a entero, None si vacío."""
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


# ── PROCESAMIENTO CSAT ─────────────────────────────────────────────────────────

def procesar_csat(responses: list, token_map: dict) -> tuple[list, list, list]:
    """
    Procesa respuestas CSAT → (fClientes, fÁreas, fClientes_mes).
    fClientes:    una fila por pregunta numérica por respondente.
    fÁreas:       una fila por área evaluada por respondente.
    fClientes_mes: una fila por respondente (nota general + comentarios).
    """
    f_clientes    = []
    f_areas       = []
    f_clientes_mes = []
    sin_mapeo = 0

    for resp in responses:
        token = resp.get("token", "")
        info  = token_map.get(token)
        if not info:
            sin_mapeo += 1
            log.debug(f"  Token sin mapeo: {token[:12]}...")
            continue

        cliente       = info["cliente"]
        mes, anio     = _fecha(resp.get("submitdate", ""))

        # fClientes — preguntas numéricas individuales
        for campo, (num, desc) in PREGUNTAS_CSAT.items():
            nota = _nota_int(resp.get(campo, ""))
            if nota is None:
                continue
            f_clientes.append({
                "Cliente":             cliente,
                "N° Pregunta":         num,
                "Descripción pregunta": desc,
                "Nota":                nota,
                "Mes":                 mes,
                "Año":                 anio,
                "Tipo":                "CSAT",
            })

        # fÁreas — sub-preguntas G01Q02[SQ001-SQ007]
        for campo, area in AREAS_CSAT.items():
            nota = _nota_int(resp.get(campo, ""))
            if nota is None:
                continue
            f_areas.append({
                "Cliente": cliente,
                "Área":    area,
                "Nota":    nota,
                "Mes":     mes,
                "Año":     anio,
            })

        # fClientes_mes — una fila por cliente con nota general y comentarios
        nota_gen = _nota_int(resp.get("G01Q01", ""))
        comentarios = " | ".join(filter(None, [
            resp.get("G01Q04", "").strip(),   # ¿Qué podríamos mejorar?
            resp.get("G01Q13", "").strip(),   # ¿Cómo mejorar próxima experiencia?
        ]))
        f_clientes_mes.append({
            "Cliente":      cliente,
            "Nota":         nota_gen,
            "Comentarios":  comentarios,
            "Mes":          mes,
            "Año":          anio,
            "Clasificacion": clasificar_csat(resp.get("G01Q01", "")),
            "Tipo":         "CSAT",
        })

    if sin_mapeo:
        log.warning(f"  {sin_mapeo} respuestas CSAT sin mapeo de cliente — revisar tokens_csat.csv")
    return f_clientes, f_areas, f_clientes_mes


# ── PROCESAMIENTO NPS ──────────────────────────────────────────────────────────

def procesar_nps(responses: list, token_map: dict) -> tuple[list, list]:
    """
    Procesa respuestas NPS → (fClientes, fClientes_mes).
    Una fila por respondente.
    """
    f_clientes    = []
    f_clientes_mes = []
    sin_mapeo = 0

    for resp in responses:
        token = resp.get("token", "")
        info  = token_map.get(token)
        if not info:
            sin_mapeo += 1
            continue

        cliente   = info["cliente"]
        mes, anio = _fecha(resp.get("submitdate", ""))

        # NPS score — buscar campo estándar o fallbacks
        nota_raw = (resp.get("Q00") or resp.get("G01Q01") or
                    resp.get("NPS1") or "")
        nota = _nota_int(nota_raw)

        if nota is not None:
            f_clientes.append({
                "Cliente":              cliente,
                "N° Pregunta":          0,
                "Descripción pregunta": "¿Qué tan probable es que recomiende Egakat? (0-10)",
                "Nota":                 nota,
                "Mes":                  mes,
                "Año":                  anio,
                "Tipo":                 "NPS",
            })

        f_clientes_mes.append({
            "Cliente":       cliente,
            "Nota":          nota,
            "Comentarios":   resp.get("G01Q03", "").strip(),
            "Mes":           mes,
            "Año":           anio,
            "Clasificacion": clasificar_nps(nota_raw),
            "Tipo":          "NPS",
        })

    if sin_mapeo:
        log.warning(f"  {sin_mapeo} respuestas NPS sin mapeo — revisar tokens_nps.csv")
    return f_clientes, f_clientes_mes


# ── GENERADOR EXCEL POWER BI ───────────────────────────────────────────────────

# Paleta
AZUL    = "1F497D"
BLANCO  = "FFFFFF"
GRIS    = "F2F2F2"


def _celda_header(ws, row, col, val, width=18):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", bold=True, color=BLANCO, size=10)
    c.fill      = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _celda_dato(ws, row, col, val, bg=BLANCO):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=10, color="404040")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    return c


def _escribir_hoja(wb, nombre: str, headers: list, filas: list, widths: dict = None):
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    widths = widths or {}
    for i, h in enumerate(headers, 1):
        _celda_header(ws, 1, i, h, width=widths.get(h, 18))

    for r_idx, fila in enumerate(filas, 2):
        bg = GRIS if r_idx % 2 == 0 else BLANCO
        for c_idx, h in enumerate(headers, 1):
            _celda_dato(ws, r_idx, c_idx, fila.get(h, ""), bg=bg)

    if filas:
        ws.auto_filter.ref = ws.dimensions

    log.info(f"  Hoja '{nombre}': {len(filas)} filas")
    return ws


def generar_excel_pbi(f_clientes: list, f_areas: list, f_clientes_mes: list,
                       contactos: dict, ruta: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # fClientes — una fila por pregunta numérica por respondente
    _escribir_hoja(wb, "fClientes",
        headers=["Cliente", "N° Pregunta", "Descripción pregunta", "Nota", "Mes", "Año", "Tipo"],
        filas=f_clientes,
        widths={
            "Cliente": 32, "N° Pregunta": 12, "Descripción pregunta": 42,
            "Nota": 8, "Mes": 12, "Año": 8, "Tipo": 8,
        })

    # fÁreas — una fila por área por respondente
    _escribir_hoja(wb, "fÁreas",
        headers=["Cliente", "Área", "Nota", "Mes", "Año"],
        filas=f_areas,
        widths={"Cliente": 32, "Área": 30, "Nota": 8, "Mes": 12, "Año": 8})

    # fClientes_mes — una fila por respondente con nota general
    _escribir_hoja(wb, "fClientes_mes",
        headers=["Cliente", "Nota", "Comentarios", "Mes", "Año", "Clasificacion", "Tipo"],
        filas=f_clientes_mes,
        widths={
            "Cliente": 32, "Nota": 8, "Comentarios": 65,
            "Mes": 12, "Año": 8, "Clasificacion": 15, "Tipo": 8,
        })

    # dClientes — dimensión: un registro único por cliente
    clientes_vistos = {}
    for info in contactos.values():
        c = info.get("cliente", "")
        if c and c not in clientes_vistos:
            clientes_vistos[c] = info

    d_clientes = sorted(
        [{"Cliente":  v["cliente"],
          "Sistema":  v.get("sistema", "WMS"),
          "Contacto": v["contacto"],
          "Email":    v["email"],
          "Cargo":    v["cargo"]}
         for v in clientes_vistos.values()],
        key=lambda x: (x["Sistema"], x["Cliente"])
    )
    _escribir_hoja(wb, "dClientes",
        headers=["Cliente", "Sistema", "Contacto", "Email", "Cargo"],
        filas=d_clientes,
        widths={"Cliente": 32, "Sistema": 12, "Contacto": 28, "Email": 36, "Cargo": 36})

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    log.info(f"Excel guardado: {ruta}")


# ── ENTRY POINT ────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("NPS Descarga v2.0 — inicio")
    log.info(f"CSAT ID: {SURVEY_ID_CSAT}  |  NPS ID: {SURVEY_ID_NPS}")
    log.info(f"Salida PBI: {ARCHIVO_PBI}")
    log.info("=" * 60)

    if not LIMESURVEY_USER or not LIMESURVEY_PASSWORD:
        log.error("LIMESURVEY_USER y/o LIMESURVEY_PASSWORD no están en .env")
        sys.exit(1)

    # Cargar mapeo token → cliente
    contactos  = cargar_contactos()
    token_csat = cargar_tokens(TOKENS_CSAT, contactos)
    token_nps  = cargar_tokens(TOKENS_NPS,  contactos)

    api = LimeSurveyAPI(LIMESURVEY_URL)
    try:
        api.open(LIMESURVEY_USER, LIMESURVEY_PASSWORD)

        # ── CSAT ──────────────────────────────────────────────
        log.info(f"\n── CSAT ({SURVEY_ID_CSAT}) ──────────────────────────────")
        resp_csat = api.export_responses(SURVEY_ID_CSAT)
        log.info(f"  Respuestas CSAT completas: {len(resp_csat)}")
        if resp_csat:
            f_cli_csat, f_areas, f_mes_csat = procesar_csat(resp_csat, token_csat)
        else:
            f_cli_csat, f_areas, f_mes_csat = [], [], []
            log.warning("  CSAT sin respuestas aún")

        # ── NPS ───────────────────────────────────────────────
        log.info(f"\n── NPS ({SURVEY_ID_NPS}) ────────────────────────────────")
        resp_nps = api.export_responses(SURVEY_ID_NPS)
        log.info(f"  Respuestas NPS completas: {len(resp_nps)}")
        if resp_nps:
            f_cli_nps, f_mes_nps = procesar_nps(resp_nps, token_nps)
        else:
            f_cli_nps, f_mes_nps = [], []
            log.info("  NPS sin respuestas aún — normal (lanza 25/03)")

    finally:
        api.close()

    # Combinar y generar Excel
    f_clientes_total = f_cli_csat + f_cli_nps
    f_mes_total      = f_mes_csat + f_mes_nps

    log.info(f"\n── Generando Excel PowerBI ─────────────────────────────")
    log.info(f"  fClientes:     {len(f_clientes_total)} filas")
    log.info(f"  fÁreas:        {len(f_areas)} filas")
    log.info(f"  fClientes_mes: {len(f_mes_total)} filas")

    generar_excel_pbi(f_clientes_total, f_areas, f_mes_total, contactos, ARCHIVO_PBI)

    log.info("\n" + "=" * 60)
    log.info("NPS Descarga v2.0 — fin OK")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
