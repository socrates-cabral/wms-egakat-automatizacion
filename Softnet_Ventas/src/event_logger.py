import sys
sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from openpyxl import load_workbook, Workbook

LOG_HEADERS = [
    "fecha_deteccion", "mes_archivo", "tipo_doc", "n_cto", "rut",
    "razon_social", "tipo_cambio", "estado_anterior_actual", "fecha_pago", "monto_total",
]


def append_eventos(log_path: Path, eventos: list[dict]) -> None:
    """Agrega eventos al log_cambios_pagos.xlsx. Crea con headers si no existe."""
    if not eventos:
        return
    if log_path.exists():
        wb = load_workbook(log_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "cambios"
        ws.append(LOG_HEADERS)
    for ev in eventos:
        ws.append([ev.get(h, "") for h in LOG_HEADERS])
    wb.save(log_path)
